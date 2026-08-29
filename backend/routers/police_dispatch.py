"""小区地址库、全链条批次预处理、审核、反馈导出与腾讯发布。"""

from __future__ import annotations

import asyncio
import hmac
import io
import json
import os
import tempfile
from datetime import date, datetime, timedelta
from hashlib import sha256
from pathlib import Path
from typing import Any, Callable, Literal
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field, field_validator

from database import db_manager, get_db
from deps import require_permission, require_super_admin
from routers.query import (
    _enabled_spreadsheets,
    _load_source_row,
    _oauth_client,
    _refresh_spreadsheet,
    _row_values_match,
    _writeback_enabled,
)
from services.audit import record_admin_audit, request_audit_fields
from services.business_time import get_business_date
from services.online_source import (
    acquire_sheet_lock,
    json_value,
    rebuild_projection,
    release_sheet_lock,
    resolve_source_columns,
)
from services.online_source import source_row_hash
from services.local_source import (
    create_local_source_row,
    local_data_source_enabled,
    local_sheet_id,
)
from services.parsers import get_parser
from services.permissions import (
    POLICE_ADDRESS_MANAGE,
    POLICE_DISPATCH_MANAGE,
    permitted_communities,
)
from services.police_dispatch import (
    FINAL_ACTIONS,
    MAX_POLICE_FILE_BYTES,
    PUBLISH_OWNED_COLUMNS,
    PoliceWorkbookError,
    apply_clean_import_actions,
    apply_preprocessing_suggestions,
    build_feedback_workbook,
    build_publish_address,
    normalize_lookup,
    normalize_space,
    parse_dispatch_workbook,
    parser_business_key,
    parser_business_key_fields,
    publish_business_key,
    resolve_community,
    community_resolver,
    dispatch_field_roles,
    dispatch_values_from_raw,
    identity_digest,
    IDENTITY_PATTERN,
    normalize_identity,
    stable_json,
)
from services.police_dispatch_publish_jobs import (
    get_latest_police_publish_run,
    get_police_publish_run,
    launch_police_publish_run,
)
from services.police_import_profiles import (
    ADAPTER_VERSION,
    PROFILES,
    parse_profile,
    preview_token as import_preview_token,
    profile_payload,
    verify_preview_token as verify_import_preview_token,
)
from services.txdocs_client import TxDocsAPIError
from services.work_activity import (
    POLICE_DISPATCH_REVIEW,
    record_work_activity,
)
from config import settings


router = APIRouter(prefix="/api/police-dispatch", tags=["全链条下发"])


class AddressCreate(BaseModel):
    name: str = Field(min_length=1, max_length=300)
    detail_address: str = Field(default="", max_length=1000)
    address_type: Literal["community", "apartment", "construction_dormitory", "other"] = "community"
    pattern: str = Field(default="", max_length=200)
    community_id: int
    aliases: list[str] = Field(default_factory=list, max_length=50)
    enabled: bool = True


class AddressSearch(BaseModel):
    keyword: str = Field(default="", max_length=100)
    enabled: bool | None = None


class TaskReview(BaseModel):
    expected_version: int = Field(gt=0)
    final_action: Literal[
        "dispatch", "no_registration", "transfer", "duplicate_exclude"
    ]
    final_community_id: int | None = None
    review_note: str = Field(default="", max_length=1000)


class BulkReview(BaseModel):
    tasks: list[dict[str, int]] = Field(min_length=1, max_length=2000)
    mode: Literal["accept_suggestion", "set_action"] = "accept_suggestion"
    final_action: Literal[
        "dispatch", "no_registration", "transfer", "duplicate_exclude"
    ] | None = None
    final_community_id: int | None = None
    review_note: str = Field(default="", max_length=1000)


class TaskSearch(BaseModel):
    batch_id: int = Field(gt=0)
    status: str = Field(default="all", max_length=30)
    category: str = Field(default="all", max_length=30)
    keyword: str = Field(default="", max_length=100)
    page: int = Field(default=1, ge=1)
    page_size: int = Field(default=20, ge=1, le=500)


class TaskPublishSelection(BaseModel):
    task_ids: list[int] = Field(min_length=1, max_length=5000)

    @field_validator("task_ids")
    @classmethod
    def validate_task_ids(cls, value: list[int]) -> list[int]:
        if any(task_id <= 0 for task_id in value):
            raise ValueError("task_ids must contain positive integers")
        return list(dict.fromkeys(value))


class DuplicateGroupResolution(BaseModel):
    tasks: list[dict[str, int]] = Field(min_length=2, max_length=100)
    review_note: str = Field(default="", max_length=1000)


class TaskBusinessFieldsUpdate(BaseModel):
    expected_version: int = Field(gt=0)
    fields: dict[str, str] = Field(default_factory=dict, max_length=200)


class ConflictResolution(BaseModel):
    expected_version: int = Field(gt=0)
    strategy: Literal["adopt_tencent", "overwrite_tencent"]
    expected_row_hash: str = Field(min_length=64, max_length=64)
    confirmation: str = Field(default="", max_length=50)


class QuickDispatchCreate(BaseModel):
    """单条临时任务；按已确认的业务表适配器创建，不直接写腾讯表。"""

    request_id: str = Field(
        min_length=16,
        max_length=100,
        pattern=r"^[A-Za-z0-9][A-Za-z0-9._:-]*$",
    )
    profile: str = Field(default="fullchain_processed", max_length=80)
    fields: dict[str, str] = Field(default_factory=dict, max_length=80)
    source_name: str = Field(default="", max_length=300)
    community_id: int | None = Field(default=None, gt=0)
    person_name: str = Field(default="", max_length=200)
    identity_number: str = Field(default="", max_length=50)
    phone: str = Field(default="", max_length=200)
    original_address: str = Field(default="", max_length=1500)
    registration_status: str = Field(default="", max_length=100)
    business_date: date
    deadline_date: date | None = None
    created_time: datetime | None = None

    @field_validator(
        "source_name", "person_name", "identity_number", "phone",
        "original_address", "registration_status",
    )
    @classmethod
    def strip_text(cls, value: str) -> str:
        return str(value).strip()


def _quick_dispatch_digest(user_id: int, request_id: str) -> str:
    return sha256(f"quick:{user_id}:{request_id}".encode("utf-8")).hexdigest()


async def _existing_quick_dispatch(cur, batch_digest: str) -> dict[str, Any] | None:
    await cur.execute("""
        SELECT b.id, t.id
        FROM _police_dispatch_batches b
        LEFT JOIN _police_dispatch_tasks t ON t.batch_id=b.id
        WHERE b.import_profile='quick_dispatch' AND b.file_sha256=%s
        ORDER BY t.id
        LIMIT 1
    """, (batch_digest,))
    row = await cur.fetchone()
    if not row:
        return None
    if row[1] is None:
        raise HTTPException(409, "快捷下发请求正在处理中，请稍后重试")
    return {
        "status": "duplicate",
        "message": "该快捷下发请求已处理，已返回原任务",
        "batch": await _batch_payload(cur, int(row[0])),
        "task_id": int(row[1]),
    }


def _clean_preview_token(file_sha256: str, filename: str, sheet_name: str, row_count: int) -> str:
    payload = f"clean:{file_sha256}:{filename}:{sheet_name}:{row_count}"
    return hmac.new(
        settings.registry_hmac_key.encode("utf-8"), payload.encode("utf-8"), sha256
    ).hexdigest()


def _verify_clean_preview_token(token: str, file_sha256: str, filename: str, sheet_name: str, row_count: int) -> bool:
    return hmac.compare_digest(
        token,
        _clean_preview_token(file_sha256, filename, sheet_name, row_count),
    )


def _mask_preview_phone(value: Any) -> str:
    text = str(value or "").strip()
    if len(text) >= 7:
        return f"{text[:3]}{'*' * (len(text) - 7)}{text[-4:]}"
    return "*" * len(text)


def _mask_preview_identity(value: Any) -> str:
    text = str(value or "").strip()
    if len(text) >= 10:
        return f"{text[:6]}{'*' * (len(text) - 10)}{text[-4:]}"
    return "*" * len(text)


def _clean_preview_summary(tasks: list[dict[str, Any]]) -> dict[str, Any]:
    counts = {
        "total": len(tasks), "dispatch": 0,
        "manual_review": 0, "invalid": 0, "duplicate": 0,
    }
    distribution: dict[int, dict[str, Any]] = {}
    rows: list[dict[str, Any]] = []
    for item in tasks:
        action = str(item.get("auto_final_action") or "")
        if action == "dispatch":
            counts["dispatch"] += 1
            community_id = item.get("auto_final_community_id")
            if community_id:
                current = distribution.setdefault(int(community_id), {
                    "community_id": int(community_id),
                    "community_name": str(item.get("community_name") or ""),
                    "count": 0,
                })
                current["count"] += 1
        else:
            counts["manual_review"] += 1
        if item.get("duplicate_group_key"):
            counts["duplicate"] += 1
        if str(item.get("allocation_mode") or "") == "conflict":
            counts["invalid"] += 1
        if len(rows) < 100:
            rows.append({
                "source_row": int(item.get("source_row") or 0),
                "person_name": str(item.get("person_name") or ""),
                "identity_number": _mask_preview_identity(item.get("identity_number")),
                "phone": _mask_preview_phone(item.get("phone")),
                "community_name": str(item.get("community_name") or ""),
                "registration_status": str(item.get("registration_status") or ""),
                "result": action or "manual",
                "reason": str(item.get("suggestion_reason") or ""),
            })
    return {
        "counts": counts,
        "community_distribution": list(distribution.values()),
        "rows": rows,
        "rows_truncated": len(tasks) > len(rows),
    }


ALLOWED_POLICE_POSITIONS = {"基础管控", "中队长", "所队领导"}
FULLCHAIN_ARCHIVE_POSITIONS = ALLOWED_POLICE_POSITIONS


def _permission_group_codes(user: dict) -> set[str]:
    codes = {
        str(group.get("code") or "")
        for group in user.get("permission_groups") or []
        if group.get("code")
    }
    primary = (user.get("permission_group") or {}).get("code")
    if primary:
        codes.add(str(primary))
    return codes


def require_police_access(permission: str) -> Callable:
    base_dependency = require_permission(permission)

    async def dependency(user: dict = Depends(base_dependency)) -> dict:
        permission_scope = (user.get("permission_scopes") or {}).get(
            permission, user.get("data_scope"),
        )
        if permission_scope != "all":
            raise HTTPException(403, "数据预处理必须使用全所数据范围，请联系超级管理员修正权限组")
        group_codes = _permission_group_codes(user)
        member = user.get("member")
        if member:
            if str(member.get("position") or "") in ALLOWED_POLICE_POSITIONS:
                return user
            raise HTTPException(403, "当前人员岗位不能进入数据预处理工作台")
        if group_codes.intersection({"admin", "super_admin"}) or (
            not group_codes and user.get("role") in {"admin", "super_admin"}
        ):
            return user
        raise HTTPException(403, "数据预处理仅向基础管控、中队长、所队领导和系统管理员开放")

    dependency.__name__ = f"require_police_{permission.replace('.', '_')}"
    return dependency


require_police_dispatch = require_police_access(POLICE_DISPATCH_MANAGE)


def _dispatch_import_dir() -> Path:
    root = Path(settings.POLICE_DISPATCH_IMPORT_DIR).resolve()
    root.mkdir(parents=True, exist_ok=True)
    return root


def _fullchain_standard_values(item: dict[str, Any]) -> dict[str, str]:
    return {
        "下发日期": str(item.get("created_time") or ""), "截止日期": "", "核查人": "",
        "社区": str(item.get("community_name") or ""), "来源": str(item.get("source_name") or ""),
        "姓名": str(item.get("person_name") or ""), "身份证号": str(item.get("identity_number") or ""),
        "电话号码": str(item.get("phone") or ""), "地址": str(item.get("original_address") or ""),
        "登记情况": str(item.get("registration_status") or ""),
        "创建时间": str(item.get("created_time") or ""), "现住址": "", "核查结果": "",
        "研判": "", "二次反馈": "",
    }


async def _parse_import_upload(
    *, profile_key: str, content: bytes, filename: str, business_date: date,
    communities: list[dict[str, Any]], addresses: list[dict[str, Any]],
) -> dict[str, Any]:
    profile = PROFILES.get(profile_key)
    if not profile:
        raise PoliceWorkbookError("未知导入类型")
    if not profile.enabled:
        raise PoliceWorkbookError(profile.description)
    if profile_key in {"fullchain_raw", "fullchain_processed"}:
        sheet_name, parsed = await asyncio.to_thread(
            parse_dispatch_workbook, content, filename,
            profile_key == "fullchain_processed",
        )
        tasks = [{
            "source_row": item.source_row, "source_name": item.source_name,
            "community_name": item.community_name, "person_name": item.person_name,
            "identity_number": item.identity_number, "phone": item.phone,
            "original_address": item.original_address,
            "registration_status": item.registration_status,
            "created_time": item.created_time, "transfer_note": item.transfer_note,
            "raw_values": item.raw_values,
        } for item in parsed]
        if profile_key == "fullchain_processed":
            apply_clean_import_actions(tasks, communities)
        else:
            apply_preprocessing_suggestions(tasks, communities, addresses)
        for item in tasks:
            item["standard_values"] = _fullchain_standard_values(item)
            key_payload = "\x1f".join((item["identity_number"], item["phone"], item["created_time"] or business_date.isoformat()))
            item["business_key_hmac"] = hmac.new(
                settings.registry_hmac_key.encode(), key_payload.encode(), sha256,
            ).hexdigest()
            item["validation_issues"] = [] if item.get("suggested_action") != "manual" else [{
                "field": "记录", "type": "conflict", "value": str(item.get("suggestion_reason") or "需要人工复核")[:200],
            }]
        summary = _clean_preview_summary(tasks)
        return {
            "profile": profile, "adapter_version": ADAPTER_VERSION,
            "sheet_name": sheet_name, "rows": tasks,
            "counts": {
                "total": len(tasks),
                "importable": sum(item.get("suggested_action") != "manual" for item in tasks),
                "missing_key": 0,
                "duplicate": summary["counts"]["duplicate"],
                "identity_invalid": summary["counts"]["invalid"],
                "community_invalid": sum("社区无法匹配" in str(item.get("suggestion_reason")) for item in tasks),
                "conflict": summary["counts"]["manual_review"],
            },
            "community_distribution": summary["community_distribution"],
            "preview_rows": [{
                "source_row": row["source_row"], "person_name": row["person_name"],
                "identity_number": row["identity_number"], "phone": row["phone"],
                "community_name": row["community_name"], "business_key": "",
                "result": "importable" if row["result"] != "manual" else "problem",
                "issues": [] if row["result"] != "manual" else [{"field": "记录", "type": "conflict", "value": row["reason"][:200]}],
            } for row in summary["rows"]],
            "rows_truncated": summary["rows_truncated"],
        }
    return await asyncio.to_thread(
        parse_profile, profile_key, content, filename, business_date, communities,
    )


@router.get("/import-profiles")
async def get_import_profiles(
    _user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    data = []
    async with conn.cursor() as cur:
        for profile in PROFILES.values():
            payload = profile_payload(profile)
            spreadsheets = await _enabled_spreadsheets(cur, profile.target_parser)
            payload["target_configured"] = len(spreadsheets) == 1
            data.append(payload)
    return {"data": data, "adapter_version": ADAPTER_VERSION}


@router.get("/quick-dispatch/options")
async def quick_dispatch_options(
    _user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    """返回快捷下发可选业务表及启用社区。"""
    quick_profiles = _quick_dispatch_profiles()
    async with conn.cursor() as cur:
        communities = await _communities(cur)
    return {
        "businesses": [dict({"key": key}, **value) for key, value in quick_profiles.items()],
        "communities": [
            {"id": item["id"], "name": item["name"]}
            for item in communities if item.get("enabled")
        ],
    }


def _quick_dispatch_profiles() -> dict[str, dict[str, Any]]:
    """返回快捷下发支持的业务表及其真实字段。"""
    text = lambda key, label, required=False, placeholder="": {
        "key": key, "label": label, "required": required, "type": "text", "placeholder": placeholder,
    }
    area = lambda key, label, required=False, placeholder="": {
        "key": key, "label": label, "required": required, "type": "textarea", "placeholder": placeholder,
    }
    registration = lambda key, label="登记情况", required=False: {
        "key": key, "label": label, "required": required, "type": "registration",
    }
    return {
        "fullchain_processed": {
            "label": "全链条", "target_parser": "全链条", "business_type": "fullchain", "police_subtype": "",
            "fields": [text("来源", "来源", True, "例如：基础管控临时指令"), text("姓名", "姓名", True), text("身份证号", "身份证号", True), text("电话号码", "电话号码", True), area("地址", "地址", True), registration("登记情况", required=True)],
        },
        "rental_processed": {
            "label": "出租房屋核查", "target_parser": "出租房屋核查", "business_type": "rental", "police_subtype": "",
            "fields": [text("姓名", "姓名", True), text("身份证号", "身份证号", True), text("手机号码", "手机号码", True), area("房屋地址", "房屋地址", True), text("入住方式", "入住方式", False, "自购、房东出租、中介出租等")],
        },
        "police_internal_processed": {
            "label": "所内涉警", "target_parser": "涉警统计", "business_type": "police", "police_subtype": "internal",
            "fields": [text("序号", "接警编号/序号", True), area("简要警情及处理结果", "简要警情及处理结果", True), text("是否开户", "是否开户"), area("现住址", "现住址"), text("房屋属性", "房屋属性"), text("居住时间", "居住时间"), text("房东信息", "房东信息"), text("二房东信息", "二房东信息"), area("备注", "备注"), text("房东是否处罚", "房东是否处罚")],
        },
        "police_suzhou_processed": {
            "label": "苏州涉警", "target_parser": "苏州涉警", "business_type": "police", "police_subtype": "suzhou",
            "fields": [text("姓名", "姓名", True), text("身份证号", "身份证号", True), text("联系号码", "联系号码", True), area("疑似现住址", "疑似现住址", True), text("接警编号", "接警编号"), text("出警日期", "出警日期"), text("出警类别", "出警类别"), area("出警内容", "出警内容"), text("出警单位", "出警单位"), text("参考派出所", "参考派出所"), area("备注", "备注")],
        },
        "police_traffic_processed": {
            "label": "交通涉警", "target_parser": "交通涉警", "business_type": "police", "police_subtype": "traffic",
            "fields": [text("姓名", "姓名", True), text("身份证号", "身份证号", True), text("联系号码", "联系号码", True), area("地址1", "地址1", True), area("备注", "备注")],
        },
        "delivery_processed": {
            "label": "寄递业", "target_parser": "寄递业", "business_type": "delivery", "police_subtype": "",
            "fields": [text("姓名", "姓名", True), text("身份证号", "身份证号", True), area("地址1", "地址1", True), text("手机号码", "手机号码", True), text("参考姓名", "参考姓名"), text("参考身份证号码", "参考身份证号码")],
        },
        "suspect_return_processed": {
            "label": "疑似返苏", "target_parser": "疑似返苏", "business_type": "suspect_return", "police_subtype": "",
            "fields": [text("姓名", "姓名", True), text("身份证号码", "身份证号码", True), text("联系号码", "联系号码", True), area("高频抓拍小区", "高频抓拍小区", True)],
        },
    }


@router.post("/quick-dispatch")
async def create_quick_dispatch(
    data: QuickDispatchCreate,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    """创建一条临时业务任务，交由既有发布工作台最终写入腾讯表。"""
    profile = _quick_dispatch_profiles().get(data.profile)
    if not profile:
        raise HTTPException(400, "不支持的快捷下发业务表")
    target_parser = str(profile["target_parser"])
    parser = get_parser(target_parser)
    allowed = set(parser.COLUMNS) | {"来源"}
    provided = {str(key): normalize_space(value) for key, value in (data.fields or {}).items() if str(key) in allowed}
    if data.profile == "fullchain_processed" and not provided:
        provided = {
            "来源": data.source_name, "姓名": data.person_name,
            "身份证号": data.identity_number, "电话号码": data.phone,
            "地址": data.original_address, "登记情况": data.registration_status,
        }
    required_fields = {
        "fullchain_processed": ("姓名", "身份证号", "电话号码", "地址", "登记情况"),
        "rental_processed": ("姓名", "身份证号", "手机号码", "房屋地址"),
        "police_internal_processed": ("序号", "简要警情及处理结果"),
        "police_suzhou_processed": ("姓名", "身份证号", "联系号码", "疑似现住址"),
        "police_traffic_processed": ("姓名", "身份证号", "联系号码", "地址1"),
        "delivery_processed": ("姓名", "身份证号", "地址1", "手机号码"),
        "suspect_return_processed": ("姓名", "身份证号码", "联系号码", "高频抓拍小区"),
    }[data.profile]
    missing = [field for field in required_fields if not provided.get(field)]
    if missing:
        raise HTTPException(400, f"快捷下发缺少必要字段：{'、'.join(missing)}")
    identity_field = "身份证号" if "身份证号" in provided else "身份证号码"
    identity_number = normalize_identity(provided.get(identity_field, ""))
    if identity_number and not IDENTITY_PATTERN.fullmatch(identity_number):
        raise HTTPException(400, "身份证号格式不正确")
    deadline = data.deadline_date or (data.business_date + timedelta(days=3))
    if deadline < data.business_date:
        raise HTTPException(400, "截止日期不能早于业务日期")
    created = data.created_time or datetime.now()
    created_text = created.strftime("%Y-%m-%d %H:%M:%S")
    dispatch_date = data.business_date.strftime("%m-%d")
    deadline_text = deadline.strftime("%m-%d")
    raw_values = dict(provided)
    standard_values = {column: "" for column in parser.COLUMNS}
    standard_values.update(provided)
    for key in ("下发日期", "下发时间"):
        if key in standard_values:
            standard_values[key] = dispatch_date
    for key in ("截止日期", "截止时间"):
        if key in standard_values:
            standard_values[key] = deadline_text
    if "创建时间" in standard_values and not standard_values["创建时间"]:
        standard_values["创建时间"] = created_text
    if "日期" in standard_values and not standard_values["日期"]:
        standard_values["日期"] = data.business_date.isoformat()
    person_name = standard_values.get("姓名", "")
    phone = standard_values.get("电话号码") or standard_values.get("手机号码") or standard_values.get("联系号码") or ""
    original_address = standard_values.get("地址") or standard_values.get("房屋地址") or standard_values.get("地址1") or standard_values.get("疑似现住址") or standard_values.get("简要警情及处理结果") or ""
    source_name = provided.get("来源") or data.source_name or str(profile["label"])
    created_value = standard_values.get("创建时间") or standard_values.get("下发日期") or standard_values.get("下发时间") or standard_values.get("日期") or created_text
    key_fields = parser_business_key_fields(parser, standard_values)
    if any(not normalize_space(standard_values.get(field, "")) for field in key_fields):
        raise HTTPException(400, f"快捷下发缺少业务主键字段：{'、'.join(key_fields)}")
    business_key = parser_business_key(parser, standard_values)
    batch_digest = _quick_dispatch_digest(int(user["id"]), data.request_id)
    reviewer_name = str(
        user.get("display_name")
        or (user.get("member") or {}).get("name")
        or user.get("username")
        or ""
    )[:100]

    await conn.begin()
    try:
        async with conn.cursor() as cur:
            existing = await _existing_quick_dispatch(cur, batch_digest)
            if existing:
                await conn.commit()
                return existing
            communities = await _communities(cur)
            community = next(
                (item for item in communities
                if data.community_id and int(item["id"]) == data.community_id and item.get("enabled")),
                None,
            )
            if not community:
                raise HTTPException(400, "请选择启用中的社区")
            community_name = str(community["name"])
            raw_values["社区"] = community_name
            standard_values["社区"] = community_name
            await cur.execute("""
                INSERT INTO _police_dispatch_batches (
                    file_name, file_sha256, sheet_name, import_mode, status,
                    total_count, counts_json, imported_by, business_type,
                    police_subtype, import_profile, adapter_version, target_parser, business_date,
                    source_summary_json, storage_key
                ) VALUES (%s,%s,%s,'quick','ready_to_publish',1,JSON_OBJECT(),%s,
                          %s,%s,'quick_dispatch',%s,%s,%s,%s,'')
            """, (
                f"快捷下发-{profile['label']}-{data.business_date.isoformat()}", batch_digest,
                "快捷下发", user["id"], str(profile["business_type"]), str(profile["police_subtype"]),
                ADAPTER_VERSION, target_parser, data.business_date,
                stable_json({"source": "quick", "row_count": 1, "profile": data.profile}),
            ))
            batch_id = int(cur.lastrowid)
            await cur.execute("""
                INSERT INTO _police_dispatch_tasks (
                    batch_id, source_row, source_name, person_name, identity_number,
                    identity_hash, phone, original_address, source_created_time,
                    raw_values_json, suggested_action, suggested_community_id,
                    suggestion_reason, allocation_mode, final_action,
                    final_community_id, review_note, reviewed_by, reviewer_name,
                    reviewed_at, task_status, publish_status, standard_values_json,
                    business_key_hmac, validation_issues_json
                ) VALUES (%s,1,%s,%s,%s,%s,%s,%s,%s,%s,'dispatch',%s,%s,
                          'quick_dispatch','dispatch',%s,%s,%s,%s,UTC_TIMESTAMP(),
                          'pending_publish','pending',%s,%s,JSON_ARRAY())
            """, (
                batch_id, source_name, person_name, identity_number,
                identity_digest(identity_number), phone, original_address,
                created_value, stable_json(raw_values), data.community_id,
                "快捷下发：已由创建人确认，等待发布", data.community_id,
                "快捷下发", user["id"], reviewer_name,
                stable_json(standard_values), business_key,
            ))
            task_id = int(cur.lastrowid)
            await _refresh_batch_status(cur, batch_id)
            payload = await _batch_payload(cur, batch_id)
        await conn.commit()
    except Exception as exc:
        await conn.rollback()
        if getattr(exc, "args", [None])[0] == 1062:
            async with conn.cursor() as cur:
                existing = await _existing_quick_dispatch(cur, batch_digest)
            if existing:
                return existing
        raise

    if local_data_source_enabled():
        local_result = await _execute_local_publish_selection(
            batch_id,
            TaskPublishSelection(task_ids=[task_id]),
            request,
            user,
            conn,
        )
        async with conn.cursor() as cur:
            payload = await _batch_payload(cur, batch_id)
        return {
            "status": "success" if not local_result["failed_count"] else "conflict",
            "message": "快捷下发已写入本地任务池" if not local_result["failed_count"] else "快捷下发存在本地业务键冲突",
            "batch": payload,
            "task_id": task_id,
            "publish": local_result,
        }

    await record_admin_audit(
        user,
        "police_dispatch.quick_create",
        target_type="police_dispatch_batch",
        target_name=str(batch_id),
        detail={"row_count": 1, "business_type": str(profile["business_type"]), "target_parser": target_parser},
        **request_audit_fields(request),
    )
    return {
        "status": "success",
        "message": "快捷下发任务已加入任务池，请到发布工作台完成发布",
        "batch": payload,
        "task_id": task_id,
    }


@router.post("/imports/preview")
async def preview_dispatch_import(
    request: Request,
    file: UploadFile = File(...),
    profile: str = Form(...),
    business_date: date = Form(...),
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    filename, content = await _read_upload(file)
    digest = sha256(content).hexdigest()
    try:
        async with conn.cursor() as cur:
            communities = await _communities(cur)
            addresses = await _address_entries(cur, enabled_only=True)
        parsed = await _parse_import_upload(
            profile_key=profile, content=content, filename=filename,
            business_date=business_date, communities=communities, addresses=addresses,
        )
    except (PoliceWorkbookError, MemoryError) as exc:
        raise HTTPException(400 if not isinstance(exc, MemoryError) else 413, str(exc) or "文件过大") from exc
    token = import_preview_token(
        user_id=int(user["id"]), file_sha256=digest, profile_key=profile,
        business_date=business_date, row_count=len(parsed["rows"]),
        sheet_name=parsed["sheet_name"],
    )
    await record_admin_audit(
        user, "police_dispatch.import_preview", target_type="police_dispatch_import",
        target_name=profile, detail={"row_count": len(parsed["rows"]), "profile": profile},
        **request_audit_fields(request),
    )
    return {
        "status": "preview", "preview_token": token, "file_sha256": digest,
        "preview": {
            "file_name": filename, "sheet_name": parsed["sheet_name"],
            "row_count": len(parsed["rows"]), "profile": profile_payload(parsed["profile"]),
            "business_date": business_date.isoformat(), "counts": parsed["counts"],
            "community_distribution": parsed["community_distribution"],
            "rows": parsed["preview_rows"], "rows_truncated": parsed["rows_truncated"],
        },
    }


@router.post("/imports/confirm")
async def confirm_dispatch_import(
    request: Request,
    file: UploadFile = File(...),
    profile: str = Form(...),
    business_date: date = Form(...),
    preview_token: str = Form(...),
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    filename, content = await _read_upload(file)
    digest = sha256(content).hexdigest()
    async with conn.cursor() as cur:
        communities = await _communities(cur)
        addresses = await _address_entries(cur, enabled_only=True)
    try:
        parsed = await _parse_import_upload(
            profile_key=profile, content=content, filename=filename,
            business_date=business_date, communities=communities, addresses=addresses,
        )
    except PoliceWorkbookError as exc:
        raise HTTPException(400, str(exc)) from exc
    if not verify_import_preview_token(
        preview_token, user_id=int(user["id"]), file_sha256=digest,
        profile_key=profile, business_date=business_date,
        row_count=len(parsed["rows"]), sheet_name=parsed["sheet_name"],
    ):
        raise HTTPException(409, "预览已过期、文件已变化或导入配置已变化，请重新预览")
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT id FROM _police_dispatch_batches WHERE file_sha256=%s AND import_profile=%s",
            (digest, profile),
        )
        duplicate = await cur.fetchone()
        if duplicate:
            return {"status": "duplicate", "message": "同一文件已按此业务类型导入", "batch": await _batch_payload(cur, int(duplicate[0]))}

    profile_meta = parsed["profile"]
    suffix = Path(filename).suffix.lower()
    storage_key = f"{profile}/{digest}{suffix}"
    target = (_dispatch_import_dir() / storage_key).resolve()
    if _dispatch_import_dir() not in target.parents:
        raise HTTPException(400, "文件存储路径无效")
    target.parent.mkdir(parents=True, exist_ok=True)
    staged_path = None
    created_file = False
    if not target.is_file():
        with tempfile.NamedTemporaryFile(dir=target.parent, prefix=".import-", suffix=".tmp", delete=False) as staged:
            staged.write(content)
            staged_path = staged.name
        os.replace(staged_path, target)
        staged_path = None
        created_file = True

    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("""
                INSERT INTO _police_dispatch_batches (
                    file_name,file_sha256,sheet_name,import_mode,status,total_count,
                    counts_json,imported_by,business_type,police_subtype,import_profile,
                    adapter_version,target_parser,business_date,source_summary_json,storage_key
                ) VALUES (%s,%s,%s,'processed','reviewing',%s,JSON_OBJECT(),%s,
                          %s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                filename, digest, parsed["sheet_name"], len(parsed["rows"]), user["id"],
                profile_meta.business_type, profile_meta.police_subtype, profile,
                parsed["adapter_version"], profile_meta.target_parser, business_date,
                stable_json(parsed["counts"]), storage_key,
            ))
            batch_id = int(cur.lastrowid)
            for item in parsed["rows"]:
                await cur.execute("""
                    INSERT INTO _police_dispatch_tasks (
                        batch_id,source_row,source_name,person_name,identity_number,
                        identity_hash,phone,original_address,source_created_time,
                        transfer_note,raw_values_json,duplicate_group_key,duplicate_kind,
                        suggested_action,suggested_community_id,suggestion_reason,
                        allocation_mode,standard_values_json,business_key_hmac,
                        validation_issues_json
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    batch_id, item["source_row"], item.get("source_name", ""),
                    item.get("person_name", ""), item.get("identity_number", ""),
                    identity_digest(item.get("identity_number", "")), item.get("phone", ""),
                    item.get("original_address", ""), item.get("created_time", ""),
                    item.get("transfer_note", ""), stable_json(item.get("raw_values", {})),
                    item.get("duplicate_group_key", ""), item.get("duplicate_kind", ""),
                    item.get("suggested_action", "manual"), item.get("suggested_community_id"),
                    item.get("suggestion_reason", ""), item.get("allocation_mode", "conflict"),
                    stable_json({k: v for k, v in item.get("standard_values", {}).items() if not k.startswith("__")}),
                    item.get("business_key_hmac", ""), stable_json(item.get("validation_issues", [])),
                ))
                task_id = int(cur.lastrowid)
                for issue in item.get("validation_issues", []):
                    await cur.execute("""
                        INSERT INTO _police_dispatch_import_issues
                        (batch_id,task_id,source_row,field_name,issue_type,safe_value)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (batch_id, task_id, item["source_row"], issue.get("field", ""), issue.get("type", "conflict"), str(issue.get("value", ""))[:200]))
            await _refresh_batch_status(cur, batch_id)
        await conn.commit()
    except Exception:
        await conn.rollback()
        if created_file:
            target.unlink(missing_ok=True)
        raise
    async with conn.cursor() as cur:
        payload = await _batch_payload(cur, batch_id)
    await record_admin_audit(
        user, "police_dispatch.import_confirm", target_type="police_dispatch_batch",
        target_name=str(batch_id), detail={"row_count": len(parsed["rows"]), "profile": profile},
        **request_audit_fields(request),
    )
    return {"status": "success", "message": "数据已导入下发工作台，请审核后发布", "batch": payload}


@router.get("/batches/{batch_id}/source-file")
async def download_dispatch_source_file(
    batch_id: int,
    _user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    async with conn.cursor() as cur:
        await cur.execute("SELECT file_name,storage_key FROM _police_dispatch_batches WHERE id=%s", (batch_id,))
        row = await cur.fetchone()
    if not row:
        raise HTTPException(404, "批次不存在")
    if not row[1]:
        raise HTTPException(404, "旧批次没有保存原始文件")
    path = (_dispatch_import_dir() / str(row[1])).resolve()
    if _dispatch_import_dir() not in path.parents or not path.is_file():
        raise HTTPException(410, "原始文件已丢失，请联系超级管理员")
    return FileResponse(path, media_type="application/octet-stream", filename=str(row[0]))


def require_fullchain_archive_access() -> Callable:
    """归档面板允许内勤岗位，以及管理员/超级管理员显式进入。"""
    base_dependency = require_permission(POLICE_DISPATCH_MANAGE)

    async def dependency(user: dict = Depends(base_dependency)) -> dict:
        permission_scope = (user.get("permission_scopes") or {}).get(
            POLICE_DISPATCH_MANAGE,
            user.get("data_scope"),
        )
        if permission_scope != "all":
            raise HTTPException(403, "全链条归档必须使用全所数据范围")
        group_codes = _permission_group_codes(user)
        if group_codes.intersection({"admin", "super_admin"}) or user.get("role") in {
            "admin", "super_admin",
        }:
            return user
        member = user.get("member") or {}
        if str(member.get("position") or "") in FULLCHAIN_ARCHIVE_POSITIONS:
            return user
        raise HTTPException(403, "全链条归档仅向基础管控、中队长、所队领导和系统管理员开放")

    dependency.__name__ = "require_fullchain_archive_manage"
    return dependency


require_fullchain_archive = require_fullchain_archive_access()


def require_police_address_access() -> Callable:
    base_dependency = require_permission(POLICE_ADDRESS_MANAGE)

    async def dependency(user: dict = Depends(base_dependency)) -> dict:
        member = user.get("member") or {}
        position = str(member.get("position") or "")
        if position in {"组长", "组员"}:
            return user
        if position in ALLOWED_POLICE_POSITIONS:
            permission_scope = (user.get("permission_scopes") or {}).get(
                POLICE_ADDRESS_MANAGE,
                user.get("data_scope"),
            )
            if permission_scope == "all":
                return user
            raise HTTPException(403, "小区管理岗位需要全所数据范围")
        group_codes = _permission_group_codes(user)
        if group_codes.intersection({"admin", "super_admin"}) or (
            not member and user.get("role") in {"admin", "super_admin"}
        ):
            return user
        raise HTTPException(403, "当前人员岗位不能进入小区管理")

    dependency.__name__ = "require_police_address_manage"
    return dependency


require_police_address = require_police_address_access()


def _safe_filename(filename: str | None, fallback: str) -> str:
    return Path((filename or fallback).replace("\\", "/")).name[:255]


async def _read_upload(file: UploadFile, *, allow_xls: bool = True) -> tuple[str, bytes]:
    filename = _safe_filename(file.filename, "下发数据.xlsx")
    suffixes = {".xlsx", ".xls"} if allow_xls else {".xlsx"}
    if Path(filename).suffix.lower() not in suffixes:
        raise HTTPException(400, "只支持 .xls 或 .xlsx 文件")
    content = await file.read(MAX_POLICE_FILE_BYTES + 1)
    await file.close()
    if not content:
        raise HTTPException(400, "上传文件为空")
    if len(content) > MAX_POLICE_FILE_BYTES:
        raise HTTPException(413, "Excel 文件不能超过 30MB")
    return filename, content


async def _communities(cur) -> list[dict[str, Any]]:
    await cur.execute("""
        SELECT community.id, community.name,
               MAX(CASE WHEN community.is_active=1
                              AND department.is_active=1
                        THEN 1 ELSE 0 END) AS enabled
        FROM _communities AS community
        LEFT JOIN _departments AS department
          ON department.community_id=community.id
         AND department.department_type='community'
        GROUP BY community.id, community.name
        ORDER BY community.id
    """)
    result = [
        {
            "id": int(row[0]),
            "name": str(row[1]),
            "enabled": bool(row[2]),
            "sort_order": int(row[0]),
            "aliases": [],
        }
        for row in await cur.fetchall()
    ]
    by_id = {item["id"]: item for item in result}
    await cur.execute("SELECT community_id, alias FROM _community_aliases ORDER BY id")
    for community_id, alias in await cur.fetchall():
        if int(community_id) in by_id:
            by_id[int(community_id)]["aliases"].append(str(alias))
    return result


def _address_scope_community_ids(
    user: dict,
    communities: list[dict[str, Any]],
) -> list[int] | None:
    position = str((user.get("member") or {}).get("position") or "")
    if position not in {"组长", "组员"}:
        return None
    names = permitted_communities(user, POLICE_ADDRESS_MANAGE) or []
    normalized_names = {
        normalize_lookup(name.removesuffix("社区"))
        for name in names
        if str(name).strip()
    }
    return [
        int(community["id"])
        for community in communities
        if normalize_lookup(str(community["name"]).removesuffix("社区"))
        in normalized_names
    ]


def _filter_address_rows(
    data: list[dict[str, Any]],
    *,
    keyword: str = "",
    enabled: bool | None = None,
) -> list[dict[str, Any]]:
    term = normalize_lookup(keyword)
    if term:
        data = [item for item in data if term in normalize_lookup(
            " ".join((item["name"], item["detail_address"], item["community_name"], *item["aliases"]))
        )]
    if enabled is not None:
        data = [item for item in data if item["enabled"] is enabled]
    return data


async def _address_entries(
    cur,
    *,
    enabled_only: bool = False,
    community_ids: list[int] | None = None,
) -> list[dict[str, Any]]:
    where: list[str] = []
    params: list[Any] = []
    if enabled_only:
        where.append("entry.enabled=1")
    if community_ids is not None:
        if not community_ids:
            where.append("1=0")
        else:
            where.append("entry.community_id IN (" + ",".join(["%s"] * len(community_ids)) + ")")
            params.extend(community_ids)
    clause = "WHERE " + " AND ".join(where) if where else ""
    await cur.execute(f"""
        SELECT entry.id, entry.name, entry.detail_address,
               entry.address_type, entry.pattern, entry.community_id,
               community.name, entry.aliases_json, entry.source_flags,
               entry.enabled, entry.created_at, entry.updated_at
        FROM _police_address_entries AS entry
        LEFT JOIN _communities AS community ON community.id=entry.community_id
        {clause}
        ORDER BY entry.enabled DESC, community.id, entry.name, entry.id
    """, tuple(params))
    return [
        {
            "id": int(row[0]),
            "name": str(row[1]),
            "detail_address": str(row[2] or ""),
            "address_type": str(row[3]),
            "pattern": str(row[4] or ""),
            "community_id": int(row[5]) if row[5] is not None else None,
            "community_name": str(row[6] or ""),
            "aliases": json_value(row[7], []),
            "sources": json_value(row[8], []),
            "enabled": bool(row[9]),
            "created_at": row[10].isoformat() + "Z" if row[10] else None,
            "updated_at": row[11].isoformat() + "Z" if row[11] else None,
        }
        for row in await cur.fetchall()
    ]


def _address_payload(item: AddressCreate) -> tuple:
    aliases = sorted({alias.strip() for alias in item.aliases if alias.strip()})
    return (
        item.name.strip(),
        normalize_lookup(item.name),
        item.detail_address.strip(),
        item.address_type,
        item.pattern.strip(),
        item.community_id,
        stable_json(aliases),
        1 if item.enabled else 0,
    )


async def _assert_community(cur, community_id: int, *, require_enabled: bool = False) -> None:
    if require_enabled:
        await cur.execute("""
            SELECT community.id
            FROM _communities AS community
            JOIN _departments AS department
              ON department.community_id=community.id
             AND department.department_type='community'
             AND department.is_active=1
            WHERE community.id=%s AND community.is_active=1
            LIMIT 1
        """, (community_id,))
    else:
        await cur.execute("SELECT id FROM _communities WHERE id=%s", (community_id,))
    if not await cur.fetchone():
        raise HTTPException(400, "所选社区不存在或已停用")


def _assert_address_scope(community_id: int, allowed_ids: list[int] | None) -> None:
    if allowed_ids is not None and community_id not in allowed_ids:
        raise HTTPException(403, "只能管理本人所属社区的小区")


async def _address_page_data(
    cur,
    user: dict,
    *,
    keyword: str = "",
    enabled: bool | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[int] | None]:
    communities = await _communities(cur)
    allowed_ids = _address_scope_community_ids(user, communities)
    visible_communities = (
        communities
        if allowed_ids is None
        else [item for item in communities if int(item["id"]) in allowed_ids]
    )
    data = await _address_entries(cur, community_ids=allowed_ids)
    return _filter_address_rows(data, keyword=keyword, enabled=enabled), visible_communities, allowed_ids


@router.get("/addresses")
async def list_addresses(
    keyword: str = Query("", max_length=100),
    enabled: bool | None = Query(None),
    user: dict = Depends(require_police_address),
    conn=Depends(get_db),
):
    async with conn.cursor() as cur:
        data, communities, allowed_ids = await _address_page_data(
            cur, user, keyword=keyword, enabled=enabled,
        )
    return {
        "data": data,
        "total": len(data),
        "communities": communities,
        "community_locked": allowed_ids is not None,
    }


@router.post("/addresses/export")
async def export_addresses(
    data: AddressSearch,
    request: Request,
    user: dict = Depends(require_police_address),
    conn=Depends(get_db),
):
    async with conn.cursor() as cur:
        rows, _, allowed_ids = await _address_page_data(
            cur, user, keyword=data.keyword, enabled=data.enabled,
        )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "小区管理"
    headers = ["名称", "正式社区", "类型", "详细地址", "模式", "别名", "状态"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1677FF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    type_labels = {
        "community": "居民小区",
        "apartment": "公寓",
        "construction_dormitory": "工地宿舍",
        "other": "其他",
    }
    for item in rows:
        sheet.append([
            item["name"], item["community_name"],
            type_labels.get(item["address_type"], item["address_type"]),
            item["detail_address"], item["pattern"], "，".join(item["aliases"]),
            "启用" if item["enabled"] else "停用",
        ])
    widths = {"A": 28, "B": 16, "C": 12, "D": 48, "E": 24, "F": 42, "G": 10}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str):
                cell.number_format = "@"
                cell.data_type = "s"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    await record_admin_audit(
        user,
        "police_address.export",
        target_type="police_address",
        target_name="community-scope" if allowed_ids is not None else "all",
        detail={"row_count": len(rows), "file_format": "XLSX"},
        **request_audit_fields(request),
    )
    filename = f"小区管理-{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/addresses")
async def create_address(
    data: AddressCreate,
    request: Request,
    user: dict = Depends(require_police_address),
    conn=Depends(get_db),
):
    async with conn.cursor() as cur:
        communities = await _communities(cur)
        allowed_ids = _address_scope_community_ids(user, communities)
        _assert_address_scope(data.community_id, allowed_ids)
        await _assert_community(cur, data.community_id, require_enabled=True)
        try:
            await cur.execute("""
                INSERT INTO _police_address_entries (
                    name, normalized_name, detail_address, address_type,
                    pattern, community_id, aliases_json, source_flags,
                    enabled, created_by, updated_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, JSON_ARRAY('manual'), %s, %s, %s)
            """, (*_address_payload(data), user["id"], user["id"]))
        except Exception as exc:
            if getattr(exc, "args", [None])[0] == 1062:
                raise HTTPException(409, "同一社区已经存在同名地址记录") from exc
            raise
        entry_id = int(cur.lastrowid)
    await record_admin_audit(
        user, "police_address.create", target_type="police_address",
        target_name=str(entry_id), detail={"address_type": data.address_type},
        **request_audit_fields(request),
    )
    return {"id": entry_id, "message": "地址记录已创建"}


@router.put("/addresses/{entry_id}")
async def update_address(
    entry_id: int,
    data: AddressCreate,
    request: Request,
    user: dict = Depends(require_police_address),
    conn=Depends(get_db),
):
    async with conn.cursor() as cur:
        communities = await _communities(cur)
        allowed_ids = _address_scope_community_ids(user, communities)
        _assert_address_scope(data.community_id, allowed_ids)
        await _assert_community(cur, data.community_id, require_enabled=True)
        await cur.execute(
            "SELECT id, community_id FROM _police_address_entries WHERE id=%s",
            (entry_id,),
        )
        current = await cur.fetchone()
        if not current:
            raise HTTPException(404, "地址记录不存在")
        _assert_address_scope(int(current[1]), allowed_ids)
        try:
            await cur.execute("""
                UPDATE _police_address_entries SET
                    name=%s, normalized_name=%s, detail_address=%s,
                    address_type=%s, pattern=%s, community_id=%s,
                    aliases_json=%s, enabled=%s, updated_by=%s
                WHERE id=%s
            """, (*_address_payload(data), user["id"], entry_id))
        except Exception as exc:
            if getattr(exc, "args", [None])[0] == 1062:
                raise HTTPException(409, "同一社区已经存在同名地址记录") from exc
            raise
    await record_admin_audit(
        user, "police_address.update", target_type="police_address",
        target_name=str(entry_id), detail={"enabled": data.enabled},
        **request_audit_fields(request),
    )
    return {"message": "地址记录已更新"}


@router.delete("/addresses/{entry_id}")
async def delete_address(
    entry_id: int,
    request: Request,
    user: dict = Depends(require_police_address),
    conn=Depends(get_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            communities = await _communities(cur)
            allowed_ids = _address_scope_community_ids(user, communities)
            await cur.execute(
                "SELECT community_id FROM _police_address_entries WHERE id=%s FOR UPDATE",
                (entry_id,),
            )
            current = await cur.fetchone()
            if not current:
                raise HTTPException(404, "地址记录不存在")
            _assert_address_scope(int(current[0]), allowed_ids)
            await cur.execute(
                "DELETE FROM _police_address_sources WHERE entry_id=%s",
                (entry_id,),
            )
            source_count = int(cur.rowcount)
            await cur.execute(
                "DELETE FROM _police_address_entries WHERE id=%s",
                (entry_id,),
            )
            if cur.rowcount != 1:
                raise HTTPException(409, "地址记录已发生变化，请刷新后重试")
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "police_address.delete", target_type="police_address",
        target_name=str(entry_id), detail={"source_links_removed": source_count},
        **request_audit_fields(request),
    )
    return {"message": "地址记录已删除"}


def _task_counts(rows: list[tuple]) -> dict[str, int]:
    result = {
        "total": 0, "pending_review": 0, "reviewed": 0, "no_registration": 0,
        "transfer": 0, "dispatch": 0, "balanced": 0, "duplicate": 0,
        "abnormal": 0, "pending_publish": 0, "published": 0,
        "retryable": 0, "needs_reconciliation": 0, "conflict": 0,
        "cache_pending": 0, "publishable": 0, "partial_publishable": 0,
    }
    for row in rows:
        result["total"] += int(row[0] or 0)
        result["pending_review"] += int(row[1] or 0)
        result["reviewed"] += int(row[2] or 0)
        result["no_registration"] += int(row[3] or 0)
        result["transfer"] += int(row[4] or 0)
        result["dispatch"] += int(row[5] or 0)
        result["balanced"] += int(row[6] or 0)
        result["duplicate"] += int(row[7] or 0)
        result["abnormal"] += int(row[8] or 0)
        result["pending_publish"] += int(row[9] or 0)
        result["published"] += int(row[10] or 0)
        if len(row) > 11:
            result["retryable"] += int(row[11] or 0)
            result["needs_reconciliation"] += int(row[12] or 0)
            result["conflict"] += int(row[13] or 0)
            result["cache_pending"] += int(row[14] or 0)
        if len(row) > 15:
            result["publishable"] += int(row[15] or 0)
        if len(row) > 16:
            result["partial_publishable"] += int(row[16] or 0)
    return result


async def _batch_counts(cur, batch_id: int) -> dict[str, int]:
    await cur.execute("""
        SELECT COUNT(*),
               SUM(task_status='pending_review'),
               SUM(final_action<>''),
               SUM(COALESCE(NULLIF(final_action, ''), suggested_action)='no_registration'),
               SUM(COALESCE(NULLIF(final_action, ''), suggested_action)='transfer'),
               SUM(COALESCE(NULLIF(final_action, ''), suggested_action)='dispatch'),
               SUM(allocation_mode='balanced'),
               SUM(duplicate_group_key<>''),
               SUM(suggested_action='manual'),
               SUM(publish_status IN ('pending', 'publishing', 'retryable',
                                      'needs_reconciliation', 'conflict')),
               SUM(publish_status='success'),
               SUM(publish_status='retryable'),
               SUM(publish_status='needs_reconciliation'),
               SUM(publish_status='conflict'),
               SUM(cache_pending=1),
               SUM(publish_status IN ('pending', 'retryable')),
               SUM(publish_status IN ('pending', 'retryable') AND duplicate_group_key='')
        FROM _police_dispatch_tasks WHERE batch_id=%s
    """, (batch_id,))
    return _task_counts([await cur.fetchone()])


async def _refresh_batch_status(cur, batch_id: int) -> dict[str, int]:
    counts = await _batch_counts(cur, batch_id)
    if counts["total"] and counts["pending_review"] == 0:
        if counts["pending_publish"] == 0:
            status = "completed"
        elif counts["needs_reconciliation"] or counts["conflict"]:
            status = "reconciling"
        else:
            status = "ready_to_publish"
    else:
        status = "reviewing"
    await cur.execute("""
        UPDATE _police_dispatch_batches SET status=%s, counts_json=%s,
            completed_at=CASE WHEN %s='completed' THEN COALESCE(completed_at, UTC_TIMESTAMP()) ELSE NULL END
        WHERE id=%s
    """, (status, stable_json(counts), status, batch_id))
    return counts


async def _batch_payload(cur, batch_id: int) -> dict[str, Any]:
    payloads = await _batch_payloads(cur, [batch_id])
    if not payloads:
        raise HTTPException(404, "批次不存在")
    return payloads[0]


async def _batch_payloads(cur, batch_ids: list[int]) -> list[dict[str, Any]]:
    if not batch_ids:
        return []
    placeholders = ",".join(["%s"] * len(batch_ids))
    await cur.execute(f"""
        SELECT batch.id, batch.file_name, batch.sheet_name, batch.status,
               batch.total_count, batch.counts_json, batch.first_publish_date,
               batch.last_error, batch.created_at, batch.updated_at,
               user.display_name, user.username, batch.import_mode,
               batch.business_type, batch.police_subtype, batch.import_profile,
               batch.adapter_version, batch.target_parser, batch.business_date,
               batch.source_summary_json, batch.storage_key
        FROM _police_dispatch_batches AS batch
        LEFT JOIN _users AS user ON user.id=batch.imported_by
        WHERE batch.id IN ({placeholders})
    """, batch_ids)
    batch_rows = {int(row[0]): row for row in await cur.fetchall()}
    await cur.execute(f"""
        SELECT batch_id, COUNT(*),
               SUM(task_status='pending_review'),
               SUM(final_action<>''),
               SUM(COALESCE(NULLIF(final_action, ''), suggested_action)='no_registration'),
               SUM(COALESCE(NULLIF(final_action, ''), suggested_action)='transfer'),
               SUM(COALESCE(NULLIF(final_action, ''), suggested_action)='dispatch'),
               SUM(allocation_mode='balanced'),
               SUM(duplicate_group_key<>''),
               SUM(suggested_action='manual'),
               SUM(publish_status IN ('pending', 'publishing', 'retryable',
                                      'needs_reconciliation', 'conflict')),
               SUM(publish_status='success'),
               SUM(publish_status='retryable'),
               SUM(publish_status='needs_reconciliation'),
               SUM(publish_status='conflict'),
               SUM(cache_pending=1),
               SUM(publish_status IN ('pending', 'retryable')),
               SUM(publish_status IN ('pending', 'retryable') AND duplicate_group_key='')
        FROM _police_dispatch_tasks
        WHERE batch_id IN ({placeholders}) GROUP BY batch_id
    """, batch_ids)
    counts_by_batch = {
        int(row[0]): _task_counts([row[1:]])
        for row in await cur.fetchall()
    }
    await cur.execute(f"""
        SELECT task.batch_id, community.id, community.name, COUNT(*)
        FROM _police_dispatch_tasks AS task
        JOIN _communities AS community
          ON community.id=COALESCE(task.final_community_id, task.suggested_community_id)
        WHERE task.batch_id IN ({placeholders})
          AND COALESCE(NULLIF(task.final_action, ''), task.suggested_action)='dispatch'
        GROUP BY task.batch_id, community.id, community.name
        ORDER BY task.batch_id, community.id
    """, batch_ids)
    distribution_by_batch: dict[int, list[dict[str, Any]]] = {}
    for batch_id, community_id, community_name, count in await cur.fetchall():
        distribution_by_batch.setdefault(int(batch_id), []).append({
            "community_id": int(community_id),
            "community_name": str(community_name),
            "count": int(count),
        })
    result = []
    for batch_id in batch_ids:
        row = batch_rows.get(int(batch_id))
        if not row:
            continue
        counts = counts_by_batch.get(int(batch_id), _task_counts([]))
        result.append({
            "id": int(row[0]), "file_name": str(row[1]), "sheet_name": str(row[2]),
            "status": str(row[3]), "total_count": int(row[4]), "counts": counts,
            "import_mode": str(row[12] or "raw") if len(row) > 12 else "raw",
            "business_type": str(row[13] or "fullchain") if len(row) > 13 else "fullchain",
            "police_subtype": str(row[14] or "") if len(row) > 14 else "",
            "import_profile": str(row[15] or "fullchain_raw") if len(row) > 15 else "fullchain_raw",
            "adapter_version": str(row[16] or "") if len(row) > 16 else "",
            "target_parser": str(row[17] or "全链条") if len(row) > 17 else "全链条",
            "business_date": row[18].isoformat() if len(row) > 18 and row[18] else None,
            "source_summary": json_value(row[19], {}) if len(row) > 19 else {},
            "source_file_available": bool(row[20]) if len(row) > 20 else False,
            "first_publish_date": row[6].isoformat() if row[6] else None,
            "last_error": str(row[7] or ""),
            "created_at": row[8].isoformat() + "Z", "updated_at": row[9].isoformat() + "Z",
            "imported_by": str(row[10] or row[11] or ""),
            "reviewed_count": counts["reviewed"],
            "community_distribution": distribution_by_batch.get(int(batch_id), []),
        })
    return result


@router.post("/batches")
async def upload_dispatch_batch(
    request: Request,
    file: UploadFile = File(...),
    import_mode: Literal["raw", "clean"] = Form("raw"),
    confirm: bool = Form(False),
    preview_token: str = Form(""),
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    filename, content = await _read_upload(file)
    digest = sha256(content).hexdigest()
    legacy_profile = "fullchain_processed" if import_mode == "clean" else "fullchain_raw"
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT id FROM _police_dispatch_batches "
            "WHERE file_sha256=%s AND import_profile=%s",
            (digest, legacy_profile),
        )
        duplicate = await cur.fetchone()
        if duplicate:
            payload = await _batch_payload(cur, int(duplicate[0]))
            return {"status": "duplicate", "message": "同一文件已经导入", "batch": payload}
    try:
        sheet_name, parsed = await asyncio.to_thread(
            parse_dispatch_workbook,
            content,
            filename,
            import_mode == "clean",
        )
    except MemoryError as exc:
        raise HTTPException(
            413,
            "工作簿展开后占用内存过大，请拆分文件后重新上传",
        ) from exc
    except PoliceWorkbookError as exc:
        raise HTTPException(400, str(exc)) from exc
    async with conn.cursor() as cur:
        communities = await _communities(cur)
        addresses = await _address_entries(cur, enabled_only=True)
    tasks = [
        {
            "source_row": item.source_row, "source_name": item.source_name,
            "community_name": item.community_name,
            "person_name": item.person_name, "identity_number": item.identity_number,
            "phone": item.phone, "original_address": item.original_address,
            "registration_status": item.registration_status,
            "created_time": item.created_time, "transfer_note": item.transfer_note,
            "raw_values": item.raw_values,
        }
        for item in parsed
    ]
    if import_mode == "clean":
        apply_clean_import_actions(tasks, communities)
    else:
        apply_preprocessing_suggestions(tasks, communities, addresses)
    if import_mode == "clean" and not confirm:
        summary = _clean_preview_summary(tasks)
        await record_admin_audit(
            user,
            "police_dispatch.preview",
            target_type="police_dispatch_preview",
            target_name=filename,
            detail={
                "row_count": len(tasks),
                "import_mode": "clean",
                "dispatch_count": summary["counts"]["dispatch"],
                "manual_review_count": summary["counts"]["manual_review"],
            },
            **request_audit_fields(request),
        )
        return {
            "status": "preview",
            "message": "已生成预览，请确认后再导入",
            "preview_token": _clean_preview_token(digest, filename, sheet_name, len(tasks)),
            "preview": {
                "file_name": filename,
                "sheet_name": sheet_name,
                "row_count": len(tasks),
                **summary,
            },
        }
    if import_mode == "clean":
        if not preview_token or not _verify_clean_preview_token(
            preview_token, digest, filename, sheet_name, len(tasks)
        ):
            raise HTTPException(409, "预览已失效或文件内容已变化，请重新预览")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("""
                INSERT INTO _police_dispatch_batches (
                    file_name, file_sha256, sheet_name, import_mode, status,
                    total_count, counts_json, imported_by, import_profile,
                    adapter_version
                ) VALUES (%s, %s, %s, %s, 'reviewing', %s, JSON_OBJECT(), %s, %s, %s)
            """, (
                filename, digest, sheet_name, import_mode, len(tasks), user["id"],
                legacy_profile, ADAPTER_VERSION,
            ))
            batch_id = int(cur.lastrowid)
            for item in tasks:
                await cur.execute("""
                    INSERT INTO _police_dispatch_tasks (
                        batch_id, source_row, source_name, person_name,
                        identity_number, identity_hash, phone, original_address,
                        source_created_time, transfer_note, raw_values_json,
                        duplicate_group_key, duplicate_kind, suggested_action,
                        suggested_community_id, suggestion_reason, allocation_mode
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s
                    )
                """, (
                    batch_id, item["source_row"], item["source_name"], item["person_name"],
                    item["identity_number"], item.get("identity_hash", ""), item["phone"],
                    item["original_address"], item["created_time"], item["transfer_note"],
                    stable_json(item["raw_values"]), item.get("duplicate_group_key", ""),
                    item.get("duplicate_kind", ""), item["suggested_action"],
                    item.get("suggested_community_id"), item["suggestion_reason"],
                    item["allocation_mode"],
                ))
                if import_mode == "clean" and item.get("auto_final_action"):
                    reviewer_name = str(
                        user.get("display_name")
                        or (user.get("member") or {}).get("name")
                        or user.get("username")
                        or ""
                    )[:100]
                    final_action = str(item["auto_final_action"])
                    await cur.execute("""
                        UPDATE _police_dispatch_tasks SET
                            final_action=%s, final_community_id=%s,
                            review_note=%s, reviewed_by=%s, reviewer_name=%s,
                            reviewed_at=UTC_TIMESTAMP(), task_status=%s,
                            publish_status=%s, version=version+1
                        WHERE batch_id=%s AND source_row=%s
                    """, (
                        final_action,
                        item.get("auto_final_community_id"),
                        "已处理数据导入：按登记情况自动生成",
                        user["id"], reviewer_name,
                        "pending_publish" if final_action == "dispatch" else "completed",
                        "pending" if final_action == "dispatch" else "not_required",
                        batch_id, item["source_row"],
                    ))
            await _refresh_batch_status(cur, batch_id)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    async with conn.cursor() as cur:
        payload = await _batch_payload(cur, batch_id)
    await record_admin_audit(
        user, "police_dispatch.import", target_type="police_dispatch_batch",
        target_name=str(batch_id), detail={"row_count": len(tasks), "import_mode": import_mode},
        **request_audit_fields(request),
    )
    message = (
        "已处理数据导入完成，可直接发布；异常记录仍需人工审核"
        if import_mode == "clean"
        else "文件已导入，所有建议等待人工审核"
    )
    return {"status": "success", "message": message, "batch": payload}


@router.get("/batches")
async def list_batches(
    file_name: str = Query("", max_length=100),
    upload_date: str = Query("", max_length=10),
    status: str = Query("all", max_length=30),
    business_type: str = Query("all", max_length=30),
    police_subtype: str = Query("all", max_length=30),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    del user
    where = ["1=1"]
    params: list[Any] = []
    if file_name.strip():
        where.append("file_name LIKE %s")
        params.append(f"%{file_name.strip()}%")
    if upload_date.strip():
        try:
            datetime.strptime(upload_date.strip(), "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(400, "上传日期格式必须为 YYYY-MM-DD") from exc
        where.append("DATE(created_at)=%s")
        params.append(upload_date.strip())
    if status != "all":
        allowed = {"reviewing", "ready_to_publish", "publishing", "reconciling", "completed"}
        if status not in allowed:
            raise HTTPException(400, "批次状态筛选无效")
        where.append("status=%s")
        params.append(status)
    if business_type != "all":
        if business_type not in {"fullchain", "rental", "police", "delivery", "suspect_return"}:
            raise HTTPException(400, "业务类型筛选无效")
        where.append("business_type=%s")
        params.append(business_type)
    if police_subtype != "all":
        if police_subtype not in {"internal", "suzhou", "traffic"}:
            raise HTTPException(400, "涉警子类型筛选无效")
        where.append("police_subtype=%s")
        params.append(police_subtype)
    where_sql = " AND ".join(where)
    async with conn.cursor() as cur:
        await cur.execute(
            f"SELECT COUNT(*) FROM _police_dispatch_batches WHERE {where_sql}",
            params,
        )
        total = int((await cur.fetchone())[0] or 0)
        await cur.execute(
            f"""
            SELECT id FROM _police_dispatch_batches
            WHERE {where_sql}
            ORDER BY (status='completed') ASC,
                     CASE WHEN status<>'completed' THEN created_at END ASC,
                     CASE WHEN status='completed' THEN created_at END DESC,
                     id DESC
            LIMIT %s OFFSET %s
            """,
            params + [page_size, (page - 1) * page_size],
        )
        ids = [int(row[0]) for row in await cur.fetchall()]
        data = await _batch_payloads(cur, ids)
    return {"data": data, "total": total, "page": page, "page_size": page_size}


@router.get("/batches/{batch_id}")
async def get_batch(
    batch_id: int,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    del user
    async with conn.cursor() as cur:
        batch = await _batch_payload(cur, batch_id)
        communities = await _communities(cur)
    return {"batch": batch, "communities": communities}


@router.delete("/batches/{batch_id}")
async def delete_batch(
    batch_id: int,
    request: Request,
    user: dict = Depends(require_super_admin),
    conn=Depends(get_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("""
                SELECT id, first_publish_date, publish_started_at
                FROM _police_dispatch_batches
                WHERE id=%s
                FOR UPDATE
            """, (batch_id,))
            batch = await cur.fetchone()
            if not batch:
                raise HTTPException(404, "批次不存在")
            if batch[1] is not None or batch[2] is not None:
                raise HTTPException(
                    409,
                    "该批次已经开始发布，不能删除；腾讯表格中的外部结果无法随批次撤销",
                )

            await cur.execute("""
                SELECT COUNT(task.id),
                       SUM(result.id IS NOT NULL),
                       SUM(task.publish_status IN (
                           'publishing', 'retryable', 'needs_reconciliation',
                           'conflict', 'success'
                       ) OR task.published_row IS NOT NULL
                         OR task.linked_source_id IS NOT NULL)
                FROM _police_dispatch_tasks AS task
                LEFT JOIN _police_dispatch_publish_results AS result
                  ON result.task_id=task.id
                WHERE task.batch_id=%s
            """, (batch_id,))
            task_count, result_count, external_state_count = await cur.fetchone()
            if int(result_count or 0) or int(external_state_count or 0):
                raise HTTPException(
                    409,
                    "该批次已经存在发布记录或腾讯来源关联，不能删除",
                )

            await cur.execute(
                "DELETE FROM _police_dispatch_tasks WHERE batch_id=%s",
                (batch_id,),
            )
            await cur.execute(
                "DELETE FROM _police_dispatch_batches WHERE id=%s",
                (batch_id,),
            )
            if not cur.rowcount:
                raise HTTPException(409, "批次状态已经变化，请刷新后重试")
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise

    await record_admin_audit(
        user,
        "police_dispatch.delete",
        target_type="police_dispatch_batch",
        target_name=str(batch_id),
        detail={"row_count": int(task_count or 0)},
        **request_audit_fields(request),
    )
    return {
        "message": "批次已删除",
        "deleted_task_count": int(task_count or 0),
    }


def _task_payload(row: tuple) -> dict[str, Any]:
    raw_values = json_value(row[27], {}) if len(row) > 27 else {}
    requested_values = json_value(row[32], {}) if len(row) > 32 else {}
    conflict_values = json_value(row[30], {}) if len(row) > 30 else {}
    standard_values = json_value(row[33], {}) if len(row) > 33 else {}
    validation_issues = json_value(row[34], []) if len(row) > 34 else []
    conflict_diff = [
        {
            "field": field,
            "platform": str(requested_values.get(field, "") or ""),
            "tencent": str(conflict_values.get(field, "") or ""),
        }
        for field in sorted(set(requested_values) | set(conflict_values))
        if str(requested_values.get(field, "") or "")
        != str(conflict_values.get(field, "") or "")
    ]
    return {
        "id": int(row[0]), "batch_id": int(row[1]), "source_row": int(row[2]),
        "source_name": str(row[3] or ""), "person_name": str(row[4] or ""),
        "identity_number": str(row[5] or ""), "phone": str(row[6] or ""),
        "original_address": str(row[7] or ""), "created_time": str(row[8] or ""),
        "transfer_note": str(row[9] or ""), "duplicate_group_key": str(row[10] or ""),
        "duplicate_kind": str(row[11] or ""), "suggested_action": str(row[12] or ""),
        "suggested_community_id": int(row[13]) if row[13] else None,
        "suggested_community_name": str(row[14] or ""),
        "suggestion_reason": str(row[15] or ""), "allocation_mode": str(row[16] or ""),
        "final_action": str(row[17] or ""),
        "final_community_id": int(row[18]) if row[18] else None,
        "final_community_name": str(row[19] or ""), "review_note": str(row[20] or ""),
        "reviewer_name": str(row[21] or ""),
        "reviewed_at": row[22].isoformat() + "Z" if row[22] else None,
        "version": int(row[23]), "task_status": str(row[24]),
        "publish_status": str(row[25]), "publish_error": str(row[26] or ""),
        "raw_values": raw_values,
        "field_roles": dispatch_field_roles(raw_values),
        "linked_source_id": int(row[28]) if len(row) > 28 and row[28] else None,
        "linked_row_hash": str(row[29] or "") if len(row) > 29 else "",
        "conflict_values": conflict_values,
        "requested_values": requested_values,
        "conflict_diff": conflict_diff,
        "cache_pending": bool(row[31]) if len(row) > 31 else False,
        "standard_values": standard_values,
        "validation_issues": validation_issues if isinstance(validation_issues, list) else [],
        "business_key_hmac": str(row[35] or "") if len(row) > 35 else "",
        "target_parser": str(row[36] or "全链条") if len(row) > 36 else "全链条",
        "business_type": str(row[37] or "fullchain") if len(row) > 37 else "fullchain",
        "police_subtype": str(row[38] or "") if len(row) > 38 else "",
        "import_profile": str(row[39] or "fullchain_raw") if len(row) > 39 else "fullchain_raw",
    }


TASK_SELECT = """
    SELECT task.id, task.batch_id, task.source_row, task.source_name,
           task.person_name, task.identity_number, task.phone,
           task.original_address, task.source_created_time,
           task.transfer_note, task.duplicate_group_key, task.duplicate_kind,
           task.suggested_action, task.suggested_community_id,
           suggested.name, task.suggestion_reason, task.allocation_mode,
           task.final_action, task.final_community_id, final.name,
           task.review_note, task.reviewer_name, task.reviewed_at,
           task.version, task.task_status, task.publish_status,
           task.publish_error, task.raw_values_json, task.linked_source_id,
           task.linked_row_hash, task.conflict_values_json,
           task.cache_pending, result.request_values_json,
           task.standard_values_json, task.validation_issues_json,
           task.business_key_hmac, batch.target_parser,
           batch.business_type, batch.police_subtype, batch.import_profile
    FROM _police_dispatch_tasks AS task
    JOIN _police_dispatch_batches AS batch ON batch.id=task.batch_id
    LEFT JOIN _communities AS suggested ON suggested.id=task.suggested_community_id
    LEFT JOIN _communities AS final ON final.id=task.final_community_id
    LEFT JOIN _police_dispatch_publish_results AS result ON result.task_id=task.id
"""


def _task_search_where(search: TaskSearch) -> tuple[list[str], list[Any]]:
    where = ["task.batch_id=%s"]
    params: list[Any] = [search.batch_id]
    if search.status == "pending_review":
        where.append("task.task_status='pending_review'")
    elif search.status == "pending_publish":
        where.append("task.publish_status IN ('pending', 'publishing')")
    elif search.status == "retryable":
        where.append("task.publish_status='retryable'")
    elif search.status == "needs_reconciliation":
        where.append("task.publish_status='needs_reconciliation'")
    elif search.status == "conflict":
        where.append("task.publish_status='conflict'")
    elif search.status == "completed":
        where.append("task.task_status='completed'")
    elif search.status != "all":
        raise HTTPException(400, "任务状态筛选无效")
    if search.category != "all":
        if search.category == "duplicate":
            where.append("task.duplicate_group_key<>''")
        elif search.category == "balanced":
            where.append("task.allocation_mode='balanced'")
        elif search.category in {"dispatch", "no_registration", "transfer", "manual"}:
            where.append("COALESCE(NULLIF(task.final_action, ''), task.suggested_action)=%s")
            params.append(search.category)
        else:
            raise HTTPException(400, "任务分类筛选无效")
    if search.keyword.strip():
        where.append("CONCAT_WS(' ', task.person_name, task.identity_number, task.phone, task.original_address) LIKE %s")
        params.append(f"%{search.keyword.strip()}%")
    return where, params


async def _search_tasks(cur, search: TaskSearch) -> dict[str, Any]:
    where, params = _task_search_where(search)
    where_sql = " AND ".join(where)
    await cur.execute(f"SELECT COUNT(*) FROM _police_dispatch_tasks AS task WHERE {where_sql}", params)
    total = int((await cur.fetchone())[0] or 0)
    await cur.execute(
        f"{TASK_SELECT} WHERE {where_sql} ORDER BY task.source_row, task.id LIMIT %s OFFSET %s",
        params + [search.page_size, (search.page - 1) * search.page_size],
    )
    data = [_task_payload(row) for row in await cur.fetchall()]
    return {"data": data, "total": total, "page": search.page, "page_size": search.page_size}


@router.get("/tasks")
async def list_tasks(
    request: Request,
    batch_id: int = Query(...),
    status: str = Query("all", max_length=30),
    category: str = Query("all", max_length=30),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=500),
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    del user
    if "keyword" in request.query_params:
        raise HTTPException(400, "敏感关键词必须通过 POST 搜索接口提交")
    async with conn.cursor() as cur:
        return await _search_tasks(cur, TaskSearch(
            batch_id=batch_id, status=status, category=category,
            page=page, page_size=page_size,
        ))


@router.post("/tasks/search")
async def search_tasks(
    data: TaskSearch,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    del user
    async with conn.cursor() as cur:
        return await _search_tasks(cur, data)


@router.post("/tasks/publishable-selection")
async def publishable_task_selection(
    data: TaskSearch,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    del user
    where, params = _task_search_where(data)
    where.extend([
        "task.final_action='dispatch'",
        "task.task_status='pending_publish'",
        "task.publish_status IN ('pending', 'retryable')",
        "(batch.target_parser='涉警统计' OR TRIM(COALESCE(task.phone, ''))<>'')",
        "(task.duplicate_group_key='' OR NOT EXISTS ("
        "SELECT 1 FROM _police_dispatch_tasks AS grouped "
        "WHERE grouped.batch_id=task.batch_id "
        "AND grouped.duplicate_group_key=task.duplicate_group_key "
        "GROUP BY grouped.duplicate_group_key "
        "HAVING SUM(grouped.final_action<>'duplicate_exclude')<>1))",
    ])
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT task.id FROM _police_dispatch_tasks AS task "
            "JOIN _police_dispatch_batches AS batch ON batch.id=task.batch_id "
            f"WHERE {' AND '.join(where)} ORDER BY task.source_row, task.id LIMIT 5001",
            params,
        )
        task_ids = [int(row[0]) for row in await cur.fetchall()]
    if len(task_ids) > 5000:
        raise HTTPException(409, "当前筛选可发布任务超过 5000 条，请缩小筛选范围后再全选")
    return {"task_ids": task_ids, "total": len(task_ids)}


@router.get("/tasks/{task_id}")
async def get_task(
    task_id: int,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    del user
    async with conn.cursor() as cur:
        await cur.execute(f"{TASK_SELECT} WHERE task.id=%s", (task_id,))
        row = await cur.fetchone()
        if not row:
            raise HTTPException(404, "任务不存在")
        task = _task_payload(row)
        siblings = []
        if task["duplicate_group_key"]:
            await cur.execute(
                f"{TASK_SELECT} WHERE task.batch_id=%s AND task.duplicate_group_key=%s ORDER BY task.source_row",
                (task["batch_id"], task["duplicate_group_key"]),
            )
            siblings = [_task_payload(item) for item in await cur.fetchall()]
        communities = await _communities(cur)
    duplicate_differences = []
    if len(siblings) > 1:
        headers = sorted({
            header
            for sibling in siblings
            for header in sibling.get("raw_values", {})
        })
        differing_headers = [
            header for header in headers
            if len({str(item.get("raw_values", {}).get(header, "")) for item in siblings}) > 1
        ]
        duplicate_differences = [
            {
                "task_id": item["id"],
                "source_row": item["source_row"],
                "fields": [
                    {"field": header, "value": str(item.get("raw_values", {}).get(header, ""))}
                    for header in differing_headers
                ],
            }
            for item in siblings
        ]
    return {
        "task": task,
        "duplicates": siblings,
        "duplicate_differences": duplicate_differences,
        "communities": communities,
    }


async def _review_one(cur, task_id: int, data: TaskReview, user: dict) -> int:
    await cur.execute("""
        SELECT batch_id, version, publish_status, duplicate_group_key,
               person_name, identity_number, phone, original_address,
               batch.target_parser, task.standard_values_json,
               task.validation_issues_json
        FROM _police_dispatch_tasks AS task
        JOIN _police_dispatch_batches AS batch ON batch.id=task.batch_id
        WHERE task.id=%s FOR UPDATE
    """, (task_id,))
    row = await cur.fetchone()
    if not row:
        raise HTTPException(404, "任务不存在")
    if int(row[1]) != data.expected_version:
        raise HTTPException(409, "任务已被其他人修改，请刷新后重试")
    if str(row[2]) in {"success", "publishing", "needs_reconciliation", "conflict"}:
        raise HTTPException(409, "任务已写入或正在对账，不能直接重新审核")
    if data.final_action == "duplicate_exclude" and not str(row[3] or ""):
        raise HTTPException(400, "只有同批重复人员记录才能标记为重复排除")
    community_id = data.final_community_id if data.final_action == "dispatch" else None
    if data.final_action == "dispatch":
        target_parser = str(row[8] or "全链条") if len(row) > 8 else "全链条"
        standard_values = json_value(row[9], {}) if len(row) > 9 else {}
        validation_issues = json_value(row[10], []) if len(row) > 10 else []
        if target_parser == "涉警统计":
            required_values = (
                (standard_values.get("序号", ""), "接警编号"),
                (standard_values.get("简要警情及处理结果", ""), "简要警情及处理结果"),
            )
        else:
            required_values = (
                (row[4] if len(row) > 4 else "", "姓名"),
                (row[5] if len(row) > 5 else "", "身份证号"),
                (row[6] if len(row) > 6 else "", "手机号"),
                (row[7] if len(row) > 7 else "", "地址"),
            )
        missing = [label for value, label in required_values if not str(value or "").strip()]
        if missing:
            raise HTTPException(400, f"下发前必须补齐：{'、'.join(missing)}")
        blocking_issues = {
            str(issue.get("type") or "")
            for issue in validation_issues
            if isinstance(issue, dict)
        } & {"missing_key", "identity_invalid"}
        if blocking_issues:
            raise HTTPException(400, "文件中的业务主键或身份证格式异常，请修正原文件后重新导入")
        if not community_id:
            raise HTTPException(400, "下发任务必须选择社区")
        await _assert_community(cur, community_id, require_enabled=True)
    task_status = "pending_publish" if data.final_action == "dispatch" else "completed"
    publish_status = "pending" if data.final_action == "dispatch" else "not_required"
    reviewer_name = str(user.get("display_name") or (user.get("member") or {}).get("name") or user.get("username") or "")
    await cur.execute("""
        UPDATE _police_dispatch_tasks SET
            final_action=%s, final_community_id=%s, review_note=%s,
            reviewed_by=%s, reviewer_name=%s, reviewed_at=UTC_TIMESTAMP(),
            version=version+1, task_status=%s, publish_status=%s,
            publish_error=''
        WHERE id=%s AND version=%s
    """, (
        data.final_action, community_id, data.review_note.strip(), user["id"],
        reviewer_name[:100], task_status, publish_status, task_id, data.expected_version,
    ))
    if cur.rowcount != 1:
        raise HTTPException(409, "任务已被其他人修改，请刷新后重试")
    return int(row[0])


async def _recalculate_batch_tasks(cur, batch_id: int, edited_task_id: int) -> set[int]:
    """重新计算整批建议、平均分配和重复关系，并清除受影响审核。"""
    await cur.execute(
        "SELECT import_mode FROM _police_dispatch_batches WHERE id=%s",
        (batch_id,),
    )
    batch_mode_row = await cur.fetchone()
    import_mode = str(batch_mode_row[0] or "raw") if batch_mode_row else "raw"
    await cur.execute("""
        SELECT id, source_row, source_name, person_name, identity_number,
               phone, original_address, source_created_time, transfer_note,
               raw_values_json, duplicate_group_key, duplicate_kind,
               suggested_action, suggested_community_id, suggestion_reason,
               allocation_mode, final_action, publish_status
        FROM _police_dispatch_tasks
        WHERE batch_id=%s ORDER BY source_row, id FOR UPDATE
    """, (batch_id,))
    loaded = await cur.fetchall()
    rows: list[dict[str, Any]] = []
    originals: dict[int, tuple[Any, ...]] = {}
    for item in loaded:
        task_id = int(item[0])
        row = {
            "id": task_id,
            "source_row": int(item[1]),
            "source_name": str(item[2] or ""),
            "person_name": str(item[3] or ""),
            "identity_number": str(item[4] or ""),
            "phone": str(item[5] or ""),
            "original_address": str(item[6] or ""),
            "created_time": str(item[7] or ""),
            "transfer_note": str(item[8] or ""),
            "raw_values": json_value(item[9], {}),
            "duplicate_group_key": "",
            "duplicate_kind": "",
        }
        extracted = dispatch_values_from_raw(row["raw_values"])
        row["community_name"] = extracted.get("community_name", "")
        row["registration_status"] = extracted.get("registration_status", "")
        rows.append(row)
        originals[task_id] = (
            str(item[10] or ""), str(item[11] or ""), str(item[12] or ""),
            int(item[13]) if item[13] else None, str(item[14] or ""),
            str(item[15] or ""), str(item[16] or ""), str(item[17] or ""),
        )
    communities = await _communities(cur)
    addresses = await _address_entries(cur, enabled_only=True)
    if import_mode == "clean":
        apply_clean_import_actions(rows, communities)
    else:
        apply_preprocessing_suggestions(rows, communities, addresses)
    affected: set[int] = {edited_task_id}
    for row in rows:
        task_id = int(row["id"])
        previous = originals[task_id]
        derived = (
            str(row.get("duplicate_group_key") or ""),
            str(row.get("duplicate_kind") or ""),
            str(row.get("suggested_action") or ""),
            int(row["suggested_community_id"]) if row.get("suggested_community_id") else None,
            str(row.get("suggestion_reason") or ""),
            str(row.get("allocation_mode") or ""),
        )
        if derived != previous[:6]:
            affected.add(task_id)
        published = previous[7] == "success"
        clear_review = task_id in affected and not published
        await cur.execute("""
            UPDATE _police_dispatch_tasks SET
                identity_hash=%s, duplicate_group_key=%s, duplicate_kind=%s,
                suggested_action=%s, suggested_community_id=%s,
                suggestion_reason=%s, allocation_mode=%s,
                final_action=CASE WHEN %s THEN '' ELSE final_action END,
                final_community_id=CASE WHEN %s THEN NULL ELSE final_community_id END,
                review_note=CASE WHEN %s THEN '' ELSE review_note END,
                reviewed_by=CASE WHEN %s THEN NULL ELSE reviewed_by END,
                reviewer_name=CASE WHEN %s THEN '' ELSE reviewer_name END,
                reviewed_at=CASE WHEN %s THEN NULL ELSE reviewed_at END,
                task_status=CASE WHEN %s THEN 'pending_review' ELSE task_status END,
                publish_status=CASE WHEN %s THEN 'not_required' ELSE publish_status END,
                publish_error=CASE WHEN %s THEN '' ELSE publish_error END,
                version=version+CASE WHEN %s THEN 1 ELSE 0 END
            WHERE id=%s
        """, (
            identity_digest(row.get("identity_number", "")),
            derived[0], derived[1], derived[2], derived[3], derived[4], derived[5],
            clear_review, clear_review, clear_review, clear_review, clear_review,
            clear_review, clear_review, clear_review, clear_review, clear_review,
            task_id,
        ))
    return affected


@router.patch("/tasks/{task_id}/business-fields")
async def update_task_business_fields(
    task_id: int,
    data: TaskBusinessFieldsUpdate,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    if not data.fields:
        raise HTTPException(400, "没有提交需要修改的业务字段")
    if any(len(value) > 5000 for value in data.fields.values()):
        raise HTTPException(400, "单个业务字段不能超过 5000 个字符")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("""
                SELECT task.batch_id, task.version, task.publish_status,
                       task.raw_values_json, batch.target_parser
                FROM _police_dispatch_tasks AS task
                JOIN _police_dispatch_batches AS batch ON batch.id=task.batch_id
                WHERE task.id=%s FOR UPDATE
            """, (task_id,))
            row = await cur.fetchone()
            if not row:
                raise HTTPException(404, "任务不存在")
            if int(row[1]) != data.expected_version:
                raise HTTPException(409, "任务已被其他人修改，请刷新后重试")
            if str(row[2]) in {"success", "publishing", "needs_reconciliation", "conflict"}:
                raise HTTPException(409, "已写入或正在对账的任务不能修改业务字段")
            target_parser = str(row[4] or "全链条") if len(row) > 4 else "全链条"
            if target_parser != "全链条":
                raise HTTPException(409, "当前业务请在对应任务表单中维护，不能修改导入原始字段")
            raw_values = json_value(row[3], {})
            unknown = sorted(set(data.fields) - set(raw_values))
            if unknown:
                raise HTTPException(400, f"不能修改不存在的导入字段：{'、'.join(unknown[:5])}")
            changed = {
                field: value.strip()
                for field, value in data.fields.items()
                if str(raw_values.get(field, "")) != value.strip()
            }
            if not changed:
                await conn.rollback()
                return {"message": "业务字段没有变化", "version": data.expected_version}
            raw_values.update(changed)
            extracted = dispatch_values_from_raw(raw_values)
            await cur.execute("""
                UPDATE _police_dispatch_tasks SET
                    source_name=%s, person_name=%s, identity_number=%s,
                    identity_hash=%s, phone=%s, original_address=%s,
                    source_created_time=%s, transfer_note=%s,
                    raw_values_json=%s
                WHERE id=%s AND version=%s
            """, (
                extracted["source_name"], extracted["person_name"],
                extracted["identity_number"], identity_digest(extracted["identity_number"]),
                extracted["phone"], extracted["original_address"],
                extracted["created_time"], extracted["transfer_note"],
                stable_json(extracted["raw_values"]), task_id, data.expected_version,
            ))
            if cur.rowcount != 1:
                raise HTTPException(409, "任务已被其他人修改，请刷新后重试")
            batch_id = int(row[0])
            affected = await _recalculate_batch_tasks(cur, batch_id, task_id)
            await _refresh_batch_status(cur, batch_id)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "police_dispatch.business_fields.update",
        target_type="police_dispatch_task", target_name=str(task_id),
        detail={
            "changed_fields": sorted(changed),
            "change_digest": sha256(stable_json(changed).encode("utf-8")).hexdigest(),
            "affected_count": len(affected),
        },
        **request_audit_fields(request),
    )
    return {
        "message": "业务字段已保存，相关建议和审核状态已重新计算",
        "affected_count": len(affected),
    }


@router.patch("/tasks/{task_id}")
async def review_task(
    task_id: int,
    data: TaskReview,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            batch_id = await _review_one(cur, task_id, data, user)
            await _refresh_batch_status(cur, batch_id)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    audit_id = await record_admin_audit(
        user, "police_dispatch.review", target_type="police_dispatch_task",
        target_name=str(task_id), detail={"action": data.final_action, "batch_id": batch_id},
        **request_audit_fields(request),
    )
    await record_work_activity(
        user,
        POLICE_DISPATCH_REVIEW,
        event_key=f"admin-audit:{audit_id}",
    )
    return {"message": "审核结果已保存", "version": data.expected_version + 1}


@router.post("/tasks/{keep_task_id}/resolve-duplicate")
async def resolve_duplicate_group(
    keep_task_id: int,
    data: DuplicateGroupResolution,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    versions = {
        int(item.get("id") or 0): int(item.get("version") or 0)
        for item in data.tasks
    }
    if len(versions) != len(data.tasks) or any(
        task_id <= 0 or version <= 0 for task_id, version in versions.items()
    ):
        raise HTTPException(400, "重复任务版本列表无效")
    if keep_task_id not in versions:
        raise HTTPException(400, "保留任务不在当前重复组中")

    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("""
                SELECT batch_id, duplicate_group_key
                FROM _police_dispatch_tasks
                WHERE id=%s FOR UPDATE
            """, (keep_task_id,))
            anchor = await cur.fetchone()
            if not anchor:
                raise HTTPException(404, "任务不存在")
            batch_id = int(anchor[0])
            duplicate_group_key = str(anchor[1] or "")
            if not duplicate_group_key:
                raise HTTPException(400, "该任务不属于重复人员组")

            await cur.execute("""
                SELECT id, version, final_community_id, suggested_community_id
                FROM _police_dispatch_tasks
                WHERE batch_id=%s AND duplicate_group_key=%s
                ORDER BY id FOR UPDATE
            """, (batch_id, duplicate_group_key))
            group_rows = await cur.fetchall()
            group_ids = {int(row[0]) for row in group_rows}
            if len(group_rows) < 2 or group_ids != set(versions):
                raise HTTPException(409, "重复组已经变化，请刷新后重新选择")
            if any(int(row[1]) != versions[int(row[0])] for row in group_rows):
                raise HTTPException(409, "重复任务已被其他人修改，请刷新后重试")

            keep_row = next(row for row in group_rows if int(row[0]) == keep_task_id)
            community_id = (
                int(keep_row[2]) if keep_row[2]
                else int(keep_row[3]) if keep_row[3]
                else None
            )
            await _review_one(cur, keep_task_id, TaskReview(
                expected_version=versions[keep_task_id],
                final_action="dispatch",
                final_community_id=community_id,
                review_note=data.review_note,
            ), user)
            for row in group_rows:
                task_id = int(row[0])
                if task_id == keep_task_id:
                    continue
                await _review_one(cur, task_id, TaskReview(
                    expected_version=versions[task_id],
                    final_action="duplicate_exclude",
                    review_note=data.review_note,
                ), user)
            await _refresh_batch_status(cur, batch_id)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise

    excluded_count = len(group_rows) - 1
    audit_id = await record_admin_audit(
        user, "police_dispatch.duplicate.resolve",
        target_type="police_dispatch_batch", target_name=str(batch_id),
        detail={
            "keep_task_id": keep_task_id,
            "excluded_count": excluded_count,
            "group_size": len(group_rows),
        },
        **request_audit_fields(request),
    )
    await record_work_activity(
        user,
        POLICE_DISPATCH_REVIEW,
        event_key=f"admin-audit:{audit_id}",
        units=len(group_rows),
    )
    return {
        "message": f"已保留 1 条并排除 {excluded_count} 条重复任务",
        "keep_task_id": keep_task_id,
        "excluded_count": excluded_count,
    }


@router.post("/tasks/bulk-review")
async def bulk_review_tasks(
    data: BulkReview,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    ids = [int(item.get("id") or 0) for item in data.tasks]
    versions = {int(item.get("id") or 0): int(item.get("version") or 0) for item in data.tasks}
    if len(set(ids)) != len(ids) or any(item <= 0 for item in ids):
        raise HTTPException(400, "批量任务列表无效")
    placeholders = ",".join(["%s"] * len(ids))
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                f"SELECT id, batch_id, version, suggested_action, suggested_community_id "
                f"FROM _police_dispatch_tasks WHERE id IN ({placeholders}) FOR UPDATE",
                ids,
            )
            rows = await cur.fetchall()
            if len(rows) != len(ids):
                raise HTTPException(404, "部分任务不存在")
            batch_ids = {int(row[1]) for row in rows}
            if len(batch_ids) != 1:
                raise HTTPException(400, "一次只能批量审核同一批次")
            for row in rows:
                task_id, _, version, suggested_action, suggested_community_id = row
                if int(version) != versions[int(task_id)]:
                    raise HTTPException(409, "部分任务已被其他人修改，请刷新后重试")
                action = str(suggested_action) if data.mode == "accept_suggestion" else data.final_action
                community_id = int(suggested_community_id) if suggested_community_id else None
                if data.mode == "set_action":
                    community_id = data.final_community_id
                if action not in FINAL_ACTIONS:
                    raise HTTPException(400, "筛选结果中含有必须逐条人工判断的任务")
                await _review_one(cur, int(task_id), TaskReview(
                    expected_version=int(version), final_action=action,
                    final_community_id=community_id, review_note=data.review_note,
                ), user)
            batch_id = next(iter(batch_ids))
            await _refresh_batch_status(cur, batch_id)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    audit_id = await record_admin_audit(
        user, "police_dispatch.bulk_review", target_type="police_dispatch_batch",
        target_name=str(batch_id), detail={"count": len(ids), "mode": data.mode},
        **request_audit_fields(request),
    )
    await record_work_activity(
        user,
        POLICE_DISPATCH_REVIEW,
        event_key=f"admin-audit:{audit_id}",
        units=len(ids),
    )
    return {"message": f"已审核 {len(ids)} 条任务", "count": len(ids)}


@router.post("/tasks/{task_id}/resolve-conflict")
async def resolve_publish_conflict(
    task_id: int,
    data: ConflictResolution,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    async with conn.cursor() as cur:
        await cur.execute("""
            SELECT task.batch_id, task.version, task.publish_status,
                   task.linked_source_id, task.linked_row_hash,
                   task.raw_values_json, result.request_values_json,
                   result.spreadsheet_id, result.physical_row,
                   batch.target_parser, task.standard_values_json,
                   task.business_key_hmac
            FROM _police_dispatch_tasks AS task
            JOIN _police_dispatch_batches AS batch ON batch.id=task.batch_id
            JOIN _police_dispatch_publish_results AS result
              ON result.task_id=task.id
            WHERE task.id=%s
        """, (task_id,))
        row = await cur.fetchone()
        if not row:
            raise HTTPException(404, "冲突任务不存在")
        if int(row[1]) != data.expected_version:
            raise HTTPException(409, "任务已变化，请刷新冲突详情")
        if str(row[2]) != "conflict":
            raise HTTPException(409, "任务当前不处于内容冲突状态")
        if str(row[4] or "") != data.expected_row_hash:
            raise HTTPException(
                409,
                "本地来源记录已变化，请刷新后重试"
                if local_data_source_enabled()
                else "腾讯来源行已变化，请等待同步后刷新",
            )
        if not row[3]:
            raise HTTPException(
                409,
                "尚未定位本地来源记录，请刷新后重试"
                if local_data_source_enabled()
                else "尚未定位腾讯来源行，请等待一次正常同步",
            )
        parser = get_parser(str(row[9] or "全链条"))
        source = await _load_source_row(cur, parser.parser_type, int(row[3]))
        if source["row_hash"] != data.expected_row_hash:
            raise HTTPException(
                409,
                "本地来源记录已变化，请刷新后重新选择"
                if local_data_source_enabled()
                else "腾讯来源行已变化，请刷新后重新选择",
            )
        batch_id = int(row[0])
        platform_values = json_value(row[6], {})

    if data.strategy == "adopt_tencent":
        values = {
            column: str(source["values"].get(column, "") or "").strip()
            for column in parser.COLUMNS
        }
        await conn.begin()
        try:
            async with conn.cursor() as cur:
                await cur.execute(
                    "SELECT version, linked_row_hash FROM _police_dispatch_tasks "
                    "WHERE id=%s FOR UPDATE",
                    (task_id,),
                )
                locked = await cur.fetchone()
                if not locked or int(locked[0]) != data.expected_version \
                        or str(locked[1] or "") != data.expected_row_hash:
                    raise HTTPException(409, "冲突任务已变化，请刷新后重试")
                communities = await _communities(cur)
                community = resolve_community(
                    values.get("社区", ""), community_resolver(communities),
                )
                if not community or not community.get("enabled", False):
                    raise HTTPException(
                        409,
                        "本地现有记录中的社区无法映射为启用中的正式社区"
                        if local_data_source_enabled()
                        else "腾讯内容中的社区无法映射为启用中的正式社区",
                    )
                raw_values = json_value(row[5], {})
                if parser.parser_type == "全链条":
                    roles = dispatch_field_roles(raw_values)
                    replacements = {
                        "source": values.get("来源", ""),
                        "name": values.get("姓名", ""),
                        "identity": values.get("身份证号", ""),
                        "phone": values.get("电话号码", ""),
                        "address": values.get("地址", ""),
                        "created": values.get("创建时间", ""),
                        "transfer_note": "",
                    }
                    for field, value in replacements.items():
                        header = roles.get(field)
                        if header:
                            raw_values[header] = value
                else:
                    replacements = {
                        "source": str(values.get("来源") or parser.parser_type),
                        "name": str(values.get("姓名") or ""),
                        "identity": str(values.get("身份证号") or values.get("身份证号码") or ""),
                        "phone": str(values.get("电话号码") or values.get("手机号码") or values.get("联系号码") or ""),
                        "address": str(values.get("地址") or values.get("地址1") or values.get("疑似现住址") or values.get("高频抓拍小区") or ""),
                        "created": str(values.get("创建时间") or values.get("下发日期") or values.get("下发时间") or values.get("日期") or ""),
                        "transfer_note": str(values.get("出警内容") or ""),
                    }
                await cur.execute("""
                    UPDATE _police_dispatch_tasks SET
                        source_name=%s, person_name=%s, identity_number=%s,
                        identity_hash=%s, phone=%s, original_address=%s,
                        source_created_time=%s, transfer_note=%s, raw_values_json=%s,
                        standard_values_json=%s,
                        final_action='dispatch', final_community_id=%s,
                        suggested_action='dispatch', suggested_community_id=%s,
                        suggestion_reason=%s, allocation_mode='matched',
                        publish_status='success', task_status='completed',
                        publish_error='', conflict_values_json=NULL,
                        cache_pending=0, published_at=COALESCE(published_at, UTC_TIMESTAMP()),
                        version=version+1
                    WHERE id=%s AND version=%s
                """, (
                    replacements["source"], replacements["name"],
                    replacements["identity"], identity_digest(replacements["identity"]),
                    replacements["phone"], replacements["address"],
                    replacements["created"], replacements["transfer_note"],
                    stable_json(raw_values), stable_json(values),
                    "已采用本地现有内容" if local_data_source_enabled() else "已采用腾讯现有内容",
                    community["id"],
                    community["id"], task_id, data.expected_version,
                ))
                if cur.rowcount != 1:
                    raise HTTPException(409, "任务已变化，请刷新后重试")
                await cur.execute("""
                    UPDATE _police_dispatch_publish_results SET
                        status='success', resolution='adopt_tencent',
                        verified_values_json=%s, expected_row_hash=%s,
                        error_code='', error_message='', cache_pending=0
                    WHERE task_id=%s
                """, (stable_json(values), data.expected_row_hash, task_id))
                if parser.parser_type == "全链条":
                    await _recalculate_batch_tasks(cur, batch_id, task_id)
                await _refresh_batch_status(cur, batch_id)
            await conn.commit()
        except Exception:
            await conn.rollback()
            raise
    else:
        if local_data_source_enabled():
            raise HTTPException(
                409,
                "本地数据源不支持覆盖已有业务记录，请修改或撤回当前任务后重新发布",
            )
        if data.confirmation != "覆盖腾讯内容":
            raise HTTPException(400, "请输入“覆盖腾讯内容”完成二次确认")
        async with conn.cursor() as cur:
            if not await _writeback_enabled(cur):
                raise HTTPException(503, "在线回写已由超级管理员暂停")
            spreadsheets = await _enabled_spreadsheets(cur, parser.parser_type)
            spreadsheet = next(
                (item for item in spreadsheets if item["id"] == int(row[7])),
                None,
            )
            if not spreadsheet:
                raise HTTPException(409, "原腾讯来源表已停用，不能覆盖")
            if not await acquire_sheet_lock(cur, spreadsheet["id"], timeout=2):
                raise HTTPException(409, f"{parser.parser_type}表格正在同步或被他人编辑，请稍后重试")
        client = None
        cache_pending = False
        request_sent = False
        requested: dict[str, str] = {}
        try:
            async with conn.cursor() as cur:
                client = await _oauth_client(cur)
            source_columns = await resolve_source_columns(client, spreadsheet, parser)
            comparison_columns = _publish_comparison_columns(parser, source_columns)
            live = await client.read_source_row(
                spreadsheet["file_id"], spreadsheet["data_sheet_id"],
                int(row[8]), source_columns,
            )
            live_values = parser.normalize_source_row(live["values"])
            if source_row_hash(live_values) != data.expected_row_hash:
                raise HTTPException(409, "腾讯行在确认后再次变化，请刷新冲突详情")
            requested = _publish_request_values(platform_values, comparison_columns)
            request_sent = True
            await client.batch_update(
                spreadsheet["file_id"],
                [client.build_update_range_request(
                    spreadsheet["data_sheet_id"], int(row[8]) - 1, 0,
                    [parser.source_row_values(
                        {**live["values"], **requested}, source_columns
                    )],
                )],
            )
            verified = await client.read_source_row(
                spreadsheet["file_id"], spreadsheet["data_sheet_id"],
                int(row[8]), source_columns,
            )
            verified["values"] = parser.normalize_source_row(verified["values"])
            if not _row_values_match(
                requested, verified["values"], verified.get("cell_meta") or {},
                comparison_columns,
            ):
                raise HTTPException(502, "腾讯覆盖后回读不一致")
            verified_values = {
                column: str(verified["values"].get(column, "") or "").strip()
                for column in parser.COLUMNS
            }
            verified_hash = source_row_hash(verified_values)
            try:
                await _refresh_spreadsheet(conn, client, spreadsheet)
            except Exception:
                cache_pending = True
            await conn.begin()
            try:
                async with conn.cursor() as cur:
                    await cur.execute("""
                        UPDATE _police_dispatch_tasks SET
                            publish_status='success', task_status='completed',
                            publish_error='', linked_row_hash=%s,
                            conflict_values_json=NULL, cache_pending=%s,
                            published_at=COALESCE(published_at, UTC_TIMESTAMP()),
                            version=version+1
                        WHERE id=%s AND version=%s AND publish_status='conflict'
                    """, (
                        verified_hash, 1 if cache_pending else 0,
                        task_id, data.expected_version,
                    ))
                    if cur.rowcount != 1:
                        raise HTTPException(409, "任务已变化，请刷新后重试")
                    await cur.execute("""
                        UPDATE _police_dispatch_publish_results SET
                            status='success', resolution='overwrite_tencent',
                            verified_values_json=%s, expected_row_hash=%s,
                            error_code='', error_message='', cache_pending=%s
                        WHERE task_id=%s
                    """, (
                        stable_json(verified_values), verified_hash,
                        1 if cache_pending else 0, task_id,
                    ))
                    await _refresh_batch_status(cur, batch_id)
                await conn.commit()
            except Exception:
                await conn.rollback()
                raise
        except Exception as exc:
            if request_sent and requested:
                safe_error = "腾讯覆盖请求结果尚未确认，等待下次正常同步对账"
                async with conn.cursor() as cur:
                    marked_for_reconciliation = await _mark_overwrite_uncertain(
                        cur,
                        task_id=task_id,
                        batch_id=batch_id,
                        spreadsheet=spreadsheet,
                        physical_row=int(row[8]),
                        requested=requested,
                        business_key=str(row[11] or "") or parser_business_key(parser, platform_values),
                        error=safe_error,
                    )
                if marked_for_reconciliation:
                    raise HTTPException(502, safe_error) from exc
            raise
        finally:
            try:
                if client:
                    await client.close()
            finally:
                async with conn.cursor() as cur:
                    await release_sheet_lock(cur, spreadsheet["id"])

    changed_fields = sorted({
        column for column in parser.COLUMNS
        if str(platform_values.get(column, "") or "")
        != str(source["values"].get(column, "") or "")
    })
    audit_action = (
        "police_dispatch.conflict.adopt_local"
        if local_data_source_enabled() and data.strategy == "adopt_tencent"
        else f"police_dispatch.conflict.{data.strategy}"
    )
    await record_admin_audit(
        user, audit_action,
        target_type="police_dispatch_task", target_name=str(task_id),
        detail={
            "changed_fields": changed_fields,
            "row_hash": data.expected_row_hash,
        },
        **request_audit_fields(request),
    )
    return {
        "message": (
            "已采用本地现有内容"
            if local_data_source_enabled() and data.strategy == "adopt_tencent"
            else "已采用腾讯内容"
            if data.strategy == "adopt_tencent"
            else "已用平台内容覆盖腾讯现有行"
        ),
        "cache_pending": cache_pending if data.strategy == "overwrite_tencent" else False,
    }


@router.get("/workbench/home")
async def workbench_home(
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    del user
    async with conn.cursor() as cur:
        await cur.execute("""
            SELECT id FROM _police_dispatch_batches
            WHERE status<>'completed'
            ORDER BY created_at, id LIMIT 1
        """)
        active_row = await cur.fetchone()
        await cur.execute("""
            SELECT id FROM _police_dispatch_batches
            ORDER BY created_at DESC, id DESC LIMIT 8
        """)
        ids = [int(row[0]) for row in await cur.fetchall()]
        if active_row and int(active_row[0]) not in ids:
            ids.insert(0, int(active_row[0]))
        batches = await _batch_payloads(cur, ids)
        communities = await _communities(cur)
    active_id = int(active_row[0]) if active_row else None
    active = next((item for item in batches if item["id"] == active_id), None)
    if active is None and batches:
        active = batches[0]
    return {"active_batch": active, "batches": batches, "communities": communities}


@router.get("/batches/{batch_id}/feedback.xlsx")
async def export_feedback(
    batch_id: int,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    async with conn.cursor() as cur:
        batch = await _batch_payload(cur, batch_id)
        target_parser = str(batch.get("target_parser") or "全链条")
        action_filter = (
            "AND final_action IN ('no_registration', 'transfer')"
            if target_parser == "全链条" else ""
        )
        await cur.execute(f"""
            SELECT source_row, source_name, person_name, identity_number,
                   phone, original_address, final_action, review_note,
                   suggestion_reason, reviewer_name, reviewed_at,
                   standard_values_json, validation_issues_json,
                   task_status, publish_status, publish_error
            FROM _police_dispatch_tasks
            WHERE batch_id=%s {action_filter}
            ORDER BY source_row
        """, (batch_id,))
        tasks = [
            {
                "source_row": row[0], "source_name": row[1], "person_name": row[2],
                "identity_number": row[3], "phone": row[4], "original_address": row[5],
                "final_action": row[6], "review_note": row[7], "suggestion_reason": row[8],
                "reviewer_name": row[9],
                "reviewed_at_text": row[10].strftime("%Y-%m-%d %H:%M:%S") if row[10] else "",
                "standard_values": json_value(row[11], {}),
                "validation_issues": json_value(row[12], []),
                "task_status": str(row[13] or ""),
                "publish_status": str(row[14] or ""),
                "publish_error": str(row[15] or ""),
            }
            for row in await cur.fetchall()
        ]
    content = await asyncio.to_thread(build_feedback_workbook, batch, tasks, datetime.now())
    filename = f"下发批次-{batch_id}-反馈.xlsx"
    await record_admin_audit(
        user,
        "police_dispatch.feedback.export",
        target_type="police_dispatch_batch",
        target_name=str(batch_id),
        detail={
            "batch_id": batch_id,
            "file_format": "XLSX",
            "row_count": len(tasks),
        },
        **request_audit_fields(request),
    )
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def _publish_values(task: dict[str, Any], community: str, publish_date) -> dict[str, str]:
    standard = task.get("standard_values") or {}
    if standard:
        values = {str(key): str(value or "") for key, value in standard.items()}
        if "社区" in values:
            values["社区"] = community
        if "地址" in values:
            values["地址"] = build_publish_address(
                task.get("original_address") or values.get("地址", ""),
                task.get("transfer_note", ""),
            )
        if not values.get("下发日期") and "下发日期" in values:
            values["下发日期"] = publish_date.strftime("%m-%d")
        if not values.get("下发时间") and "下发时间" in values:
            values["下发时间"] = publish_date.strftime("%m-%d")
        return values
    deadline = publish_date + timedelta(days=3)
    return {
        "下发日期": publish_date.strftime("%m-%d"),
        "截止日期": deadline.strftime("%m-%d"),
        "核查人": "",
        "社区": community,
        "来源": task["source_name"],
        "姓名": task["person_name"],
        "身份证号": task["identity_number"],
        "电话号码": task["phone"],
        "地址": build_publish_address(task["original_address"], task["transfer_note"]),
        "登记情况": task.get("registration_status", ""),
        "创建时间": task["created_time"],
        "现住址": "", "核查结果": "", "研判": "", "二次反馈": "",
    }


def _publish_comparison_columns(parser, source_columns: list[str]) -> list[str]:
    """只校验腾讯物理表实际存在且由下发流程拥有的列。"""
    source_column_set = set(source_columns)
    if parser.parser_type == "全链条":
        owned = set(PUBLISH_OWNED_COLUMNS)
    else:
        owned = set(parser.COLUMNS) - {
            "核查人", "现住址", "核查结果", "核查反馈", "研判",
            "二次反馈", "二次核查结果", "入住方式", "是否开户",
            "房屋属性", "居住时间", "房东信息", "二房东信息", "备注",
            "房东是否处罚",
        }
    return [column for column in parser.COLUMNS if column in source_column_set and column in owned]


def _publish_request_values(
    values: dict[str, Any],
    comparison_columns: list[str],
) -> dict[str, str]:
    """冻结本次物理布局真正写入的值，供失败后的只读对账使用。"""
    return {
        column: str(values.get(column, "") or "").strip()
        for column in comparison_columns
    }


async def _save_publish_result(
    cur,
    *,
    task_id: int,
    spreadsheet: dict[str, Any],
    business_key: str,
    request_values: dict[str, str],
    status: str,
    physical_row: int | None = None,
    verified_values: dict[str, str] | None = None,
    row_hash: str = "",
    error_code: str = "",
    error_message: str = "",
    cache_pending: bool = False,
    count_attempt: bool = True,
) -> None:
    await cur.execute("""
        INSERT INTO _police_dispatch_publish_results (
            task_id, spreadsheet_id, sheet_id, physical_row,
            business_key, request_values_json, verified_values_json,
            expected_row_hash, cache_pending, status, error_code,
            error_message, attempt_count, last_attempt_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                  UTC_TIMESTAMP())
        ON DUPLICATE KEY UPDATE
            spreadsheet_id=VALUES(spreadsheet_id), sheet_id=VALUES(sheet_id),
            physical_row=VALUES(physical_row), business_key=VALUES(business_key),
            request_values_json=VALUES(request_values_json),
            verified_values_json=VALUES(verified_values_json),
            expected_row_hash=VALUES(expected_row_hash),
            cache_pending=VALUES(cache_pending), status=VALUES(status),
            error_code=VALUES(error_code), error_message=VALUES(error_message),
            attempt_count=attempt_count+VALUES(attempt_count),
            last_attempt_at=UTC_TIMESTAMP()
    """, (
        task_id, spreadsheet["id"], spreadsheet["data_sheet_id"], physical_row,
        business_key, stable_json(request_values),
        stable_json(verified_values) if verified_values is not None else None,
        row_hash, 1 if cache_pending else 0, status, error_code,
        error_message[:500], 1 if count_attempt else 0,
    ))


async def _set_publish_run_phase(
    cur,
    run_id: int | None,
    phase: str,
) -> None:
    if run_id is None:
        return
    await cur.execute(
        "UPDATE _police_dispatch_publish_runs SET phase=%s WHERE id=%s",
        (phase, run_id),
    )


async def _set_publish_run_item(
    cur,
    run_id: int | None,
    task_id: int,
    status: str,
    *,
    physical_row: int | None = None,
    error_code: str = "",
) -> None:
    if run_id is None:
        return
    await cur.execute("""
        UPDATE _police_dispatch_publish_run_items
        SET status=%s,physical_row=COALESCE(%s,physical_row),error_code=%s
        WHERE run_id=%s AND task_id=%s
    """, (status, physical_row, error_code[:100], run_id, task_id))


async def _refresh_publish_run_progress(cur, run_id: int | None) -> None:
    if run_id is None:
        return
    await cur.execute("""
        UPDATE _police_dispatch_publish_runs AS run SET
            processed_count=(SELECT COUNT(*) FROM _police_dispatch_publish_run_items
                             WHERE run_id=run.id
                               AND status IN ('success','conflict','needs_reconciliation','retryable')),
            success_count=(SELECT COUNT(*) FROM _police_dispatch_publish_run_items
                           WHERE run_id=run.id AND status='success'),
            conflict_count=(SELECT COUNT(*) FROM _police_dispatch_publish_run_items
                            WHERE run_id=run.id AND status='conflict'),
            reconciliation_count=(SELECT COUNT(*) FROM _police_dispatch_publish_run_items
                                  WHERE run_id=run.id AND status='needs_reconciliation'),
            retryable_count=(SELECT COUNT(*) FROM _police_dispatch_publish_run_items
                             WHERE run_id=run.id AND status='retryable')
        WHERE run.id=%s
    """, (run_id,))


async def _set_task_publish_state(
    cur,
    *,
    task_id: int,
    status: str,
    business_key: str,
    error: str = "",
    physical_row: int | None = None,
    linked_source_id: int | None = None,
    linked_row_hash: str = "",
    conflict_values: dict[str, str] | None = None,
    cache_pending: bool = False,
) -> None:
    task_status = "completed" if status == "success" else "publish_failed"
    if status in {"pending", "publishing", "retryable"}:
        task_status = "pending_publish"
    await cur.execute("""
        UPDATE _police_dispatch_tasks SET
            publish_status=%s, task_status=%s, publish_key=%s,
            publish_error=%s, published_row=COALESCE(%s, published_row),
            linked_source_id=%s, linked_row_hash=%s,
            conflict_values_json=%s, cache_pending=%s,
            published_at=CASE WHEN %s='success'
                THEN COALESCE(published_at, UTC_TIMESTAMP()) ELSE published_at END,
            version=version+1
        WHERE id=%s
    """, (
        status, task_status, business_key, error[:500], physical_row,
        linked_source_id, linked_row_hash,
        stable_json(conflict_values) if conflict_values is not None else None,
        1 if cache_pending else 0, status, task_id,
    ))


async def _mark_overwrite_uncertain(
    cur,
    *,
    task_id: int,
    batch_id: int,
    spreadsheet: dict[str, Any],
    physical_row: int,
    requested: dict[str, str],
    business_key: str = "",
    error: str,
) -> bool:
    """覆盖请求可能已到腾讯时锁定重试，交由正常同步只读对账。"""
    if not business_key:
        business_key = publish_business_key(
            requested.get("身份证号", ""),
            requested.get("电话号码", ""),
            requested.get("下发日期", ""),
        )
    await cur.execute("""
        UPDATE _police_dispatch_tasks SET
            publish_status='needs_reconciliation',
            task_status='publish_failed', publish_error=%s,
            publish_key=%s, published_row=%s,
            conflict_values_json=NULL, cache_pending=0,
            version=version+1
        WHERE id=%s AND publish_status='conflict'
    """, (error, business_key, physical_row, task_id))
    if not cur.rowcount:
        return False
    await _save_publish_result(
        cur,
        task_id=task_id,
        spreadsheet=spreadsheet,
        business_key=business_key,
        request_values=requested,
        status="needs_reconciliation",
        physical_row=physical_row,
        error_code="overwrite_uncertain",
        error_message=error,
    )
    await _refresh_batch_status(cur, batch_id)
    return True


async def _execute_local_publish_selection(
    batch_id: int,
    data: TaskPublishSelection,
    request: Request | None,
    user: dict[str, Any],
    conn,
    *,
    run_id: int | None = None,
) -> dict[str, Any]:
    """把已审核任务直接发布到本地业务表，不经过腾讯写回或回读。"""
    selected_ids = data.task_ids
    placeholders = ",".join(["%s"] * len(selected_ids))
    pseudo_spreadsheet = {
        "id": 0,
        "data_sheet_id": local_sheet_id("local"),
        "parser_type": "local",
    }
    success_count = 0
    failed_count = 0
    counts: dict[str, Any] = {}
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            batch = await _batch_payload(cur, batch_id)
            parser = get_parser(batch.get("target_parser") or "全链条")
            await cur.execute(f"""
                SELECT task.id, task.source_row, task.source_name, task.person_name,
                       task.identity_number, task.phone, task.original_address,
                       task.source_created_time, task.transfer_note,
                       task.raw_values_json, community.name,
                       task.standard_values_json, task.business_key_hmac
                FROM _police_dispatch_tasks AS task
                JOIN _communities AS community ON community.id=task.final_community_id
                WHERE task.batch_id=%s AND task.id IN ({placeholders})
                  AND task.final_action='dispatch' AND task.task_status IN ('pending_publish','publish_failed')
                  AND task.publish_status IN ('pending','publishing','retryable')
                ORDER BY task.source_row, task.id
                FOR UPDATE
            """, [batch_id, *selected_ids])
            rows = await cur.fetchall()
            if len(rows) != len(selected_ids):
                raise HTTPException(409, "部分所选任务已不可发布，请刷新列表后重新选择")
            if run_id is not None:
                await _set_publish_run_phase(cur, run_id, "writing_local")

            publish_date = await get_business_date(cur)
            for row in rows:
                task = {
                    "id": int(row[0]),
                    "source_row": int(row[1]),
                    "source_name": str(row[2] or ""),
                    "person_name": str(row[3] or ""),
                    "identity_number": str(row[4] or ""),
                    "phone": str(row[5] or ""),
                    "original_address": str(row[6] or ""),
                    "created_time": str(row[7] or ""),
                    "transfer_note": str(row[8] or ""),
                    "registration_status": dispatch_values_from_raw(json_value(row[9], {})).get("registration_status", ""),
                    "community": str(row[10] or ""),
                    "standard_values": json_value(row[11], {}),
                    "business_key_hmac": str(row[12] or ""),
                }
                values = _publish_values(task, task["community"], publish_date)
                key = task.get("business_key_hmac") or parser_business_key(parser, values, legacy=True)
                try:
                    local_source = await create_local_source_row(
                        cur,
                        parser.parser_type,
                        values,
                        source_kind="local_dispatch",
                        source_ref=f"police_dispatch_task:{task['id']}",
                    )
                except ValueError as exc:
                    if str(exc) != "local_business_key_conflict":
                        raise
                    await cur.execute(
                        "SELECT source.id, source.values_json, source.row_hash "
                        "FROM _online_source_rows AS source "
                        "WHERE source.spreadsheet_id=0 AND source.parser_type=%s "
                        "AND source.row_key=%s AND source.archived_at IS NULL LIMIT 1 FOR UPDATE",
                        (parser.parser_type, parser.make_row_key(values)),
                    )
                    conflict = await cur.fetchone()
                    conflict_values = json_value(conflict[1], {}) if conflict else {}
                    await _set_task_publish_state(
                        cur, task_id=task["id"], status="conflict", business_key=key,
                        error="本地业务表已存在相同业务主键但内容不同",
                        linked_source_id=int(conflict[0]) if conflict else None,
                        linked_row_hash=str(conflict[2] or "") if conflict else "",
                        conflict_values=conflict_values,
                    )
                    await _save_publish_result(
                        cur, task_id=task["id"], spreadsheet=pseudo_spreadsheet,
                        business_key=key, request_values=values, status="conflict",
                        physical_row=int(conflict[0]) if conflict else None,
                        verified_values=conflict_values,
                        row_hash=str(conflict[2] or "") if conflict else "",
                        error_code="local_content_conflict",
                        error_message="本地业务表已存在相同业务主键但内容不同",
                    )
                    await _set_publish_run_item(
                        cur, run_id, task["id"], "conflict",
                        physical_row=int(conflict[0]) if conflict else None,
                        error_code="local_content_conflict",
                    )
                    failed_count += 1
                    continue

                await _set_task_publish_state(
                    cur, task_id=task["id"], status="success", business_key=key,
                    physical_row=local_source["local_task_id"],
                    linked_source_id=local_source["id"],
                    linked_row_hash=local_source["row_hash"],
                )
                await _save_publish_result(
                    cur, task_id=task["id"], spreadsheet=pseudo_spreadsheet,
                    business_key=key, request_values=values, status="success",
                    physical_row=local_source["local_task_id"],
                    verified_values=local_source["values"],
                    row_hash=local_source["row_hash"],
                )
                await _set_publish_run_item(
                    cur, run_id, task["id"], "success",
                    physical_row=local_source["local_task_id"],
                )
                success_count += 1

            # Local publishing does not have a Tencent sync cycle to rebuild
            # the task projection.  Rebuild it in the same transaction so a
            # successful batch is immediately visible to 指令核查.
            await rebuild_projection(cur, parser.parser_type)
            await _refresh_publish_run_progress(cur, run_id)
            counts = await _refresh_batch_status(cur, batch_id)
            await cur.execute(
                "UPDATE _police_dispatch_batches SET last_error=%s WHERE id=%s",
                ("部分任务存在本地业务键冲突" if failed_count else "", batch_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise

    await record_admin_audit(
        user,
        "police_dispatch.publish_local",
        target_type="police_dispatch_batch",
        target_name=str(batch_id),
        detail={"selected": len(selected_ids), "success": success_count, "failed": failed_count},
        result="partial" if failed_count else "success",
        **(request_audit_fields(request) if request is not None else {}),
    )
    return {
        "message": "本地发布完成" if not failed_count else "部分任务存在本地业务键冲突",
        "success_count": success_count,
        "failed_count": failed_count,
        "counts": counts,
    }


async def _execute_publish_selection(
    batch_id: int,
    data: TaskPublishSelection,
    request: Request | None,
    user: dict,
    conn,
    *,
    run_id: int | None = None,
):
    if local_data_source_enabled():
        return await _execute_local_publish_selection(
            batch_id, data, request, user, conn, run_id=run_id,
        )
    selected_ids = data.task_ids
    placeholders = ",".join(["%s"] * len(selected_ids))
    eligible_publish_status = (
        "task.publish_status='publishing'"
        if run_id is not None
        else "task.publish_status IN ('pending', 'retryable')"
    )
    spreadsheet: dict[str, Any] | None = None
    sheet_lock_acquired = False
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            batch = await _batch_payload(cur, batch_id)
            parser = get_parser(batch.get("target_parser") or "全链条")
            if not await _writeback_enabled(cur):
                raise HTTPException(503, "在线回写已由超级管理员暂停")
            spreadsheets = await _enabled_spreadsheets(cur, parser.parser_type)
            if len(spreadsheets) != 1:
                raise HTTPException(409, f"{parser.parser_type}业务没有唯一启用的腾讯来源表")
            spreadsheet = spreadsheets[0]
            await cur.execute(f"""
                SELECT task.id, task.source_row, task.source_name, task.person_name,
                       task.identity_number, task.phone, task.original_address,
                       task.source_created_time, task.transfer_note,
                       task.raw_values_json, community.name,
                       task.standard_values_json, task.business_key_hmac
                FROM _police_dispatch_tasks AS task
                JOIN _communities AS community ON community.id=task.final_community_id
                WHERE task.batch_id=%s AND task.id IN ({placeholders})
                  AND task.final_action='dispatch' AND task.task_status='pending_publish'
                  AND {eligible_publish_status}
                ORDER BY task.source_row, task.id
                FOR UPDATE
            """, [batch_id, *selected_ids])
            pending = [
                {
                    "id": int(row[0]), "source_row": int(row[1]),
                    "source_name": str(row[2] or ""), "person_name": str(row[3] or ""),
                    "identity_number": str(row[4] or ""), "phone": str(row[5] or ""),
                    "original_address": str(row[6] or ""), "created_time": str(row[7] or ""),
                    "transfer_note": str(row[8] or ""),
                    "registration_status": dispatch_values_from_raw(
                        json_value(row[9], {})
                    ).get("registration_status", ""),
                    "community": str(row[10]),
                    "standard_values": json_value(row[11], {}),
                    "business_key_hmac": str(row[12] or ""),
                }
                for row in await cur.fetchall()
            ]
            if len(pending) != len(selected_ids):
                raise HTTPException(409, "部分所选任务已不可发布，请刷新列表后重新选择")
            await cur.execute(f"""
                SELECT grouped.duplicate_group_key,
                       SUM(grouped.final_action<>'duplicate_exclude') AS kept
                FROM _police_dispatch_tasks AS grouped
                WHERE grouped.batch_id=%s
                  AND grouped.duplicate_group_key IN (
                      SELECT selected.duplicate_group_key
                      FROM _police_dispatch_tasks AS selected
                      WHERE selected.batch_id=%s AND selected.id IN ({placeholders})
                        AND selected.duplicate_group_key<>''
                  )
                GROUP BY grouped.duplicate_group_key
                HAVING kept<>1
                LIMIT 1
            """, [batch_id, batch_id, *selected_ids])
            if await cur.fetchone():
                raise HTTPException(409, "所选任务中存在尚未正确处理的重复人员组")
            missing_phone_count = sum(not item["phone"].strip() for item in pending) \
                if parser.parser_type != "涉警统计" else 0
            if missing_phone_count:
                raise HTTPException(
                    409,
                    f"有 {missing_phone_count} 条待下发任务缺少手机号，请先研判或补齐手机号",
                )
            lock_timeout = 30 if run_id is not None else 2
            if not await acquire_sheet_lock(cur, spreadsheet["id"], timeout=lock_timeout):
                raise HTTPException(409, f"{parser.parser_type}表格正在同步或被他人编辑，请稍后重试")
            sheet_lock_acquired = True
            publish_date = await get_business_date(cur)
            await cur.execute("""
                UPDATE _police_dispatch_batches
                SET first_publish_date=COALESCE(first_publish_date, %s),
                    publish_started_at=COALESCE(publish_started_at, UTC_TIMESTAMP()),
                    status='publishing', last_error=''
                WHERE id=%s
            """, (publish_date, batch_id))
            await cur.execute(
                "SELECT first_publish_date FROM _police_dispatch_batches WHERE id=%s",
                (batch_id,),
            )
            publish_date = (await cur.fetchone())[0]
            pending_ids = [item["id"] for item in pending]
            pending_placeholders = ",".join(["%s"] * len(pending_ids))
            if run_id is None:
                await cur.execute(f"""
                    UPDATE _police_dispatch_tasks
                    SET publish_status='publishing', task_status='pending_publish', publish_error=''
                    WHERE batch_id=%s AND id IN ({pending_placeholders})
                      AND final_action='dispatch' AND task_status='pending_publish'
                      AND publish_status IN ('pending', 'retryable')
                """, [batch_id, *pending_ids])
                if cur.rowcount != len(pending_ids):
                    raise HTTPException(409, "部分所选任务状态已经变化，请刷新后重新选择")
            if run_id is not None:
                await cur.execute("""
                    UPDATE _police_dispatch_publish_run_items
                    SET status='checking',error_code=''
                    WHERE run_id=%s
                """, (run_id,))
        await conn.commit()
    except Exception:
        await conn.rollback()
        if sheet_lock_acquired and spreadsheet is not None:
            async with conn.cursor() as cur:
                await release_sheet_lock(cur, spreadsheet["id"])
        raise

    assert spreadsheet is not None

    client = None
    success_count = 0
    failed_count = 0
    try:
        async with conn.cursor() as cur:
            await _set_publish_run_phase(cur, run_id, "reading_source")
            client = await _oauth_client(cur)
        source_columns = await resolve_source_columns(client, spreadsheet, parser)
        comparison_columns = _publish_comparison_columns(parser, source_columns)
        all_rows = await client.read_all_source_rows(
            spreadsheet["file_id"], spreadsheet["data_sheet_id"],
            spreadsheet["header_row"], source_columns,
            include_detected_headers=True,
        )
        source_rows = [row for row in all_rows if not row.get("is_header")]
        existing_by_key: dict[str, list[dict[str, Any]]] = {}
        for source in source_rows:
            for key in {
                parser_business_key(parser, source["values"]),
                parser_business_key(parser, source["values"], legacy=True),
            }:
                existing_by_key.setdefault(key, []).append(source)
        next_row = max([spreadsheet["header_row"], *[row["physical_row"] for row in all_rows]]) + 1
        ready: list[tuple[dict, dict, dict, str]] = []
        async with conn.cursor() as cur:
            for task in pending:
                values = _publish_values(task, task["community"], publish_date)
                request_values = _publish_request_values(values, comparison_columns)
                key = task.get("business_key_hmac") or parser_business_key(
                    parser, values, legacy=True
                )
                candidates = existing_by_key.get(key, [])
                if candidates:
                    exact = next((
                        item for item in candidates
                        if _row_values_match(
                            values, item["values"], item.get("cell_meta") or {},
                            comparison_columns,
                        )
                    ), None)
                    candidate = exact or candidates[0]
                    row_hash = source_row_hash({
                        column: str(candidate["values"].get(column, "") or "").strip()
                        for column in parser.COLUMNS
                    })
                    if exact:
                        success_count += 1
                        await _set_task_publish_state(
                            cur, task_id=task["id"], status="success",
                            business_key=key, physical_row=int(candidate["physical_row"]),
                            linked_row_hash=row_hash, cache_pending=True,
                        )
                        await _save_publish_result(
                            cur, task_id=task["id"], spreadsheet=spreadsheet,
                            business_key=key, request_values=request_values, status="success",
                            physical_row=int(candidate["physical_row"]),
                            verified_values=candidate["values"], row_hash=row_hash,
                            cache_pending=True,
                        )
                        await _set_publish_run_item(
                            cur, run_id, task["id"], "success",
                            physical_row=int(candidate["physical_row"]),
                        )
                    else:
                        failed_count += 1
                        await _set_task_publish_state(
                            cur, task_id=task["id"], status="conflict",
                            business_key=key,
                            error="腾讯表格已存在相同业务主键但内容不同",
                            physical_row=int(candidate["physical_row"]),
                            linked_row_hash=row_hash,
                            conflict_values=candidate["values"],
                        )
                        await _save_publish_result(
                            cur, task_id=task["id"], spreadsheet=spreadsheet,
                            business_key=key, request_values=request_values, status="conflict",
                            physical_row=int(candidate["physical_row"]),
                            verified_values=candidate["values"], row_hash=row_hash,
                            error_code="content_conflict",
                            error_message="同主键内容不同",
                        )
                        await _set_publish_run_item(
                            cur, run_id, task["id"], "conflict",
                            physical_row=int(candidate["physical_row"]),
                            error_code="content_conflict",
                        )
                    continue
                existing_by_key[key] = []
                ready.append((task, values, request_values, key))
            await _refresh_publish_run_progress(cur, run_id)

        async with conn.cursor() as cur:
            await _set_publish_run_phase(cur, run_id, "publishing")
        for offset in range(0, len(ready), 50):
            chunk = ready[offset:offset + 50]
            start_row = next_row
            rows = [
                parser.source_row_values(values, source_columns)
                for _, values, _, _ in chunk
            ]
            async with conn.cursor() as cur:
                for index, (task, _values, request_values, key) in enumerate(chunk):
                    physical_row = start_row + index
                    await _save_publish_result(
                        cur, task_id=task["id"], spreadsheet=spreadsheet,
                        business_key=key, request_values=request_values,
                        status="needs_reconciliation", physical_row=physical_row,
                        error_code="sending", error_message="腾讯写入请求正在处理",
                    )
                    await _set_publish_run_item(
                        cur, run_id, task["id"], "sending",
                        physical_row=physical_row,
                    )
            chunk_positions = {
                task["id"]: start_row + index
                for index, (task, _values, _request_values, _key) in enumerate(chunk)
            }
            completed_chunk_ids: set[int] = set()
            try:
                await client.batch_update(
                    spreadsheet["file_id"],
                    [client.build_update_range_request(
                        spreadsheet["data_sheet_id"], start_row - 1, 0, rows
                    )],
                )
                try:
                    verified_rows = await client.read_source_rows(
                        spreadsheet["file_id"], spreadsheet["data_sheet_id"],
                        start_row, start_row + len(chunk) - 1, source_columns,
                    )
                    verified_by_row = {
                        int(item["physical_row"]): item
                        for item in verified_rows
                    }
                except Exception:
                    verified_by_row = {}
                for index, (task, values, request_values, key) in enumerate(chunk):
                    physical_row = start_row + index
                    verified = verified_by_row.get(physical_row)
                    try:
                        if verified is None:
                            raise ValueError("腾讯范围回读缺少目标行")
                        verified["values"] = parser.normalize_source_row(
                            verified["values"]
                        )
                        matched = _row_values_match(
                            values, verified["values"], verified.get("cell_meta") or {},
                            comparison_columns,
                        )
                    except Exception:
                        verified = None
                        matched = False
                    async with conn.cursor() as cur:
                        if matched and verified is not None:
                            verified_values = {
                                column: str(verified["values"].get(column, "") or "").strip()
                                for column in parser.COLUMNS
                            }
                            verified_hash = source_row_hash(verified_values)
                            await _set_task_publish_state(
                                cur, task_id=task["id"], status="success",
                                business_key=key, physical_row=physical_row,
                                linked_row_hash=verified_hash, cache_pending=True,
                            )
                            await _save_publish_result(
                                cur, task_id=task["id"], spreadsheet=spreadsheet,
                                business_key=key, request_values=request_values,
                                status="success", physical_row=physical_row,
                                verified_values=verified_values,
                                row_hash=verified_hash, cache_pending=True,
                                count_attempt=False,
                            )
                            await _set_publish_run_item(
                                cur, run_id, task["id"], "success",
                                physical_row=physical_row,
                            )
                            success_count += 1
                        else:
                            error = "腾讯已收到写入请求，但回读结果尚未确认"
                            await _set_task_publish_state(
                                cur, task_id=task["id"],
                                status="needs_reconciliation",
                                business_key=key, error=error,
                                physical_row=physical_row,
                            )
                            await _save_publish_result(
                                cur, task_id=task["id"], spreadsheet=spreadsheet,
                                business_key=key, request_values=request_values,
                                status="needs_reconciliation",
                                physical_row=physical_row,
                                verified_values=(verified or {}).get("values"),
                                error_code="verification_uncertain",
                                error_message=error,
                                count_attempt=False,
                            )
                            await _set_publish_run_item(
                                cur, run_id, task["id"], "needs_reconciliation",
                                physical_row=physical_row,
                                error_code="verification_uncertain",
                            )
                            failed_count += 1
                        completed_chunk_ids.add(task["id"])
                        await _refresh_publish_run_progress(cur, run_id)
                next_row += len(chunk)
            except (Exception, asyncio.CancelledError) as exc:
                remaining_chunk = [
                    item for item in chunk if item[0]["id"] not in completed_chunk_ids
                ]
                failed_count += len(remaining_chunk)
                safe_error = "腾讯写入请求结果不确定，等待下次正常同步对账"
                if isinstance(exc, HTTPException):
                    safe_error = str(exc.detail)[:500]
                elif isinstance(exc, TxDocsAPIError):
                    safe_error = str(exc)[:500]
                async with conn.cursor() as cur:
                    for task, values, request_values, key in remaining_chunk:
                        await _set_task_publish_state(
                            cur, task_id=task["id"],
                            status="needs_reconciliation",
                            business_key=key, error=safe_error,
                        )
                        await _save_publish_result(
                            cur, task_id=task["id"], spreadsheet=spreadsheet,
                            business_key=key, request_values=request_values,
                            status="needs_reconciliation",
                            physical_row=chunk_positions[task["id"]],
                            error_code="request_uncertain",
                            error_message=safe_error,
                            count_attempt=False,
                        )
                        await _set_publish_run_item(
                            cur, run_id, task["id"], "needs_reconciliation",
                            physical_row=chunk_positions[task["id"]],
                            error_code="request_uncertain",
                        )
                    await _refresh_publish_run_progress(cur, run_id)
                if isinstance(exc, asyncio.CancelledError):
                    raise
                break
        try:
            async with conn.cursor() as cur:
                await _set_publish_run_phase(cur, run_id, "refreshing_cache")
            await _refresh_spreadsheet(conn, client, spreadsheet)
            async with conn.cursor() as cur:
                await cur.execute(f"""
                    UPDATE _police_dispatch_tasks AS task
                    JOIN _police_dispatch_publish_results AS result
                      ON result.task_id=task.id
                    JOIN _online_source_rows AS source
                      ON source.spreadsheet_id=result.spreadsheet_id
                     AND source.sheet_id=result.sheet_id
                     AND source.physical_row=result.physical_row
                    SET task.linked_source_id=source.id,
                        task.linked_row_hash=source.row_hash,
                        result.source_row_id=source.id,
                        result.expected_row_hash=source.row_hash
                    WHERE task.batch_id=%s AND task.id IN ({pending_placeholders})
                """, [batch_id, *pending_ids])
                if run_id is not None:
                    await cur.execute("""
                        UPDATE _police_dispatch_publish_run_items AS item
                        JOIN _police_dispatch_tasks AS task ON task.id=item.task_id
                        SET item.status=CASE task.publish_status
                            WHEN 'success' THEN 'success'
                            WHEN 'conflict' THEN 'conflict'
                            WHEN 'needs_reconciliation' THEN 'needs_reconciliation'
                            WHEN 'retryable' THEN 'retryable'
                            ELSE item.status END
                        WHERE item.run_id=%s
                    """, (run_id,))
                    await _refresh_publish_run_progress(cur, run_id)
                await cur.execute(f"""
                    UPDATE _police_dispatch_tasks AS task
                    JOIN _police_dispatch_publish_results AS result
                      ON result.task_id=task.id
                    SET task.cache_pending=0, result.cache_pending=0
                    WHERE task.batch_id=%s AND task.publish_status='success'
                      AND task.id IN ({pending_placeholders})
                """, [batch_id, *pending_ids])
        except Exception:
            async with conn.cursor() as cur:
                await cur.execute(f"""
                    UPDATE _police_dispatch_tasks AS task
                    JOIN _police_dispatch_publish_results AS result
                      ON result.task_id=task.id
                    SET task.cache_pending=1, result.cache_pending=1
                    WHERE task.batch_id=%s AND task.publish_status='success'
                      AND task.id IN ({pending_placeholders})
                """, [batch_id, *pending_ids])
    finally:
        try:
            if client:
                await client.close()
        finally:
            async with conn.cursor() as cur:
                if run_id is None:
                    await cur.execute(f"""
                        UPDATE _police_dispatch_tasks
                        SET publish_status='retryable', task_status='pending_publish',
                            publish_error='尚未向腾讯发送，可安全重试'
                        WHERE batch_id=%s AND publish_status='publishing'
                          AND id IN ({pending_placeholders})
                    """, [batch_id, *pending_ids])
                else:
                    await cur.execute("""
                        UPDATE _police_dispatch_tasks AS task
                        JOIN _police_dispatch_publish_run_items AS item
                          ON item.task_id=task.id AND item.run_id=%s
                        SET task.publish_status='retryable',task.task_status='pending_publish',
                            task.publish_error='尚未向腾讯发送，可安全重试',
                            task.version=task.version+1,
                            item.status='retryable',item.error_code='not_sent'
                        WHERE task.batch_id=%s AND task.publish_status='publishing'
                          AND item.status IN ('queued','checking')
                    """, (run_id, batch_id))
                    await cur.execute("""
                        UPDATE _police_dispatch_tasks AS task
                        JOIN _police_dispatch_publish_run_items AS item
                          ON item.task_id=task.id AND item.run_id=%s
                        SET task.publish_status='needs_reconciliation',
                            task.task_status='publish_failed',
                            task.publish_error='腾讯请求可能已经送达，等待同步对账',
                            task.version=task.version+1,
                            item.status='needs_reconciliation',
                            item.error_code='request_uncertain'
                        WHERE task.batch_id=%s AND task.publish_status='publishing'
                          AND item.status='sending'
                    """, (run_id, batch_id))
                    await _refresh_publish_run_progress(cur, run_id)
                counts = await _refresh_batch_status(cur, batch_id)
                await cur.execute("""
                    UPDATE _police_dispatch_batches SET
                        last_error=CASE
                            WHEN %s>0 THEN '部分任务等待同步对账或存在内容冲突'
                            ELSE '' END
                    WHERE id=%s
                """, (failed_count, batch_id))
                await release_sheet_lock(cur, spreadsheet["id"])
    await record_admin_audit(
        user, "police_dispatch.publish", target_type="police_dispatch_batch",
        target_name=str(batch_id), detail={
            "selected": len(selected_ids), "success": success_count, "failed": failed_count,
        },
        result="partial" if failed_count else "success",
        **(request_audit_fields(request) if request is not None else {}),
    )
    return {
        "message": "发布完成" if not failed_count else "部分任务需要等待同步对账或人工处理冲突",
        "success_count": success_count, "failed_count": failed_count,
        "counts": counts,
    }


POLICE_PUBLISH_CREATE_LOCK = "binhu_police_dispatch_publish_create"


async def _finish_police_publish_run(
    cur,
    run_id: int,
    *,
    status: str,
    error_code: str = "",
    error_message: str = "",
) -> None:
    await cur.execute("""
        UPDATE _police_dispatch_tasks AS task
        JOIN _police_dispatch_publish_run_items AS item
          ON item.task_id=task.id AND item.run_id=%s
        SET task.publish_status='retryable',task.task_status='pending_publish',
            task.publish_error='尚未向腾讯发送，可安全重试',
            task.version=task.version+1
        WHERE task.publish_status='publishing'
          AND item.status IN ('queued','checking')
    """, (run_id,))
    await cur.execute("""
        UPDATE _police_dispatch_tasks AS task
        JOIN _police_dispatch_publish_run_items AS item
          ON item.task_id=task.id AND item.run_id=%s
        SET task.publish_status='needs_reconciliation',task.task_status='publish_failed',
            task.publish_error='腾讯请求可能已经送达，等待同步对账',
            task.version=task.version+1
        WHERE task.publish_status='publishing' AND item.status='sending'
    """, (run_id,))
    await cur.execute("""
        UPDATE _police_dispatch_publish_run_items
        SET status='retryable',error_code=CASE
                WHEN error_code='' THEN 'not_sent' ELSE error_code END
        WHERE run_id=%s AND status IN ('queued','checking')
    """, (run_id,))
    await _refresh_publish_run_progress(cur, run_id)
    await cur.execute("""
        UPDATE _police_dispatch_publish_runs
        SET status=%s,phase='finished',error_code=%s,error_message=%s,
            finished_at=UTC_TIMESTAMP()
        WHERE id=%s
    """, (status, error_code[:100], error_message[:500], run_id))
    await cur.execute(
        "SELECT batch_id FROM _police_dispatch_publish_runs WHERE id=%s",
        (run_id,),
    )
    batch_row = await cur.fetchone()
    if batch_row:
        await _refresh_batch_status(cur, int(batch_row[0]))


async def _run_police_publish_job(run_id: int) -> None:
    pool = db_manager.get_pool("online_data")
    async with pool.acquire() as conn:
        async with conn.cursor() as cur:
            await cur.execute("""
                SELECT batch_id,requested_by,requested_username
                FROM _police_dispatch_publish_runs WHERE id=%s
            """, (run_id,))
            row = await cur.fetchone()
            if not row:
                return
            batch_id = int(row[0])
            user = {"id": int(row[1]) if row[1] is not None else None,
                    "username": str(row[2] or "")}
            await cur.execute("""
                SELECT task_id FROM _police_dispatch_publish_run_items
                WHERE run_id=%s ORDER BY item_order
            """, (run_id,))
            task_ids = [int(item[0]) for item in await cur.fetchall()]
            await cur.execute("""
                UPDATE _police_dispatch_publish_runs
                SET status='running',phase='preparing',
                    started_at=COALESCE(started_at,UTC_TIMESTAMP()),
                    error_code='',error_message=''
                WHERE id=%s AND status='pending'
            """, (run_id,))
        if not task_ids:
            async with conn.cursor() as cur:
                await _finish_police_publish_run(
                    cur, run_id, status="failed", error_code="empty_selection",
                    error_message="发布任务没有可处理的选中项",
                )
            return
        try:
            result = await _execute_publish_selection(
                batch_id,
                TaskPublishSelection(task_ids=task_ids),
                None,
                user,
                conn,
                run_id=run_id,
            )
        except asyncio.CancelledError:
            async with conn.cursor() as cur:
                await _finish_police_publish_run(
                    cur, run_id, status="failed", error_code="service_stopping",
                    error_message="服务停止，未发送任务可重试，可能已发送任务等待同步对账",
                )
            raise
        except Exception as exc:
            if isinstance(exc, HTTPException):
                safe_error = str(exc.detail)[:500]
            elif isinstance(exc, TxDocsAPIError):
                safe_error = str(exc)[:500]
            else:
                safe_error = "后台发布任务执行失败，未发送任务可安全重试"
            async with conn.cursor() as cur:
                await _finish_police_publish_run(
                    cur, run_id, status="failed", error_code="publish_failed",
                    error_message=safe_error,
                )
            await record_admin_audit(
                user, "police_dispatch.publish",
                target_type="police_dispatch_publish_run", target_name=str(run_id),
                result="failed", detail={"run_id": run_id, "batch_id": batch_id},
            )
            return
        async with conn.cursor() as cur:
            await _finish_police_publish_run(
                cur,
                run_id,
                status="partial" if result["failed_count"] else "completed",
                error_code="partial" if result["failed_count"] else "",
                error_message=(
                    "部分任务等待同步对账或需要处理内容冲突"
                    if result["failed_count"] else ""
                ),
            )


@router.get("/batches/{batch_id}/publish-runs/latest")
async def latest_publish_run(
    batch_id: int,
    _user: dict = Depends(require_police_dispatch),
    _conn=Depends(get_db),
):
    return {"data": await get_latest_police_publish_run(batch_id)}


@router.get("/publish-runs/{run_id}")
async def publish_run_detail(
    run_id: int,
    _user: dict = Depends(require_police_dispatch),
    _conn=Depends(get_db),
):
    run = await get_police_publish_run(run_id)
    if not run:
        raise HTTPException(404, "发布任务不存在")
    return run


@router.post("/batches/{batch_id}/publish-selected", status_code=202)
async def publish_selected_tasks(
    batch_id: int,
    data: TaskPublishSelection,
    request: Request,
    user: dict = Depends(require_police_dispatch),
    conn=Depends(get_db),
):
    selected_ids = data.task_ids
    placeholders = ",".join(["%s"] * len(selected_ids))
    run_id = 0
    lock_acquired = False
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("SELECT GET_LOCK(%s, 5)", (POLICE_PUBLISH_CREATE_LOCK,))
            lock_row = await cur.fetchone()
            lock_acquired = bool(lock_row and lock_row[0] == 1)
            if not lock_acquired:
                raise HTTPException(409, "发布任务正在创建，请稍后重试")
            batch = await _batch_payload(cur, batch_id)
            target_parser = str(batch.get("target_parser") or "全链条")
            local_mode = local_data_source_enabled()
            if local_mode:
                spreadsheets = []
                publish_spreadsheet_id = 0
            else:
                if not await _writeback_enabled(cur):
                    raise HTTPException(503, "在线回写已由超级管理员暂停")
                spreadsheets = await _enabled_spreadsheets(cur, target_parser)
                if len(spreadsheets) != 1:
                    raise HTTPException(409, f"{target_parser}业务没有唯一启用的腾讯来源表")
                publish_spreadsheet_id = int(spreadsheets[0]["id"])
            await cur.execute("""
                SELECT id FROM _police_dispatch_publish_runs
                WHERE spreadsheet_id=%s AND status IN ('pending','running')
                ORDER BY id DESC LIMIT 1
            """, (publish_spreadsheet_id,))
            active = await cur.fetchone()
            if active:
                raise HTTPException(
                    409,
                    f"已有发布任务 #{int(active[0])} 正在处理，请先查看当前进度",
                )
            await cur.execute(f"""
                SELECT task.id,task.phone
                FROM _police_dispatch_tasks AS task
                WHERE task.batch_id=%s AND task.id IN ({placeholders})
                  AND task.final_action='dispatch' AND task.task_status='pending_publish'
                  AND task.publish_status IN ('pending','retryable')
                ORDER BY task.source_row,task.id
                FOR UPDATE
            """, [batch_id, *selected_ids])
            eligible = [(int(row[0]), str(row[1] or "")) for row in await cur.fetchall()]
            if len(eligible) != len(selected_ids):
                raise HTTPException(409, "部分所选任务已不可发布，请刷新列表后重新选择")
            missing_phone_count = sum(not phone.strip() for _, phone in eligible) \
                if target_parser != "涉警统计" else 0
            if missing_phone_count:
                raise HTTPException(
                    409,
                    f"有 {missing_phone_count} 条待下发任务缺少手机号，请先研判或补齐手机号",
                )
            await cur.execute(f"""
                SELECT grouped.duplicate_group_key,
                       SUM(grouped.final_action<>'duplicate_exclude') AS kept
                FROM _police_dispatch_tasks AS grouped
                WHERE grouped.batch_id=%s
                  AND grouped.duplicate_group_key IN (
                      SELECT selected.duplicate_group_key
                      FROM _police_dispatch_tasks AS selected
                      WHERE selected.batch_id=%s AND selected.id IN ({placeholders})
                        AND selected.duplicate_group_key<>''
                  )
                GROUP BY grouped.duplicate_group_key
                HAVING kept<>1
                LIMIT 1
            """, [batch_id, batch_id, *selected_ids])
            if await cur.fetchone():
                raise HTTPException(409, "所选任务中存在尚未正确处理的重复人员组")
            await cur.execute("""
                INSERT INTO _police_dispatch_publish_runs (
                    batch_id,spreadsheet_id,status,phase,total_count,
                    requested_by,requested_username
                ) VALUES (%s,%s,'pending','queued',%s,%s,%s)
            """, (
                batch_id, publish_spreadsheet_id, len(selected_ids), user.get("id"),
                str(user.get("username") or "")[:50],
            ))
            run_id = int(cur.lastrowid)
            ordered_ids = [task_id for task_id, _phone in eligible]
            await cur.executemany("""
                INSERT INTO _police_dispatch_publish_run_items (
                    run_id,task_id,item_order,status
                ) VALUES (%s,%s,%s,'queued')
            """, [
                (run_id, task_id, index)
                for index, task_id in enumerate(ordered_ids, start=1)
            ])
            publish_date = await get_business_date(cur)
            await cur.execute("""
                UPDATE _police_dispatch_batches
                SET first_publish_date=COALESCE(first_publish_date,%s),
                    publish_started_at=COALESCE(publish_started_at,UTC_TIMESTAMP()),
                    status='publishing',last_error=''
                WHERE id=%s
            """, (publish_date, batch_id))
            await cur.execute(f"""
                UPDATE _police_dispatch_tasks
                SET publish_status='publishing',task_status='pending_publish',publish_error=''
                WHERE batch_id=%s AND id IN ({placeholders})
                  AND final_action='dispatch' AND task_status='pending_publish'
                  AND publish_status IN ('pending','retryable')
            """, [batch_id, *selected_ids])
            if cur.rowcount != len(selected_ids):
                raise HTTPException(409, "部分所选任务状态已经变化，请刷新后重新选择")
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    finally:
        if lock_acquired:
            async with conn.cursor() as cur:
                await cur.execute("SELECT RELEASE_LOCK(%s)", (POLICE_PUBLISH_CREATE_LOCK,))
                await cur.fetchone()
    launch_police_publish_run(run_id, _run_police_publish_job)
    await record_admin_audit(
        user, "police_dispatch.publish", target_type="police_dispatch_publish_run",
        target_name=str(run_id), result="pending",
        detail={"run_id": run_id, "batch_id": batch_id, "selected": len(selected_ids)},
        **request_audit_fields(request),
    )
    run = await get_police_publish_run(run_id)
    if not run:
        raise HTTPException(500, "发布任务创建后无法重新定位")
    return {**run, "message": "发布任务已进入后台处理，可以离开本页面"}
