"""辖区档案、机构、变更审核、合并历史和人员标记扩展接口。"""

from __future__ import annotations

import hashlib
import json
from io import BytesIO
from datetime import datetime
from typing import Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from config import settings
from deps import require_permission
from routers.registry import _allowed_community_ids, _allowed_community_names, get_registry_db
from services.audit import record_admin_audit, request_audit_fields
from services.permissions import (
    REGISTRY_IMPORT_MANAGE,
    REGISTRY_PROPERTY_MANAGE,
    REGISTRY_PROPERTY_VIEW,
    REGISTRY_WATCH_MANAGE,
    REGISTRY_WATCH_VIEW,
)
from services.registry_security import hmac_digest, normalize_identity, normalize_phone
from services.watch_matching import backfill_assignment_snapshots
from services.registry_import import (
    ISSUE_CERTIFICATE_CONTENT_CONFLICT,
    ISSUE_CERTIFICATE_DUPLICATE,
    ISSUE_CERTIFICATE_NON_RENTAL,
    ISSUE_HOUSEHOLD_DUPLICATE,
    ISSUE_HOUSEHOLD_COMMUNITY_UNRESOLVED,
    ISSUE_HOUSEHOLD_MISSING_TYPE,
    classify_certificate_rows,
    classify_household_rows,
    household_community_candidates,
    issue_problem_details,
    normalize_address,
    normalize_community,
    normalize_text,
)
from services.registry_certificate_source import fetch_certificate_rows
from services.registry_certificate_apply import apply_certificate_batch
from services.registry_certificate_status import certificate_status_summary
from services.registry_certificate_jobs import (
    create_certificate_source_run,
    get_certificate_source_run,
    get_latest_certificate_source_run,
    retry_certificate_source_run,
)
from services.visit_source import VisitSourceError


router = APIRouter(prefix="/api/registry", tags=["辖区档案"])

REGISTRY_IMPORT_WRITE_CHUNK = 500


def _can_view_identity(user: dict) -> bool:
    return user.get("role") == "super_admin"


def _redact_sensitive_payload(value):
    if isinstance(value, dict):
        return {
            key: _redact_sensitive_payload(item)
            for key, item in value.items()
            if key not in {
                "identity_number", "identity_hmac", "phone", "phone_hmac",
                "czrzjhm", "sjczrzjhm", "sfzh", "landlord_identity_number",
                "actual_renter_identity_number",
            }
        }
    if isinstance(value, list):
        return [_redact_sensitive_payload(item) for item in value]
    return value


def _json(value, default):
    if isinstance(value, type(default)):
        return value
    if isinstance(value, (bytes, bytearray)):
        value = value.decode("utf-8", errors="replace")
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, type(default)) else default
        except (TypeError, ValueError):
            return default
    return default


def _iso(value) -> str | None:
    return value.isoformat() + ("Z" if isinstance(value, datetime) else "") if value else None


async def _property_scope(cur, property_id: int, user: dict, permission: str):
    await cur.execute(
        "SELECT community_id FROM registry_properties WHERE id=%s",
        (property_id,),
    )
    row = await cur.fetchone()
    if not row:
        raise HTTPException(404, "房屋档案不存在")
    allowed = await _allowed_community_ids(user, permission)
    if allowed is not None and row[0] not in allowed:
        raise HTTPException(403, "无权访问该社区档案")
    return row[0]


async def _organization_scope(cur, organization_id: int, user: dict, permission: str) -> None:
    allowed = await _allowed_community_ids(user, permission)
    if allowed is None:
        return
    if not allowed:
        raise HTTPException(403, "无权查看该机构档案")
    placeholders = ",".join(["%s"] * len(allowed))
    await cur.execute(
        "SELECT 1 FROM registry_property_organization_roles relation "
        "JOIN registry_properties property ON property.id=relation.property_id "
        "WHERE relation.organization_id=%s "
        f"AND property.community_id IN ({placeholders}) LIMIT 1",
        (organization_id, *allowed),
    )
    if not await cur.fetchone():
        raise HTTPException(403, "无权查看该机构档案")


async def _canonical_community(
    cur,
    community_id: int | None,
    snapshot: str | None = None,
    *,
    require_active: bool = True,
) -> tuple[int | None, str]:
    """Return the server-owned community id/name pair.

    RegistryData keeps a snapshot for historical display, but callers must not
    be able to forge that snapshot or select a disabled/non-existent community.
    A name-only request may use a formal name or an enabled alias.
    """
    requested = str(snapshot or "").strip()
    if community_id is not None:
        await cur.execute(
            "SELECT id, name, is_active FROM _communities WHERE id=%s",
            (community_id,),
        )
        row = await cur.fetchone()
        if not row:
            raise HTTPException(422, "所属社区不存在")
        if require_active and not bool(row[2]):
            raise HTTPException(409, "所属社区已停用")
        return int(row[0]), str(row[1]).strip()
    if not requested:
        return None, ""
    await cur.execute(
        "SELECT community.id, community.name, community.is_active "
        "FROM _communities community WHERE community.name=%s "
        "UNION ALL "
        "SELECT community.id, community.name, community.is_active "
        "FROM _community_aliases alias "
        "JOIN _communities community ON community.id=alias.community_id "
        "WHERE alias.alias=%s AND community.is_active=1 LIMIT 1",
        (requested, requested),
    )
    row = await cur.fetchone()
    if not row:
        raise HTTPException(422, "所属社区或别名不存在或已停用")
    if require_active and not bool(row[2]):
        raise HTTPException(409, "所属社区已停用")
    return int(row[0]), str(row[1]).strip()


async def _household_import_community(cur, snapshot: str) -> tuple[int | None, str]:
    """Resolve an import label without weakening the global community rules."""
    last_error: HTTPException | None = None
    for candidate in household_community_candidates(snapshot):
        try:
            return await _canonical_community(cur, None, candidate)
        except HTTPException as exc:
            if exc.status_code != 422:
                raise
            last_error = exc
    if last_error:
        raise last_error
    raise HTTPException(422, "所属社区为空")


async def _housing_person_scope(cur, person_id: int, user: dict, permission: str) -> None:
    """Require a housing person to have a property in the user's scope."""
    allowed = await _allowed_community_ids(user, permission)
    if allowed is None:
        return
    if not allowed:
        raise HTTPException(403, "无权访问该人员档案")
    placeholders = ",".join(["%s"] * len(allowed))
    await cur.execute(
        "SELECT 1 FROM registry_property_person_roles relation "
        "JOIN registry_properties property ON property.id=relation.property_id "
        f"WHERE relation.person_id=%s AND property.community_id IN ({placeholders}) LIMIT 1",
        (person_id, *allowed),
    )
    if not await cur.fetchone():
        raise HTTPException(403, "无权访问该人员档案")


def _intervals_overlap(
    first_from: datetime | None,
    first_to: datetime | None,
    second_from: datetime | None,
    second_to: datetime | None,
) -> bool:
    """Compare nullable validity intervals without relying on SQL NULL ordering."""
    floor = datetime.min
    ceiling = datetime.max
    return max(first_from or floor, second_from or floor) <= min(first_to or ceiling, second_to or ceiling)


async def _ensure_relation_interval_available(
    cur,
    table: str,
    key_clause: str,
    key_params: tuple[object, ...],
    valid_from: datetime | None,
    valid_to: datetime | None,
    exclude_id: int | None = None,
) -> None:
    """Reject overlapping validity intervals for one relation key."""
    if valid_from and valid_to and valid_to < valid_from:
        raise HTTPException(422, "关系结束时间不能早于生效时间")
    await cur.execute(
        f"SELECT id, valid_from, valid_to FROM {table} WHERE {key_clause}",
        key_params,
    )
    for row in await cur.fetchall():
        if exclude_id is not None and int(row[0]) == exclude_id:
            continue
        if _intervals_overlap(row[1], row[2], valid_from, valid_to):
            raise HTTPException(409, "相同关系的有效时间已重叠")


async def _watch_person_scope(cur, person_id: int, user: dict) -> None:
    allowed_names = _allowed_community_names(user, REGISTRY_WATCH_VIEW)
    if allowed_names is None:
        return
    if not allowed_names:
        raise HTTPException(403, "无权查看该人员标记")
    online_schema = settings.MYSQL_ONLINE_DATA_DB.replace("`", "")
    placeholders = ",".join(["%s"] * len(allowed_names))
    await cur.execute(
        "SELECT 1 FROM watch_assignments scoped_assignment "
        "JOIN online_task_watch_snapshots scoped_snapshot "
        "ON scoped_snapshot.assignment_id=scoped_assignment.id "
        f"JOIN `{online_schema}`._online_source_projection scoped_projection "
        "ON scoped_projection.parser_type=scoped_snapshot.parser_type "
        "AND scoped_projection.row_key=scoped_snapshot.row_key "
        "WHERE scoped_assignment.person_id=%s "
        f"AND scoped_projection.community IN ({placeholders}) LIMIT 1",
        (person_id, *allowed_names),
    )
    if not await cur.fetchone():
        raise HTTPException(403, "该人员标记无法归属到你的社区")


class PropertyUpdate(BaseModel):
    street: str = Field(default="", max_length=200)
    community_id: int | None = None
    community_name_snapshot: str = Field(default="", max_length=200)
    natural_address: str = Field(default="", max_length=500)
    building: str = Field(default="", max_length=100)
    room: str = Field(default="", max_length=100)
    housing_type: str = Field(default="", max_length=50)
    residence_type: str = Field(default="", max_length=100)
    source_house_no: str = Field(default="", max_length=100)
    source_updated_at: datetime | None = None
    source_type: str | None = Field(default=None, max_length=30)
    source_ref: str | None = Field(default=None, max_length=190)
    normalized_address: str = Field(default="", max_length=1000)
    change_reason: str = Field(default="", max_length=500)


class StatusChange(BaseModel):
    status: Literal["active", "inactive"]
    reason: str = Field(default="", max_length=500)


class AliasCreate(BaseModel):
    alias: str = Field(min_length=1, max_length=500)
    community_id: int | None = None


class PhoneCreate(BaseModel):
    phone: str = Field(min_length=1, max_length=200)
    is_primary: bool = False
    verified: bool = False
    valid_from: datetime | None = None
    valid_to: datetime | None = None


class HousingPersonUpdate(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    identity_number: str | None = Field(default=None, max_length=50)
    is_temporary: bool = False
    verification_status: Literal["unverified", "pending", "verified"] = "unverified"


class OrganizationPayload(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    organization_type: Literal["agency", "rental_platform", "property", "other"] = "other"
    license_number: str = Field(default="", max_length=100)
    notes: str = Field(default="", max_length=1000)


class OrganizationMemberCreate(BaseModel):
    person_id: int = Field(gt=0)
    title: str = Field(default="", max_length=100)
    valid_from: datetime | None = None
    valid_to: datetime | None = None
    verified: bool = False


class PropertyOrganizationRoleCreate(BaseModel):
    organization_id: int = Field(gt=0)
    role_type_id: int = Field(gt=0)
    valid_from: datetime | None = None
    valid_to: datetime | None = None
    verified: bool = False


class RelationUpdate(BaseModel):
    valid_from: datetime | None = None
    valid_to: datetime | None = None
    verified: bool = False


class OrganizationMembershipUpdate(RelationUpdate):
    title: str = Field(default="", max_length=100)


class CandidateCreate(BaseModel):
    entity_type: str = Field(min_length=1, max_length=30)
    entity_id: int | None = None
    change_type: str = Field(min_length=1, max_length=30)
    payload: dict
    reason: str = Field(default="", max_length=500)
    source_type: Literal["online_task", "visit", "dispatch", "manual"] = "manual"
    source_ref: str = Field(default="", max_length=190)


class ReviewDecision(BaseModel):
    action: Literal["accept", "reject"]
    reason: str = Field(default="", max_length=500)


class RegistryIssueItem(BaseModel):
    issue_type: Literal[
        ISSUE_CERTIFICATE_DUPLICATE,
        ISSUE_CERTIFICATE_CONTENT_CONFLICT,
        ISSUE_CERTIFICATE_NON_RENTAL,
        ISSUE_HOUSEHOLD_DUPLICATE,
        ISSUE_HOUSEHOLD_MISSING_TYPE,
    ]
    source_type: str = Field(default="external", max_length=30)
    source_ref: str = Field(default="", max_length=190)
    entity_key: str = Field(default="", max_length=500)
    payload: dict = Field(default_factory=dict)
    reason: str = Field(default="", max_length=500)


class RegistryIssueSearch(BaseModel):
    keyword: str = Field(default="", max_length=200)
    status: Literal["", "pending", "resolved", "dismissed"] = "pending"
    issue_type: str = Field(default="", max_length=60)
    source_type: Literal["", "household", "certificate"] = ""
    community_id: int | None = None
    housing_category: Literal["", "rental", "self_owned", "other", "unmarked"] = ""
    page: int = Field(default=1, ge=1)
    page_size: int = Field(default=50, ge=1, le=500)


class RegistryIssueBulkCreate(BaseModel):
    batch_id: int | None = Field(default=None, gt=0)
    items: list[RegistryIssueItem] = Field(default_factory=list, max_length=50000)


class RegistryCertificateImport(BaseModel):
    source_name: str = Field(default="房东责任告知书", max_length=100)
    rows: list[dict] = Field(default_factory=list, max_length=50000)


class RegistryCertificateSourceRetry(BaseModel):
    restart: bool = False


class MergeRequest(BaseModel):
    target_person_id: int = Field(gt=0)
    reason: str = Field(default="", max_length=500)


class WatchCategoryUpdate(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    parent_id: int | None = None
    color: str = Field(default="#1677ff", max_length=20)
    alert_level: Literal["normal", "notice", "warning", "critical"] = "normal"
    is_active: bool = True
    description: str = Field(default="", max_length=1000)


class WatchPersonUpdate(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    identity_number: str | None = Field(default=None, max_length=50)
    verification_status: Literal["unverified", "pending", "verified"] = "unverified"
    status: Literal["active", "inactive"] = "active"


class WatchAssignmentUpdate(BaseModel):
    valid_from: datetime
    valid_to: datetime | None = None
    released_at: datetime | None = None
    basis: str = Field(default="", max_length=1000)
    status: Literal["active", "released", "inactive"] = "active"


async def _apply_candidate_payload(
    cur,
    *,
    entity_type: str,
    entity_id: int | None,
    change_type: str,
    payload: dict,
    user_id: int,
    allow_identity: bool,
) -> int | None:
    """将审核通过的候选值写入正式档案；字段和目标表使用固定白名单。"""
    if change_type == "deactivate" and entity_id:
        table = {
            "property": "registry_properties",
            "person": "registry_housing_people",
            "organization": "registry_organizations",
        }.get(entity_type)
        if not table:
            raise HTTPException(422, "该候选变更类型不支持停用")
        await cur.execute(
            f"UPDATE `{table}` SET status='inactive', updated_by=%s WHERE id=%s",
            (user_id, entity_id),
        )
        if cur.rowcount != 1:
            raise HTTPException(404, "候选变更对应的正式档案不存在")
        return entity_id

    if entity_type == "property":
        community_id, community_name = await _canonical_community(
            cur,
            int(payload["community_id"]) if payload.get("community_id") is not None else None,
            payload.get("community_name_snapshot"),
        )
        values = (
            str(payload.get("street") or "")[:200], community_id,
            community_name,
            str(payload.get("natural_address") or "")[:500],
            str(payload.get("building") or "")[:100], str(payload.get("room") or "")[:100],
            str(payload.get("housing_type") or "")[:50], str(payload.get("residence_type") or "")[:100],
            str(payload.get("source_house_no") or "")[:100], payload.get("source_updated_at"),
            str(payload.get("source_type") or "candidate_review")[:30], str(payload.get("source_ref") or "")[:190],
            str(payload.get("normalized_address") or payload.get("natural_address") or "")[:1000],
        )
        if entity_id:
            await cur.execute(
                "UPDATE registry_properties SET street=%s, community_id=%s, community_name_snapshot=%s, "
                "natural_address=%s, building=%s, room=%s, housing_type=%s, residence_type=%s, source_house_no=%s, "
                "source_updated_at=%s, source_type=%s, source_ref=%s, normalized_address=%s, "
                "current_version=current_version+1, updated_by=%s WHERE id=%s",
                (*values, user_id, entity_id),
            )
            if cur.rowcount != 1:
                raise HTTPException(404, "候选变更对应的房屋不存在")
            return entity_id
        await cur.execute(
            "INSERT INTO registry_properties "
            "(street, community_id, community_name_snapshot, natural_address, building, room, "
            "housing_type, residence_type, source_house_no, source_updated_at, source_type, source_ref, normalized_address, created_by, updated_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (*values, user_id, user_id),
        )
        return int(cur.lastrowid)

    if entity_type == "person":
        identity = normalize_identity(payload.get("identity_number"))
        if identity and not allow_identity:
            raise HTTPException(403, "身份证号只能由超级管理员审核写入")
        identity_hmac, hmac_version = hmac_digest(identity, kind="identity")
        values = (
            str(payload.get("name") or "")[:100], identity or None, identity_hmac, hmac_version,
            int(bool(payload.get("is_temporary"))),
            str(payload.get("verification_status") or "unverified")[:20],
        )
        if not values[0]:
            raise HTTPException(422, "候选人员姓名不能为空")
        if entity_id:
            await cur.execute(
                "UPDATE registry_housing_people SET name=%s, identity_number=%s, identity_hmac=%s, "
                "identity_hmac_version=%s, is_temporary=%s, verification_status=%s, updated_by=%s WHERE id=%s",
                (*values, user_id, entity_id),
            )
            if cur.rowcount != 1:
                raise HTTPException(404, "候选变更对应的人员不存在")
            return entity_id
        await cur.execute(
            "INSERT INTO registry_housing_people "
            "(name, identity_number, identity_hmac, identity_hmac_version, is_temporary, "
            "verification_status, source_type, created_by, updated_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,'candidate_review',%s,%s)",
            (*values, user_id, user_id),
        )
        return int(cur.lastrowid)

    if entity_type == "phone":
        person_id = int(payload.get("person_id") or entity_id or 0)
        phone = normalize_phone(payload.get("phone"))
        phone_hmac, hmac_version = hmac_digest(phone, kind="phone")
        if not person_id or not phone or not phone_hmac:
            raise HTTPException(422, "候选电话缺少人员或号码")
        await cur.execute(
            "INSERT INTO registry_person_phones "
            "(person_id, phone, phone_hmac, hmac_version, is_primary, verified, valid_from, valid_to, "
            "source_type, created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'candidate_review',%s)",
            (person_id, phone, phone_hmac, hmac_version, int(bool(payload.get("is_primary"))),
             int(bool(payload.get("verified"))), payload.get("valid_from"), payload.get("valid_to"), user_id),
        )
        return int(cur.lastrowid)

    if entity_type == "organization":
        values = (
            str(payload.get("name") or "")[:200],
            str(payload.get("organization_type") or "other")[:30],
            str(payload.get("license_number") or "")[:100],
            str(payload.get("notes") or "")[:1000],
        )
        if not values[0]:
            raise HTTPException(422, "候选机构名称不能为空")
        if entity_id:
            await cur.execute(
                "UPDATE registry_organizations SET name=%s, organization_type=%s, license_number=%s, "
                "notes=%s, updated_by=%s WHERE id=%s",
                (*values, user_id, entity_id),
            )
            if cur.rowcount != 1:
                raise HTTPException(404, "候选变更对应的机构不存在")
            return entity_id
        await cur.execute(
            "INSERT INTO registry_organizations "
            "(name, organization_type, license_number, notes, source_type, created_by, updated_by) "
            "VALUES (%s,%s,%s,%s,'candidate_review',%s,%s)",
            (*values, user_id, user_id),
        )
        return int(cur.lastrowid)

    raise HTTPException(422, "该候选变更暂不支持自动采用，请人工维护正式档案")


@router.get("/properties/{property_id}")
async def get_property_detail(
    property_id: int,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_VIEW)),
    conn=Depends(get_registry_db),
):
    async with conn.cursor() as cur:
        await _property_scope(cur, property_id, user, REGISTRY_PROPERTY_VIEW)
        await cur.execute(
            "SELECT id, street, community_id, community_name_snapshot, natural_address, building, room, "
            "housing_type, residence_type, source_house_no, source_updated_at, source_type, source_ref, "
            "normalized_address, status, current_version, created_at, updated_at "
            "FROM registry_properties WHERE id=%s",
            (property_id,),
        )
        row = await cur.fetchone()
        await cur.execute(
            "SELECT id, alias, community_id, enabled, source_type, created_at "
            "FROM registry_address_aliases WHERE property_id=%s ORDER BY enabled DESC, id DESC",
            (property_id,),
        )
        aliases = await cur.fetchall()
        await cur.execute(
            "SELECT version_no, street, natural_address, building, room, normalized_address, "
            "effective_from, effective_to, source_type, change_reason, created_at "
            "FROM registry_property_address_versions WHERE property_id=%s ORDER BY version_no DESC",
            (property_id,),
        )
        versions = await cur.fetchall()
        await cur.execute(
            "SELECT rel.id, person.id, person.name, role.id, role.name, rel.valid_from, rel.valid_to, rel.verified "
            "FROM registry_property_person_roles rel "
            "JOIN registry_housing_people person ON person.id=rel.person_id "
            "JOIN registry_role_types role ON role.id=rel.role_type_id "
            "WHERE rel.property_id=%s ORDER BY rel.valid_to IS NULL DESC, rel.id DESC",
            (property_id,),
        )
        people = await cur.fetchall()
        await cur.execute(
            "SELECT rel.id, org.id, org.name, role.id, role.name, rel.valid_from, rel.valid_to, rel.verified "
            "FROM registry_property_organization_roles rel "
            "JOIN registry_organizations org ON org.id=rel.organization_id "
            "JOIN registry_role_types role ON role.id=rel.role_type_id "
            "WHERE rel.property_id=%s ORDER BY rel.valid_to IS NULL DESC, rel.id DESC",
            (property_id,),
        )
        organizations = await cur.fetchall()
        await cur.execute(
            "SELECT id, source_ref, source_row, community_snapshot, address_snapshot, landlord_name, "
            "landlord_identity_number, actual_renter_name, actual_renter_identity_number, signed_status, sign_type, sign_time, document_ref, created_at "
            ", source_last_seen_at, updated_at FROM registry_property_certificates "
            "WHERE property_id=%s ORDER BY updated_at DESC, id DESC",
            (property_id,),
        )
        certificates = await cur.fetchall()
    reveal_sensitive = REGISTRY_IMPORT_MANAGE in set(user.get("permissions") or [])
    certificate_items = [
        {
            "id": int(item[0]), "source_ref": item[1], "source_row": item[2],
            "community": item[3], "address": item[4],
            "landlord_name": item[5] if reveal_sensitive else (str(item[5] or "")[:1] + "***" if item[5] else ""),
            "landlord_identity_number": item[6] if reveal_sensitive else "",
            "actual_renter_name": item[7] if reveal_sensitive else (str(item[7] or "")[:1] + "***" if item[7] else ""),
            "actual_renter_identity_number": item[8] if reveal_sensitive else "",
            "signed_status": item[9], "sign_type": item[10], "sign_time": _iso(item[11]),
            "document_ref": item[12], "created_at": _iso(item[13]),
            "source_last_seen_at": _iso(item[14]), "updated_at": _iso(item[15]),
        }
        for item in certificates
    ]
    latest_certificate = certificates[0] if certificates else None
    responsibility_identity = "未确认"
    if latest_certificate and latest_certificate[7]:
        role_names = {
            str(item[4] or "")
            for item in people
            if item[2] == latest_certificate[7] and item[6] is None
        }
        is_sublessor = any("二房东" in name for name in role_names)
        is_manager = any("管家" in name for name in role_names)
        if is_sublessor and is_manager:
            responsibility_identity = "既是二房东又是管家"
        elif is_sublessor:
            responsibility_identity = "二房东"
        elif is_manager:
            responsibility_identity = "管家"
        elif latest_certificate[5] and latest_certificate[5] == latest_certificate[7]:
            responsibility_identity = "普通房东"
    certificate_summary = certificate_status_summary(
        housing_type=row[7],
        certificate_count=len(certificates),
        landlord_name=latest_certificate[5] if latest_certificate else "",
        actual_renter_name=latest_certificate[7] if latest_certificate else "",
        signed_status=latest_certificate[9] if latest_certificate else "",
        sign_type=latest_certificate[10] if latest_certificate else "",
        updated_at=_iso(latest_certificate[14] or latest_certificate[15]) if latest_certificate else None,
        responsibility_identity=responsibility_identity,
    )
    return {
        "id": int(row[0]), "street": row[1], "community_id": row[2],
        "community_name": row[3], "natural_address": row[4], "building": row[5],
        "room": row[6], "housing_type": row[7], "residence_type": row[8],
        "source_house_no": row[9], "source_updated_at": _iso(row[10]),
        "source_type": row[11], "source_ref": row[12], "normalized_address": row[13], "status": row[14],
        "version": int(row[15]), "created_at": _iso(row[16]), "updated_at": _iso(row[17]),
        "aliases": [
            {"id": int(item[0]), "alias": item[1], "community_id": item[2], "enabled": bool(item[3]),
             "source_type": item[4], "created_at": _iso(item[5])} for item in aliases
        ],
        "versions": [
            {"version": int(item[0]), "street": item[1], "natural_address": item[2],
             "building": item[3], "room": item[4], "normalized_address": item[5],
             "effective_from": _iso(item[6]), "effective_to": _iso(item[7]),
             "source_type": item[8], "reason": item[9], "created_at": _iso(item[10])}
            for item in versions
        ],
        "people": [
            {"relation_id": int(item[0]), "person_id": int(item[1]), "person_name": item[2],
             "role_type_id": int(item[3]), "role_name": item[4], "valid_from": _iso(item[5]),
             "valid_to": _iso(item[6]), "verified": bool(item[7])} for item in people
        ],
        "organizations": [
            {"relation_id": int(item[0]), "organization_id": int(item[1]), "organization_name": item[2],
             "role_type_id": int(item[3]), "role_name": item[4], "valid_from": _iso(item[5]),
             "valid_to": _iso(item[6]), "verified": bool(item[7])} for item in organizations
        ],
        "certificate_summary": certificate_summary,
        "certificates": certificate_items,
    }


@router.put("/properties/{property_id}")
async def update_property(
    property_id: int,
    data: PropertyUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    allowed = await _allowed_community_ids(user, REGISTRY_PROPERTY_MANAGE)
    normalized = data.normalized_address.strip() or " ".join(
        part.strip() for part in (data.street, data.natural_address, data.building, data.room) if part.strip()
    )
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            community_id, community_name = await _canonical_community(
                cur, data.community_id, data.community_name_snapshot
            )
            if allowed is not None and community_id not in allowed:
                raise HTTPException(403, "只能维护所属社区的房屋档案")
            await _property_scope(cur, property_id, user, REGISTRY_PROPERTY_MANAGE)
            await cur.execute(
                "SELECT street, natural_address, building, room, normalized_address, current_version "
                "FROM registry_properties WHERE id=%s FOR UPDATE",
                (property_id,),
            )
            before = await cur.fetchone()
            if not before:
                raise HTTPException(404, "房屋档案不存在")
            previous_version = int(before[5] or 0)
            next_version = max(previous_version + 1, 1)
            now = datetime.utcnow()
            # 兼容早期没有首个版本记录的房屋：先补齐当前旧值，再关闭旧版本。
            await cur.execute(
                "SELECT id FROM registry_property_address_versions "
                "WHERE property_id=%s AND version_no=%s FOR UPDATE",
                (property_id, previous_version),
            )
            if not await cur.fetchone():
                await cur.execute(
                    "INSERT INTO registry_property_address_versions "
                    "(property_id, version_no, street, natural_address, building, room, normalized_address, "
                    "effective_from, source_type, change_reason, changed_by) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'manual','历史版本补齐',%s)",
                    (property_id, previous_version, before[0], before[1], before[2], before[3], before[4],
                     now, user["id"]),
                )
            await cur.execute(
                "UPDATE registry_property_address_versions SET effective_to=%s "
                "WHERE property_id=%s AND version_no=%s AND effective_to IS NULL",
                (now, property_id, previous_version),
            )
            await cur.execute(
                "INSERT INTO registry_property_address_versions "
                "(property_id, version_no, street, natural_address, building, room, normalized_address, "
                "effective_from, effective_to, source_type, change_reason, changed_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NULL,'manual',%s,%s)",
                (property_id, next_version, data.street.strip(), data.natural_address.strip(), data.building.strip(),
                 data.room.strip(), normalized, now, data.change_reason.strip(), user["id"]),
            )
            await cur.execute(
                "UPDATE registry_properties SET street=%s, community_id=%s, community_name_snapshot=%s, "
                "natural_address=%s, building=%s, room=%s, housing_type=%s, residence_type=%s, source_house_no=%s, "
                "source_updated_at=COALESCE(%s,source_updated_at), source_type=COALESCE(%s,source_type), "
                "source_ref=COALESCE(%s,source_ref), normalized_address=%s, current_version=%s, "
                "updated_by=%s WHERE id=%s",
                (data.street.strip(), community_id, community_name,
                 data.natural_address.strip(), data.building.strip(), data.room.strip(), data.housing_type.strip(),
                 data.residence_type.strip(), data.source_house_no.strip(), data.source_updated_at,
                 data.source_type.strip() if data.source_type else None,
                 data.source_ref.strip() if data.source_ref is not None else None, normalized,
                 next_version, user["id"], property_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.property.update", target_type="registry_property",
        target_name=str(property_id), detail={"community_id": community_id},
        **request_audit_fields(request),
    )
    return {"message": "房屋档案已更新"}


@router.post("/properties/{property_id}/status")
async def change_property_status(
    property_id: int,
    data: StatusChange,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    async with conn.cursor() as cur:
        await _property_scope(cur, property_id, user, REGISTRY_PROPERTY_MANAGE)
        if data.status == "inactive":
            await cur.execute(
                "SELECT 1 FROM registry_property_person_roles "
                "WHERE property_id=%s AND (valid_to IS NULL OR valid_to>=UTC_TIMESTAMP()) LIMIT 1",
                (property_id,),
            )
            if await cur.fetchone():
                raise HTTPException(409, "房屋仍有人事关系，不能停用房屋档案")
            await cur.execute(
                "SELECT 1 FROM registry_property_organization_roles "
                "WHERE property_id=%s AND (valid_to IS NULL OR valid_to>=UTC_TIMESTAMP()) LIMIT 1",
                (property_id,),
            )
            if await cur.fetchone():
                raise HTTPException(409, "房屋仍有机构关系，不能停用房屋档案")
        await cur.execute(
            "UPDATE registry_properties SET status=%s, updated_by=%s WHERE id=%s",
            (data.status, user["id"], property_id),
        )
    await record_admin_audit(
        user, "registry.property.status", target_type="registry_property",
        target_name=str(property_id), detail={"status": data.status, "reason_length": len(data.reason)},
        **request_audit_fields(request),
    )
    return {"message": "房屋档案状态已更新"}


@router.post("/properties/{property_id}/aliases")
async def create_property_alias(
    property_id: int,
    data: AliasCreate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    alias = " ".join(data.alias.strip().split())
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            property_community = await _property_scope(cur, property_id, user, REGISTRY_PROPERTY_MANAGE)
            requested_community = data.community_id if data.community_id is not None else property_community
            community_id, _ = await _canonical_community(cur, requested_community)
            if property_community is not None and community_id != property_community:
                raise HTTPException(422, "地址别名必须与房屋所属社区一致")
            await cur.execute(
                "INSERT INTO registry_address_aliases "
                "(property_id, alias, normalized_alias, community_id, source_type, created_by) "
                "VALUES (%s,%s,%s,%s,'manual',%s)",
                (property_id, data.alias.strip(), alias.lower(), community_id, user["id"]),
            )
            alias_id = int(cur.lastrowid)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.alias.create", target_type="registry_address_alias",
        target_name=str(alias_id), detail={"property_id": property_id}, **request_audit_fields(request),
    )
    return {"id": alias_id, "message": "地址别名已添加"}


@router.put("/aliases/{alias_id}/status")
async def change_alias_status(
    alias_id: int,
    data: StatusChange,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("SELECT property_id FROM registry_address_aliases WHERE id=%s FOR UPDATE", (alias_id,))
            row = await cur.fetchone()
            if not row:
                raise HTTPException(404, "地址别名不存在")
            await _property_scope(cur, int(row[0]), user, REGISTRY_PROPERTY_MANAGE)
            await cur.execute(
                "UPDATE registry_address_aliases SET enabled=%s WHERE id=%s",
                (1 if data.status == "active" else 0, alias_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.alias.status", target_type="registry_address_alias",
        target_name=str(alias_id), detail={"status": data.status}, **request_audit_fields(request),
    )
    return {"message": "地址别名状态已更新"}


@router.get("/people/{person_id}")
async def get_housing_person_detail(
    person_id: int,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_VIEW)),
    conn=Depends(get_registry_db),
):
    allowed = await _allowed_community_ids(user, REGISTRY_PROPERTY_VIEW)
    async with conn.cursor() as cur:
        if allowed is not None:
            if not allowed:
                raise HTTPException(403, "无权查看该人员档案")
            placeholders = ",".join(["%s"] * len(allowed))
            await cur.execute(
                "SELECT 1 FROM registry_property_person_roles rel JOIN registry_properties prop ON prop.id=rel.property_id "
                f"WHERE rel.person_id=%s AND prop.community_id IN ({placeholders}) LIMIT 1",
                (person_id, *allowed),
            )
            if not await cur.fetchone():
                raise HTTPException(403, "无权查看该人员档案")
        await cur.execute(
            "SELECT id, name, identity_number, is_temporary, verification_status, status, merged_into_id, "
            "source_type, created_at, updated_at FROM registry_housing_people WHERE id=%s",
            (person_id,),
        )
        row = await cur.fetchone()
        if not row:
            raise HTTPException(404, "辖区人员档案不存在")
        await cur.execute(
            "SELECT id, phone, is_primary, verified, valid_from, valid_to, source_type, created_at "
            "FROM registry_person_phones WHERE person_id=%s ORDER BY is_primary DESC, id DESC",
            (person_id,),
        )
        phones = await cur.fetchall()
        await cur.execute(
            "SELECT rel.id, prop.id, prop.normalized_address, role.name, rel.valid_from, rel.valid_to, rel.verified "
            "FROM registry_property_person_roles rel JOIN registry_properties prop ON prop.id=rel.property_id "
            "JOIN registry_role_types role ON role.id=rel.role_type_id WHERE rel.person_id=%s ORDER BY rel.id DESC",
            (person_id,),
        )
        properties = await cur.fetchall()
    include_identity = _can_view_identity(user)
    return {
        "id": int(row[0]), "name": row[1],
        "identity_number": row[2] or "" if include_identity else "",
        "has_identity": bool(row[2]) if include_identity else False,
        "is_temporary": bool(row[3]), "verification_status": row[4], "status": row[5],
        "merged_into_id": row[6], "source_type": row[7], "created_at": _iso(row[8]),
        "updated_at": _iso(row[9]),
        "phones": [
            {"id": int(item[0]), "phone": item[1], "is_primary": bool(item[2]),
             "verified": bool(item[3]), "valid_from": _iso(item[4]), "valid_to": _iso(item[5]),
             "source_type": item[6], "created_at": _iso(item[7])} for item in phones
        ],
        "properties": [
            {"relation_id": int(item[0]), "property_id": int(item[1]), "address": item[2],
             "role_name": item[3], "valid_from": _iso(item[4]), "valid_to": _iso(item[5]),
             "verified": bool(item[6])} for item in properties
        ],
    }


@router.put("/people/{person_id}")
async def update_housing_person(
    person_id: int,
    data: HousingPersonUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    identity = normalize_identity(data.identity_number)
    if identity and not _can_view_identity(user):
        raise HTTPException(403, "身份证号只能由超级管理员编辑")
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT id, identity_number, identity_hmac, identity_hmac_version "
            "FROM registry_housing_people WHERE id=%s",
            (person_id,),
        )
        existing = await cur.fetchone()
        if not existing:
            raise HTTPException(404, "辖区人员档案不存在")
        if _can_view_identity(user):
            digest, version = hmac_digest(identity, kind="identity")
        else:
            identity = str(existing[1] or "")
            digest = existing[2]
            version = int(existing[3] or 1)
        if digest:
            await cur.execute(
                "SELECT id FROM registry_housing_people WHERE identity_hmac=%s AND id<>%s",
                (digest, person_id),
            )
            if await cur.fetchone():
                raise HTTPException(409, "该身份证号已存在其他辖区人员档案")
        await cur.execute(
            "UPDATE registry_housing_people SET name=%s, identity_number=%s, identity_hmac=%s, "
            "identity_hmac_version=%s, is_temporary=%s, verification_status=%s, updated_by=%s WHERE id=%s",
            (data.name.strip(), identity or None, digest, version, int(data.is_temporary),
             data.verification_status, user["id"], person_id),
        )
    await record_admin_audit(
        user, "registry.person.update", target_type="registry_housing_person",
        target_name=str(person_id), detail={}, **request_audit_fields(request),
    )
    return {"message": "辖区人员档案已更新"}


@router.post("/people/{person_id}/phones")
async def add_person_phone(
    person_id: int,
    data: PhoneCreate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    if data.valid_to and data.valid_from and data.valid_to < data.valid_from:
        raise HTTPException(422, "号码失效时间不能早于生效时间")
    phone = normalize_phone(data.phone)
    if not phone:
        raise HTTPException(422, "手机号不能为空")
    digest, version = hmac_digest(phone, kind="phone")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("SELECT id FROM registry_housing_people WHERE id=%s FOR UPDATE", (person_id,))
            if not await cur.fetchone():
                raise HTTPException(404, "辖区人员档案不存在")
            if data.is_primary:
                await cur.execute("UPDATE registry_person_phones SET is_primary=0 WHERE person_id=%s", (person_id,))
            await cur.execute(
                "INSERT INTO registry_person_phones "
                "(person_id, phone, phone_hmac, hmac_version, is_primary, verified, valid_from, valid_to, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (person_id, phone, digest, version, int(data.is_primary), int(data.verified),
                 data.valid_from, data.valid_to, user["id"]),
            )
            phone_id = int(cur.lastrowid)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.phone.create", target_type="registry_person_phone",
        target_name=str(phone_id), detail={"person_id": person_id}, **request_audit_fields(request),
    )
    return {"id": phone_id, "message": "联系电话已添加"}


@router.get("/organizations")
async def list_organizations(
    keyword: str = Query(default="", max_length=100),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_VIEW)),
    conn=Depends(get_registry_db),
):
    allowed = await _allowed_community_ids(user, REGISTRY_PROPERTY_VIEW)
    where = " WHERE status='active'"
    params: list[object] = []
    if allowed is not None:
        if not allowed:
            return {"total": 0, "page": page, "page_size": page_size, "data": []}
        where += (
            " AND EXISTS (SELECT 1 FROM registry_property_organization_roles relation "
            "JOIN registry_properties property ON property.id=relation.property_id "
            "WHERE relation.organization_id=registry_organizations.id AND property.community_id IN ("
            + ",".join(["%s"] * len(allowed)) + "))"
        )
        params.extend(allowed)
    if keyword.strip():
        where += " AND (name LIKE %s OR license_number LIKE %s)"
        params.extend([f"%{keyword.strip()}%", f"%{keyword.strip()}%"])
    async with conn.cursor() as cur:
        await cur.execute(f"SELECT COUNT(*) FROM registry_organizations{where}", tuple(params))
        total = int((await cur.fetchone())[0])
        await cur.execute(
            "SELECT id, name, organization_type, license_number, status, notes, created_at, updated_at "
            f"FROM registry_organizations{where} ORDER BY id DESC LIMIT %s OFFSET %s",
            tuple(params) + (page_size, (page - 1) * page_size),
        )
        rows = await cur.fetchall()
    return {"total": total, "page": page, "page_size": page_size, "data": [
        {"id": int(row[0]), "name": row[1], "organization_type": row[2],
         "license_number": row[3], "status": row[4], "notes": row[5],
         "created_at": _iso(row[6]), "updated_at": _iso(row[7])} for row in rows
    ]}


@router.get("/organizations/{organization_id}")
async def get_organization_detail(
    organization_id: int,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_VIEW)),
    conn=Depends(get_registry_db),
):
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT id, name, organization_type, license_number, status, notes, created_at, updated_at "
            "FROM registry_organizations WHERE id=%s",
            (organization_id,),
        )
        row = await cur.fetchone()
        if not row:
            raise HTTPException(404, "机构档案不存在")
        await _organization_scope(cur, organization_id, user, REGISTRY_PROPERTY_VIEW)
        await cur.execute(
            "SELECT membership.id, person.id, person.name, membership.title, membership.valid_from, "
            "membership.valid_to, membership.verified "
            "FROM registry_organization_memberships membership "
            "JOIN registry_housing_people person ON person.id=membership.person_id "
            "WHERE membership.organization_id=%s ORDER BY membership.valid_to IS NULL DESC, membership.id DESC",
            (organization_id,),
        )
        members = await cur.fetchall()
        await cur.execute(
            "SELECT relation.id, property.id, property.normalized_address, role.id, role.name, "
            "relation.valid_from, relation.valid_to, relation.verified "
            "FROM registry_property_organization_roles relation "
            "JOIN registry_properties property ON property.id=relation.property_id "
            "JOIN registry_role_types role ON role.id=relation.role_type_id "
            "WHERE relation.organization_id=%s ORDER BY relation.valid_to IS NULL DESC, relation.id DESC",
            (organization_id,),
        )
        properties = await cur.fetchall()
    return {
        "id": int(row[0]), "name": row[1], "organization_type": row[2],
        "license_number": row[3], "status": row[4], "notes": row[5],
        "created_at": _iso(row[6]), "updated_at": _iso(row[7]),
        "members": [
            {"membership_id": int(item[0]), "person_id": int(item[1]), "person_name": item[2],
             "title": item[3], "valid_from": _iso(item[4]), "valid_to": _iso(item[5]),
             "verified": bool(item[6])} for item in members
        ],
        "properties": [
            {"relation_id": int(item[0]), "property_id": int(item[1]), "normalized_address": item[2],
             "role_type_id": int(item[3]), "role_name": item[4], "valid_from": _iso(item[5]),
             "valid_to": _iso(item[6]), "verified": bool(item[7])} for item in properties
        ],
    }


@router.post("/organizations")
async def create_organization(
    data: OrganizationPayload,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "INSERT INTO registry_organizations "
                "(name, organization_type, license_number, notes, created_by, updated_by) VALUES (%s,%s,%s,%s,%s,%s)",
                (data.name.strip(), data.organization_type, data.license_number.strip(), data.notes.strip(),
                 user["id"], user["id"]),
            )
            organization_id = int(cur.lastrowid)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.organization.create", target_type="registry_organization",
        target_name=str(organization_id), detail={"type": data.organization_type}, **request_audit_fields(request),
    )
    return {"id": organization_id, "message": "机构档案已创建"}


@router.put("/organizations/{organization_id}")
async def update_organization(
    organization_id: int,
    data: OrganizationPayload,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await _organization_scope(cur, organization_id, user, REGISTRY_PROPERTY_MANAGE)
            await cur.execute(
                "UPDATE registry_organizations SET name=%s, organization_type=%s, license_number=%s, notes=%s, "
                "updated_by=%s WHERE id=%s",
                (data.name.strip(), data.organization_type, data.license_number.strip(), data.notes.strip(),
                 user["id"], organization_id),
            )
            if cur.rowcount != 1:
                raise HTTPException(404, "机构档案不存在")
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.organization.update", target_type="registry_organization",
        target_name=str(organization_id), detail={"type": data.organization_type}, **request_audit_fields(request),
    )
    return {"message": "机构档案已更新"}


@router.post("/organizations/{organization_id}/members")
async def attach_organization_member(
    organization_id: int,
    data: OrganizationMemberCreate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    if data.valid_to and data.valid_from and data.valid_to < data.valid_from:
        raise HTTPException(422, "任职结束时间不能早于生效时间")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await _organization_scope(cur, organization_id, user, REGISTRY_PROPERTY_MANAGE)
            await cur.execute("SELECT id FROM registry_organizations WHERE id=%s AND status='active' FOR UPDATE", (organization_id,))
            if not await cur.fetchone():
                raise HTTPException(404, "机构档案不存在")
            await cur.execute("SELECT id FROM registry_housing_people WHERE id=%s AND status='active' FOR UPDATE", (data.person_id,))
            if not await cur.fetchone():
                raise HTTPException(404, "辖区人员档案不存在")
            await _ensure_relation_interval_available(
                cur,
                "registry_organization_memberships",
                "organization_id=%s AND person_id=%s",
                (organization_id, data.person_id),
                data.valid_from,
                data.valid_to,
            )
            await cur.execute(
                "INSERT INTO registry_organization_memberships "
                "(organization_id, person_id, title, valid_from, valid_to, verified, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (organization_id, data.person_id, data.title.strip(), data.valid_from, data.valid_to,
                 int(data.verified), user["id"]),
            )
            relation_id = int(cur.lastrowid)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.organization_member.attach", target_type="registry_organization_membership",
        target_name=str(relation_id), detail={"organization_id": organization_id, "person_id": data.person_id},
        **request_audit_fields(request),
    )
    return {"id": relation_id, "message": "机构经办人关系已保存"}


@router.post("/properties/{property_id}/organizations")
async def attach_property_organization(
    property_id: int,
    data: PropertyOrganizationRoleCreate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    if data.valid_to and data.valid_from and data.valid_to < data.valid_from:
        raise HTTPException(422, "关系结束时间不能早于生效时间")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await _property_scope(cur, property_id, user, REGISTRY_PROPERTY_MANAGE)
            await cur.execute("SELECT id FROM registry_organizations WHERE id=%s AND status='active' FOR UPDATE", (data.organization_id,))
            if not await cur.fetchone():
                raise HTTPException(404, "机构档案不存在")
            await cur.execute(
                "SELECT id FROM registry_role_types WHERE id=%s AND subject_type='organization' AND is_active=1",
                (data.role_type_id,),
            )
            if not await cur.fetchone():
                raise HTTPException(400, "该角色不能关联机构")
            await _ensure_relation_interval_available(
                cur,
                "registry_property_organization_roles",
                "property_id=%s AND organization_id=%s AND role_type_id=%s",
                (property_id, data.organization_id, data.role_type_id),
                data.valid_from,
                data.valid_to,
            )
            await cur.execute(
                "INSERT INTO registry_property_organization_roles "
                "(property_id, organization_id, role_type_id, valid_from, valid_to, verified, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (property_id, data.organization_id, data.role_type_id, data.valid_from, data.valid_to,
                 int(data.verified), user["id"]),
            )
            relation_id = int(cur.lastrowid)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.property_organization.attach", target_type="registry_property_organization_role",
        target_name=str(relation_id), detail={"property_id": property_id, "organization_id": data.organization_id},
        **request_audit_fields(request),
    )
    return {"id": relation_id, "message": "房屋机构关系已保存"}


@router.put("/property-person-relations/{relation_id}")
async def update_property_person_relation(
    relation_id: int,
    data: RelationUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    if data.valid_to and data.valid_from and data.valid_to < data.valid_from:
        raise HTTPException(422, "关系结束时间不能早于生效时间")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("SELECT property_id, person_id, role_type_id FROM registry_property_person_roles WHERE id=%s FOR UPDATE", (relation_id,))
            row = await cur.fetchone()
            if not row:
                raise HTTPException(404, "房屋人员关系不存在")
            await _property_scope(cur, int(row[0]), user, REGISTRY_PROPERTY_MANAGE)
            await _ensure_relation_interval_available(
                cur,
                "registry_property_person_roles",
                "property_id=%s AND person_id=%s AND role_type_id=%s",
                (row[0], row[1], row[2]),
                data.valid_from,
                data.valid_to,
                exclude_id=relation_id,
            )
            await cur.execute(
                "UPDATE registry_property_person_roles SET valid_from=%s, valid_to=%s, verified=%s WHERE id=%s",
                (data.valid_from, data.valid_to, int(data.verified), relation_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.property_person.update", target_type="registry_property_person_role",
        target_name=str(relation_id), detail={"has_end": bool(data.valid_to)}, **request_audit_fields(request),
    )
    return {"message": "房屋人员关系已更新"}


@router.put("/property-organization-relations/{relation_id}")
async def update_property_organization_relation(
    relation_id: int,
    data: RelationUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    if data.valid_to and data.valid_from and data.valid_to < data.valid_from:
        raise HTTPException(422, "关系结束时间不能早于生效时间")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("SELECT property_id, organization_id, role_type_id FROM registry_property_organization_roles WHERE id=%s FOR UPDATE", (relation_id,))
            row = await cur.fetchone()
            if not row:
                raise HTTPException(404, "房屋机构关系不存在")
            await _property_scope(cur, int(row[0]), user, REGISTRY_PROPERTY_MANAGE)
            await _ensure_relation_interval_available(
                cur,
                "registry_property_organization_roles",
                "property_id=%s AND organization_id=%s AND role_type_id=%s",
                (row[0], row[1], row[2]),
                data.valid_from,
                data.valid_to,
                exclude_id=relation_id,
            )
            await cur.execute(
                "UPDATE registry_property_organization_roles SET valid_from=%s, valid_to=%s, verified=%s WHERE id=%s",
                (data.valid_from, data.valid_to, int(data.verified), relation_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.property_organization.update", target_type="registry_property_organization_role",
        target_name=str(relation_id), detail={"has_end": bool(data.valid_to)}, **request_audit_fields(request),
    )
    return {"message": "房屋机构关系已更新"}


@router.put("/organization-memberships/{membership_id}")
async def update_organization_membership(
    membership_id: int,
    data: OrganizationMembershipUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    if data.valid_to and data.valid_from and data.valid_to < data.valid_from:
        raise HTTPException(422, "任职结束时间不能早于生效时间")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute("SELECT id, organization_id FROM registry_organization_memberships WHERE id=%s FOR UPDATE", (membership_id,))
            membership = await cur.fetchone()
            if not membership:
                raise HTTPException(404, "机构经办人关系不存在")
            await _organization_scope(cur, int(membership[1]), user, REGISTRY_PROPERTY_MANAGE)
            await cur.execute(
                "SELECT person_id FROM registry_organization_memberships WHERE id=%s",
                (membership_id,),
            )
            membership_person = await cur.fetchone()
            await _ensure_relation_interval_available(
                cur,
                "registry_organization_memberships",
                "organization_id=%s AND person_id=%s",
                (membership[1], membership_person[0]),
                data.valid_from,
                data.valid_to,
                exclude_id=membership_id,
            )
            await cur.execute(
                "UPDATE registry_organization_memberships SET title=%s, valid_from=%s, valid_to=%s, verified=%s WHERE id=%s",
                (data.title.strip(), data.valid_from, data.valid_to, int(data.verified), membership_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.organization_member.update", target_type="registry_organization_membership",
        target_name=str(membership_id), detail={"has_end": bool(data.valid_to)}, **request_audit_fields(request),
    )
    return {"message": "机构经办人关系已更新"}


def _find_header(headers: list[str], *names: str) -> int | None:
    normalized = [normalize_text(item).replace(" ", "") for item in headers]
    wanted = {normalize_text(name).replace(" ", "") for name in names}
    for index, value in enumerate(normalized):
        if value in wanted:
            return index
    return None


def _parse_household_workbook(content: bytes) -> list[dict]:
    """读取户号表，按中文表头解析，所有标识字段都先转为文本。"""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(422, "户号表不是可读取的 XLSX 文件") from exc
    rows: list[dict] = []
    for sheet in workbook.worksheets:
        iterator = iter(sheet.iter_rows(values_only=True))
        header_row = None
        indexes: dict[str, int] = {}
        for row_number, raw in enumerate(iterator, start=1):
            headers = [normalize_text(value) for value in raw]
            address_index = _find_header(headers, "出租屋地址", "详细地址", "标准详细地址")
            type_index = _find_header(headers, "住房类型", "房屋类型", "类型")
            if address_index is not None and type_index is not None:
                header_row = row_number
                def optional_index(*names: str) -> int:
                    found = _find_header(headers, *names)
                    return found if found is not None else -1
                indexes = {
                    "community": optional_index("社区名称", "社区", "所属社区"),
                    "police_station": optional_index("派出所名称", "派出所"),
                    "community_code": optional_index("社区代码", "社区编号"),
                    "house_no": optional_index("居住房屋编号", "房屋编号", "户号"),
                    "landlord": optional_index("房主", "房东"),
                    "address": address_index,
                    "housing_type": type_index,
                    "residence_type": optional_index("居住处所", "居住场所"),
                    "resident_count": optional_index("居住人数", "人数"),
                    "updated_at": optional_index("更新时间", "更新日期"),
                }
                break
            if row_number >= 20:
                break
        if header_row is None:
            continue
        for physical_row, raw in enumerate(iterator, start=header_row + 1):
            if not any(value not in (None, "") for value in raw):
                continue
            def cell(key: str):
                index = indexes[key]
                return raw[index] if index >= 0 and index < len(raw) else ""
            updated_at = cell("updated_at")
            if isinstance(updated_at, datetime):
                parsed_updated_at = updated_at
            else:
                text_value = normalize_text(updated_at)
                try:
                    parsed_updated_at = datetime.fromisoformat(text_value.replace("/", "-")) if text_value else None
                except ValueError:
                    parsed_updated_at = None
            rows.append({
                "source_sheet": sheet.title,
                "source_row": physical_row,
                "community": cell("community"),
                "police_station": cell("police_station"),
                "community_code": cell("community_code"),
                "house_no": cell("house_no"),
                "landlord": cell("landlord"),
                "address": cell("address"),
                "housing_type": cell("housing_type"),
                "residence_type": cell("residence_type"),
                "resident_count": cell("resident_count"),
                "updated_at": parsed_updated_at.isoformat() if parsed_updated_at else normalize_text(updated_at),
            })
    if not rows:
        raise HTTPException(422, "未找到包含出租屋地址和住房类型的户号表表头")
    workbook.close()
    return rows


def _issue_payload(payload: dict) -> dict:
    """避免把 datetime 等对象直接交给 JSON 编码器。"""
    return {key: value.isoformat() if isinstance(value, datetime) else value for key, value in payload.items()}


def _chunked(values: list[tuple], size: int = REGISTRY_IMPORT_WRITE_CHUNK):
    for offset in range(0, len(values), size):
        yield values[offset:offset + size]


async def _bulk_insert_source_records(cur, values: list[tuple]) -> None:
    sql = (
        "INSERT INTO registry_source_records "
        "(batch_id, source_ref, entity_type, payload_json) VALUES (%s,%s,%s,%s)"
    )
    await _executemany_chunked(cur, sql, values)


async def _bulk_insert_import_issues(cur, values: list[tuple]) -> None:
    sql = (
        "INSERT INTO registry_import_issues "
        "(batch_id, issue_type, source_type, source_ref, entity_key, payload_json, reason) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)"
    )
    await _executemany_chunked(cur, sql, values)


async def _executemany_chunked(cur, sql: str, values: list[tuple]) -> None:
    for chunk in _chunked(values):
        await cur.executemany(sql, chunk)


def _household_source_ref(row: dict) -> str:
    sheet = normalize_text(row.get("source_sheet"))
    physical_row = normalize_text(row.get("source_row"))
    return f"{sheet}:{physical_row}"[:190] if sheet else physical_row[:190]


def _source_datetime(value):
    """Only pass valid datetimes to DATETIME columns; preserve invalid source text in preview payload."""
    if isinstance(value, datetime):
        return value
    text = normalize_text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("/", "-"))
    except ValueError:
        return None


def _issue_public(
    row,
    include_payload: bool = False,
    reveal_sensitive: bool = False,
    group_payloads: list[dict] | None = None,
) -> dict:
    payload = _json(row[6], {})
    safe_payload = payload if reveal_sensitive else _redact_sensitive_payload(payload)
    safe_group_payloads = group_payloads or []
    if not reveal_sensitive:
        safe_group_payloads = [_redact_sensitive_payload(item) for item in safe_group_payloads]
    return {
        "id": int(row[0]), "batch_id": row[1], "issue_type": row[2], "source_type": row[3],
        "source_ref": row[4], "entity_key": row[5], "payload": safe_payload if include_payload else {},
        "reason": row[7], "status": row[8], "review_note": row[9], "reviewed_by": row[10],
        "reviewed_at": _iso(row[11]), "created_at": _iso(row[12]),
        "problem_details": issue_problem_details(
            str(row[2]),
            safe_payload,
            entity_key=str(row[5] or ""),
            group_payloads=safe_group_payloads or None,
        ),
    }


@router.post("/imports/households/preview")
async def preview_household_import(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(413, "户号表不能超过 50MB")
    rows = _parse_household_workbook(content)
    classified = classify_household_rows(rows)
    file_hash = hashlib.sha256(content).hexdigest()
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT id, status FROM registry_source_batches WHERE source_type='household' AND file_sha256=%s",
                (file_hash,),
            )
            existing = await cur.fetchone()
            if existing:
                batch_id = int(existing[0])
                await conn.rollback()
                return {
                    "batch_id": batch_id, "status": existing[1], "idempotent": True,
                    "total_count": len(rows), "normal_count": classified["normal_count"],
                    "issue_count": classified["issue_count"], "duplicate_groups": classified["duplicate_groups"],
                    "other_type_count": classified["other_type_count"],
                }
            await cur.execute(
                "INSERT INTO registry_source_batches (source_type, file_name, file_sha256, status, imported_count, candidate_count, conflict_count, created_by) "
                "VALUES ('household',%s,%s,'preview',0,%s,%s,%s)",
                (normalize_text(file.filename)[:255], file_hash, classified["normal_count"], classified["issue_count"], user["id"]),
            )
            batch_id = int(cur.lastrowid)
            await _bulk_insert_source_records(cur, [
                (
                    batch_id,
                    _household_source_ref(row),
                    "household_property",
                    json.dumps(_issue_payload(row), ensure_ascii=False),
                )
                for row in classified["rows"]
            ])
            await _bulk_insert_import_issues(cur, [
                (
                    batch_id,
                    issue["issue_type"],
                    "household",
                    _household_source_ref(issue["payload"]),
                    issue["entity_key"],
                    json.dumps(_issue_payload(issue["payload"]), ensure_ascii=False),
                    issue["reason"],
                )
                for issue in classified["issues"]
            ])
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    result = {
        "batch_id": batch_id, "status": "preview", "idempotent": False,
        "total_count": len(rows), "normal_count": classified["normal_count"],
        "issue_count": classified["issue_count"], "duplicate_groups": classified["duplicate_groups"],
        "other_type_count": classified["other_type_count"],
        "issue_breakdown": {
            ISSUE_HOUSEHOLD_DUPLICATE: sum(1 for item in classified["issues"] if item["issue_type"] == ISSUE_HOUSEHOLD_DUPLICATE),
            ISSUE_HOUSEHOLD_MISSING_TYPE: sum(1 for item in classified["issues"] if item["issue_type"] == ISSUE_HOUSEHOLD_MISSING_TYPE),
        },
    }
    await record_admin_audit(
        user,
        "registry.household_import.preview",
        target_type="registry_source_batch",
        target_name=str(batch_id),
        detail={
            "total_count": len(rows),
            "normal_count": classified["normal_count"],
            "issue_count": classified["issue_count"],
            "duplicate_groups": classified["duplicate_groups"],
        },
        **request_audit_fields(request),
    )
    return result


@router.post("/imports/households/{batch_id}/confirm")
async def confirm_household_import(
    batch_id: int,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    imported = 0
    try:
        async with conn.cursor() as cur:
            await cur.execute("SELECT status FROM registry_source_batches WHERE id=%s AND source_type='household' FOR UPDATE", (batch_id,))
            batch = await cur.fetchone()
            if not batch:
                raise HTTPException(404, "户号表导入批次不存在")
            if str(batch[0]) == "imported":
                await conn.rollback()
                return {"batch_id": batch_id, "status": "imported", "imported_count": 0, "idempotent": True}
            await cur.execute(
                "SELECT id, source_ref, payload_json FROM registry_source_records WHERE batch_id=%s AND entity_type='household_property' ORDER BY id",
                (batch_id,),
            )
            records = await cur.fetchall()
            await cur.execute("SELECT source_ref FROM registry_import_issues WHERE batch_id=%s AND status='pending'", (batch_id,))
            blocked_refs = {str(row[0]) for row in await cur.fetchall()}
            import_rows: list[tuple[int, str, dict]] = []
            for record_id, source_ref, payload_json in records:
                if str(source_ref) in blocked_refs:
                    continue
                payload = _json(payload_json, {})
                address = normalize_text(payload.get("address"))
                normalized = normalize_address(address)
                if not normalized:
                    continue
                import_rows.append((int(record_id), str(source_ref), payload))

            # Resolve aliases once per source community instead of once per row.
            community_cache: dict[str, tuple[int | None, str] | None] = {}
            unresolved_community_issues: list[tuple] = []
            resolved_rows: list[tuple[int, str, dict, int | None, str, str]] = []
            for record_id, source_ref, payload in import_rows:
                community_name = normalize_community(payload.get("community"))
                if community_name not in community_cache:
                    try:
                        community_cache[community_name] = await _household_import_community(
                            cur,
                            community_name,
                        )
                    except HTTPException as exc:
                        if exc.status_code not in {409, 422}:
                            raise
                        community_cache[community_name] = None
                resolved_community = community_cache[community_name]
                if resolved_community is None:
                    unresolved_community_issues.append((
                        batch_id,
                        ISSUE_HOUSEHOLD_COMMUNITY_UNRESOLVED,
                        "household",
                        source_ref,
                        normalize_address(payload.get("address")) or source_ref,
                        json.dumps(_issue_payload(payload), ensure_ascii=False),
                        "所属社区无法唯一对应到启用的正式社区或别名，需人工核对",
                    ))
                    continue
                community_id, canonical_name = resolved_community
                address = normalize_text(payload.get("address"))
                resolved_rows.append((record_id, source_ref, payload, community_id, canonical_name, normalize_address(address)))
            await _bulk_insert_import_issues(cur, unresolved_community_issues)

            # Fetch and lock existing properties in chunks to avoid an N+1 query
            # for the large historical workbook.
            existing_cache: dict[tuple[str, int | None], tuple] = {}
            for offset in range(0, len(resolved_rows), 500):
                chunk = resolved_rows[offset:offset + 500]
                keys = sorted({item[5] for item in chunk if item[5]})
                if not keys:
                    continue
                placeholders = ",".join(["%s"] * len(keys))
                await cur.execute(
                    "SELECT id, community_id, street, natural_address, building, room, normalized_address "
                    f"FROM registry_properties WHERE normalized_address IN ({placeholders}) FOR UPDATE",
                    tuple(keys),
                )
                for existing_row in await cur.fetchall():
                    existing_cache[(str(existing_row[6]), existing_row[1])] = existing_row

            new_rows: list[tuple[int, str, dict, int | None, str, str]] = []
            for record_id, source_ref, payload, community_id, canonical_name, normalized in resolved_rows:
                address = normalize_text(payload.get("address"))
                existing = existing_cache.get((normalized, community_id))
                if existing:
                    await cur.execute(
                        "UPDATE registry_properties SET community_id=%s, community_name_snapshot=%s, natural_address=%s, "
                        "housing_type=%s, residence_type=%s, source_house_no=%s, source_updated_at=%s, source_type=%s, source_ref=%s, "
                        "updated_by=%s WHERE id=%s",
                        (community_id, canonical_name, address, payload.get("housing_type") or "", payload.get("residence_type") or "",
                         payload.get("house_no") or "", _source_datetime(payload.get("updated_at")), "household", str(source_ref), user["id"], existing[0]),
                    )
                    property_id = int(existing[0])
                    previous_address = (str(existing[2] or ""), str(existing[3] or ""), str(existing[4] or ""), str(existing[5] or ""))
                    if previous_address != ("", address, "", ""):
                        await cur.execute(
                            "SELECT COALESCE(MAX(version_no), 0) FROM registry_property_address_versions WHERE property_id=%s",
                            (property_id,),
                        )
                        next_version = int((await cur.fetchone())[0]) + 1
                        await cur.execute(
                            "UPDATE registry_property_address_versions SET effective_to=UTC_TIMESTAMP() "
                            "WHERE property_id=%s AND effective_to IS NULL",
                            (property_id,),
                        )
                        await cur.execute(
                            "INSERT INTO registry_property_address_versions "
                            "(property_id, version_no, street, natural_address, building, room, normalized_address, "
                            "effective_from, source_type, source_ref, change_reason, changed_by) "
                            "VALUES (%s,%s,'',%s,'','',%s,UTC_TIMESTAMP(),'household',%s,'户号表导入地址更新',%s)",
                            (property_id, next_version, address, normalized, str(source_ref), user["id"]),
                        )
                    else:
                        await cur.execute(
                            "SELECT 1 FROM registry_property_address_versions WHERE property_id=%s LIMIT 1",
                            (property_id,),
                        )
                        if not await cur.fetchone():
                            await cur.execute(
                                "INSERT INTO registry_property_address_versions "
                                "(property_id, version_no, street, natural_address, building, room, normalized_address, "
                                "effective_from, source_type, source_ref, change_reason, changed_by) "
                                "VALUES (%s,1,'',%s,'','',%s,UTC_TIMESTAMP(),'household',%s,'户号表导入补齐初始地址',%s)",
                                (property_id, address, normalized, str(source_ref), user["id"]),
                            )
                    await cur.execute("UPDATE registry_source_records SET entity_id=%s WHERE id=%s", (property_id, record_id))
                    imported += 1
                else:
                    new_rows.append((record_id, source_ref, payload, community_id, canonical_name, normalized))

            await _executemany_chunked(
                cur,
                "INSERT INTO registry_properties (street, community_id, community_name_snapshot, natural_address, building, room, "
                "housing_type, residence_type, source_house_no, source_updated_at, source_type, source_ref, normalized_address, created_by, updated_by) "
                "VALUES ('',%s,%s,%s,'','',%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                [
                    (
                        community_id,
                        canonical_name,
                        normalize_text(payload.get("address")),
                        payload.get("housing_type") or "",
                        payload.get("residence_type") or "",
                        payload.get("house_no") or "",
                        _source_datetime(payload.get("updated_at")),
                        "household",
                        str(source_ref),
                        normalized,
                        user["id"],
                        user["id"],
                    )
                    for _, source_ref, payload, community_id, canonical_name, normalized in new_rows
                ],
            )

            inserted_cache: dict[tuple[str, int | None], int] = {}
            for offset in range(0, len(new_rows), REGISTRY_IMPORT_WRITE_CHUNK):
                chunk = new_rows[offset:offset + REGISTRY_IMPORT_WRITE_CHUNK]
                keys = sorted({item[5] for item in chunk if item[5]})
                if not keys:
                    continue
                placeholders = ",".join(["%s"] * len(keys))
                await cur.execute(
                    "SELECT id, community_id, normalized_address FROM registry_properties "
                    f"WHERE normalized_address IN ({placeholders}) ORDER BY id",
                    tuple(keys),
                )
                for property_id, community_id, normalized in await cur.fetchall():
                    inserted_cache[(str(normalized), community_id)] = int(property_id)

            address_versions: list[tuple] = []
            source_links: list[tuple] = []
            for record_id, source_ref, payload, community_id, _, normalized in new_rows:
                property_id = inserted_cache.get((normalized, community_id))
                if property_id is None:
                    raise RuntimeError("户号表导入后无法重新定位新建房屋")
                address_versions.append((
                    property_id,
                    normalize_text(payload.get("address")),
                    normalized,
                    str(source_ref),
                    user["id"],
                ))
                source_links.append((property_id, record_id))
            await _executemany_chunked(
                cur,
                "INSERT INTO registry_property_address_versions "
                "(property_id, version_no, street, natural_address, building, room, normalized_address, "
                "effective_from, source_type, source_ref, change_reason, changed_by) "
                "VALUES (%s,1,'',%s,'','',%s,UTC_TIMESTAMP(),'household',%s,'户号表导入初始地址',%s)",
                address_versions,
            )
            await _executemany_chunked(
                cur,
                "UPDATE registry_source_records SET entity_id=%s WHERE id=%s",
                source_links,
            )
            imported += len(new_rows)
            await cur.execute("SELECT COUNT(*) FROM registry_import_issues WHERE batch_id=%s AND status='pending'", (batch_id,))
            pending_issue_count = int((await cur.fetchone())[0])
            await cur.execute(
                "UPDATE registry_source_batches SET status=%s, imported_count=%s WHERE id=%s",
                ("partially_imported" if pending_issue_count else "imported", imported, batch_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    result = {
        "batch_id": batch_id,
        "status": "partially_imported" if pending_issue_count else "imported",
        "imported_count": imported,
        "idempotent": False,
        "pending_issue_count": pending_issue_count,
    }
    await record_admin_audit(
        user,
        "registry.household_import.confirm",
        target_type="registry_source_batch",
        target_name=str(batch_id),
        detail={"imported_count": imported, "pending_issue_count": pending_issue_count},
        **request_audit_fields(request),
    )
    return result


@router.post("/imports/certificates/preview")
async def preview_certificate_import(
    data: RegistryCertificateImport,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    if not data.rows:
        raise HTTPException(422, "没有可预览的告知书记录")
    classified = classify_certificate_rows(data.rows)
    canonical = json.dumps(data.rows, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    file_hash = hashlib.sha256(canonical).hexdigest()
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT id, status FROM registry_source_batches WHERE source_type='certificate' AND file_sha256=%s",
                (file_hash,),
            )
            existing = await cur.fetchone()
            if existing:
                await conn.rollback()
                return {
                    "batch_id": int(existing[0]), "status": existing[1], "idempotent": True,
                    "total_count": len(data.rows), "normal_count": classified["normal_count"],
                    "issue_count": classified["issue_count"], "problem_row_count": classified["problem_row_count"],
                    "duplicate_groups": classified["duplicate_groups"],
                    "conflict_groups": classified["conflict_groups"],
                }
            await cur.execute(
                "INSERT INTO registry_source_batches (source_type, file_name, file_sha256, status, imported_count, candidate_count, conflict_count, created_by) "
                "VALUES ('certificate',%s,%s,'preview',0,%s,%s,%s)",
                (data.source_name[:255], file_hash, classified["normal_count"], classified["problem_row_count"], classified["conflict_groups"]),
            )
            batch_id = int(cur.lastrowid)
            await _bulk_insert_source_records(cur, [
                (
                    batch_id,
                    f"{data.source_name}:{row.get('source_row') or ''}"[:190],
                    "property_certificate",
                    json.dumps(_issue_payload(row), ensure_ascii=False, default=str),
                )
                for row in classified["rows"]
            ])
            await _bulk_insert_import_issues(cur, [
                (
                    batch_id,
                    issue["issue_type"],
                    "certificate",
                    f"{data.source_name}:{issue['source_ref']}"[:190],
                    issue["entity_key"],
                    json.dumps(_issue_payload(issue["payload"]), ensure_ascii=False, default=str),
                    issue["reason"],
                )
                for issue in classified["issues"]
            ])
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    result = {
        "batch_id": batch_id, "status": "preview", "idempotent": False,
        "total_count": len(data.rows), "normal_count": classified["normal_count"],
        "issue_count": classified["issue_count"], "problem_row_count": classified["problem_row_count"],
        "duplicate_groups": classified["duplicate_groups"],
        "conflict_groups": classified["conflict_groups"],
    }
    await record_admin_audit(
        user,
        "registry.certificate_import.preview",
        target_type="registry_source_batch",
        target_name=str(batch_id),
        detail={key: result[key] for key in ("total_count", "normal_count", "problem_row_count", "duplicate_groups", "conflict_groups")},
        **request_audit_fields(request),
    )
    return result


@router.post("/imports/certificates/source-runs", status_code=202)
async def start_certificate_source_run(
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    _conn=Depends(get_registry_db),
):
    try:
        run, reused = await create_certificate_source_run(int(user["id"]))
    except RuntimeError as exc:
        raise HTTPException(409, str(exc)) from exc
    await record_admin_audit(
        user,
        "registry.certificate_import.preview",
        target_type="registry_certificate_source_run",
        target_name=str(run["id"]),
        result="pending" if not reused else "existing",
        detail={"run_id": run["id"], "status": run["status"], "reused": reused},
        **request_audit_fields(request),
    )
    return {**run, "reused": reused}


@router.get("/imports/certificates/source-runs/latest")
async def latest_certificate_source_run(
    _user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    _conn=Depends(get_registry_db),
):
    return {"data": await get_latest_certificate_source_run()}


@router.get("/imports/certificates/source-runs/{run_id}")
async def certificate_source_run_detail(
    run_id: int,
    _user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    _conn=Depends(get_registry_db),
):
    run = await get_certificate_source_run(run_id)
    if not run:
        raise HTTPException(404, "告知书读取任务不存在")
    return run


@router.post("/imports/certificates/source-runs/{run_id}/retry", status_code=202)
async def retry_certificate_source_run_endpoint(
    run_id: int,
    payload: RegistryCertificateSourceRetry,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    _conn=Depends(get_registry_db),
):
    try:
        run = await retry_certificate_source_run(run_id, restart=payload.restart)
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(409, str(exc)) from exc
    await record_admin_audit(
        user,
        "registry.certificate_import.preview",
        target_type="registry_certificate_source_run",
        target_name=str(run_id),
        result="pending",
        detail={"run_id": run_id, "restart": payload.restart},
        **request_audit_fields(request),
    )
    return run


@router.post("/imports/certificates/source-preview")
async def preview_certificate_source(
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    try:
        source = await fetch_certificate_rows()
    except VisitSourceError as exc:
        raise HTTPException(502, exc.message) from exc
    result = await preview_certificate_import(
        RegistryCertificateImport(source_name="房东责任告知书只读接口", rows=source["rows"]),
        request,
        user,
        conn,
    )
    result["source_record_count"] = source["record_count"]
    result["source_rejected_count"] = source["issue_count"]
    return result


@router.post("/imports/certificates/{batch_id}/confirm")
async def confirm_certificate_import(
    batch_id: int,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    try:
        result = await apply_certificate_batch(conn, batch_id, int(user["id"]))
    except LookupError as exc:
        raise HTTPException(404, str(exc)) from exc
    await record_admin_audit(
        user,
        "registry.certificate_import.confirm",
        target_type="registry_source_batch",
        target_name=str(batch_id),
        detail={
            key: result[key]
            for key in (
                "inserted_count", "updated_count", "unchanged_count",
                "skipped_count", "pending_issue_count",
            )
            if key in result
        },
        **request_audit_fields(request),
    )
    return result


async def _registry_import_issue_search_result(
    data: RegistryIssueSearch,
    user: dict,
    conn,
) -> dict:
    where: list[str] = []
    params: list[object] = []
    if data.status:
        where.append("status=%s")
        params.append(data.status)
    if data.issue_type:
        where.append("issue_type=%s")
        params.append(data.issue_type)
    if data.source_type:
        where.append("source_type=%s")
        params.append(data.source_type)

    allowed = await _allowed_community_ids(user, REGISTRY_PROPERTY_VIEW)
    if allowed is not None and data.community_id is not None and data.community_id not in allowed:
        raise HTTPException(403, "无权查看该社区问题数据")
    scoped_communities = [data.community_id] if data.community_id is not None else allowed
    community_expr = (
        "COALESCE(NULLIF(JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.community')), ''), "
        "NULLIF(JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.community_name')), ''), "
        "NULLIF(JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.社区名称')), ''), "
        "JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.sssq')))"
    )
    if scoped_communities is not None:
        if not scoped_communities:
            return {"total": 0, "page": data.page, "page_size": data.page_size, "data": []}
        placeholders = ",".join(["%s"] * len(scoped_communities))
        where.append(
            "EXISTS (SELECT 1 FROM _communities c LEFT JOIN _community_aliases a ON a.community_id=c.id "
            f"WHERE c.id IN ({placeholders}) AND (c.name={community_expr} OR a.alias={community_expr}))"
        )
        params.extend(scoped_communities)

    housing_expr = (
        "COALESCE(NULLIF(JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.housing_type')), ''), "
        "JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.住房类型')), '')"
    )
    if data.housing_category == "rental":
        where.append(f"{housing_expr} IN (%s,%s)")
        params.extend(["个人出租", "单位出租"])
    elif data.housing_category == "self_owned":
        where.append(f"{housing_expr}=%s")
        params.append("自购房屋")
    elif data.housing_category == "other":
        where.append(f"{housing_expr}<>'' AND {housing_expr} NOT IN (%s,%s,%s)")
        params.extend(["个人出租", "单位出租", "自购房屋"])
    elif data.housing_category == "unmarked":
        where.append(f"{housing_expr}='' ")

    keyword = data.keyword.strip()
    if keyword:
        address_expr = (
            "COALESCE(NULLIF(JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.address')), ''), "
            "NULLIF(JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.normalized_address')), ''), "
            "JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.详细地址')), '')"
        )
        house_no_expr = (
            "COALESCE(NULLIF(JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.house_no')), ''), "
            "JSON_UNQUOTE(JSON_EXTRACT(payload_json, '$.户号')), '')"
        )
        like_value = f"%{keyword}%"
        where.append(
            f"(source_ref LIKE %s OR entity_key LIKE %s OR reason LIKE %s OR {community_expr} LIKE %s "
            f"OR {address_expr} LIKE %s OR {house_no_expr} LIKE %s OR {housing_expr} LIKE %s)"
        )
        params.extend([like_value] * 7)

    clause = " AND ".join(where) if where else "1=1"
    async with conn.cursor() as cur:
        await cur.execute(f"SELECT COUNT(*) FROM registry_import_issues WHERE {clause}", tuple(params))
        total = int((await cur.fetchone())[0])
        await cur.execute(
            "SELECT id, batch_id, issue_type, source_type, source_ref, entity_key, payload_json, "
            "reason, status, review_note, reviewed_by, reviewed_at, created_at "
            f"FROM registry_import_issues WHERE {clause} ORDER BY id DESC LIMIT %s OFFSET %s",
            tuple(params) + (data.page_size, (data.page - 1) * data.page_size),
        )
        rows = await cur.fetchall()

        conflict_batches = sorted({
            int(row[1]) for row in rows
            if row[1] is not None and str(row[2]) == ISSUE_CERTIFICATE_CONTENT_CONFLICT
        })
        conflict_groups: dict[tuple[int, str], list[dict]] = {}
        if conflict_batches:
            placeholders = ",".join(["%s"] * len(conflict_batches))
            await cur.execute(
                "SELECT batch_id, entity_key, payload_json FROM registry_import_issues "
                f"WHERE issue_type=%s AND batch_id IN ({placeholders})",
                (ISSUE_CERTIFICATE_CONTENT_CONFLICT, *conflict_batches),
            )
            for batch_id, entity_key, payload_json in await cur.fetchall():
                conflict_groups.setdefault((int(batch_id), str(entity_key or "")), []).append(
                    _json(payload_json, {})
                )

    reveal_sensitive = REGISTRY_IMPORT_MANAGE in set(user.get("permissions") or [])
    return {
        "total": total,
        "page": data.page,
        "page_size": data.page_size,
        "data": [
            _issue_public(
                row,
                include_payload=True,
                reveal_sensitive=reveal_sensitive,
                group_payloads=conflict_groups.get((int(row[1]), str(row[5] or "")))
                if row[1] is not None else None,
            )
            for row in rows
        ],
    }


@router.get("/import/issues")
async def list_registry_import_issues(
    status: str = Query(default="pending", max_length=20),
    issue_type: str | None = Query(default=None, max_length=60),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=100, ge=1, le=500),
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_VIEW)),
    conn=Depends(get_registry_db),
):
    return await _registry_import_issue_search_result(
        RegistryIssueSearch(
            status=status if status in {"pending", "resolved", "dismissed"} else "",
            issue_type=issue_type or "",
            page=page,
            page_size=page_size,
        ),
        user,
        conn,
    )


@router.post("/import/issues/search")
async def search_registry_import_issues(
    data: RegistryIssueSearch,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_VIEW)),
    conn=Depends(get_registry_db),
):
    """搜索问题房屋；地址、户号和错误值关键词不进入访问日志 URL。"""
    return await _registry_import_issue_search_result(data, user, conn)


@router.post("/import/issues/bulk")
async def create_registry_import_issues(
    data: RegistryIssueBulkCreate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    if not data.items:
        return {"created_count": 0}
    await conn.begin()
    created_count = 0
    try:
        async with conn.cursor() as cur:
            for item in data.items:
                await cur.execute(
                    "SELECT id FROM registry_import_issues WHERE issue_type=%s AND source_type=%s AND source_ref=%s AND status='pending' LIMIT 1",
                    (item.issue_type, item.source_type, item.source_ref),
                )
                if await cur.fetchone():
                    continue
                await cur.execute(
                    "INSERT INTO registry_import_issues (batch_id, issue_type, source_type, source_ref, entity_key, payload_json, reason) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (data.batch_id, item.issue_type, item.source_type, item.source_ref, item.entity_key,
                     json.dumps(item.payload, ensure_ascii=False), item.reason.strip()),
                )
                created_count += 1
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user,
        "registry.import_issue.bulk_create",
        target_type="registry_import_issue",
        target_name=str(data.batch_id or ""),
        detail={"requested_count": len(data.items), "created_count": created_count},
        **request_audit_fields(request),
    )
    return {"created_count": created_count}


@router.post("/import/issues/{issue_id}/review")
async def review_registry_import_issue(
    issue_id: int,
    data: ReviewDecision,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    status = "resolved" if data.action == "accept" else "dismissed"
    async with conn.cursor() as cur:
        await cur.execute(
            "UPDATE registry_import_issues SET status=%s, review_note=%s, reviewed_by=%s, reviewed_at=UTC_TIMESTAMP() "
            "WHERE id=%s AND status='pending'",
            (status, data.reason.strip(), user["id"], issue_id),
        )
        if cur.rowcount != 1:
            raise HTTPException(404, "问题数据不存在或已经处理")
    await record_admin_audit(
        user,
        "registry.import_issue.review",
        target_type="registry_import_issue",
        target_name=str(issue_id),
        detail={"action": data.action, "status": status},
        **request_audit_fields(request),
    )
    return {"id": issue_id, "status": status}


@router.get("/change-candidates")
async def list_change_candidates(
    status: str = Query(default="pending", max_length=20),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    async with conn.cursor() as cur:
        await cur.execute("SELECT COUNT(*) FROM registry_change_candidates WHERE status=%s", (status,))
        total = int((await cur.fetchone())[0])
        await cur.execute(
            "SELECT id, entity_type, entity_id, change_type, payload_json, reason, status, reviewed_by, "
            "reviewed_at, created_at FROM registry_change_candidates WHERE status=%s "
            "ORDER BY id DESC LIMIT %s OFFSET %s",
            (status, page_size, (page - 1) * page_size),
        )
        rows = await cur.fetchall()
    return {"total": total, "page": page, "page_size": page_size, "data": [
        {"id": int(row[0]), "entity_type": row[1], "entity_id": row[2], "change_type": row[3],
         "payload": _json(row[4], {}) if user.get("role") == "super_admin" else _redact_sensitive_payload(_json(row[4], {})),
         "reason": row[5], "status": row[6], "reviewed_by": row[7],
         "reviewed_at": _iso(row[8]), "created_at": _iso(row[9])} for row in rows
    ]}


@router.post("/change-candidates")
async def create_change_candidate(
    data: CandidateCreate,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    safe_payload = dict(data.payload)
    safe_payload.pop("identity_hmac", None)
    safe_payload.pop("phone_hmac", None)
    batch_hash = hashlib.sha256(
        f"{data.source_type}:{data.source_ref}:{datetime.utcnow().isoformat()}".encode("utf-8")
    ).hexdigest()
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "INSERT INTO registry_source_batches (source_type, file_sha256, status, created_by) "
                "VALUES (%s,%s,'review',%s)",
                (data.source_type, batch_hash, user["id"]),
            )
            batch_id = int(cur.lastrowid)
            await cur.execute(
                "INSERT INTO registry_change_candidates "
                "(batch_id, entity_type, entity_id, change_type, payload_json, reason) VALUES (%s,%s,%s,%s,%s,%s)",
                (batch_id, data.entity_type, data.entity_id, data.change_type,
                 json.dumps(safe_payload, ensure_ascii=False), data.reason.strip()),
            )
            candidate_id = int(cur.lastrowid)
            await cur.execute(
                "INSERT INTO registry_source_records (batch_id, source_ref, entity_type, entity_id, payload_json) "
                "VALUES (%s,%s,%s,%s,%s)",
                (batch_id, data.source_ref.strip(), data.entity_type, data.entity_id,
                 json.dumps(safe_payload, ensure_ascii=False)),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    return {"id": candidate_id, "message": "待审核变更已创建"}


@router.post("/change-candidates/{candidate_id}/review")
async def review_change_candidate(
    candidate_id: int,
    data: ReviewDecision,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    applied_entity_id = None
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT entity_type, entity_id, change_type, payload_json, status "
                "FROM registry_change_candidates WHERE id=%s FOR UPDATE",
                (candidate_id,),
            )
            row = await cur.fetchone()
            if not row:
                raise HTTPException(404, "待审核变更不存在")
            if row[4] != "pending":
                raise HTTPException(409, "该变更已经处理")
            if data.action == "accept":
                if str(row[0]) == "property":
                    if row[1] is not None:
                        await _property_scope(cur, int(row[1]), user, REGISTRY_PROPERTY_MANAGE)
                    else:
                        candidate_payload = _json(row[3], {})
                        allowed = await _allowed_community_ids(user, REGISTRY_PROPERTY_MANAGE)
                        community_id, _ = await _canonical_community(
                            cur,
                            int(candidate_payload["community_id"]) if candidate_payload.get("community_id") is not None else None,
                            candidate_payload.get("community_name_snapshot"),
                        )
                        if allowed is not None and community_id not in allowed:
                            raise HTTPException(403, "无权审核该社区的房屋档案")
                elif str(row[0]) == "person" and row[1] is not None:
                    await _housing_person_scope(cur, int(row[1]), user, REGISTRY_PROPERTY_MANAGE)
                elif str(row[0]) == "organization" and row[1] is not None:
                    await _organization_scope(cur, int(row[1]), user, REGISTRY_PROPERTY_MANAGE)
                applied_entity_id = await _apply_candidate_payload(
                    cur,
                    entity_type=str(row[0]),
                    entity_id=int(row[1]) if row[1] is not None else None,
                    change_type=str(row[2]),
                    payload=_json(row[3], {}),
                    user_id=int(user["id"]),
                    allow_identity=_can_view_identity(user),
                )
            await cur.execute(
                "UPDATE registry_change_candidates SET status=%s, entity_id=COALESCE(entity_id,%s), "
                "reviewed_by=%s, reviewed_at=UTC_TIMESTAMP(), reason=%s WHERE id=%s",
                ("accepted" if data.action == "accept" else "rejected", applied_entity_id,
                 user["id"], data.reason.strip(), candidate_id),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.candidate.review", target_type="registry_change_candidate",
        target_name=str(candidate_id), detail={"action": data.action, "applied": bool(applied_entity_id)}, **request_audit_fields(request),
    )
    return {"message": "待审核变更已处理"}


@router.get("/conflicts")
async def list_conflicts(
    status: str = Query(default="pending", max_length=20),
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    del user
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT id, entity_type, entity_key, conflict_type, details_json, status, resolved_by, "
            "resolved_at, created_at FROM registry_conflicts WHERE status=%s ORDER BY id DESC",
            (status,),
        )
        rows = await cur.fetchall()
    return {"data": [
        {"id": int(row[0]), "entity_type": row[1], "entity_key": row[2], "conflict_type": row[3],
         "details": _json(row[4], {}), "status": row[5], "resolved_by": row[6],
         "resolved_at": _iso(row[7]), "created_at": _iso(row[8])} for row in rows
    ]}


@router.post("/conflicts/{conflict_id}/review")
async def review_conflict(
    conflict_id: int,
    data: ReviewDecision,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_IMPORT_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "UPDATE registry_conflicts SET status=%s, resolved_by=%s, resolved_at=UTC_TIMESTAMP() "
                "WHERE id=%s AND status='pending'",
                ("resolved" if data.action == "accept" else "dismissed", user["id"], conflict_id),
            )
            if cur.rowcount != 1:
                raise HTTPException(404, "冲突不存在或已处理")
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.conflict.review", target_type="registry_conflict",
        target_name=str(conflict_id), detail={"action": data.action}, **request_audit_fields(request),
    )
    return {"message": "冲突已处理"}


@router.post("/people/{source_person_id}/merge")
async def merge_housing_people(
    source_person_id: int,
    data: MergeRequest,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    if source_person_id == data.target_person_id:
        raise HTTPException(422, "不能合并到同一档案")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT id, status, identity_hmac FROM registry_housing_people WHERE id IN (%s,%s) FOR UPDATE",
                (source_person_id, data.target_person_id),
            )
            rows = {int(row[0]): (row[1], row[2]) for row in await cur.fetchall()}
            await _housing_person_scope(cur, source_person_id, user, REGISTRY_PROPERTY_MANAGE)
            await _housing_person_scope(cur, data.target_person_id, user, REGISTRY_PROPERTY_MANAGE)
            if rows.get(source_person_id, (None, None))[0] != "active" or rows.get(data.target_person_id, (None, None))[0] != "active":
                raise HTTPException(409, "源档案和目标档案都必须处于启用状态")
            source_identity = rows[source_person_id][1]
            target_identity = rows[data.target_person_id][1]
            if source_identity and target_identity and source_identity != target_identity:
                raise HTTPException(409, "源档案和目标档案身份证号不一致，不能自动合并")
            await cur.execute("SELECT id FROM registry_person_phones WHERE person_id=%s", (source_person_id,))
            phone_ids = [int(row[0]) for row in await cur.fetchall()]
            await cur.execute("SELECT id FROM registry_property_person_roles WHERE person_id=%s", (source_person_id,))
            role_ids = [int(row[0]) for row in await cur.fetchall()]
            await cur.execute("SELECT id FROM registry_organization_memberships WHERE person_id=%s", (source_person_id,))
            membership_ids = [int(row[0]) for row in await cur.fetchall()]
            snapshot = {"phone_ids": phone_ids, "role_ids": role_ids, "membership_ids": membership_ids}
            await cur.execute("UPDATE registry_person_phones SET person_id=%s WHERE person_id=%s", (data.target_person_id, source_person_id))
            await cur.execute("UPDATE registry_property_person_roles SET person_id=%s WHERE person_id=%s", (data.target_person_id, source_person_id))
            await cur.execute("UPDATE registry_organization_memberships SET person_id=%s WHERE person_id=%s", (data.target_person_id, source_person_id))
            await cur.execute(
                "UPDATE registry_housing_people SET status='merged', merged_into_id=%s, updated_by=%s WHERE id=%s",
                (data.target_person_id, user["id"], source_person_id),
            )
            await cur.execute(
                "INSERT INTO registry_merge_history "
                "(source_person_id, target_person_id, action, relation_snapshot, reason, changed_by) "
                "VALUES (%s,%s,'merge',%s,%s,%s)",
                (source_person_id, data.target_person_id, json.dumps(snapshot, ensure_ascii=False), data.reason.strip(), user["id"]),
            )
            merge_id = int(cur.lastrowid)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.person.merge", target_type="registry_merge",
        target_name=str(merge_id), detail={"source_id": source_person_id, "target_id": data.target_person_id},
        **request_audit_fields(request),
    )
    return {"id": merge_id, "message": "人员档案已合并，可从合并历史撤销"}


@router.post("/merges/{merge_id}/undo")
async def undo_housing_person_merge(
    merge_id: int,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT source_person_id, target_person_id, action, relation_snapshot "
                "FROM registry_merge_history WHERE id=%s FOR UPDATE",
                (merge_id,),
            )
            row = await cur.fetchone()
            if row and row[2] == "merge":
                await _housing_person_scope(cur, int(row[0]), user, REGISTRY_PROPERTY_MANAGE)
                await _housing_person_scope(cur, int(row[1]), user, REGISTRY_PROPERTY_MANAGE)
            if not row or row[2] != "merge":
                raise HTTPException(404, "可撤销的合并记录不存在")
            await cur.execute(
                "SELECT 1 FROM registry_merge_history WHERE source_person_id=%s AND action='undo' AND id>%s LIMIT 1",
                (row[0], merge_id),
            )
            if await cur.fetchone():
                raise HTTPException(409, "该合并已经撤销")
            snapshot = _json(row[3], {})
            phone_ids = [int(value) for value in snapshot.get("phone_ids", [])]
            role_ids = [int(value) for value in snapshot.get("role_ids", [])]
            membership_ids = [int(value) for value in snapshot.get("membership_ids", [])]
            if phone_ids:
                placeholders = ",".join(["%s"] * len(phone_ids))
                await cur.execute(f"UPDATE registry_person_phones SET person_id=%s WHERE id IN ({placeholders}) AND person_id=%s", (row[0], *phone_ids, row[1]))
            if role_ids:
                placeholders = ",".join(["%s"] * len(role_ids))
                await cur.execute(f"UPDATE registry_property_person_roles SET person_id=%s WHERE id IN ({placeholders}) AND person_id=%s", (row[0], *role_ids, row[1]))
            if membership_ids:
                placeholders = ",".join(["%s"] * len(membership_ids))
                await cur.execute(f"UPDATE registry_organization_memberships SET person_id=%s WHERE id IN ({placeholders}) AND person_id=%s", (row[0], *membership_ids, row[1]))
            await cur.execute(
                "UPDATE registry_housing_people SET status='active', merged_into_id=NULL, updated_by=%s WHERE id=%s",
                (user["id"], row[0]),
            )
            await cur.execute(
                "INSERT INTO registry_merge_history "
                "(source_person_id, target_person_id, action, relation_snapshot, reason, changed_by) "
                "VALUES (%s,%s,'undo',%s,'撤销合并',%s)",
                (row[0], row[1], json.dumps(snapshot, ensure_ascii=False), user["id"]),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.person.merge_undo", target_type="registry_merge",
        target_name=str(merge_id), detail={}, **request_audit_fields(request),
    )
    return {"message": "人员档案合并已撤销"}


@router.get("/merges")
async def list_housing_person_merges(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    user: dict = Depends(require_permission(REGISTRY_PROPERTY_MANAGE)),
    conn=Depends(get_registry_db),
):
    allowed = await _allowed_community_ids(user, REGISTRY_PROPERTY_MANAGE)
    if allowed == []:
        return {"total": 0, "page": page, "page_size": page_size, "data": []}
    where = ["history.action='merge'"]
    params: list[object] = []
    if allowed is not None:
        placeholders = ",".join(["%s"] * len(allowed))
        where.append(
            "EXISTS (SELECT 1 FROM registry_property_person_roles source_rel "
            "JOIN registry_properties source_property ON source_property.id=source_rel.property_id "
            f"WHERE source_rel.person_id=history.source_person_id AND source_property.community_id IN ({placeholders}))"
        )
        params.extend(allowed)
        where.append(
            "EXISTS (SELECT 1 FROM registry_property_person_roles target_rel "
            "JOIN registry_properties target_property ON target_property.id=target_rel.property_id "
            f"WHERE target_rel.person_id=history.target_person_id AND target_property.community_id IN ({placeholders}))"
        )
        params.extend(allowed)
    clause = " AND ".join(where)
    offset = (page - 1) * page_size
    async with conn.cursor() as cur:
        await cur.execute(f"SELECT COUNT(*) FROM registry_merge_history history WHERE {clause}", tuple(params))
        total = int((await cur.fetchone())[0])
        await cur.execute(
            "SELECT history.id, history.source_person_id, source.name, history.target_person_id, target.name, "
            "history.reason, history.created_at, "
            "EXISTS (SELECT 1 FROM registry_merge_history undo WHERE undo.source_person_id=history.source_person_id "
            "AND undo.action='undo' AND undo.id>history.id) AS undone "
            "FROM registry_merge_history history "
            "JOIN registry_housing_people source ON source.id=history.source_person_id "
            "JOIN registry_housing_people target ON target.id=history.target_person_id "
            f"WHERE {clause} ORDER BY history.id DESC LIMIT %s OFFSET %s",
            tuple(params) + (page_size, offset),
        )
        rows = await cur.fetchall()
    return {"total": total, "page": page, "page_size": page_size, "data": [
        {"id": int(row[0]), "source_person_id": int(row[1]), "source_name": row[2],
         "target_person_id": int(row[3]), "target_name": row[4], "reason": row[5],
         "created_at": _iso(row[6]), "undone": bool(row[7])} for row in rows
    ]}


@router.get("/watch/people/{person_id}")
async def get_watch_person_detail(
    person_id: int,
    user: dict = Depends(require_permission(REGISTRY_WATCH_VIEW)),
    conn=Depends(get_registry_db),
):
    async with conn.cursor() as cur:
        await _watch_person_scope(cur, person_id, user)
        await cur.execute(
            "SELECT id, name, identity_number, verification_status, status, source_type, created_at, updated_at "
            "FROM watch_people WHERE id=%s",
            (person_id,),
        )
        row = await cur.fetchone()
        if not row:
            raise HTTPException(404, "人员标记档案不存在")
        await cur.execute(
            "SELECT assignment.id, category.id, category.name, category.color, category.alert_level, "
            "assignment.valid_from, assignment.valid_to, assignment.released_at, assignment.basis, "
            "assignment.status, assignment.source_type, assignment.created_at "
            "FROM watch_assignments assignment JOIN watch_categories category ON category.id=assignment.category_id "
            "WHERE assignment.person_id=%s ORDER BY assignment.id DESC",
            (person_id,),
        )
        assignments = await cur.fetchall()
        await cur.execute(
            "SELECT version.assignment_id, version.version_no, version.snapshot_json, version.changed_by, version.created_at "
            "FROM watch_assignment_versions version JOIN watch_assignments assignment ON assignment.id=version.assignment_id "
            "WHERE assignment.person_id=%s ORDER BY version.assignment_id, version.version_no DESC",
            (person_id,),
        )
        assignment_versions = await cur.fetchall()
    return {
        "id": int(row[0]), "name": row[1],
        "identity_number": (row[2] or "") if user.get("role") == "super_admin" else "",
        "verification_status": row[3], "status": row[4], "source_type": row[5],
        "created_at": _iso(row[6]), "updated_at": _iso(row[7]),
        "assignments": [
            {"id": int(item[0]), "category_id": int(item[1]), "category_name": item[2],
             "color": item[3], "alert_level": item[4], "valid_from": _iso(item[5]),
             "valid_to": _iso(item[6]), "released_at": _iso(item[7]), "basis": item[8],
             "status": item[9], "source_type": item[10], "created_at": _iso(item[11])}
            for item in assignments
        ],
        "assignment_versions": [
            {"assignment_id": int(item[0]), "version": int(item[1]), "snapshot": _json(item[2], {}),
             "changed_by": item[3], "created_at": _iso(item[4])} for item in assignment_versions
        ],
    }


@router.put("/watch/categories/{category_id}")
async def update_watch_category(
    category_id: int,
    data: WatchCategoryUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_WATCH_MANAGE)),
    conn=Depends(get_registry_db),
):
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "UPDATE watch_categories SET name=%s, parent_id=%s, color=%s, alert_level=%s, is_active=%s, "
                "description=%s WHERE id=%s",
                (data.name.strip(), data.parent_id, data.color, data.alert_level, int(data.is_active),
                 data.description.strip(), category_id),
            )
            if cur.rowcount != 1:
                raise HTTPException(404, "标记分类不存在")
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.watch_category.update", target_type="watch_category",
        target_name=str(category_id), detail={"active": data.is_active}, **request_audit_fields(request),
    )
    return {"message": "标记分类已更新"}


@router.put("/watch/people/{person_id}")
async def update_watch_person(
    person_id: int,
    data: WatchPersonUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_WATCH_MANAGE)),
    conn=Depends(get_registry_db),
):
    identity = normalize_identity(data.identity_number)
    if identity and not _can_view_identity(user):
        raise HTTPException(403, "身份证号只能由超级管理员编辑")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT identity_number, identity_hmac, identity_hmac_version "
                "FROM watch_people WHERE id=%s FOR UPDATE",
                (person_id,),
            )
            existing = await cur.fetchone()
            if not existing:
                raise HTTPException(404, "人员标记档案不存在")
            if _can_view_identity(user):
                digest, version = hmac_digest(identity, kind="identity")
            else:
                identity = str(existing[0] or "")
                digest = existing[1]
                version = int(existing[2] or 1)
            if digest:
                await cur.execute("SELECT id FROM watch_people WHERE identity_hmac=%s AND id<>%s", (digest, person_id))
                if await cur.fetchone():
                    raise HTTPException(409, "该身份证号已存在其他人员标记档案")
            await cur.execute(
                "UPDATE watch_people SET name=%s, identity_number=%s, identity_hmac=%s, identity_hmac_version=%s, "
                "verification_status=%s, status=%s, updated_by=%s WHERE id=%s",
                (data.name.strip(), identity or None, digest, version, data.verification_status,
                 data.status, user["id"], person_id),
            )
            if cur.rowcount != 1:
                raise HTTPException(404, "人员标记档案不存在")
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.watch_person.update", target_type="watch_person",
        target_name=str(person_id), detail={}, **request_audit_fields(request),
    )
    return {"message": "人员标记档案已更新"}


@router.put("/watch/assignments/{assignment_id}")
async def update_watch_assignment(
    assignment_id: int,
    data: WatchAssignmentUpdate,
    request: Request,
    user: dict = Depends(require_permission(REGISTRY_WATCH_MANAGE)),
    conn=Depends(get_registry_db),
):
    if data.valid_to and data.valid_to < data.valid_from:
        raise HTTPException(422, "标记结束时间不能早于生效时间")
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT person_id, category_id, source_type, source_ref, status, valid_from, valid_to, released_at, basis "
                "FROM watch_assignments WHERE id=%s FOR UPDATE",
                (assignment_id,),
            )
            before = await cur.fetchone()
            if not before:
                raise HTTPException(404, "人员标记不存在")
            await cur.execute(
                "SELECT COALESCE(MAX(version_no),0)+1 FROM watch_assignment_versions WHERE assignment_id=%s",
                (assignment_id,),
            )
            version_no = int((await cur.fetchone())[0])
            snapshot = {
                "person_id": before[0], "category_id": before[1], "source_type": before[2],
                "source_ref": before[3], "status": before[4], "valid_from": _iso(before[5]),
                "valid_to": _iso(before[6]), "released_at": _iso(before[7]), "basis": before[8],
            }
            await cur.execute(
                "INSERT INTO watch_assignment_versions (assignment_id, version_no, snapshot_json, changed_by) "
                "VALUES (%s,%s,%s,%s)",
                (assignment_id, version_no, json.dumps(snapshot, ensure_ascii=False), user["id"]),
            )
            await cur.execute(
                "UPDATE watch_assignments SET valid_from=%s, valid_to=%s, released_at=%s, basis=%s, status=%s, "
                "updated_by=%s WHERE id=%s",
                (data.valid_from, data.valid_to, data.released_at, data.basis.strip(), data.status,
                 user["id"], assignment_id),
            )
        # 有效期可能覆盖历史任务；回填采用 INSERT IGNORE，解除不会删除既有快照。
        backfilled = await backfill_assignment_snapshots(conn, assignment_id)
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    await record_admin_audit(
        user, "registry.watch_assignment.update", target_type="watch_assignment",
        target_name=str(assignment_id), detail={"status": data.status, "backfilled": int(backfilled)}, **request_audit_fields(request),
    )
    return {"message": "人员标记已更新"}
