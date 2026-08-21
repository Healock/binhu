"""全链条反馈导出、原始公安网数据比对和腾讯来源行归档。"""

from __future__ import annotations

import io
import json
import os
import re
from datetime import datetime
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from config import settings
from services.police_dispatch import (
    MAX_POLICE_FILE_BYTES,
    _cell_text,
    _header_index,
    _locate_header,
    normalize_identity,
    read_workbook_rows,
)
from services.registry_security import hmac_digest


RAW_ALIASES = {
    "identity": ("身份证号", "身份证号码", "身份证", "证件号码"),
    "name": ("姓名", "人员姓名", "名字"),
    "result": ("核查结果", "反馈结果", "登记情况", "结果"),
}


def archive_dir() -> Path:
    path = Path(settings.FULLCHAIN_ARCHIVE_DIR)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _safe_filename(filename: str, fallback: str) -> str:
    return Path((filename or fallback).replace("\\", "/")).name[:180]


def parse_police_raw(content: bytes, filename: str) -> dict[str, Any]:
    sheets = read_workbook_rows(content, filename)
    sheet_name, rows, header_row, mapping = _locate_header(
        sheets, RAW_ALIASES, {"identity"}
    )
    identities: list[str] = []
    preview: list[dict[str, str]] = []
    invalid = 0
    duplicate = 0
    seen: set[str] = set()
    for row_number, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        values = {field: _cell_text(row[index]) if index < len(row) else "" for field, index in mapping.items()}
        identity = normalize_identity(values.get("identity"))
        if not identity:
            continue
        digest, _ = hmac_digest(identity, kind="identity")
        if not digest:
            invalid += 1
            continue
        if digest in seen:
            duplicate += 1
        seen.add(digest)
        identities.append(digest)
        if len(preview) < 30:
            masked = identity[:6] + "*" * max(len(identity) - 10, 0) + identity[-4:]
            preview.append({
                "row": str(row_number),
                "name": values.get("name", ""),
                "identity": masked,
                "result": values.get("result", ""),
            })
    if not identities:
        raise ValueError("未读取到有效身份证号，不能作为公安网原始数据比对")
    return {
        "filename": _safe_filename(filename, "公安网原始数据.xlsx"),
        "file_sha256": sha256(content).hexdigest(),
        "sheet_name": sheet_name,
        "header_row": header_row + 1,
        "row_count": len(identities),
        "invalid_count": invalid,
        "duplicate_count": duplicate,
        "preview": preview,
        "identity_hmacs": list(dict.fromkeys(identities)),
    }


def build_archive_workbook(rows: list[dict[str, Any]], export_id: int, exported_at: datetime) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    summary.append(["全链条反馈归档"])
    summary.append(["导出编号", export_id])
    summary.append(["导出时间", exported_at.strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["总数", len(rows)])
    headers = ["来源行", "来源", "姓名", "身份证号", "电话号码", "地址", "登记情况", "核查结果", "归档类别"]
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        groups.setdefault(str(row.get("category") or "其他"), []).append(row)
    for category, items in groups.items():
        sheet = workbook.create_sheet(category[:31])
        sheet.append(headers)
        for item in items:
            sheet.append([item.get(key, "") for key in ("physical_row", "source", "name", "identity", "phone", "address", "registration", "result", "category")])
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.number_format = "@"
                if isinstance(cell.value, str):
                    cell.data_type = "s"
                if cell.column in {4, 5}:
                    cell.quotePrefix = True
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[chr(64 + column)].width = 22
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
