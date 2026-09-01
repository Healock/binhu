"""辖区自购自住人员资产资料。

资料导入时校验身份证并保存 HMAC，姓名、身份证号和个人联系电话进入受权限
保护的人员档案；匹配当前及后续模型三任务后将空白/无法核实结果更新为
“近期返吴”。不调用全民防写接口，也不保存原始名单行或工作簿。
"""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, Iterable, TypeVar
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from config import settings
from services.qmf_registration import valid_identity, normalize_identity
from services.registry_security import hmac_digest, normalize_phone
from services.online_source import source_row_hash, rebuild_projection_rows
from services.parsers import get_parser
from services.watch_matching import sync_current_task_snapshots_for_keys


MODEL_THREE_PARSER = "疑似未注销模型三"
RULE_VERSION = "self-owned-v2"
MAX_ZIP_BYTES = 100 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 500 * 1024 * 1024
MAX_FILES = 30
PROJECTION_REFRESH_BATCH_SIZE = 500
MAX_ROWS = 150_000
WRITE_CHUNK = 800

T = TypeVar("T")
REGISTRY_SCHEMA = settings.MYSQL_REGISTRY_DB.replace("`", "``")


def _registry_table(name: str) -> str:
    if name not in {
        "registry_housing_people",
        "registry_person_phones",
        "watch_people",
        "watch_person_phones",
        "watch_categories",
        "watch_assignments",
        "watch_assignment_versions",
    }:
        raise ValueError("unsupported registry table")
    return f"`{REGISTRY_SCHEMA}`.`{name}`"


class SelfOwnedImportError(ValueError):
    pass


@dataclass(frozen=True)
class SelfOwnedPerson:
    identity_number: str
    identity_hmac: str
    identity_hmac_version: int
    name: str = ""
    phone: str = ""
    phone_hmac: str = ""
    phone_hmac_version: int = 1


@dataclass(frozen=True)
class ParsedSelfOwned:
    file_sha256: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    duplicate_rows: int
    identities: tuple[tuple[str, int], ...]
    workbook_count: int
    names: tuple[tuple[str, str], ...] = ()
    people: tuple[SelfOwnedPerson, ...] = ()


def _chunks(values: list[T], size: int = WRITE_CHUNK) -> Iterable[list[T]]:
    for start in range(0, len(values), size):
        yield values[start:start + size]


def _text(value: Any) -> str:
    return str(value or "").lstrip("\ufeff").strip()


def _header_index(headers: list[Any], *names: str) -> int | None:
    normalized = {_text(value): index for index, value in enumerate(headers)}
    for name in names:
        if name in normalized:
            return normalized[name]
    return None


def parse_self_owned_zip(content: bytes) -> ParsedSelfOwned:
    if not content:
        raise SelfOwnedImportError("自购自住名单文件为空")
    if len(content) > MAX_ZIP_BYTES:
        raise SelfOwnedImportError("自购自住名单 ZIP 超过 100MB")

    people: dict[str, SelfOwnedPerson] = {}
    total_rows = invalid_rows = duplicate_rows = 0
    workbook_count = 0
    try:
        archive = ZipFile(BytesIO(content))
    except BadZipFile as exc:
        raise SelfOwnedImportError("自购自住名单必须是有效 ZIP 文件") from exc
    with archive:
        members = [item for item in archive.infolist() if not item.is_dir()]
        xlsx_members = [
            item for item in members
            if PurePosixPath(item.filename).suffix.lower() == ".xlsx"
        ]
        if not xlsx_members:
            raise SelfOwnedImportError("ZIP 中没有 XLSX 文件")
        if len(xlsx_members) > MAX_FILES:
            raise SelfOwnedImportError("ZIP 中 XLSX 文件数量超过保护阈值")
        if sum(item.file_size for item in xlsx_members) > MAX_UNCOMPRESSED_BYTES:
            raise SelfOwnedImportError("ZIP 解压后的 XLSX 总大小超过保护阈值")
        for member in xlsx_members:
            workbook_count += 1
            try:
                workbook = load_workbook(
                    filename=BytesIO(archive.read(member)),
                    read_only=True,
                    data_only=True,
                )
            except Exception as exc:  # noqa: BLE001 - hide workbook details
                raise SelfOwnedImportError("名单中存在无法读取的 XLSX 文件") from exc
            try:
                for sheet in workbook.worksheets:
                    rows = sheet.iter_rows(values_only=True)
                    headers = list(next(rows, ()))
                    if not any(_text(value) for value in headers):
                        continue
                    identity_index = _header_index(headers, "居民证号", "身份证号", "身份证号码")
                    if identity_index is None:
                        raise SelfOwnedImportError("XLSX 缺少“居民证号”列")
                    name_index = _header_index(headers, "姓名", "居民姓名", "人员姓名")
                    phone_index = _header_index(
                        headers,
                        "个人联系电话",
                        "联系电话",
                        "联系号码",
                        "手机号",
                        "手机号码",
                    )
                    for row in rows:
                        if not any(value not in (None, "") for value in row):
                            continue
                        total_rows += 1
                        if total_rows > MAX_ROWS:
                            raise SelfOwnedImportError("名单记录数超过 150000 条保护阈值")
                        identity = normalize_identity(row[identity_index] if identity_index < len(row) else "")
                        if not valid_identity(identity):
                            invalid_rows += 1
                            continue
                        digest, version = hmac_digest(identity, kind="identity")
                        if not digest:
                            invalid_rows += 1
                            continue
                        name = ""
                        if name_index is not None and name_index < len(row):
                            name = _text(row[name_index])[:100]
                        phone = ""
                        phone_digest = None
                        phone_version = 1
                        if phone_index is not None and phone_index < len(row):
                            phone = normalize_phone(_text(row[phone_index]))[:200]
                            phone_digest, phone_version = hmac_digest(phone, kind="phone")
                        person = SelfOwnedPerson(
                            identity_number=identity,
                            identity_hmac=digest,
                            identity_hmac_version=version,
                            name=name,
                            phone=phone if phone_digest else "",
                            phone_hmac=phone_digest or "",
                            phone_hmac_version=phone_version,
                        )
                        if digest in people:
                            duplicate_rows += 1
                            existing = people[digest]
                            people[digest] = SelfOwnedPerson(
                                identity_number=existing.identity_number,
                                identity_hmac=existing.identity_hmac,
                                identity_hmac_version=existing.identity_hmac_version,
                                name=existing.name or person.name,
                                phone=existing.phone or person.phone,
                                phone_hmac=existing.phone_hmac or person.phone_hmac,
                                phone_hmac_version=(
                                    existing.phone_hmac_version
                                    if existing.phone_hmac
                                    else person.phone_hmac_version
                                ),
                            )
                            continue
                        people[digest] = person
            finally:
                workbook.close()
    if not people:
        raise SelfOwnedImportError("名单中没有有效身份证记录")
    ordered_people = tuple(sorted(people.values(), key=lambda item: item.identity_hmac))
    return ParsedSelfOwned(
        file_sha256=hashlib.sha256(content).hexdigest(),
        total_rows=total_rows,
        valid_rows=len(people),
        invalid_rows=invalid_rows,
        duplicate_rows=duplicate_rows,
        identities=tuple(
            (item.identity_hmac, item.identity_hmac_version)
            for item in ordered_people
        ),
        names=tuple(
            (item.identity_hmac, item.name)
            for item in ordered_people
            if item.name
        ),
        people=ordered_people,
        workbook_count=workbook_count,
    )


async def ensure_self_owned_schema(cur) -> None:
    await cur.execute("""
        CREATE TABLE IF NOT EXISTS _qmf_self_owned_batches (
            id BIGINT AUTO_INCREMENT PRIMARY KEY,
            file_name VARCHAR(255) NOT NULL DEFAULT '',
            file_sha256 CHAR(64) NOT NULL,
            rule_version VARCHAR(40) NOT NULL DEFAULT 'self-owned-v1',
            status VARCHAR(20) NOT NULL DEFAULT 'completed',
            workbook_count INT NOT NULL DEFAULT 0,
            total_rows INT NOT NULL DEFAULT 0,
            valid_rows INT NOT NULL DEFAULT 0,
            invalid_rows INT NOT NULL DEFAULT 0,
            duplicate_rows INT NOT NULL DEFAULT 0,
            matched_tasks INT NOT NULL DEFAULT 0,
            updated_tasks INT NOT NULL DEFAULT 0,
            skipped_tasks INT NOT NULL DEFAULT 0,
            created_by BIGINT DEFAULT NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            completed_at DATETIME DEFAULT NULL,
            error_message VARCHAR(500) NOT NULL DEFAULT '',
            UNIQUE KEY uk_qmf_self_owned_sha (file_sha256),
            INDEX idx_qmf_self_owned_created (created_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    """)
    await cur.execute("""
        CREATE TABLE IF NOT EXISTS _qmf_self_owned_identities (
            batch_id BIGINT NOT NULL,
            identity_hmac CHAR(64) NOT NULL,
            identity_hmac_version SMALLINT UNSIGNED NOT NULL DEFAULT 1,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (batch_id, identity_hmac),
            INDEX idx_qmf_self_owned_identity (identity_hmac, batch_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
    """)


def _public_batch(row: tuple | None) -> dict[str, Any] | None:
    if not row:
        return None
    keys = (
        "id", "file_name", "file_sha256", "rule_version", "status", "workbook_count",
        "total_rows", "valid_rows", "invalid_rows", "duplicate_rows", "matched_tasks",
        "updated_tasks", "skipped_tasks", "created_by", "created_at", "completed_at",
        "error_message",
    )
    data = dict(zip(keys, row))
    for key in ("id", "workbook_count", "total_rows", "valid_rows", "invalid_rows", "duplicate_rows", "matched_tasks", "updated_tasks", "skipped_tasks"):
        data[key] = int(data[key] or 0)
    for key in ("created_at", "completed_at"):
        if data[key] is not None:
            data[key] = data[key].isoformat() + "Z"
    # created_by is an internal account id and is not part of the safe summary.
    data.pop("created_by", None)
    return data


def should_apply_self_owned_result(
    current_result: Any,
    has_local_change: bool,
    *,
    source_count: int = 1,
    conflict: bool = False,
) -> bool:
    """只填充来源唯一的空白/无法核实结果，且不覆盖平台本地编辑。"""
    if has_local_change or source_count != 1 or conflict:
        return False
    normalized = str(current_result or "").strip()
    return not normalized or normalized == "无法核实"


async def _load_registry_people(cur, digests: list[str], *, for_update: bool = False) -> dict[str, tuple[int, str, str]]:
    result: dict[str, tuple[int, str, str]] = {}
    for chunk in _chunks(digests):
        placeholders = ",".join(["%s"] * len(chunk))
        await cur.execute(
            f"SELECT id,identity_hmac,name,identity_number FROM {_registry_table('registry_housing_people')} "
            f"WHERE identity_hmac IN ({placeholders})" + (" FOR UPDATE" if for_update else ""),
            tuple(chunk),
        )
        for person_id, digest, name, identity_number in await cur.fetchall():
            result[str(digest)] = (int(person_id), str(name or ""), str(identity_number or ""))
    return result


async def _load_watch_people(cur, digests: list[str], *, for_update: bool = False) -> dict[str, tuple[int, str, str, str]]:
    result: dict[str, tuple[int, str, str, str]] = {}
    for chunk in _chunks(digests):
        placeholders = ",".join(["%s"] * len(chunk))
        await cur.execute(
            f"SELECT id,identity_hmac,name,identity_number,status FROM {_registry_table('watch_people')} "
            f"WHERE identity_hmac IN ({placeholders})" + (" FOR UPDATE" if for_update else ""),
            tuple(chunk),
        )
        for person_id, digest, name, identity_number, status in await cur.fetchall():
            result[str(digest)] = (
                int(person_id), str(name or ""), str(identity_number or ""), str(status or "")
            )
    return result


async def _store_people_and_tags(
    cur,
    *,
    people: tuple[SelfOwnedPerson, ...],
    batch_id: int,
    user_id: int | None,
    category_id: int,
) -> dict[str, int]:
    """Bulk-upsert archive people, compatibility tag people, phones and assignments."""
    digests = [item.identity_hmac for item in people]
    source_ref = f"batch:{batch_id}"
    registry_existing = await _load_registry_people(cur, digests, for_update=True)
    registry_new = [item for item in people if item.identity_hmac not in registry_existing]
    for chunk in _chunks(registry_new):
        await cur.executemany(
            f"INSERT INTO {_registry_table('registry_housing_people')} "
            "(name,identity_number,identity_hmac,identity_hmac_version,verification_status,"
            "source_type,source_ref,created_by,updated_by) "
            "VALUES (%s,%s,%s,%s,'verified','self_owned_asset',%s,%s,%s)",
            [
                (
                    item.name or "自购自住人员", item.identity_number, item.identity_hmac,
                    item.identity_hmac_version, source_ref, user_id, user_id,
                )
                for item in chunk
            ],
        )
    registry_people = await _load_registry_people(cur, digests, for_update=True)
    registry_updates = []
    for item in people:
        person_id, current_name, current_identity = registry_people[item.identity_hmac]
        if (not current_identity) or current_name in {"", "自购自住人员", "未命名人员"}:
            registry_updates.append((
                item.name or "自购自住人员", item.identity_number,
                item.identity_hmac_version, user_id, person_id,
            ))
    for chunk in _chunks(registry_updates):
        await cur.executemany(
            f"UPDATE {_registry_table('registry_housing_people')} SET "
            "name=IF(TRIM(COALESCE(name,''))='' OR name IN ('自购自住人员','未命名人员'),%s,name),"
            "identity_number=IF(TRIM(COALESCE(identity_number,''))='',%s,identity_number),"
            "identity_hmac_version=COALESCE(identity_hmac_version,%s),updated_by=%s WHERE id=%s",
            chunk,
        )

    watch_existing = await _load_watch_people(cur, digests, for_update=True)
    watch_new = [item for item in people if item.identity_hmac not in watch_existing]
    for chunk in _chunks(watch_new):
        await cur.executemany(
            f"INSERT INTO {_registry_table('watch_people')} "
            "(name,identity_number,identity_hmac,identity_hmac_version,registry_person_id,"
            "verification_status,source_type,source_ref,created_by,updated_by) "
            "VALUES (%s,%s,%s,%s,%s,'verified','self_owned_asset',%s,%s,%s)",
            [
                (
                    item.name or "自购自住人员", item.identity_number, item.identity_hmac,
                    item.identity_hmac_version, registry_people[item.identity_hmac][0],
                    source_ref, user_id, user_id,
                )
                for item in chunk
            ],
        )
    watch_people = await _load_watch_people(cur, digests, for_update=True)
    watch_updates = []
    for item in people:
        person_id, current_name, current_identity, _status = watch_people[item.identity_hmac]
        watch_updates.append((
            registry_people[item.identity_hmac][0], item.name or "自购自住人员",
            item.identity_number, item.identity_hmac_version, user_id, person_id,
        ))
    for chunk in _chunks(watch_updates):
        await cur.executemany(
            f"UPDATE {_registry_table('watch_people')} SET registry_person_id=COALESCE(registry_person_id,%s),"
            "name=IF(TRIM(COALESCE(name,''))='' OR name IN ('自购自住人员','未命名人员'),%s,name),"
            "identity_number=IF(TRIM(COALESCE(identity_number,''))='',%s,identity_number),"
            "identity_hmac_version=COALESCE(identity_hmac_version,%s),updated_by=%s WHERE id=%s",
            chunk,
        )

    phone_people = [item for item in people if item.phone and item.phone_hmac]
    watch_ids = [watch_people[item.identity_hmac][0] for item in phone_people]
    active_watch_phones: set[tuple[int, str]] = set()
    watch_with_phone: set[int] = set()
    for chunk in _chunks(watch_ids):
        placeholders = ",".join(["%s"] * len(chunk))
        await cur.execute(
            f"SELECT person_id,phone_hmac FROM {_registry_table('watch_person_phones')} "
            f"WHERE person_id IN ({placeholders}) AND valid_to IS NULL",
            tuple(chunk),
        )
        for person_id, phone_hmac in await cur.fetchall():
            active_watch_phones.add((int(person_id), str(phone_hmac)))
            watch_with_phone.add(int(person_id))
    watch_phone_rows = []
    for item in phone_people:
        person_id = watch_people[item.identity_hmac][0]
        if (person_id, item.phone_hmac) in active_watch_phones:
            continue
        watch_phone_rows.append((
            person_id, item.phone, item.phone_hmac, item.phone_hmac_version,
            0 if person_id in watch_with_phone else 1, source_ref, user_id,
        ))
        watch_with_phone.add(person_id)
    for chunk in _chunks(watch_phone_rows):
        await cur.executemany(
            f"INSERT INTO {_registry_table('watch_person_phones')} "
            "(person_id,phone,phone_hmac,hmac_version,is_primary,source_type,source_ref,created_by) "
            "VALUES (%s,%s,%s,%s,%s,'self_owned_asset',%s,%s)",
            chunk,
        )

    registry_ids = [registry_people[item.identity_hmac][0] for item in phone_people]
    active_registry_phones: set[tuple[int, str]] = set()
    registry_with_phone: set[int] = set()
    for chunk in _chunks(registry_ids):
        placeholders = ",".join(["%s"] * len(chunk))
        await cur.execute(
            f"SELECT person_id,phone_hmac FROM {_registry_table('registry_person_phones')} "
            f"WHERE person_id IN ({placeholders}) AND valid_to IS NULL",
            tuple(chunk),
        )
        for person_id, phone_hmac in await cur.fetchall():
            active_registry_phones.add((int(person_id), str(phone_hmac)))
            registry_with_phone.add(int(person_id))
    registry_phone_rows = []
    for item in phone_people:
        person_id = registry_people[item.identity_hmac][0]
        if (person_id, item.phone_hmac) in active_registry_phones:
            continue
        registry_phone_rows.append((
            person_id, item.phone, item.phone_hmac, item.phone_hmac_version,
            0 if person_id in registry_with_phone else 1, source_ref, user_id,
        ))
        registry_with_phone.add(person_id)
    for chunk in _chunks(registry_phone_rows):
        await cur.executemany(
            f"INSERT INTO {_registry_table('registry_person_phones')} "
            "(person_id,phone,phone_hmac,hmac_version,is_primary,verified,source_type,source_ref,created_by) "
            "VALUES (%s,%s,%s,%s,%s,1,'self_owned_asset',%s,%s)",
            chunk,
        )

    active_watch_ids = [
        row[0] for row in watch_people.values() if row[3] == "active"
    ]
    existing_assignments: set[int] = set()
    for chunk in _chunks(active_watch_ids):
        placeholders = ",".join(["%s"] * len(chunk))
        await cur.execute(
            f"SELECT person_id FROM {_registry_table('watch_assignments')} "
            f"WHERE person_id IN ({placeholders}) AND category_id=%s AND status='active' "
            "AND valid_from<=UTC_TIMESTAMP() AND (valid_to IS NULL OR valid_to>=UTC_TIMESTAMP()) "
            "AND (released_at IS NULL OR released_at>UTC_TIMESTAMP()) FOR UPDATE",
            tuple(chunk) + (category_id,),
        )
        existing_assignments.update(int(row[0]) for row in await cur.fetchall())
    new_assignment_people = [person_id for person_id in active_watch_ids if person_id not in existing_assignments]
    now = datetime.utcnow()
    for chunk in _chunks(new_assignment_people):
        await cur.executemany(
            f"INSERT INTO {_registry_table('watch_assignments')} "
            "(person_id,category_id,valid_from,source_type,source_ref,basis,created_by,updated_by) "
            "VALUES (%s,%s,%s,'self_owned_asset',%s,%s,%s,%s)",
            [
                (person_id, category_id, now, source_ref, "辖区资产资料：自购自住", user_id, user_id)
                for person_id in chunk
            ],
        )
    if new_assignment_people:
        for chunk in _chunks(new_assignment_people):
            placeholders = ",".join(["%s"] * len(chunk))
            await cur.execute(
                f"SELECT id FROM {_registry_table('watch_assignments')} "
                f"WHERE person_id IN ({placeholders}) AND category_id=%s AND source_ref=%s",
                tuple(chunk) + (category_id, source_ref),
            )
            versions = [(int(row[0]), "{}", user_id) for row in await cur.fetchall()]
            if versions:
                await cur.executemany(
                    f"INSERT IGNORE INTO {_registry_table('watch_assignment_versions')} "
                    "(assignment_id,version_no,snapshot_json,changed_by) VALUES (%s,1,%s,%s)",
                    versions,
                )
    return {
        "registry_people_created": len(registry_new),
        "registry_people_reused": len(people) - len(registry_new),
        "tag_people_created": len(watch_new),
        "tag_people_reused": len(people) - len(watch_new),
        "watch_phones_created": len(watch_phone_rows),
        "registry_phones_created": len(registry_phone_rows),
        "tag_assignments_created": len(new_assignment_people),
    }


async def apply_self_owned_matches(cur, *, batch_id: int | None = None) -> dict[str, Any]:
    """Apply the retained roster to current model-three tasks without external writes."""
    parser = get_parser(MODEL_THREE_PARSER)
    roster_filter = "AND roster.batch_id=%s" if batch_id is not None else "AND batch.status='completed'"
    params: tuple[Any, ...] = (MODEL_THREE_PARSER, batch_id) if batch_id is not None else (MODEL_THREE_PARSER,)
    await cur.execute(
        "SELECT projection.row_key,projection.values_json,projection.source_count,projection.conflict "
        "FROM _online_source_projection projection "
        "WHERE projection.parser_type=%s AND EXISTS ("
        "SELECT 1 FROM _qmf_self_owned_identities roster "
        "JOIN _qmf_self_owned_batches batch ON batch.id=roster.batch_id "
        "WHERE roster.identity_hmac=projection.identity_hmac " + roster_filter + ")",
        params,
    )
    matched = await cur.fetchall()
    updated = skipped = 0
    updated_row_keys: list[str] = []
    matched_row_keys: list[str] = []
    for row_key, raw_values, source_count, conflict in matched:
        matched_row_keys.append(str(row_key))
        values = raw_values if isinstance(raw_values, dict) else json.loads(raw_values or "{}")
        current = str(values.get("核查结果") or "").strip()
        await cur.execute(
            "SELECT 1 FROM _online_local_changes WHERE parser_type=%s AND row_key=%s "
            "AND field_name='核查结果' AND status IN ('pending','processing','retry','conflict') LIMIT 1",
            (MODEL_THREE_PARSER, str(row_key)),
        )
        if not should_apply_self_owned_result(
            current,
            bool(await cur.fetchone()),
            source_count=int(source_count or 0),
            conflict=bool(conflict),
        ):
            skipped += 1
            continue
        await cur.execute(
            "SELECT id,values_json FROM _online_source_rows WHERE parser_type=%s AND row_key=%s",
            (MODEL_THREE_PARSER, str(row_key)),
        )
        source_rows = await cur.fetchall()
        if len(source_rows) != 1:
            skipped += 1
            continue
        for source_id, source_json in source_rows:
            source_values = source_json if isinstance(source_json, dict) else json.loads(source_json or "{}")
            source_values["核查结果"] = "近期返吴"
            encoded = json.dumps(source_values, ensure_ascii=False, sort_keys=True)
            await cur.execute(
                "UPDATE _online_source_rows SET values_json=%s,row_hash=%s,revision=revision+1,"
                "refreshed_at=UTC_TIMESTAMP() WHERE id=%s",
                (encoded, source_row_hash(source_values), int(source_id)),
            )
        await cur.execute(
            f"UPDATE `{parser.table_name}` SET `核查结果`=%s WHERE `_row_key`=%s",
            ("近期返吴", str(row_key)),
        )
        updated += 1
        updated_row_keys.append(str(row_key))
    return {
        "matched_tasks": len(matched),
        "updated_tasks": updated,
        "skipped_tasks": skipped,
        "updated_row_keys": updated_row_keys,
        "matched_row_keys": matched_row_keys,
    }


async def latest_batch(cur) -> dict[str, Any] | None:
    await cur.execute(
        "SELECT id,file_name,file_sha256,rule_version,status,workbook_count,total_rows,valid_rows,"
        "invalid_rows,duplicate_rows,matched_tasks,updated_tasks,skipped_tasks,created_by,created_at,"
        "completed_at,error_message FROM _qmf_self_owned_batches ORDER BY id DESC LIMIT 1"
    )
    return _public_batch(await cur.fetchone())


async def apply_self_owned_import(conn, *, parsed: ParsedSelfOwned, file_name: str, user_id: int | None) -> dict[str, Any]:
    await conn.begin()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT id,rule_version,status FROM _qmf_self_owned_batches "
                "WHERE file_sha256=%s LIMIT 1 FOR UPDATE",
                (parsed.file_sha256,),
            )
            existing = await cur.fetchone()
            if existing:
                batch_id = int(existing[0])
                previous_rule_version = str(existing[1] or "")
                if str(existing[2] or "") == "processing":
                    raise SelfOwnedImportError("该名单文件正在处理中，请勿重复提交")
                if previous_rule_version == RULE_VERSION:
                    raise SelfOwnedImportError("该名单文件已经按当前规则导入过")
                await cur.execute(
                    "UPDATE _qmf_self_owned_batches SET file_name=%s,rule_version=%s,status='processing',"
                    "workbook_count=%s,total_rows=%s,valid_rows=%s,invalid_rows=%s,duplicate_rows=%s,"
                    "created_by=%s,completed_at=NULL,error_message='' WHERE id=%s",
                    (
                        file_name[:255], RULE_VERSION, parsed.workbook_count, parsed.total_rows,
                        parsed.valid_rows, parsed.invalid_rows, parsed.duplicate_rows, user_id, batch_id,
                    ),
                )
                upgraded = True
            else:
                await cur.execute(
                    "INSERT INTO _qmf_self_owned_batches "
                    "(file_name,file_sha256,rule_version,status,workbook_count,total_rows,valid_rows,"
                    "invalid_rows,duplicate_rows,created_by,completed_at) "
                    "VALUES (%s,%s,%s,'processing',%s,%s,%s,%s,%s,%s,NULL)",
                    (
                        file_name[:255], parsed.file_sha256, RULE_VERSION, parsed.workbook_count,
                        parsed.total_rows, parsed.valid_rows, parsed.invalid_rows,
                        parsed.duplicate_rows, user_id,
                    ),
                )
                batch_id = int(cur.lastrowid)
                upgraded = False
            identity_rows = [
                (batch_id, digest, version) for digest, version in parsed.identities
            ]
            for chunk in _chunks(identity_rows):
                await cur.executemany(
                    "INSERT IGNORE INTO _qmf_self_owned_identities "
                    "(batch_id,identity_hmac,identity_hmac_version) VALUES (%s,%s,%s)",
                    chunk,
                )
            await cur.execute(
                f"SELECT id FROM {_registry_table('watch_categories')} "
                "WHERE code=%s AND is_active=1 FOR UPDATE",
                ("self_owned_resident",),
            )
            category_row = await cur.fetchone()
            if not category_row:
                raise SelfOwnedImportError("自购自住人员标签分类未初始化")
            self_owned_category_id = int(category_row[0])
            people_stats = await _store_people_and_tags(
                cur,
                people=parsed.people,
                batch_id=batch_id,
                user_id=user_id,
                category_id=self_owned_category_id,
            )
            match_stats = await apply_self_owned_matches(cur, batch_id=batch_id)
            # The roster matcher already updates only the affected model-three
            # rows. Rebuilding the entire business projection here made large
            # ZIP imports exceed the request timeout; refresh just those rows.
            matched_keys = list(match_stats.pop("updated_row_keys", []))
            # A self-owned tag is also a current personnel-archive fact.  If
            # a task already had a result, the result update is correctly
            # skipped, but its tag snapshot still needs to be materialized so
            # the model-three list can display the tag immediately.
            matched_snapshot_keys = list(match_stats.pop("matched_row_keys", []))
            if matched_snapshot_keys:
                await sync_current_task_snapshots_for_keys(
                    cur, MODEL_THREE_PARSER, matched_snapshot_keys
                )
            for offset in range(0, len(matched_keys), PROJECTION_REFRESH_BATCH_SIZE):
                await rebuild_projection_rows(
                    cur,
                    MODEL_THREE_PARSER,
                    matched_keys[offset:offset + PROJECTION_REFRESH_BATCH_SIZE],
                    reconcile_graph=False,
                )
            await cur.execute(
                "UPDATE _qmf_self_owned_batches SET status='completed',matched_tasks=%s,updated_tasks=%s,skipped_tasks=%s,completed_at=UTC_TIMESTAMP() WHERE id=%s",
                (
                    match_stats["matched_tasks"], match_stats["updated_tasks"],
                    match_stats["skipped_tasks"], batch_id,
                ),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise
    return {
        "batch_id": batch_id,
        "status": "completed",
        "upgraded": upgraded,
        "rule_version": RULE_VERSION,
        "workbook_count": parsed.workbook_count,
        "total_rows": parsed.total_rows,
        "valid_rows": parsed.valid_rows,
        "invalid_rows": parsed.invalid_rows,
        "duplicate_rows": parsed.duplicate_rows,
        **match_stats,
        **people_stats,
    }
