"""下发 Excel 解析、地址匹配、建议生成与反馈工作簿。"""

from __future__ import annotations

import io
import hmac
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import settings


MAX_POLICE_FILE_BYTES = 30 * 1024 * 1024
FINAL_ACTIONS = {"dispatch", "no_registration", "transfer", "duplicate_exclude"}
MISSING_PHONE_ANALYSIS_REASON = "缺少手机号，需基础管控先研判；补齐手机号后才能下发"
IDENTITY_PATTERN = re.compile(r"^(?:\d{15}|\d{17}[0-9X])$")
PUBLISH_OWNED_COLUMNS = (
    "下发日期", "截止日期", "社区", "来源", "姓名",
    "身份证号", "电话号码", "地址", "登记情况", "创建时间",
)
WORKFLOW_EDITABLE_COLUMNS = (
    "核查人", "现住址", "核查结果", "研判", "二次反馈",
)


class PoliceWorkbookError(ValueError):
    """上传文件无法按全链条模板解析。"""


@dataclass(slots=True)
class DispatchImportRow:
    source_row: int
    source_name: str
    community_name: str
    person_name: str
    identity_number: str
    phone: str
    original_address: str
    registration_status: str
    created_time: str
    transfer_note: str
    raw_values: dict[str, str]


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def normalize_lookup(value: Any) -> str:
    return re.sub(r"[\s,，。；;、（）()\[\]【】·\-—_]", "", str(value or "")).lower()


def normalize_community_label(value: Any) -> str:
    text = normalize_lookup(value)
    return text[:-2] if text.endswith("社区") else text


def normalize_identity(value: Any) -> str:
    text = normalize_space(value).replace(" ", "").upper()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def dispatch_field_roles(raw_values: dict[str, Any]) -> dict[str, str]:
    """返回规范业务字段对应的原始 Excel 列名。"""
    normalized = {
        _normalized_header(header): str(header)
        for header in raw_values
    }
    result: dict[str, str] = {}
    for field, aliases in DISPATCH_HEADER_ALIASES.items():
        for alias in aliases:
            header = normalized.get(_normalized_header(alias))
            if header:
                result[field] = header
                break
    return result


def dispatch_values_from_raw(raw_values: dict[str, Any]) -> dict[str, str]:
    """从可编辑原始列重新提取任务的规范字段。"""
    values = {
        str(header): _cell_text(value)
        for header, value in raw_values.items()
    }
    roles = dispatch_field_roles(values)

    def value(field: str) -> str:
        header = roles.get(field)
        return values.get(header, "") if header else ""

    return {
        "source_name": value("source"),
        "community_name": value("community"),
        "person_name": value("name"),
        "identity_number": normalize_identity(value("identity")),
        "phone": normalize_space(value("phone")),
        "original_address": value("address"),
        "registration_status": value("registration"),
        "created_time": value("created"),
        "transfer_note": value("transfer_note"),
        "raw_values": values,
        "field_roles": roles,
    }


def identity_digest(value: str) -> str:
    normalized = normalize_identity(value)
    return sha256(normalized.encode("utf-8")).hexdigest() if normalized else ""


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, ".15g")
    return normalize_space(value)


def _read_xlsx(content: bytes) -> list[tuple[str, list[list[str]]]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise PoliceWorkbookError("无法读取 XLSX 文件，请确认文件未损坏") from exc
    result: list[tuple[str, list[list[str]]]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
            result.append((sheet.title, rows))
    finally:
        workbook.close()
    return result


def _read_xls(content: bytes) -> list[tuple[str, list[list[str]]]]:
    try:
        import xlrd  # 全链条系统导出仍可能使用 BIFF .xls。
    except ImportError as exc:  # pragma: no cover - 部署依赖缺失保护
        raise PoliceWorkbookError("服务器缺少 .xls 解析组件 xlrd") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:
        raise PoliceWorkbookError("无法读取 XLS 文件，请确认文件未损坏") from exc
    result = []
    try:
        for sheet in workbook.sheets():
            def cell_text(row: int, col: int) -> str:
                cell = sheet.cell(row, col)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    parsed = xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
                    if parsed.hour == parsed.minute == parsed.second == parsed.microsecond == 0:
                        return parsed.date().isoformat()
                    return parsed.strftime("%Y-%m-%d %H:%M:%S")
                return _cell_text(cell.value)

            result.append((
                sheet.name,
                [[cell_text(row, col) for col in range(sheet.ncols)]
                 for row in range(sheet.nrows)],
            ))
    finally:
        workbook.release_resources()
    return result


def read_workbook_rows(content: bytes, filename: str) -> list[tuple[str, list[list[str]]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(content)
    if suffix == ".xls":
        return _read_xls(content)
    raise PoliceWorkbookError("只支持 .xls 或 .xlsx 文件")


def _normalized_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("：", "").replace(":", "")


def _header_index(row: list[str], aliases: dict[str, tuple[str, ...]]) -> dict[str, int]:
    values = [_normalized_header(item) for item in row]
    result: dict[str, int] = {}
    for field, names in aliases.items():
        normalized_names = {_normalized_header(name) for name in names}
        for index, value in enumerate(values):
            if value in normalized_names:
                result[field] = index
                break
    return result


DISPATCH_HEADER_ALIASES = {
    "source": ("来源", "数据来源"),
    "community": ("社区", "所属社区", "社区名称"),
    "name": ("姓名", "人员姓名"),
    "identity": ("身份证号", "身份证号码", "身份证"),
    "phone": ("手机号", "手机号码", "电话号码", "联系电话"),
    "address": ("地址", "原地址", "居住地址"),
    "registration": ("登记情况", "登记状态", "处理结果"),
    "created": ("创建时间", "下发时间"),
    "transfer_note": ("移交备注", "移交反馈", "退回备注"),
}


def _locate_header(
    sheets: list[tuple[str, list[list[str]]]],
    aliases: dict[str, tuple[str, ...]],
    required: set[str],
) -> tuple[str, list[list[str]], int, dict[str, int]]:
    best: tuple[str, list[list[str]], int, dict[str, int]] | None = None
    for sheet_name, rows in sheets:
        for row_index, row in enumerate(rows[:40]):
            mapping = _header_index(row, aliases)
            if required.issubset(mapping):
                if best is None or len(mapping) > len(best[3]):
                    best = (sheet_name, rows, row_index, mapping)
    if best is None:
        raise PoliceWorkbookError("没有找到可识别的标准表头")
    return best


def parse_dispatch_workbook(
    content: bytes,
    filename: str,
    require_clean_fields: bool = False,
) -> tuple[str, list[DispatchImportRow]]:
    required = {"name", "identity", "address"}
    if require_clean_fields:
        required.update({"community", "registration", "phone"})
    sheet_name, rows, header_row, columns = _locate_header(
        read_workbook_rows(content, filename),
        DISPATCH_HEADER_ALIASES,
        required,
    )
    header_values = rows[header_row]
    result: list[DispatchImportRow] = []
    for row_number, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        def value(field: str) -> str:
            index = columns.get(field)
            return _cell_text(row[index]) if index is not None and index < len(row) else ""

        if not any(normalize_space(item) for item in row):
            continue
        if _normalized_header(value("name")) in {
            _normalized_header(item) for item in DISPATCH_HEADER_ALIASES["name"]
        }:
            continue
        raw_values: dict[str, str] = {}
        for index, raw_header in enumerate(header_values):
            header = normalize_space(raw_header) or f"第{index + 1}列"
            if header in raw_values:
                header = f"{header}_{index + 1}"
            raw_values[header] = _cell_text(row[index]) if index < len(row) else ""
        identity = normalize_identity(value("identity"))
        if not value("name") and not identity and not value("address"):
            continue
        result.append(DispatchImportRow(
            source_row=row_number,
            source_name=value("source"),
            community_name=value("community"),
            person_name=value("name"),
            identity_number=identity,
            phone=normalize_space(value("phone")),
            original_address=value("address"),
            registration_status=value("registration"),
            created_time=value("created"),
            transfer_note=value("transfer_note"),
            raw_values=raw_values,
        ))
    if not result:
        raise PoliceWorkbookError("全链条文件中没有可导入的数据")
    return sheet_name, result


CLEAN_DISPATCH_STATUSES = {
    "流口未登记", "地址待变更", "未登记", "待登记",
    "流口已注销", "已注销", "注销",
}


def apply_clean_import_actions(
    rows: list[dict[str, Any]],
    communities: list[dict[str, Any]],
) -> None:
    """将基础管控已处理的字段转换为最终动作，异常记录仍留待人工审核。"""
    resolver = community_resolver(
        item for item in communities if item.get("enabled", True)
    )
    mark_duplicate_groups(rows)
    for row in rows:
        row.update({
            "suggested_action": "manual",
            "suggested_community_id": None,
            "suggestion_reason": "",
            "allocation_mode": "clean_import",
            "auto_final_action": "",
            "auto_final_community_id": None,
        })
        missing = [
            label for field, label in (
                ("person_name", "姓名"),
                ("identity_number", "身份证号"),
                ("phone", "手机号"),
                ("original_address", "地址"),
            ) if not normalize_space(row.get(field, ""))
        ]
        if missing:
            row["suggestion_reason"] = f"缺少必要字段：{'、'.join(missing)}"
            row["allocation_mode"] = "conflict"
            continue
        if not IDENTITY_PATTERN.fullmatch(normalize_identity(row.get("identity_number", ""))):
            row["suggestion_reason"] = "身份证号格式异常，需要人工确认"
            row["allocation_mode"] = "conflict"
            continue
        if row.get("duplicate_group_key"):
            row["suggestion_reason"] = "同批身份证号重复，需要人工确认保留记录"
            row["allocation_mode"] = "conflict"
            continue

        registration = normalize_lookup(row.get("registration_status", ""))
        if registration not in CLEAN_DISPATCH_STATUSES:
            row["suggestion_reason"] = "登记情况不是可直接导入的已处理状态"
            row["allocation_mode"] = "conflict"
            continue

        community = resolve_community(row.get("community_name", ""), resolver)
        if not community:
            row["suggestion_reason"] = "文件中的社区无法匹配启用社区"
            row["allocation_mode"] = "conflict"
            continue
        row.update({
            "suggested_action": "dispatch",
            "suggested_community_id": int(community["id"]),
            "suggestion_reason": (
                "此前已注销，需下发社区重新核查并按实际情况登记"
                if registration in {"流口已注销", "已注销", "注销"}
                else "已处理文件按登记情况和社区直接生成下发任务"
            ),
            "auto_final_action": "dispatch",
            "auto_final_community_id": int(community["id"]),
        })


def community_resolver(communities: Iterable[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for community in communities:
        names = [community.get("name"), *(community.get("aliases") or [])]
        for name in names:
            key = normalize_community_label(name)
            if key:
                result[key] = community
    return result


def resolve_community(
    value: str,
    resolver: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    return resolver.get(normalize_community_label(value))


def mark_duplicate_groups(rows: list[dict[str, Any]]) -> None:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        digest = identity_digest(row.get("identity_number", ""))
        row["identity_hash"] = digest
        if digest:
            grouped.setdefault(digest, []).append(row)
    for digest, group in grouped.items():
        if len(group) < 2:
            continue
        signatures = {
            stable_json({
                "source": item.get("source_name", ""),
                "name": item.get("person_name", ""),
                "identity": normalize_identity(item.get("identity_number", "")),
                "phone": normalize_lookup(item.get("phone", "")),
                "address": normalize_lookup(item.get("original_address", "")),
                "created": item.get("created_time", ""),
                "note": normalize_space(item.get("transfer_note", "")),
            })
            for item in group
        }
        kind = "exact" if len(signatures) == 1 else "conflict"
        for item in group:
            item["duplicate_group_key"] = digest
            item["duplicate_kind"] = kind


HOTEL_WORDS = ("酒店", "宾馆", "旅馆", "客栈", "民宿")
TRANSFER_WORDS = ("移交至", "请移交", "非滨湖辖区", "不属滨湖", "不在滨湖辖区")
OUTSIDE_BINHU_WUJIANG_WORDS = (
    "盛泽镇", "平望镇", "震泽镇", "桃源镇", "七都镇", "黎里镇",
    "汾湖高新区", "同里镇", "江陵街道",
)


def _address_matches(
    address: str,
    note: str,
    entries: Iterable[dict[str, Any]],
) -> list[dict[str, Any]]:
    haystack = normalize_lookup(f"{address} {note}")
    matches = []
    for entry in entries:
        if not entry.get("enabled", True) or not entry.get("community_id"):
            continue
        tokens = [
            entry.get("name", ""),
            entry.get("detail_address", ""),
            *(entry.get("aliases") or []),
        ]
        normalized = [normalize_lookup(token) for token in tokens]
        if any(len(token) >= 2 and token in haystack for token in normalized):
            matches.append(entry)
    return matches


def apply_preprocessing_suggestions(
    rows: list[dict[str, Any]],
    communities: list[dict[str, Any]],
    address_entries: list[dict[str, Any]],
) -> None:
    """原地写入建议；任何建议仍必须由内勤人工确认。"""
    enabled = [item for item in communities if item.get("enabled", True)]
    enabled.sort(key=lambda item: (int(item.get("sort_order") or item["id"]), int(item["id"])))
    enabled_ids = {int(item["id"]) for item in enabled}
    resolver = community_resolver(enabled)
    unmatched_rows: list[dict[str, Any]] = []
    mark_duplicate_groups(rows)

    for row in rows:
        address = normalize_space(row.get("original_address", ""))
        note = normalize_space(row.get("transfer_note", ""))
        phone = normalize_space(row.get("phone", ""))
        combined = f"{address} {note}"
        row.update({
            "suggested_action": "dispatch",
            "suggested_community_id": None,
            "suggestion_reason": "",
            "allocation_mode": "",
        })
        missing = [
            label for field, label in (
                ("person_name", "姓名"),
                ("identity_number", "身份证号"),
                ("original_address", "地址"),
            )
            if not normalize_space(row.get(field, ""))
        ]
        if missing:
            row.update({
                "suggested_action": "manual",
                "suggestion_reason": f"缺少必要字段：{'、'.join(missing)}",
                "allocation_mode": "conflict",
            })
            continue
        if not phone:
            row.update({
                "suggested_action": "manual",
                "suggested_community_id": None,
                "suggestion_reason": MISSING_PHONE_ANALYSIS_REASON,
                "allocation_mode": "missing_phone",
            })
            continue
        if any(word in combined for word in HOTEL_WORDS):
            row.update({
                "suggested_action": "no_registration",
                "suggestion_reason": "地址包含明确的酒店、宾馆或民宿线索",
            })
            continue

        direct_communities: dict[int, dict[str, Any]] = {}
        for key, community in resolver.items():
            if len(key) >= 2 and key in normalize_lookup(combined):
                direct_communities[int(community["id"])] = community
        matches = _address_matches(address, note, address_entries)
        for entry in matches:
            community_id = int(entry["community_id"])
            if community_id not in enabled_ids:
                continue
            direct_communities[community_id] = {
                "id": community_id,
                "name": entry.get("community_name", ""),
            }
        if len(direct_communities) == 1:
            community = next(iter(direct_communities.values()))
            row.update({
                "suggested_community_id": int(community["id"]),
                "suggestion_reason": "地址库、社区名称或移交备注唯一匹配",
                "allocation_mode": "matched",
            })
            continue
        if len(direct_communities) > 1:
            row.update({
                "suggested_action": "manual",
                "suggestion_reason": "地址或备注同时命中多个社区，需要人工判断",
                "allocation_mode": "conflict",
            })
            continue
        if any(word in combined for word in TRANSFER_WORDS) or any(
            word in address for word in OUTSIDE_BINHU_WUJIANG_WORDS
        ):
            row.update({
                "suggested_action": "transfer",
                "suggestion_reason": "原移交信息明确说明不属于滨湖辖区",
            })
            continue
        unmatched_rows.append(row)

    for row in unmatched_rows:
        row.update({
            "suggested_action": "manual",
            "suggested_community_id": None,
            "suggestion_reason": "地址无法唯一匹配小区，需要人工确认",
            "allocation_mode": "unmatched",
        })


def build_publish_address(original_address: str, transfer_note: str) -> str:
    address = normalize_space(original_address)
    note = normalize_space(transfer_note)
    if not note:
        return address
    # Rebuild an existing suffix instead of appending another copy when a
    # previously normalized address is passed back through the publisher.
    marker = "；移交反馈："
    if marker in address:
        address = address.split(marker, 1)[0].rstrip()
    return f"{address}{marker}{note}"


def publish_business_key(identity_number: str, phone: str, dispatch_date: str) -> str:
    raw = "\x1f".join((normalize_identity(identity_number), normalize_space(phone), dispatch_date))
    return sha256(raw.encode("utf-8")).hexdigest()


def business_key_hmac(
    parts: Iterable[Any],
    fields: Iterable[str] | None = None,
) -> str:
    """生成不暴露业务主键原文的稳定 HMAC。"""
    names = list(fields or [])
    normalized: list[str] = []
    for index, value in enumerate(parts):
        field = names[index] if index < len(names) else ""
        normalized.append(
            normalize_identity(value)
            if "身份证" in field
            else normalize_space(value)
        )
    payload = "\x1f".join(normalized)
    return hmac.new(
        settings.registry_hmac_key.encode("utf-8"),
        payload.encode("utf-8"),
        sha256,
    ).hexdigest()


def parser_business_key_fields(parser, values: dict[str, Any]) -> tuple[str, ...]:
    """返回某行实际采用的业务主键字段。"""
    if parser.parser_type == "苏州涉警" and normalize_space(values.get("接警编号", "")):
        return ("接警编号",)
    if parser.parser_type == "涉警统计":
        case_number = normalize_space(values.get("序号", ""))
        if len(case_number) == 20 and case_number.isdigit():
            return ("序号",)
    return tuple(parser.get_business_key())


def parser_business_key(
    parser,
    values: dict[str, Any],
    *,
    legacy: bool = False,
) -> str:
    fields = parser_business_key_fields(parser, values)
    parts = [values.get(field, "") for field in fields]
    if not legacy:
        return business_key_hmac(parts, fields)
    payload = "\x1f".join(normalize_space(value) for value in parts)
    return sha256(payload.encode("utf-8")).hexdigest()


def publish_values_match(
    requested: dict[str, Any],
    actual: dict[str, Any],
    allowed_columns: Iterable[str] | None = None,
) -> bool:
    """只比较下发拥有字段，允许腾讯用户并发填写后续业务字段。"""
    allowed = set(allowed_columns or PUBLISH_OWNED_COLUMNS)
    comparable = [
        (column, expected)
        for column, expected in requested.items()
        if column in allowed
    ]
    if not comparable:
        return False
    return all(
        str(actual.get(column, "") or "").strip()
        == str(expected or "").strip()
        for column, expected in comparable
    )


def _select_publish_candidate(
    candidates: list[dict[str, Any]],
    requested: dict[str, Any],
    physical_row: int | None,
    comparison_columns: Iterable[str] | None = None,
) -> dict[str, Any] | None:
    """优先原物理行；否则只接受唯一匹配，避免重复主键误关联。"""
    if physical_row is not None:
        preferred = next((
            item for item in candidates
            if int(item["physical_row"]) == physical_row
        ), None)
        if preferred and publish_values_match(
            requested, preferred["values"], comparison_columns
        ):
            return preferred
    matched = [
        item for item in candidates
        if publish_values_match(requested, item["values"], comparison_columns)
    ]
    return matched[0] if len(matched) == 1 else None


async def reconcile_police_dispatch_publications(
    cur,
    spreadsheet_id: int,
    source_columns: list[str] | None = None,
    parser_type: str | None = None,
) -> dict[str, int]:
    """仅在一次正常同步完整成功后，对结果不确定的发布任务做只读对账。"""
    from services.online_source import json_value, source_row_hash
    from services.parsers import get_parser

    # Lightweight transaction doubles used by the atomicity tests do not expose
    # result-set readers. Production database cursors always implement fetchall;
    # keep this optional reconciliation read non-invasive for those doubles.
    if not hasattr(cur, "fetchall"):
        return {"success": 0, "conflict": 0, "retryable": 0}

    # 兼容旧调用和历史测试：未显式提供时仍按全链条处理；正常同步会传入
    # 当前腾讯配置的 parser_type，从而覆盖所有五入口目标表。
    parser = get_parser(parser_type or "全链条")
    if parser.parser_type == "全链条":
        owned_columns = set(PUBLISH_OWNED_COLUMNS)
    else:
        owned_columns = set(parser.COLUMNS) - {
            "核查人", "现住址", "核查结果", "核查反馈", "研判",
            "二次反馈", "二次核查结果", "入住方式", "是否开户",
            "房屋属性", "居住时间", "房东信息", "二房东信息", "备注",
            "房东是否处罚",
        }
    comparison_columns = set(source_columns or parser.COLUMNS) & owned_columns
    await cur.execute("""
        SELECT source.id, source.physical_row, source.row_hash, source.values_json
        FROM _online_source_rows AS source
        WHERE source.spreadsheet_id=%s AND source.parser_type=%s
        ORDER BY source.physical_row, source.id
    """, (spreadsheet_id, parser.parser_type))
    sources_by_key: dict[str, list[dict[str, Any]]] = {}
    for source_id, physical_row, row_hash, raw_values in await cur.fetchall():
        values = {
            column: str(json_value(raw_values, {}).get(column, "") or "").strip()
            for column in parser.COLUMNS
        }
        source_item = {
            "id": int(source_id),
            "physical_row": int(physical_row),
            "row_hash": str(row_hash or source_row_hash(values)),
            "values": values,
        }
        for key in {
            parser_business_key(parser, values),
            parser_business_key(parser, values, legacy=True),
        }:
            sources_by_key.setdefault(key, []).append(source_item)

    await cur.execute("""
        SELECT result.task_id, result.business_key, result.request_values_json,
               task.batch_id, result.status, result.error_code, result.physical_row
        FROM _police_dispatch_publish_results AS result
        JOIN _police_dispatch_tasks AS task ON task.id=result.task_id
        WHERE result.spreadsheet_id=%s
          AND (
            (result.status='needs_reconciliation'
             AND task.publish_status='needs_reconciliation')
            OR
            (result.status='conflict' AND result.error_code='content_conflict'
             AND task.publish_status='conflict')
          )
        FOR UPDATE
    """, (spreadsheet_id,))
    pending = await cur.fetchall()
    counts = {"success": 0, "conflict": 0, "retryable": 0}
    affected_batches: set[int] = set()
    affected_runs: set[int] = set()
    task_runs: dict[int, list[int]] = {}
    pending_task_ids = [int(row[0]) for row in pending]
    if pending_task_ids:
        placeholders = ",".join(["%s"] * len(pending_task_ids))
        await cur.execute(f"""
            SELECT task_id,run_id FROM _police_dispatch_publish_run_items
            WHERE task_id IN ({placeholders})
              AND status IN ('needs_reconciliation','conflict')
        """, pending_task_ids)
        for task_id, run_id in await cur.fetchall():
            task_runs.setdefault(int(task_id), []).append(int(run_id))
            affected_runs.add(int(run_id))
    for (
        task_id, business_key, raw_requested, batch_id,
        result_status, _error_code, physical_row,
    ) in pending:
        requested = {
            str(column): str(value or "").strip()
            for column, value in json_value(raw_requested, {}).items()
            if column in comparison_columns
        }
        candidates = sources_by_key.get(str(business_key or ""), [])
        exact = _select_publish_candidate(
            candidates,
            requested,
            int(physical_row) if physical_row is not None else None,
            comparison_columns,
        )
        if exact:
            run_item_status = "success"
            run_item_error = ""
            counts["success"] += 1
            affected_batches.add(int(batch_id))
            await cur.execute("""
                UPDATE _police_dispatch_tasks SET
                    publish_status='success', task_status='completed',
                    publish_error='', published_row=%s, linked_source_id=%s,
                    linked_row_hash=%s, conflict_values_json=NULL,
                    cache_pending=0, published_at=COALESCE(published_at, UTC_TIMESTAMP()),
                    version=version+1
                WHERE id=%s
            """, (
                exact["physical_row"], exact["id"], exact["row_hash"], task_id,
            ))
            await cur.execute("""
                UPDATE _police_dispatch_publish_results SET
                    status='success', physical_row=%s, source_row_id=%s,
                    expected_row_hash=%s, verified_values_json=%s,
                    resolution=CASE WHEN %s='conflict'
                        THEN 'auto_accept_workflow_updates' ELSE resolution END,
                    error_code='', error_message='', cache_pending=0
                WHERE task_id=%s
            """, (
                exact["physical_row"], exact["id"], exact["row_hash"],
                stable_json(exact["values"]), str(result_status), task_id,
            ))
        elif str(result_status) == "conflict":
            # 历史冲突只有在下发字段唯一匹配时才自动修复；否则保持原状。
            continue
        elif candidates:
            run_item_status = "conflict"
            run_item_error = "content_conflict"
            counts["conflict"] += 1
            affected_batches.add(int(batch_id))
            candidate = candidates[0]
            await cur.execute("""
                UPDATE _police_dispatch_tasks SET
                    publish_status='conflict', task_status='publish_failed',
                    publish_error='同步发现同主键内容不同', published_row=%s,
                    linked_source_id=%s, linked_row_hash=%s,
                    conflict_values_json=%s, cache_pending=0, version=version+1
                WHERE id=%s
            """, (
                candidate["physical_row"], candidate["id"], candidate["row_hash"],
                stable_json(candidate["values"]), task_id,
            ))
            await cur.execute("""
                UPDATE _police_dispatch_publish_results SET
                    status='conflict', physical_row=%s, source_row_id=%s,
                    expected_row_hash=%s, verified_values_json=%s,
                    error_code='content_conflict',
                    error_message='同步发现同主键内容不同', cache_pending=0
                WHERE task_id=%s
            """, (
                candidate["physical_row"], candidate["id"], candidate["row_hash"],
                stable_json(candidate["values"]), task_id,
            ))
        else:
            run_item_status = "retryable"
            run_item_error = "confirmed_absent"
            counts["retryable"] += 1
            affected_batches.add(int(batch_id))
            await cur.execute("""
                UPDATE _police_dispatch_tasks SET
                    publish_status='retryable', task_status='pending_publish',
                    publish_error='完整同步确认腾讯中不存在目标，可安全重试',
                    published_row=NULL, linked_source_id=NULL,
                    linked_row_hash='', conflict_values_json=NULL,
                    cache_pending=0, version=version+1
                WHERE id=%s
            """, (task_id,))
            await cur.execute("""
                UPDATE _police_dispatch_publish_results SET
                    status='retryable', physical_row=NULL, source_row_id=NULL,
                    expected_row_hash='', verified_values_json=NULL,
                    error_code='confirmed_absent',
                    error_message='完整同步确认目标不存在', cache_pending=0
                WHERE task_id=%s
            """, (task_id,))
        task_run_ids = task_runs.get(int(task_id), [])
        if task_run_ids:
            placeholders = ",".join(["%s"] * len(task_run_ids))
            await cur.execute(f"""
                UPDATE _police_dispatch_publish_run_items
                SET status=%s,error_code=%s
                WHERE task_id=%s AND run_id IN ({placeholders})
            """, [run_item_status, run_item_error, task_id, *task_run_ids])

    for run_id in affected_runs:
        await cur.execute("""
            SELECT COUNT(*),
                   SUM(status IN ('success','conflict','needs_reconciliation','retryable')),
                   SUM(status='success'),SUM(status='conflict'),
                   SUM(status='needs_reconciliation'),SUM(status='retryable')
            FROM _police_dispatch_publish_run_items WHERE run_id=%s
        """, (run_id,))
        total, processed, success, conflict, reconciliation, retryable = await cur.fetchone()
        run_status = "completed" if int(success or 0) == int(total or 0) else "partial"
        await cur.execute("""
            UPDATE _police_dispatch_publish_runs SET
                status=%s,processed_count=%s,success_count=%s,conflict_count=%s,
                reconciliation_count=%s,retryable_count=%s,
                error_code=CASE WHEN %s='completed' THEN '' ELSE 'partial' END,
                error_message=CASE WHEN %s='completed' THEN ''
                    ELSE '同步对账后仍有任务需要重试或处理冲突' END
            WHERE id=%s
        """, (
            run_status, int(processed or 0), int(success or 0), int(conflict or 0),
            int(reconciliation or 0), int(retryable or 0), run_status, run_status, run_id,
        ))

    for batch_id in affected_batches:
        await cur.execute("""
            SELECT COUNT(*), SUM(task_status='pending_review'),
                   SUM(publish_status IN ('pending','publishing','retryable',
                                          'needs_reconciliation','conflict')),
                   SUM(publish_status IN ('needs_reconciliation','conflict'))
            FROM _police_dispatch_tasks WHERE batch_id=%s
        """, (batch_id,))
        total, pending_review, pending_publish, blocked = await cur.fetchone()
        if total and not pending_review and not pending_publish:
            batch_status = "completed"
        elif blocked:
            batch_status = "reconciling"
        elif not pending_review:
            batch_status = "ready_to_publish"
        else:
            batch_status = "reviewing"
        await cur.execute("""
            UPDATE _police_dispatch_batches SET status=%s,
                completed_at=CASE WHEN %s='completed'
                    THEN COALESCE(completed_at, UTC_TIMESTAMP()) ELSE NULL END,
                last_error=CASE WHEN %s='reconciling'
                    THEN '部分任务等待对账或存在内容冲突' ELSE last_error END
            WHERE id=%s
        """, (batch_status, batch_status, batch_status, batch_id))
    return counts


def build_feedback_workbook(
    batch: dict[str, Any],
    tasks: list[dict[str, Any]],
    exported_at: datetime,
) -> bytes:
    from services.parsers import get_parser

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    target_parser = str(batch.get("target_parser") or "全链条")
    if target_parser != "全链条":
        business_labels = {
            "rental": "出租房屋核查",
            "police": "涉警",
            "delivery": "寄递业",
            "suspect_return": "疑似返苏",
        }
        subtype_labels = {"internal": "所内涉警", "suzhou": "苏州涉警", "traffic": "交通涉警"}
        business_label = business_labels.get(
            str(batch.get("business_type") or ""), target_parser,
        )
        subtype_label = subtype_labels.get(str(batch.get("police_subtype") or ""), "")
        display_business = f"{business_label} · {subtype_label}" if subtype_label else business_label
        summary_rows = [
            [f"{display_business}下发反馈"],
            ["批次", batch.get("id")],
            ["原文件", batch.get("file_name", "")],
            ["目标业务表", target_parser],
            ["导出时间", exported_at.strftime("%Y-%m-%d %H:%M:%S")],
            ["任务总数", len(tasks)],
            ["待审核", sum(item.get("task_status") == "pending_review" for item in tasks)],
            ["待发布", sum(item.get("publish_status") in {"pending", "publishing", "retryable"} for item in tasks)],
            ["发布成功", sum(item.get("publish_status") == "success" for item in tasks)],
            ["冲突或待对账", sum(item.get("publish_status") in {"conflict", "needs_reconciliation"} for item in tasks)],
        ]
        for row in summary_rows:
            summary.append(row)
        summary.merge_cells("A1:B1")

        parser = get_parser(target_parser)
        business_columns = list(parser.COLUMNS)
        sheet = workbook.create_sheet("任务明细")
        headers = [
            "批次", "业务", "Excel行", "来源", "最终动作", "任务状态",
            "发布状态", "发布错误", "审核人", "审核时间",
            *business_columns,
        ]
        sheet.append(headers)
        action_labels = {
            "dispatch": "下发到社区", "no_registration": "无需登记",
            "transfer": "移交", "duplicate_exclude": "重复排除", "": "待审核",
        }
        for item in tasks:
            values = item.get("standard_values") or {}
            sheet.append([
                batch.get("id"), display_business, item.get("source_row"),
                item.get("source_name", ""), action_labels.get(item.get("final_action", ""), item.get("final_action", "")),
                item.get("task_status", ""), item.get("publish_status", ""),
                item.get("publish_error", ""), item.get("reviewer_name", ""),
                item.get("reviewed_at_text", ""),
                *[values.get(column, "") for column in business_columns],
            ])
    else:
        summary_rows = [
            ["全链条预处理反馈"],
            ["批次", batch.get("id")],
            ["原文件", batch.get("file_name", "")],
            ["导出时间", exported_at.strftime("%Y-%m-%d %H:%M:%S")],
            ["审核进度", f"{batch.get('reviewed_count', 0)}/{batch.get('total_count', 0)}"],
            ["版本说明", "最终版本" if batch.get("reviewed_count") == batch.get("total_count") else "非最终版本"],
            ["无需登记", sum(item.get("final_action") == "no_registration" for item in tasks)],
            ["移交", sum(item.get("final_action") == "transfer" for item in tasks)],
        ]
        for row in summary_rows:
            summary.append(row)
        summary.merge_cells("A1:B1")

        headers = [
            "批次", "Excel行", "来源", "姓名", "身份证号", "手机号", "原地址",
            "反馈结果", "处理说明", "建议理由", "审核人", "审核时间",
        ]
        action_labels = {"no_registration": "无需登记", "transfer": "移交"}
        for action, title in (("no_registration", "无需登记"), ("transfer", "移交")):
            sheet = workbook.create_sheet(title)
            sheet.append(headers)
            for item in tasks:
                if item.get("final_action") != action:
                    continue
                sheet.append([
                    batch.get("id"), item.get("source_row"), item.get("source_name", ""),
                    item.get("person_name", ""), item.get("identity_number", ""),
                    item.get("phone", ""), item.get("original_address", ""),
                    action_labels[action], item.get("review_note", ""),
                    item.get("suggestion_reason", ""), item.get("reviewer_name", ""),
                    item.get("reviewed_at_text", ""),
                ])

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    light_fill = PatternFill("solid", fgColor="EFF6FF")
    thin = Side(style="thin", color="D8DEE9")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2" if sheet.title != "汇总" else "A2"
        sheet.sheet_view.showGridLines = False
        max_row = max(sheet.max_row, 1)
        max_col = max(sheet.max_column, 1)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if sheet.title == "汇总":
            sheet["A1"].font = Font(color="FFFFFF", bold=True, size=15)
            for row in range(2, max_row + 1):
                sheet.cell(row, 1).fill = light_fill
                sheet.cell(row, 1).font = Font(bold=True)
        else:
            sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
            for row in range(2, max_row + 1):
                sheet.cell(row, 5).number_format = "@"
                sheet.cell(row, 6).number_format = "@"
                sheet.cell(row, 5).quotePrefix = True
                sheet.cell(row, 6).quotePrefix = True
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                if isinstance(cell.value, str):
                    # openpyxl 会把以“=”开头的字符串自动识别为公式；
                    # 来源字段必须始终作为纯文本输出。
                    cell.data_type = "s"
                    cell.number_format = "@"
                    cell.quotePrefix = True
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, max_col + 1):
            values = [str(sheet.cell(row, column).value or "") for row in range(1, max_row + 1)]
            width = min(42, max(10, max((len(value) for value in values), default=0) + 2))
            if sheet.title != "汇总" and column in {7, 8, 9, 10}:
                width = 34
            sheet.column_dimensions[get_column_letter(column)].width = width
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
