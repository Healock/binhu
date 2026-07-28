"""星级评定 XLSX 解析，并在 24 小时内关联到走访明细。"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
import json
import re
from typing import Any, Iterable

from openpyxl import load_workbook

from services.visit_import import (
    HEADER_SCAN_ROWS,
    MAX_DATA_ROWS,
    MAX_FILE_BYTES,
    MAX_SCANNED_ROWS,
    ImportIssue,
    VisitWorkbookError,
    _parse_business_datetime,
    _require_text,
    _text,
    _validate_xlsx_archive,
    fail_import_batch,
    normalize_address,
    normalize_community,
)


STAR_RATING_HEADERS = (
    "派出所名称",
    "所属社区",
    "地址",
    "得分",
    "星级",
    "采集时间",
    "隐患详情",
)
MATCH_WINDOW_SECONDS = 24 * 60 * 60
STAR_LEVEL_PATTERN = re.compile(r"^[一二三四五]星出租房$")


@dataclass
class StarRatingRow:
    row_number: int
    police_station: str
    raw_community: str
    community: str
    address: str
    normalized_address: str
    address_key: str
    score: Decimal
    star_level: str
    collected_at: datetime
    collected_date: date
    raw_collected_at: str
    hazard_details: str
    preview: dict[str, Any]

    def star_values(
        self,
        *,
        canonical_community: str,
        time_difference_seconds: int,
    ) -> tuple[Any, ...]:
        return (
            self.police_station,
            self.raw_community,
            canonical_community,
            self.address,
            self.score,
            self.star_level,
            self.collected_at,
            self.collected_date,
            self.raw_collected_at,
            self.hazard_details,
            time_difference_seconds,
        )


@dataclass
class ParsedStarRatingWorkbook:
    sheet_name: str
    total_rows: int
    valid_rows: int
    rows: list[StarRatingRow]
    issues: list[ImportIssue]
    ignored_rows: int
    start_date: date | None
    end_date: date | None


@dataclass
class VisitCandidate:
    id: int
    address_key: str
    visit_at: datetime
    existing_star_values: tuple[Any, ...] | None
    existing_time_difference_seconds: int | None


@dataclass
class StarRatingMatch:
    rating: StarRatingRow
    visit: VisitCandidate
    time_difference_seconds: int


def _safe_preview(raw: dict[str, Any]) -> dict[str, Any]:
    return {name: _text(raw.get(name)) for name in STAR_RATING_HEADERS}


def _issue(
    severity: str,
    code: str,
    row: StarRatingRow | None,
    message: str,
    *,
    row_number: int | None = None,
    preview: dict[str, Any] | None = None,
) -> ImportIssue:
    return ImportIssue(
        severity=severity,
        code=code,
        row_number=row.row_number if row else int(row_number or 0),
        message=message,
        row_preview=row.preview if row else (preview or {}),
    )


def _parse_score(value: Any) -> Decimal:
    if value is None or _text(value) == "":
        raise ValueError("得分不能为空")
    if isinstance(value, bool):
        raise ValueError("得分必须是数值")
    try:
        score = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError("得分必须是数值") from None
    if not score.is_finite():
        raise ValueError("得分必须是有限数值")
    if abs(score) > Decimal("999999999999.999999"):
        raise ValueError("得分数值过大")
    return score


def _normalize_star_level(value: Any) -> str:
    star_level = re.sub(r"\s+", "", _text(value))
    if not star_level:
        raise ValueError("星级不能为空")
    if not STAR_LEVEL_PATTERN.fullmatch(star_level):
        raise ValueError("星级格式应为“一星出租房”至“五星出租房”")
    return star_level


def _find_header_sheet(workbook) -> tuple[Any, int, dict[str, int]]:
    candidates: list[tuple[Any, int, dict[str, int]]] = []
    required = set(STAR_RATING_HEADERS)
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=HEADER_SCAN_ROWS,
                values_only=True,
            ),
            start=1,
        ):
            headers = [_text(value) for value in values]
            if not required.issubset(set(headers)):
                continue
            if any(headers.count(name) != 1 for name in STAR_RATING_HEADERS):
                raise VisitWorkbookError("星级评定表头存在重复列")
            candidates.append(
                (
                    worksheet,
                    row_number,
                    {
                        name: headers.index(name)
                        for name in STAR_RATING_HEADERS
                    },
                )
            )
            break
    if not candidates:
        raise VisitWorkbookError("未找到包含完整 7 列表头的工作表")
    if len(candidates) > 1:
        names = "、".join(item[0].title for item in candidates)
        raise VisitWorkbookError(
            f"发现多个星级评定工作表：{names}，请只保留一个"
        )
    return candidates[0]


def parse_star_rating_workbook(
    content: bytes,
    timezone_name: str,
) -> ParsedStarRatingWorkbook:
    if not content:
        raise VisitWorkbookError("上传文件为空")
    if len(content) > MAX_FILE_BYTES:
        raise VisitWorkbookError("XLSX 文件不能超过 20MB")
    _validate_xlsx_archive(content)

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise VisitWorkbookError("XLSX 文件无法读取") from exc

    issues: list[ImportIssue] = []
    deduplicated: dict[str, StarRatingRow] = {}
    total_rows = 0
    valid_rows = 0
    ignored_rows = 0
    valid_dates: list[date] = []
    scanned_rows = 0
    try:
        worksheet, header_row, column_map = _find_header_sheet(workbook)
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=header_row + 1,
                values_only=True,
            ),
            start=header_row + 1,
        ):
            scanned_rows += 1
            if scanned_rows > MAX_SCANNED_ROWS:
                raise VisitWorkbookError("工作表不能超过 20 万行")
            raw = {
                name: values[index] if index < len(values) else None
                for name, index in column_map.items()
            }
            if not any(_text(raw.get(name)) for name in STAR_RATING_HEADERS):
                continue
            total_rows += 1

            try:
                police_station = _require_text(
                    raw,
                    "派出所名称",
                    max_length=200,
                    required=False,
                )
                raw_community = _require_text(
                    raw,
                    "所属社区",
                    max_length=200,
                )
                community = normalize_community(raw_community)
                if not community:
                    raise ValueError(
                        "所属社区去除“社区”或“村”后不能为空"
                    )
                address = _require_text(raw, "地址", max_length=1000)
                normalized_address = normalize_address(address)
                if not normalized_address:
                    raise ValueError("地址不能为空")
                score = _parse_score(raw.get("得分"))
                star_level = _normalize_star_level(raw.get("星级"))
                collected_at, collected_date = _parse_business_datetime(
                    raw.get("采集时间"),
                    timezone_name,
                    "采集时间",
                )
                hazard_details = _require_text(
                    raw,
                    "隐患详情",
                    max_length=20000,
                    required=False,
                )
            except ValueError as exc:
                issues.append(
                    _issue(
                        "error",
                        "invalid_star_rating_row",
                        None,
                        str(exc),
                        row_number=row_number,
                        preview=_safe_preview(raw),
                    )
                )
                continue

            if valid_rows >= MAX_DATA_ROWS:
                raise VisitWorkbookError("星级评定不能超过 10 万条有效记录")
            address_key = sha256(
                normalized_address.encode("utf-8")
            ).hexdigest()
            rating = StarRatingRow(
                row_number=row_number,
                police_station=police_station,
                raw_community=raw_community,
                community=community,
                address=address,
                normalized_address=normalized_address,
                address_key=address_key,
                score=score,
                star_level=star_level,
                collected_at=collected_at,
                collected_date=collected_date,
                raw_collected_at=_text(raw.get("采集时间")),
                hazard_details=hazard_details,
                preview=_safe_preview(raw),
            )
            valid_rows += 1
            valid_dates.append(collected_date)
            dedup_key = f"{address_key}|{collected_at.isoformat()}"
            previous = deduplicated.get(dedup_key)
            if previous:
                ignored_rows += 1
                issues.append(
                    _issue(
                        "warning",
                        "duplicate_star_rating_replaced",
                        previous,
                        "同一地址和采集时间重复，采用文件中靠后的记录",
                    )
                )
            deduplicated[dedup_key] = rating
    finally:
        workbook.close()

    return ParsedStarRatingWorkbook(
        sheet_name=worksheet.title,
        total_rows=total_rows,
        valid_rows=valid_rows,
        rows=list(deduplicated.values()),
        issues=issues,
        ignored_rows=ignored_rows,
        start_date=min(valid_dates) if valid_dates else None,
        end_date=max(valid_dates) if valid_dates else None,
    )


def choose_star_rating_matches(
    ratings: list[StarRatingRow],
    candidates: list[VisitCandidate],
) -> tuple[list[StarRatingMatch], list[ImportIssue], int, int]:
    candidates_by_address: dict[str, list[VisitCandidate]] = {}
    for candidate in candidates:
        candidates_by_address.setdefault(candidate.address_key, []).append(
            candidate
        )

    proposed_by_visit: dict[int, list[StarRatingMatch]] = {}
    issues: list[ImportIssue] = []
    unmatched_rows = 0
    ambiguous_rows = 0
    for rating in ratings:
        eligible: list[tuple[int, VisitCandidate]] = []
        for candidate in candidates_by_address.get(rating.address_key, []):
            difference = int(
                abs((rating.collected_at - candidate.visit_at).total_seconds())
            )
            if difference <= MATCH_WINDOW_SECONDS:
                eligible.append((difference, candidate))
        if not eligible:
            unmatched_rows += 1
            issues.append(
                _issue(
                    "warning",
                    "star_rating_visit_not_found",
                    rating,
                    "没有找到地址相同且采集时间前后 24 小时内的走访记录",
                )
            )
            continue

        minimum_difference = min(item[0] for item in eligible)
        nearest = [
            candidate
            for difference, candidate in eligible
            if difference == minimum_difference
        ]
        if len(nearest) > 1:
            exact_existing = [
                candidate
                for candidate in nearest
                if candidate.existing_star_values
                and candidate.existing_star_values[6]
                == rating.collected_at
            ]
            if len(exact_existing) == 1:
                nearest = exact_existing
            else:
                ambiguous_rows += 1
                issues.append(
                    _issue(
                        "warning",
                        "star_rating_visit_ambiguous",
                        rating,
                        "存在多条时间距离相同的走访记录，无法判断属于哪位网格员",
                    )
                )
                continue

        match = StarRatingMatch(
            rating=rating,
            visit=nearest[0],
            time_difference_seconds=minimum_difference,
        )
        proposed_by_visit.setdefault(nearest[0].id, []).append(match)

    selected: list[StarRatingMatch] = []
    for matches in proposed_by_visit.values():
        matches.sort(
            key=lambda item: (
                item.time_difference_seconds,
                -item.rating.collected_at.timestamp(),
                -item.rating.row_number,
            )
        )
        selected.append(matches[0])
        for ignored in matches[1:]:
            ambiguous_rows += 1
            issues.append(
                _issue(
                    "warning",
                    "multiple_star_ratings_for_visit",
                    ignored.rating,
                    "多条星级评定指向同一次走访，已采用时间最接近的一条",
                )
            )
    return selected, issues, unmatched_rows, ambiguous_rows


def _chunks(values: list[str], size: int = 800) -> Iterable[list[str]]:
    for offset in range(0, len(values), size):
        yield values[offset : offset + size]


async def _load_community_lookup(conn) -> tuple[set[str], dict[str, str]]:
    async with conn.cursor() as cur:
        await cur.execute("SELECT id, name FROM _communities")
        community_rows = await cur.fetchall()
        communities = {str(row[1]).strip() for row in community_rows}
        lookup = {
            normalize_community(row[1]): str(row[1]).strip()
            for row in community_rows
        }
        await cur.execute(
            """
            SELECT a.alias, c.name
            FROM _community_aliases AS a
            JOIN _communities AS c ON c.id = a.community_id
            """
        )
        for alias, community_name in await cur.fetchall():
            lookup[normalize_community(alias)] = str(
                community_name
            ).strip()
    return communities, lookup


async def _load_visit_candidates(
    conn,
    address_keys: list[str],
) -> list[VisitCandidate]:
    candidates: list[VisitCandidate] = []
    async with conn.cursor() as cur:
        for chunk in _chunks(address_keys):
            placeholders = ", ".join(["%s"] * len(chunk))
            await cur.execute(
                f"""
                SELECT id, `_address_key`, `入户时间`,
                       `星级派出所名称`, `星级所属社区`, `星级社区`,
                       `星级地址`, `得分`, `星级`, `星级采集时间`,
                       `星级采集日期`, `_raw_star_time`, `隐患详情`,
                       `星级时间差秒`
                FROM t_visit_details
                WHERE `_address_key` IN ({placeholders})
                """,
                chunk,
            )
            for row in await cur.fetchall():
                existing_values = (
                    tuple(row[3:13])
                    if row[9] is not None
                    else None
                )
                candidates.append(
                    VisitCandidate(
                        id=int(row[0]),
                        address_key=str(row[1]),
                        visit_at=row[2],
                        existing_star_values=existing_values,
                        existing_time_difference_seconds=(
                            int(abs((row[9] - row[2]).total_seconds()))
                            if row[9] is not None
                            else None
                        ),
                    )
                )
    return candidates


async def import_star_rating_workbook(
    conn,
    *,
    batch_id: int,
    parsed: ParsedStarRatingWorkbook,
) -> dict[str, Any]:
    issues = list(parsed.issues)
    inserted_rows = 0
    updated_rows = 0
    unchanged_rows = 0
    ignored_rows = parsed.ignored_rows
    unmatched_rows = 0
    ambiguous_rows = 0
    overlap_start_date = None
    overlap_end_date = None

    if not parsed.rows:
        await fail_import_batch(
            conn,
            batch_id,
            "没有可关联的有效星级评定记录",
            issues,
        )
        return {
            "batch_id": batch_id,
            "import_type": "rating",
            "status": "failed",
            "duplicate_file": False,
            "file_start_date": None,
            "file_end_date": None,
            "overlap_start_date": None,
            "overlap_end_date": None,
            "inserted_rows": 0,
            "updated_rows": 0,
            "unchanged_rows": 0,
            "ignored_rows": ignored_rows,
            "unmatched_rows": 0,
            "ambiguous_rows": 0,
            "error_count": sum(i.severity == "error" for i in issues),
            "warning_count": sum(i.severity == "warning" for i in issues),
            "message": "没有可关联的有效星级评定记录",
        }

    await conn.begin()
    try:
        communities, community_lookup = await _load_community_lookup(conn)
        for rating in parsed.rows:
            canonical = community_lookup.get(rating.community)
            if canonical:
                rating.community = canonical
            elif rating.community not in communities:
                issues.append(
                    _issue(
                        "warning",
                        "star_rating_community_not_found",
                        rating,
                        f"社区“{rating.community}”不在社区管理名单中",
                    )
                )

        candidates = await _load_visit_candidates(
            conn,
            sorted({row.address_key for row in parsed.rows}),
        )
        (
            matches,
            match_issues,
            unmatched_rows,
            ambiguous_rows,
        ) = choose_star_rating_matches(parsed.rows, candidates)
        issues.extend(match_issues)
        ignored_rows += unmatched_rows + ambiguous_rows

        if parsed.start_date and parsed.end_date:
            async with conn.cursor() as cur:
                await cur.execute(
                    """
                    SELECT MIN(`星级采集日期`), MAX(`星级采集日期`)
                    FROM t_visit_details
                    WHERE `星级采集日期` BETWEEN %s AND %s
                    """,
                    (parsed.start_date, parsed.end_date),
                )
                overlap_start_date, overlap_end_date = await cur.fetchone()

        updates: list[tuple[Any, ...]] = []
        for match in matches:
            incoming_values = match.rating.star_values(
                canonical_community=match.rating.community,
                time_difference_seconds=match.time_difference_seconds,
            )
            existing_values = match.visit.existing_star_values
            if existing_values == incoming_values[:-1]:
                unchanged_rows += 1
                continue
            if existing_values is None:
                inserted_rows += 1
            else:
                existing_difference = (
                    match.visit.existing_time_difference_seconds
                )
                existing_collected_at = existing_values[6]
                if (
                    existing_difference is not None
                    and match.time_difference_seconds > existing_difference
                    and match.rating.collected_at != existing_collected_at
                ):
                    ignored_rows += 1
                    issues.append(
                        _issue(
                            "warning",
                            "star_rating_farther_than_existing",
                            match.rating,
                            "该走访已有时间更接近的星级评定，本行未采用",
                        )
                    )
                    continue
                updated_rows += 1
            updates.append(
                (
                    *incoming_values,
                    batch_id,
                    match.rating.row_number,
                    match.visit.id,
                )
            )

        async with conn.cursor() as cur:
            if updates:
                await cur.executemany(
                    """
                    UPDATE t_visit_details
                    SET `星级派出所名称`=%s, `星级所属社区`=%s,
                        `星级社区`=%s, `星级地址`=%s, `得分`=%s,
                        `星级`=%s, `星级采集时间`=%s,
                        `星级采集日期`=%s, `_raw_star_time`=%s,
                        `隐患详情`=%s, `星级时间差秒`=%s,
                        star_import_batch_id=%s,
                        star_source_row_number=%s,
                        updated_at=UTC_TIMESTAMP()
                    WHERE id=%s
                    """,
                    updates,
                )
            if issues:
                await cur.executemany(
                    """
                    INSERT INTO _visit_import_issues (
                        batch_id, severity, code, source_row_number, message,
                        row_preview, created_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, UTC_TIMESTAMP())
                    """,
                    [
                        (
                            batch_id,
                            issue.severity,
                            issue.code,
                            issue.row_number,
                            issue.message,
                            json.dumps(
                                issue.row_preview,
                                ensure_ascii=False,
                                default=str,
                            ),
                        )
                        for issue in issues
                    ],
                )
            error_count = sum(
                issue.severity == "error"
                for issue in issues
            )
            warning_count = sum(
                issue.severity == "warning"
                for issue in issues
            )
            status = (
                "partial"
                if error_count or unmatched_rows or ambiguous_rows
                else "success"
            )
            await cur.execute(
                """
                UPDATE _visit_import_batches
                SET status=%s, sheet_name=%s, total_rows=%s, valid_rows=%s,
                    inserted_rows=%s, updated_rows=%s, unchanged_rows=%s,
                    ignored_rows=%s, error_count=%s, warning_count=%s,
                    file_start_date=%s, file_end_date=%s,
                    overlap_start_date=%s, overlap_end_date=%s,
                    finished_at=UTC_TIMESTAMP()
                WHERE id=%s
                """,
                (
                    status,
                    parsed.sheet_name[:100],
                    parsed.total_rows,
                    parsed.valid_rows,
                    inserted_rows,
                    updated_rows,
                    unchanged_rows,
                    ignored_rows,
                    error_count,
                    warning_count,
                    parsed.start_date,
                    parsed.end_date,
                    overlap_start_date,
                    overlap_end_date,
                    batch_id,
                ),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise

    matched_rows = inserted_rows + updated_rows + unchanged_rows
    return {
        "batch_id": batch_id,
        "import_type": "rating",
        "status": status,
        "duplicate_file": False,
        "file_start_date": (
            parsed.start_date.isoformat() if parsed.start_date else None
        ),
        "file_end_date": (
            parsed.end_date.isoformat() if parsed.end_date else None
        ),
        "overlap_start_date": (
            overlap_start_date.isoformat() if overlap_start_date else None
        ),
        "overlap_end_date": (
            overlap_end_date.isoformat() if overlap_end_date else None
        ),
        "inserted_rows": inserted_rows,
        "updated_rows": updated_rows,
        "unchanged_rows": unchanged_rows,
        "ignored_rows": ignored_rows,
        "matched_rows": matched_rows,
        "unmatched_rows": unmatched_rows,
        "ambiguous_rows": ambiguous_rows,
        "error_count": error_count,
        "warning_count": warning_count,
        "message": (
            "星级评定导入完成，部分记录未能关联"
            if status == "partial"
            else "星级评定已关联到走访记录"
        ),
    }
