"""走访明细 XLSX 解析、去重和入库。"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
import json
import re
import unicodedata
from typing import Any, Iterable
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from services.business_time import resolve_timezone
from services.privacy import mask_identity_number


VISIT_HEADERS = (
    "派出所名称",
    "村社区",
    "进入方式",
    "地址",
    "操作人",
    "操作人账号",
    "入户时间",
    "房间核查数量",
    "新增",
    "变更",
    "注销",
)
ALLOWED_ENTRY_METHODS = {"扫码", "搜索"}
MAX_FILE_BYTES = 20 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_DATA_ROWS = 100_000
MAX_SCANNED_ROWS = 200_000
MAX_UNSIGNED_INT = 4_294_967_295
HEADER_SCAN_ROWS = 20
ISSUE_PAGE_SIZE = 50
VISIT_IMPORT_LOCK_NAME = "binhu_visit_detail_import"
IDENTITY_PATTERN = re.compile(r"^(?:\d{15}|\d{17}[\dX])$")


class VisitWorkbookError(ValueError):
    """The workbook structure cannot be imported."""


@dataclass
class ImportIssue:
    severity: str
    code: str
    row_number: int
    message: str
    row_preview: dict[str, Any] = field(default_factory=dict)

    def as_dict(self) -> dict[str, Any]:
        return {
            "severity": self.severity,
            "code": self.code,
            "row_number": self.row_number,
            "message": self.message,
            "row_preview": self.row_preview,
        }


@dataclass
class VisitRow:
    row_number: int
    police_station: str
    raw_community: str
    community: str
    entry_method: str
    address: str
    normalized_address: str
    address_key: str
    row_key: str
    operator_name: str
    operator_account: str
    operator_account_valid: bool
    visit_at: datetime
    visit_date: date
    raw_visit_time: str
    room_check_count: int
    added_count: int
    changed_count: int
    cancelled_count: int
    preview: dict[str, Any]

    def business_values(self) -> tuple[Any, ...]:
        return (
            self.police_station,
            self.raw_community,
            self.community,
            self.entry_method,
            self.address,
            self.normalized_address,
            self.address_key,
            self.operator_name,
            self.operator_account,
            self.visit_at,
            self.visit_date,
            self.raw_visit_time,
            self.room_check_count,
            self.added_count,
            self.changed_count,
            self.cancelled_count,
        )


@dataclass
class ParsedVisitWorkbook:
    sheet_name: str
    total_rows: int
    valid_rows: int
    rows: list[VisitRow]
    issues: list[ImportIssue]
    ignored_rows: int
    start_date: date | None
    end_date: date | None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_community(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value)).strip()
    if text.endswith("社区"):
        text = text[:-2].strip()
    return text


def normalize_address(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value)).strip()
    return re.sub(r"\s+", " ", text)


def normalize_identity(value: Any) -> tuple[str, bool, str | None]:
    text = unicodedata.normalize("NFKC", _text(value)).replace(" ", "").upper()
    if not text:
        return "", False, "操作人账号为空，无法补齐网格员身份证号"
    if isinstance(value, (int, float, Decimal)):
        return text, False, "操作人账号是数字单元格，可能已经丢失精度，未用于补齐身份证号"
    if not IDENTITY_PATTERN.fullmatch(text):
        return text, False, "操作人账号不是有效的 15 位或 18 位身份证号，未用于补齐"
    return text, True, None


def _parse_nonnegative_int(value: Any, field_name: str) -> int:
    if value is None or _text(value) == "":
        return 0
    if isinstance(value, bool):
        raise ValueError(f"{field_name}必须是非负整数")
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name}必须是非负整数") from None
    if number < 0 or number != number.to_integral_value():
        raise ValueError(f"{field_name}必须是非负整数")
    if number > MAX_UNSIGNED_INT:
        raise ValueError(f"{field_name}不能超过 {MAX_UNSIGNED_INT}")
    return int(number)


def _parse_visit_datetime(value: Any, timezone_name: str) -> tuple[datetime, date]:
    parsed: datetime
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    elif isinstance(value, (int, float, Decimal)):
        raise ValueError("入户时间是未格式化的数字单元格，请在 Excel 中改成日期时间")
    else:
        text = _text(value)
        if not text:
            raise ValueError("入户时间不能为空")
        normalized = (
            text.replace("年", "-")
            .replace("月", "-")
            .replace("日", " ")
            .replace("/", "-")
            .strip()
        )
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            parsed = None  # type: ignore[assignment]
            for pattern in (
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%Y-%m-%d",
            ):
                try:
                    parsed = datetime.strptime(normalized, pattern)
                    break
                except ValueError:
                    continue
            if parsed is None:
                raise ValueError("入户时间格式无法识别") from None

    business_timezone = resolve_timezone(timezone_name)
    if parsed.tzinfo is None:
        local_time = parsed.replace(tzinfo=business_timezone)
    else:
        local_time = parsed.astimezone(business_timezone)
    if local_time.date() < date(1000, 1, 1):
        raise ValueError("入户时间不能早于 1000-01-01")
    utc_time = local_time.astimezone(timezone.utc).replace(tzinfo=None)
    return utc_time, local_time.date()


def _safe_preview(raw: dict[str, Any]) -> dict[str, Any]:
    preview = {name: _text(raw.get(name)) for name in VISIT_HEADERS}
    preview["操作人账号"] = mask_identity_number(preview["操作人账号"])
    return preview


def _require_text(
    raw: dict[str, Any],
    field_name: str,
    *,
    max_length: int,
    required: bool = True,
) -> str:
    value = _text(raw.get(field_name))
    if required and not value:
        raise ValueError(f"{field_name}不能为空")
    if len(value) > max_length:
        raise ValueError(f"{field_name}不能超过 {max_length} 个字符")
    return value


def _row_issue(
    severity: str,
    code: str,
    row_number: int,
    message: str,
    raw: dict[str, Any],
) -> ImportIssue:
    return ImportIssue(
        severity=severity,
        code=code,
        row_number=row_number,
        message=message,
        row_preview=_safe_preview(raw),
    )


def _find_header_sheet(workbook) -> tuple[Any, int, dict[str, int]]:
    candidates: list[tuple[Any, int, dict[str, int]]] = []
    required = set(VISIT_HEADERS)
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True),
            start=1,
        ):
            headers = [_text(value) for value in values]
            if not required.issubset(set(headers)):
                continue
            if any(headers.count(name) != 1 for name in VISIT_HEADERS):
                raise VisitWorkbookError("走访明细表头存在重复列")
            candidates.append(
                (
                    worksheet,
                    row_number,
                    {name: headers.index(name) for name in VISIT_HEADERS},
                )
            )
            break
    if not candidates:
        raise VisitWorkbookError("未找到包含完整 11 列表头的工作表")
    if len(candidates) > 1:
        names = "、".join(item[0].title for item in candidates)
        raise VisitWorkbookError(f"发现多个走访明细工作表：{names}，请只保留一个")
    return candidates[0]


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            expanded_size = sum(item.file_size for item in archive.infolist())
    except BadZipFile as exc:
        raise VisitWorkbookError("文件不是有效的 XLSX") from exc
    if expanded_size > MAX_UNCOMPRESSED_BYTES:
        raise VisitWorkbookError("XLSX 解压后的内容过大")


def parse_visit_workbook(
    content: bytes,
    timezone_name: str,
) -> ParsedVisitWorkbook:
    """Parse a visit-detail workbook without touching the database."""
    if not content:
        raise VisitWorkbookError("上传文件为空")
    if len(content) > MAX_FILE_BYTES:
        raise VisitWorkbookError("XLSX 文件不能超过 20MB")
    _validate_xlsx_archive(content)

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise VisitWorkbookError("XLSX 文件无法读取") from exc

    issues: list[ImportIssue] = []
    deduplicated: dict[str, VisitRow] = {}
    total_rows = 0
    valid_rows = 0
    ignored_rows = 0
    valid_dates: list[date] = []
    scanned_rows = 0
    try:
        worksheet, header_row, column_map = _find_header_sheet(workbook)
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            scanned_rows += 1
            if scanned_rows > MAX_SCANNED_ROWS:
                raise VisitWorkbookError("工作表不能超过 20 万行")
            raw = {
                name: values[index] if index < len(values) else None
                for name, index in column_map.items()
            }
            if not any(_text(raw.get(name)) for name in VISIT_HEADERS):
                continue
            total_rows += 1

            row_errors: list[str] = []
            try:
                police_station = _require_text(
                    raw,
                    "派出所名称",
                    max_length=200,
                    required=False,
                )
                raw_community = _require_text(raw, "村社区", max_length=200)
                community = normalize_community(raw_community)
                if not community:
                    raise ValueError("村社区去除“社区”后不能为空")
                entry_method = _require_text(raw, "进入方式", max_length=20)
                if entry_method not in ALLOWED_ENTRY_METHODS:
                    raise ValueError("进入方式只能是“扫码”或“搜索”")
                address = _require_text(raw, "地址", max_length=1000)
                normalized_address = normalize_address(address)
                if not normalized_address:
                    raise ValueError("地址不能为空")
                operator_name = _require_text(raw, "操作人", max_length=100)
                operator_account, account_valid, account_warning = normalize_identity(
                    raw.get("操作人账号")
                )
                if len(operator_account) > 50:
                    raise ValueError("操作人账号不能超过 50 个字符")
                visit_at, visit_date = _parse_visit_datetime(
                    raw.get("入户时间"),
                    timezone_name,
                )
                room_check_count = _parse_nonnegative_int(
                    raw.get("房间核查数量"),
                    "房间核查数量",
                )
                added_count = _parse_nonnegative_int(raw.get("新增"), "新增")
                changed_count = _parse_nonnegative_int(raw.get("变更"), "变更")
                cancelled_count = _parse_nonnegative_int(raw.get("注销"), "注销")
            except ValueError as exc:
                row_errors.append(str(exc))

            if row_errors:
                issues.append(
                    _row_issue(
                        "error",
                        "invalid_row",
                        row_number,
                        "；".join(row_errors),
                        raw,
                    )
                )
                continue

            if valid_rows >= MAX_DATA_ROWS:
                raise VisitWorkbookError("走访明细不能超过 10 万条有效记录")
            preview = _safe_preview(raw)
            address_key = sha256(normalized_address.encode("utf-8")).hexdigest()
            row_key = sha256(
                f"{visit_date.isoformat()}|{normalized_address}".encode("utf-8")
            ).hexdigest()
            visit_row = VisitRow(
                row_number=row_number,
                police_station=police_station,
                raw_community=raw_community,
                community=community,
                entry_method=entry_method,
                address=address,
                normalized_address=normalized_address,
                address_key=address_key,
                row_key=row_key,
                operator_name=operator_name,
                operator_account=operator_account,
                operator_account_valid=account_valid,
                visit_at=visit_at,
                visit_date=visit_date,
                raw_visit_time=_text(raw.get("入户时间")),
                room_check_count=room_check_count,
                added_count=added_count,
                changed_count=changed_count,
                cancelled_count=cancelled_count,
                preview=preview,
            )
            valid_rows += 1
            valid_dates.append(visit_date)
            if account_warning:
                issues.append(
                    _row_issue(
                        "warning",
                        "identity_not_usable",
                        row_number,
                        account_warning,
                        raw,
                    )
                )

            previous = deduplicated.get(row_key)
            if previous is None:
                deduplicated[row_key] = visit_row
                continue
            ignored_rows += 1
            if visit_row.visit_at >= previous.visit_at:
                issues.append(
                    ImportIssue(
                        severity="warning",
                        code="same_day_address_replaced",
                        row_number=previous.row_number,
                        message="同日同地址存在更晚记录，本行未采用",
                        row_preview=previous.preview,
                    )
                )
                deduplicated[row_key] = visit_row
            else:
                issues.append(
                    ImportIssue(
                        severity="warning",
                        code="same_day_address_ignored",
                        row_number=visit_row.row_number,
                        message="同日同地址已有更晚记录，本行未采用",
                        row_preview=visit_row.preview,
                    )
                )
    finally:
        workbook.close()

    return ParsedVisitWorkbook(
        sheet_name=worksheet.title,
        total_rows=total_rows,
        valid_rows=valid_rows,
        rows=list(deduplicated.values()),
        issues=issues,
        ignored_rows=ignored_rows,
        start_date=min(valid_dates) if valid_dates else None,
        end_date=max(valid_dates) if valid_dates else None,
    )


def decide_existing_action(
    existing_values: tuple[Any, ...] | None,
    incoming: VisitRow,
) -> str:
    """Return insert, update, unchanged, or ignored for one canonical row."""
    if existing_values is None:
        return "insert"
    existing_visit_at = existing_values[9]
    if incoming.visit_at < existing_visit_at:
        return "ignored"
    if incoming.business_values() == existing_values:
        return "unchanged"
    return "update"


def decide_member_identity(
    *,
    member_name: str,
    existing_identity: str,
    incoming_identity: str,
    id_owners: dict[str, str],
) -> str:
    """Return update, same, conflict, or used_by_other."""
    if existing_identity:
        return "same" if existing_identity == incoming_identity else "conflict"
    other_owner = id_owners.get(incoming_identity)
    if other_owner and other_owner != member_name:
        return "used_by_other"
    return "update"


def _chunks(values: list[str], size: int = 800) -> Iterable[list[str]]:
    for offset in range(0, len(values), size):
        yield values[offset : offset + size]


def _utc_iso(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")


async def get_visit_coverage(conn) -> dict[str, Any]:
    async with conn.cursor() as cur:
        await cur.execute(
            """
            SELECT MIN(`业务日期`), MAX(`业务日期`), COUNT(*),
                   COUNT(DISTINCT `业务日期`)
            FROM t_visit_details
            """
        )
        start_date, end_date, total_records, data_days = await cur.fetchone()
        missing_dates: list[str] = []
        if start_date and end_date:
            await cur.execute(
                "SELECT DISTINCT `业务日期` FROM t_visit_details "
                "ORDER BY `业务日期`"
            )
            existing_dates = {row[0] for row in await cur.fetchall()}
            current = start_date
            while current <= end_date:
                if current not in existing_dates:
                    missing_dates.append(current.isoformat())
                current = date.fromordinal(current.toordinal() + 1)
        await cur.execute(
            """
            SELECT MAX(finished_at)
            FROM _visit_import_batches
            WHERE status IN ('success', 'partial')
            """
        )
        last_import = (await cur.fetchone())[0]
    return {
        "start_date": start_date.isoformat() if start_date else None,
        "end_date": end_date.isoformat() if end_date else None,
        "total_records": int(total_records or 0),
        "data_days": int(data_days or 0),
        "missing_date_count": len(missing_dates),
        "missing_dates": missing_dates,
        "last_import_at": _utc_iso(last_import),
    }


async def recover_interrupted_visit_imports() -> int:
    """Close batches left running by an interrupted backend process."""
    from database import db_manager

    pool = db_manager.get_pool("online_data")
    conn = await pool.acquire()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT GET_LOCK(%s, 0)",
                (VISIT_IMPORT_LOCK_NAME,),
            )
            lock_row = await cur.fetchone()
            if not lock_row or lock_row[0] != 1:
                return 0
            try:
                await cur.execute(
                    """
                    UPDATE _visit_import_batches
                    SET status='failed',
                        error_message='服务重启，导入未完成，请重新上传原文件',
                        finished_at=UTC_TIMESTAMP()
                    WHERE status='running'
                    """
                )
                return int(cur.rowcount or 0)
            finally:
                await cur.execute(
                    "SELECT RELEASE_LOCK(%s)",
                    (VISIT_IMPORT_LOCK_NAME,),
                )
    finally:
        pool.release(conn)


async def find_duplicate_batch(conn, file_sha256: str) -> dict[str, Any] | None:
    async with conn.cursor() as cur:
        await cur.execute(
            """
            SELECT id, status, file_start_date, file_end_date,
                   inserted_rows, updated_rows, unchanged_rows, ignored_rows,
                   error_count, warning_count
            FROM _visit_import_batches
            WHERE file_sha256=%s AND status IN ('success', 'partial')
            ORDER BY id DESC LIMIT 1
            """,
            (file_sha256,),
        )
        row = await cur.fetchone()
    if not row:
        return None
    return {
        "batch_id": row[0],
        "status": "duplicate",
        "duplicate_file": True,
        "file_start_date": row[2].isoformat() if row[2] else None,
        "file_end_date": row[3].isoformat() if row[3] else None,
        "overlap_start_date": None,
        "overlap_end_date": None,
        "inserted_rows": 0,
        "updated_rows": 0,
        "unchanged_rows": 0,
        "ignored_rows": 0,
        "error_count": 0,
        "warning_count": 0,
        "message": "这个文件已经成功导入过，本次没有重复处理",
    }


async def create_import_batch(
    conn,
    *,
    filename: str,
    file_sha256: str,
    file_size: int,
    uploader_id: int,
) -> int:
    async with conn.cursor() as cur:
        await cur.execute(
            """
            INSERT INTO _visit_import_batches (
                filename, file_sha256, file_size_bytes, status, uploader_id,
                started_at, created_at
            ) VALUES (%s, %s, %s, 'running', %s, UTC_TIMESTAMP(), UTC_TIMESTAMP())
            """,
            (filename[:255], file_sha256, file_size, uploader_id),
        )
        return int(cur.lastrowid)


async def fail_import_batch(
    conn,
    batch_id: int,
    message: str,
    issues: list[ImportIssue] | None = None,
) -> None:
    issues = issues or []
    async with conn.cursor() as cur:
        if issues:
            await cur.executemany(
                """
                INSERT INTO _visit_import_issues (
                    batch_id, severity, code, source_row_number, message,
                    row_preview, created_at
                ) VALUES (%s, %s, %s, %s, %s, %s, UTC_TIMESTAMP())
                """,
                [
                    (
                        batch_id,
                        issue.severity,
                        issue.code,
                        issue.row_number,
                        issue.message,
                        json.dumps(
                            issue.row_preview,
                            ensure_ascii=False,
                            default=str,
                        ),
                    )
                    for issue in issues
                ],
            )
        await cur.execute(
            """
            UPDATE _visit_import_batches
            SET status='failed', error_count=%s, error_message=%s,
                finished_at=UTC_TIMESTAMP()
            WHERE id=%s
            """,
            (
                sum(issue.severity == "error" for issue in issues),
                message[:1000],
                batch_id,
            ),
        )


async def _load_existing_rows(conn, row_keys: list[str]) -> dict[str, tuple[Any, ...]]:
    existing: dict[str, tuple[Any, ...]] = {}
    async with conn.cursor() as cur:
        for chunk in _chunks(row_keys):
            placeholders = ", ".join(["%s"] * len(chunk))
            await cur.execute(
                f"""
                SELECT `_row_key`, `派出所名称`, `村社区`, `社区`, `进入方式`,
                       `地址`, `_normalized_address`, `_address_key`, `操作人`,
                       `操作人账号`, `入户时间`, `业务日期`, `_raw_visit_time`,
                       `房间核查数量`, `新增`, `变更`, `注销`
                FROM t_visit_details
                WHERE `_row_key` IN ({placeholders})
                """,
                chunk,
            )
            for row in await cur.fetchall():
                existing[row[0]] = tuple(row[1:])
    return existing


async def _load_reference_data(conn):
    async with conn.cursor() as cur:
        await cur.execute("SELECT name FROM _communities")
        communities = {str(row[0]).strip() for row in await cur.fetchall()}
        await cur.execute(
            "SELECT id, name, community, id_card_number FROM _grid_members"
        )
        members = {
            str(row[1]).strip(): {
                "id": row[0],
                "community": str(row[2] or "").strip(),
                "id_card_number": str(row[3] or "").strip().upper(),
            }
            for row in await cur.fetchall()
        }
    id_owners = {
        value["id_card_number"]: name
        for name, value in members.items()
        if value["id_card_number"]
    }
    return communities, members, id_owners


async def import_parsed_workbook(
    conn,
    *,
    batch_id: int,
    parsed: ParsedVisitWorkbook,
) -> dict[str, Any]:
    issues = list(parsed.issues)
    inserted_rows = 0
    updated_rows = 0
    unchanged_rows = 0
    ignored_rows = parsed.ignored_rows
    overlap_start_date = None
    overlap_end_date = None

    if parsed.start_date and parsed.end_date:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                SELECT MIN(`业务日期`), MAX(`业务日期`)
                FROM t_visit_details
                WHERE `业务日期` BETWEEN %s AND %s
                """,
                (parsed.start_date, parsed.end_date),
            )
            overlap_start_date, overlap_end_date = await cur.fetchone()

    if not parsed.rows:
        await fail_import_batch(
            conn,
            batch_id,
            "没有可入库的有效走访记录",
            issues,
        )
        async with conn.cursor() as cur:
            await cur.execute(
                """
                UPDATE _visit_import_batches
                SET sheet_name=%s, total_rows=%s, valid_rows=%s,
                    ignored_rows=%s, error_count=%s, warning_count=%s,
                    file_start_date=%s, file_end_date=%s
                WHERE id=%s
                """,
                (
                    parsed.sheet_name[:100],
                    parsed.total_rows,
                    parsed.valid_rows,
                    ignored_rows,
                    sum(i.severity == "error" for i in issues),
                    sum(i.severity == "warning" for i in issues),
                    parsed.start_date,
                    parsed.end_date,
                    batch_id,
                ),
            )
        return {
            "batch_id": batch_id,
            "status": "failed",
            "duplicate_file": False,
            "file_start_date": None,
            "file_end_date": None,
            "overlap_start_date": None,
            "overlap_end_date": None,
            "inserted_rows": 0,
            "updated_rows": 0,
            "unchanged_rows": 0,
            "ignored_rows": ignored_rows,
            "error_count": sum(i.severity == "error" for i in issues),
            "warning_count": sum(i.severity == "warning" for i in issues),
            "message": "没有可入库的有效走访记录",
        }

    await conn.begin()
    try:
        communities, members, id_owners = await _load_reference_data(conn)
        seen_reference_warnings: set[tuple[str, str]] = set()
        member_updates: list[tuple[str, int]] = []

        def warning_once(code: str, key: str, row: VisitRow, message: str) -> None:
            marker = (code, key)
            if marker in seen_reference_warnings:
                return
            seen_reference_warnings.add(marker)
            issues.append(
                ImportIssue(
                    severity="warning",
                    code=code,
                    row_number=row.row_number,
                    message=message,
                    row_preview=row.preview,
                )
            )

        for row in parsed.rows:
            if row.community not in communities:
                warning_once(
                    "community_not_found",
                    row.community,
                    row,
                    f"社区“{row.community}”不在社区管理名单中，数据已保留",
                )
            member = members.get(row.operator_name)
            if not member:
                warning_once(
                    "member_not_found",
                    row.operator_name,
                    row,
                    f"操作人“{row.operator_name}”不在网格员名单中，未自动创建",
                )
                continue
            if (
                member["community"]
                and member["community"] != row.community
            ):
                warning_once(
                    "member_community_mismatch",
                    row.operator_name,
                    row,
                    f"操作人“{row.operator_name}”的网格员社区与走访社区不同",
                )
            if not row.operator_account_valid:
                continue
            existing_identity = member["id_card_number"]
            identity_action = decide_member_identity(
                member_name=row.operator_name,
                existing_identity=existing_identity,
                incoming_identity=row.operator_account,
                id_owners=id_owners,
            )
            if identity_action == "update":
                member_updates.append((row.operator_account, member["id"]))
                member["id_card_number"] = row.operator_account
                id_owners[row.operator_account] = row.operator_name
            elif identity_action == "used_by_other":
                warning_once(
                    "identity_already_used",
                    row.operator_account,
                    row,
                    "该操作人账号已经属于另一名网格员，未自动补齐",
                )
            elif identity_action == "conflict":
                warning_once(
                    "identity_conflict",
                    row.operator_name,
                    row,
                    f"操作人“{row.operator_name}”的身份证号与现有资料不一致，未覆盖",
                )

        existing_rows = await _load_existing_rows(
            conn,
            [row.row_key for row in parsed.rows],
        )
        inserts: list[tuple[Any, ...]] = []
        updates: list[tuple[Any, ...]] = []
        for row in parsed.rows:
            action = decide_existing_action(existing_rows.get(row.row_key), row)
            if action == "insert":
                inserted_rows += 1
                inserts.append(
                    (
                        row.row_key,
                        *row.business_values(),
                        batch_id,
                        row.row_number,
                    )
                )
            elif action == "update":
                updated_rows += 1
                updates.append(
                    (
                        *row.business_values(),
                        batch_id,
                        row.row_number,
                        row.row_key,
                    )
                )
            elif action == "unchanged":
                unchanged_rows += 1
            else:
                ignored_rows += 1
                issues.append(
                    ImportIssue(
                        severity="warning",
                        code="older_than_database",
                        row_number=row.row_number,
                        message="数据库中已有同日同地址的更晚记录，本行未采用",
                        row_preview=row.preview,
                    )
                )

        async with conn.cursor() as cur:
            if member_updates:
                await cur.executemany(
                    "UPDATE _grid_members SET id_card_number=%s WHERE id=%s",
                    member_updates,
                )
            if inserts:
                await cur.executemany(
                    """
                    INSERT INTO t_visit_details (
                        `_row_key`, `派出所名称`, `村社区`, `社区`, `进入方式`,
                        `地址`, `_normalized_address`, `_address_key`, `操作人`,
                        `操作人账号`, `入户时间`, `业务日期`, `_raw_visit_time`,
                        `房间核查数量`, `新增`, `变更`, `注销`,
                        import_batch_id, source_row_number
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    """,
                    inserts,
                )
            if updates:
                await cur.executemany(
                    """
                    UPDATE t_visit_details
                    SET `派出所名称`=%s, `村社区`=%s, `社区`=%s,
                        `进入方式`=%s, `地址`=%s, `_normalized_address`=%s,
                        `_address_key`=%s, `操作人`=%s, `操作人账号`=%s,
                        `入户时间`=%s, `业务日期`=%s, `_raw_visit_time`=%s,
                        `房间核查数量`=%s, `新增`=%s, `变更`=%s, `注销`=%s,
                        import_batch_id=%s, source_row_number=%s,
                        updated_at=UTC_TIMESTAMP()
                    WHERE `_row_key`=%s
                    """,
                    updates,
                )
            if issues:
                await cur.executemany(
                    """
                    INSERT INTO _visit_import_issues (
                        batch_id, severity, code, source_row_number, message,
                        row_preview, created_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, UTC_TIMESTAMP())
                    """,
                    [
                        (
                            batch_id,
                            issue.severity,
                            issue.code,
                            issue.row_number,
                            issue.message,
                            json.dumps(
                                issue.row_preview,
                                ensure_ascii=False,
                                default=str,
                            ),
                        )
                        for issue in issues
                    ],
                )
            error_count = sum(issue.severity == "error" for issue in issues)
            warning_count = sum(issue.severity == "warning" for issue in issues)
            status = "partial" if error_count else "success"
            await cur.execute(
                """
                UPDATE _visit_import_batches
                SET status=%s, sheet_name=%s, total_rows=%s, valid_rows=%s,
                    inserted_rows=%s, updated_rows=%s, unchanged_rows=%s,
                    ignored_rows=%s, error_count=%s, warning_count=%s,
                    file_start_date=%s, file_end_date=%s,
                    overlap_start_date=%s, overlap_end_date=%s,
                    finished_at=UTC_TIMESTAMP()
                WHERE id=%s
                """,
                (
                    status,
                    parsed.sheet_name[:100],
                    parsed.total_rows,
                    parsed.valid_rows,
                    inserted_rows,
                    updated_rows,
                    unchanged_rows,
                    ignored_rows,
                    error_count,
                    warning_count,
                    parsed.start_date,
                    parsed.end_date,
                    overlap_start_date,
                    overlap_end_date,
                    batch_id,
                ),
            )
        await conn.commit()
    except Exception:
        await conn.rollback()
        raise

    return {
        "batch_id": batch_id,
        "status": status,
        "duplicate_file": False,
        "file_start_date": parsed.start_date.isoformat(),
        "file_end_date": parsed.end_date.isoformat(),
        "overlap_start_date": (
            overlap_start_date.isoformat() if overlap_start_date else None
        ),
        "overlap_end_date": (
            overlap_end_date.isoformat() if overlap_end_date else None
        ),
        "inserted_rows": inserted_rows,
        "updated_rows": updated_rows,
        "unchanged_rows": unchanged_rows,
        "ignored_rows": ignored_rows,
        "error_count": error_count,
        "warning_count": warning_count,
        "message": (
            "导入完成，部分错误行已跳过"
            if status == "partial"
            else "走访明细导入完成"
        ),
    }


async def list_import_issues(
    conn,
    batch_id: int,
    *,
    page: int,
    page_size: int,
) -> dict[str, Any]:
    offset = (page - 1) * page_size
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT COUNT(*) FROM _visit_import_issues WHERE batch_id=%s",
            (batch_id,),
        )
        total = int((await cur.fetchone())[0])
        await cur.execute(
            """
            SELECT id, severity, code, source_row_number, message, row_preview
            FROM _visit_import_issues
            WHERE batch_id=%s
            ORDER BY CASE severity WHEN 'error' THEN 0 ELSE 1 END,
                     source_row_number, id
            LIMIT %s OFFSET %s
            """,
            (batch_id, page_size, offset),
        )
        rows = await cur.fetchall()
    data = []
    for row in rows:
        preview = row[5]
        if isinstance(preview, str):
            try:
                preview = json.loads(preview)
            except json.JSONDecodeError:
                preview = {}
        if isinstance(preview, dict) and "操作人账号" in preview:
            preview["操作人账号"] = mask_identity_number(
                preview.get("操作人账号")
            )
        data.append(
            {
                "id": row[0],
                "severity": row[1],
                "code": row[2],
                "row_number": row[3],
                "message": row[4],
                "row_preview": preview or {},
            }
        )
    return {
        "data": data,
        "total": total,
        "page": page,
        "page_size": page_size,
    }
