"""滨湖网格每日工作明细 XLSX 生成。"""

from __future__ import annotations

import json
from copy import copy
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.personnel_attendance import (
    get_attendance_context,
    is_member_on_duty,
    normalize_week_start,
    period_covers,
)
from services.personnel_positions import (
    VISIT_POSITION_CONFIG_KEY,
    get_configured_positions,
    get_known_personnel_positions,
    normalized_person_name,
)
from services.report_builders.summary import get_summary
from services.visit_summary import (
    VISIT_CATEGORY_RENTAL,
    VISIT_CATEGORY_SELF_OWNED,
    get_visit_summary,
)


DEFAULT_RENTAL_TARGET = 10
DEFAULT_SELF_OWNED_TARGET = 15
MAX_VISIT_TARGET = 999
MIN_DATA_ROWS = 64
DATA_START_ROW = 4
LAST_COLUMN = 22
PREFERENCE_KEY_PREFIX = "work_log_daily_detail_targets:"
TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1]
    / "assets"
    / "work_log_daily_detail_template.xlsx"
)
DATA_FONT = Font(name="宋体", size=12)
DATA_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
DATA_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def normalize_targets(value: Any) -> dict[str, int]:
    source = value if isinstance(value, dict) else {}

    def target(key: str, default: int) -> int:
        try:
            normalized = int(source.get(key, default))
        except (TypeError, ValueError):
            normalized = default
        return min(MAX_VISIT_TARGET, max(0, normalized))

    return {
        "rental_target": target("rental_target", DEFAULT_RENTAL_TARGET),
        "self_owned_target": target(
            "self_owned_target",
            DEFAULT_SELF_OWNED_TARGET,
        ),
    }


async def load_target_preferences(conn, user_id: int) -> dict[str, int]:
    key = f"{PREFERENCE_KEY_PREFIX}{int(user_id)}"
    async with conn.cursor() as cur:
        await cur.execute(
            "SELECT config_value FROM _system_config WHERE config_key=%s",
            (key,),
        )
        row = await cur.fetchone()
    if not row or not row[0]:
        return normalize_targets({})
    try:
        return normalize_targets(json.loads(row[0]))
    except (TypeError, ValueError, json.JSONDecodeError):
        return normalize_targets({})


async def save_target_preferences(
    conn,
    user_id: int,
    targets: dict[str, int],
) -> None:
    key = f"{PREFERENCE_KEY_PREFIX}{int(user_id)}"
    normalized = normalize_targets(targets)
    async with conn.cursor() as cur:
        await cur.execute(
            """
            INSERT INTO _system_config (config_key, config_value)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE config_value=VALUES(config_value)
            """,
            (key, json.dumps(normalized, ensure_ascii=False)),
        )


async def _area_labels(conn) -> tuple[dict[str, str], dict[str, int]]:
    async with conn.cursor() as cur:
        await cur.execute(
            """
            SELECT area.name,
                   GROUP_CONCAT(DISTINCT member.name
                       ORDER BY member.name SEPARATOR '、')
            FROM _areas AS area
            LEFT JOIN _area_leader_links AS link ON link.area_id=area.id
            LEFT JOIN _grid_members AS member ON member.id=link.member_id
            GROUP BY area.id, area.name
            ORDER BY area.id
            """
        )
        rows = await cur.fetchall()
    labels: dict[str, str] = {}
    order: dict[str, int] = {}
    for index, (area_name, leaders) in enumerate(rows):
        name = str(area_name or "").strip()
        if not name:
            continue
        leader_text = str(leaders or "").strip()
        labels[name] = f"{name}\n{leader_text}" if leader_text else name
        order[name] = index
    return labels, order


async def _load_roster(conn, business_date: date) -> list[dict[str, Any]]:
    async with conn.cursor() as cur:
        await cur.execute(
            """
            SELECT member.id, member.name, member.position, member.status,
                   GROUP_CONCAT(DISTINCT community.name
                       ORDER BY link.sort_order, community.name SEPARATOR '、'),
                   GROUP_CONCAT(DISTINCT COALESCE(area.name, leader_area.name)
                       ORDER BY COALESCE(area.id, leader_area.id) SEPARATOR '、'),
                   GROUP_CONCAT(DISTINCT department.name
                       ORDER BY link.sort_order, department.name SEPARATOR '、')
            FROM _grid_members AS member
            LEFT JOIN _grid_member_department_links AS link
              ON link.member_id=member.id
            LEFT JOIN _departments AS department
              ON department.id=link.department_id
            LEFT JOIN _communities AS community
              ON community.id=department.community_id
            LEFT JOIN _areas AS area ON area.id=community.area_id
            LEFT JOIN _area_leader_links AS leader_link
              ON leader_link.member_id=member.id
            LEFT JOIN _areas AS leader_area ON leader_area.id=leader_link.area_id
            WHERE TRIM(member.name) <> '' AND member.status <> '离岗'
            GROUP BY member.id, member.name, member.position, member.status
            """
        )
        rows = await cur.fetchall()

    positions = {
        str(row[2] or "").strip() or "未设置岗位"
        for row in rows
    }
    async with conn.cursor() as cur:
        attendance = await get_attendance_context(
            cur,
            start_date=business_date,
            end_date=business_date,
            selected_positions=positions,
        )
    area_labels, area_order = await _area_labels(conn)
    missing_week = normalize_week_start(business_date) in attendance.get(
        "missing_week_starts",
        set(),
    )

    roster: list[dict[str, Any]] = []
    for member_id, name, position, status, communities, areas, departments in rows:
        member_name = str(name).strip()
        member_position = str(position or "").strip() or "未设置岗位"
        community = str(communities or "").split("、", 1)[0].strip()
        area = str(areas or "").split("、", 1)[0].strip()
        department = str(departments or "").split("、", 1)[0].strip()
        attendance_member = attendance.get("members", {}).get(member_name, {
            "id": int(member_id),
            "name": member_name,
            "community": community or "未分配社区",
            "communities": [community] if community else [],
            "position": member_position,
        })
        periods = attendance.get("periods", {}).get(int(member_id), [])
        if period_covers(business_date, periods):
            attendance_text = "否"
        elif (
            business_date.weekday() >= 5
            and member_position in attendance.get("weekend_duty_positions", set())
            and missing_week
        ):
            attendance_text = "未排班"
        else:
            attendance_text = (
                "是"
                if is_member_on_duty(attendance_member, business_date, attendance)
                else "否"
            )

        if area:
            group_key = f"area:{area}"
            group_label = area_labels.get(area, area)
            group_order = area_order.get(area, 999)
        else:
            group_key = f"internal:{member_position}"
            group_label = member_position or department or "其他"
            group_order = 1000

        roster.append({
            "id": int(member_id),
            "name": member_name,
            "position": member_position,
            "community": community,
            "area": area,
            "department": department,
            "group_key": group_key,
            "group_label": group_label,
            "group_order": group_order,
            "attendance": attendance_text,
        })

    position_order = {
        "片长": 0,
        "社区民警": 1,
        "组长": 2,
        "组员": 3,
        "自购房": 4,
        "基础管控": 10,
        "中队长": 11,
        "所队领导": 12,
    }
    roster.sort(key=lambda item: (
        int(item["group_order"]),
        str(item["group_label"]),
        str(item["community"]),
        position_order.get(str(item["position"]), 50),
        str(item["name"]),
    ))
    return roster


def _summary_by_name(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    rows = (payload.get("inspector") or {}).get("data") or []
    return {
        normalized_person_name(row.get("姓名")): row
        for row in rows
        if normalized_person_name(row.get("姓名"))
    }


async def build_daily_detail_data(
    conn,
    business_date: date,
    *,
    rental_target: int,
    self_owned_target: int,
) -> list[dict[str, Any]]:
    roster = await _load_roster(conn, business_date)

    async with conn.cursor() as cur:
        known_positions = await get_known_personnel_positions(cur)
        rental_positions = set(await get_configured_positions(
            cur,
            VISIT_POSITION_CONFIG_KEY,
        )) - {"自购房"}

    rental = await get_visit_summary(
        conn,
        business_date,
        business_date,
        category=VISIT_CATEGORY_RENTAL,
        selected_positions=rental_positions,
        known_positions=known_positions,
    )
    self_owned = await get_visit_summary(
        conn,
        business_date,
        business_date,
        category=VISIT_CATEGORY_SELF_OWNED,
        selected_positions={"自购房"},
        known_positions=known_positions,
    )
    online = await get_summary(business_date.isoformat())

    rental_rows = _summary_by_name(rental)
    self_owned_rows = _summary_by_name(self_owned)
    online_rows = _summary_by_name(online) if online.get("exists") else {}

    result: list[dict[str, Any]] = []
    for member in roster:
        normalized_name = normalized_person_name(member["name"])
        position = str(member["position"])
        if position == "自购房":
            visit_row = self_owned_rows.get(normalized_name)
            target: int | None = self_owned_target
        elif position in rental_positions:
            visit_row = rental_rows.get(normalized_name)
            target = rental_target
        else:
            visit_row = None
            target = None

        task_row = online_rows.get(normalized_name)
        result.append({
            **member,
            "visit_target": target,
            "visits": (
                int(visit_row.get("走访户数") or 0)
                if visit_row is not None
                else None
            ),
            "ratings": (
                int(visit_row.get("星级评定数") or 0)
                if visit_row is not None
                else None
            ),
            "checked_instructions": (
                int(task_row.get("已核查") or 0)
                + int(task_row.get("已完成") or 0)
                if task_row is not None
                else None
            ),
            "completed_instructions": (
                int(task_row.get("已完成") or 0)
                if task_row is not None
                else None
            ),
        })
    return result


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, LAST_COLUMN + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _merge_groups(ws, rows: list[dict[str, Any]]) -> None:
    def merge_column(column: int, key_builder) -> None:
        start = DATA_START_ROW
        current = None
        for offset, item in enumerate(rows):
            row_number = DATA_START_ROW + offset
            key = key_builder(item)
            if current is None:
                current = key
                start = row_number
                continue
            if key == current:
                continue
            if current and row_number - 1 > start:
                ws.merge_cells(
                    start_row=start,
                    start_column=column,
                    end_row=row_number - 1,
                    end_column=column,
                )
            current = key
            start = row_number
        if current and rows and DATA_START_ROW + len(rows) - 1 > start:
            ws.merge_cells(
                start_row=start,
                start_column=column,
                end_row=DATA_START_ROW + len(rows) - 1,
                end_column=column,
            )

    merge_column(2, lambda item: item.get("group_key"))
    merge_column(
        3,
        lambda item: (
            f"{item.get('group_key')}:{item.get('community')}"
            if item.get("community")
            else None
        ),
    )


def build_daily_detail_workbook(
    business_date: date,
    rows: list[dict[str, Any]],
) -> tuple[bytes, str]:
    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook["Sheet1"]
    final_data_rows = max(MIN_DATA_ROWS, len(rows))
    final_row = DATA_START_ROW + final_data_rows - 1
    if final_row > 67:
        worksheet.insert_rows(68, final_row - 67)
        for row_number in range(68, final_row + 1):
            _copy_row_style(worksheet, 67, row_number)

    for merged in list(worksheet.merged_cells.ranges):
        if merged.min_row >= DATA_START_ROW:
            worksheet.unmerge_cells(str(merged))

    worksheet["B2"] = f"{business_date.month}月{business_date.day}号"
    no_fill = PatternFill(fill_type=None)
    for row in worksheet.iter_rows(
        min_row=DATA_START_ROW,
        max_row=final_row,
        min_col=1,
        max_col=LAST_COLUMN,
    ):
        for cell in row:
            cell.value = None
            cell.fill = copy(no_fill)
            cell.font = copy(DATA_FONT)
            cell.alignment = copy(DATA_ALIGNMENT)
            cell.border = copy(DATA_BORDER)
            cell.number_format = "General"

    for index, item in enumerate(rows, start=1):
        row_number = DATA_START_ROW + index - 1
        worksheet.cell(row_number, 1, index)
        worksheet.cell(row_number, 2, item["group_label"])
        worksheet.cell(row_number, 3, item["community"] or None)
        worksheet.cell(row_number, 4, item["name"])
        worksheet.cell(row_number, 5, item["attendance"])
        worksheet.cell(row_number, 9, item["visit_target"])
        worksheet.cell(row_number, 10, item["visits"])
        worksheet.cell(row_number, 11, item["ratings"])
        worksheet.cell(row_number, 12, item["checked_instructions"])
        worksheet.cell(row_number, 13, item["completed_instructions"])

    _merge_groups(worksheet, rows)
    worksheet.print_area = f"A1:V{final_row}"
    worksheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    filename = (
        f"{business_date.strftime('%m%d')}滨湖网格工作每日明细.xlsx"
    )
    return output.getvalue(), filename


async def build_daily_detail_export(
    conn,
    business_date: date,
    *,
    rental_target: int,
    self_owned_target: int,
) -> tuple[bytes, str, int]:
    rows = await build_daily_detail_data(
        conn,
        business_date,
        rental_target=rental_target,
        self_owned_target=self_owned_target,
    )
    content, filename = build_daily_detail_workbook(business_date, rows)
    return content, filename, len(rows)
