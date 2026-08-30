from __future__ import annotations

from datetime import datetime
from io import BytesIO
import json
import os

import pytest
from fastapi import UploadFile
from openpyxl import load_workbook

os.environ.setdefault("MYSQL_PASSWORD", "test-password")
os.environ.setdefault("ENCRYPTION_KEY", "test-encryption-key")

from routers import mobile_tasks, registry
from routers.mobile_tasks import TaskSearch
from routers.registry import PropertySearch, _property_search_result
from services.xlsx_export import build_xlsx


def test_xlsx_export_escapes_formula_like_user_text():
    payload = build_xlsx("测试", ["字段"], [["=HYPERLINK(\"bad\")"], ["普通文本"]])
    sheet = load_workbook(payload, data_only=False).active

    assert sheet["A2"].value == "'=HYPERLINK(\"bad\")"
    assert sheet["A3"].value == "普通文本"


class _TaskExportCursor:
    def __init__(self):
        self.sql = ""

    async def __aenter__(self):
        return self

    async def __aexit__(self, *_args):
        return False

    async def execute(self, sql, _params=()):
        self.sql = sql

    async def fetchall(self):
        return [
            (
                "疑似返苏",
                "task-1",
                json.dumps({"姓名": "测试人员", "电话号码": "13800138000", "核查结果": "无法核实"}, ensure_ascii=False),
                1,
                0,
                "",
                "checked",
                31,
                4,
                "abc123",
            )
        ]


class _TaskExportConnection:
    def __init__(self):
        self.query_cursor = _TaskExportCursor()

    def cursor(self):
        return self.query_cursor


@pytest.mark.asyncio
async def test_mobile_task_export_reuses_list_order_and_writes_public_fields(monkeypatch):
    async def flow_context(_conn, _user):
        return {}

    async def review_flows(_cur, _rows):
        return {("疑似返苏", "task-1"): {"state": "initial_pending", "flow_version": 2}}

    monkeypatch.setattr(mobile_tasks, "_flow_context", flow_context)
    monkeypatch.setattr(mobile_tasks, "_task_where", lambda *_args: ("1=1", []))
    monkeypatch.setattr(mobile_tasks, "review_flows_by_rows", review_flows)
    conn = _TaskExportConnection()

    payload, count = await mobile_tasks._mobile_export_workbook(
        data=TaskSearch(sort="address_asc"),
        parser_type="疑似返苏",
        user={"id": 1},
        conn=conn,
    )
    sheet = load_workbook(payload, data_only=True).active
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]

    assert count == 1
    assert f"ORDER BY {mobile_tasks._task_order('疑似返苏', 'address_asc')}" in conn.query_cursor.sql
    assert "任务标识" not in headers
    assert "来源ID" not in headers
    assert "手机号" in headers
    assert "现住址" in headers
    assert values[headers.index("手机号")] == "13800138000"
    assert values[headers.index("研判阶段")] == "初步待研判"


@pytest.mark.asyncio
async def test_analysis_export_can_be_imported_with_chinese_stage(monkeypatch):
    workbook = build_xlsx(
        "研判任务",
        [
            "业务类型", "任务标识", "来源ID", "来源版本", "来源行哈希",
            "流程版本", "研判阶段", "本次研判决定", "研判意见",
        ],
        [["疑似返苏", "task-1", 31, 4, "abc123", 2, "初步待研判", "成功", "继续核查"]],
    )
    captured = {}

    async def decide(parser_type, source_id, decision, _request, _user, _conn):
        captured.update(parser_type=parser_type, source_id=source_id, decision=decision)
        return {"review_flow": {"state": "initial_extension"}}

    async def audit(*_args, **_kwargs):
        return None

    monkeypatch.setattr(mobile_tasks, "decide_mobile_task_unverifiable_review", decide)
    monkeypatch.setattr(mobile_tasks, "record_admin_audit", audit)
    monkeypatch.setattr(mobile_tasks, "request_audit_fields", lambda _request: {})
    upload = UploadFile(filename="研判任务.xlsx", file=BytesIO(workbook.getvalue()))

    result = await mobile_tasks.import_mobile_task_analysis(
        request=None,
        file=upload,
        user={"id": 1},
        conn=object(),
    )

    assert result["success_count"] == 1
    assert result["failed_count"] == 0
    assert captured["parser_type"] == "疑似返苏"
    assert captured["source_id"] == 31
    assert captured["decision"].stage == "initial_pending"
    assert captured["decision"].outcome == "success"


class _PropertySortCursor:
    def __init__(self):
        self.sql_calls: list[str] = []
        self.mode = ""

    async def __aenter__(self):
        return self

    async def __aexit__(self, *_args):
        return False

    async def execute(self, sql, _params=()):
        self.sql_calls.append(sql)
        self.mode = "count" if sql.startswith("SELECT COUNT") else "rows"

    async def fetchone(self):
        return (2,)

    async def fetchall(self):
        if self.mode != "rows":
            return []
        return [
            (1, "", 1, "长板社区", "一号", "", "", "", "", "", None,
             "manual", "", "一号", "active", 1, None, datetime(2026, 8, 1),
             0, None, None, None, None, None, 0, 0),
            (2, "", 1, "长板社区", "二号", "", "", "", "", "", None,
             "manual", "", "二号", "active", 1, None, datetime(2026, 8, 2),
             0, None, None, None, None, None, 0, 0),
        ]


class _PropertySortConnection:
    def __init__(self):
        self.query_cursor = _PropertySortCursor()

    def cursor(self):
        return self.query_cursor


@pytest.mark.asyncio
async def test_property_visit_sort_is_real_and_applied_before_pagination(monkeypatch):
    async def allowed(_user, _permission):
        return None

    async def summaries(_cur, _properties):
        return {
            1: {"visit_count": 1, "latest_visit_date": "2026-08-29", "latest_star_rating": None, "latest_star_rating_at": None},
            2: {"visit_count": 1, "latest_visit_date": "2026-08-20", "latest_star_rating": None, "latest_star_rating_at": None},
        }

    monkeypatch.setattr(registry, "_allowed_community_ids", allowed)
    monkeypatch.setattr(registry, "load_property_visit_summaries", summaries)
    conn = _PropertySortConnection()

    result = await _property_search_result(
        PropertySearch(sort="visit_desc", page=1, page_size=1),
        {"id": 1},
        conn,
    )

    row_query = next(sql for sql in conn.query_cursor.sql_calls if sql.startswith("SELECT property.id"))
    assert "LIMIT" not in row_query
    assert result["total"] == 2
    assert [row["id"] for row in result["data"]] == [1]
