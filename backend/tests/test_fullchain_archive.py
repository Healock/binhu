import io
import json
import os
import unittest
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

os.environ.setdefault("MYSQL_PASSWORD", "test-password")
os.environ.setdefault("ENCRYPTION_KEY", "test-encryption-key")

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from routers.fullchain_archive import (
    CandidateSearch,
    POLICE_RAW_RETIRED_MESSAGE,
    _candidate_rows,
    _filter_candidate_rows,
    _parse_deadline,
    _preview_token,
    confirm_police_raw,
    preview_police_raw,
    require_fullchain_archive,
)
from routers.police_dispatch import require_police_access
from services.permissions import POLICE_DISPATCH_MANAGE
from services.fullchain_archive import build_archive_workbook, parse_police_raw
from services.fullchain_archive import (
    REGISTRATION_ARCHIVE_RULE_VERSION,
    registration_archive_ready,
)
from services.fullchain_archive_jobs import (
    ArchiveStageError,
    _acquire_sheet_lock_with_retry,
    _classify_archive_error,
    _commit_platform_archive,
    _delete_source_row_once,
    _validate_registration_archive_evidence,
    _reconcile_deleted_archive_item,
    _safe_error_code,
    _stage_platform_archive,
    reconcile_deleted_archive_items,
    recover_interrupted_fullchain_exports,
)
from services.report_builders.base import BaseReportBuilder
from services.task_workflow import TASK_WORKFLOWS
from services.unverifiable_review import FINAL_UNVERIFIABLE


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def candidate_source_row(
    values,
    *,
    source_id=7,
    row_key="a" * 32,
    revision=3,
    row_hash="b" * 64,
    identity_hmac="c" * 64,
    source_count=1,
    conflict=0,
    decision="",
    note="",
    registration_status="",
    confirmed_at=None,
    registration_source_id=None,
    registration_revision=None,
    registration_row_hash="",
    registration_identity_hmac="",
    task_community="",
    property_id=None,
    property_version=None,
    property_status=None,
    current_property_version=None,
    active_writeback=0,
    active_archive=0,
    community="长板社区",
):
    return (
        source_id, row_key, revision, row_hash, 20, 4, "sheet-1",
        json.dumps(values, ensure_ascii=False), identity_hmac, source_count,
        conflict, decision, note, registration_status, confirmed_at,
        registration_source_id, registration_revision, registration_row_hash,
        registration_identity_hmac, task_community, property_id, property_version,
        property_status, current_property_version, active_writeback, active_archive,
        community,
    )


class FullchainArchiveTests(unittest.TestCase):
    def test_candidate_selection_filters_stages_keywords_and_eligibility(self):
        rows = [
            {"source_id": 10, "stage": "direct", "eligible": True, "name": "甲", "identity": "", "phone": "", "address": "一号"},
            {"source_id": 11, "stage": "direct", "eligible": False, "name": "乙", "identity": "", "phone": "", "address": "二号"},
            {"source_id": 12, "stage": "review", "eligible": True, "name": "丙", "identity": "", "phone": "", "address": "三号"},
        ]
        filtered = _filter_candidate_rows(rows, CandidateSearch(stages=["direct"], keyword="甲"))
        self.assertEqual([row["source_id"] for row in filtered], [10])
        self.assertEqual([row["source_id"] for row in filtered if row["eligible"]], [10])

    @patch("services.fullchain_archive.hmac_digest", side_effect=lambda value, kind: (f"digest-{value}", 1))
    def test_police_raw_parser_finds_identity_column_and_deduplicates(self, _digest):
        content = workbook_bytes([
            ["说明"],
            ["姓名", "身份证号码", "反馈结果"],
            ["甲", "320000199001010011", "在册"],
            ["乙", "320000199001010011", "在册"],
            ["丙", "320000199001010022", ""],
        ])
        result = parse_police_raw(content, "公安网.xlsx")
        self.assertEqual(result["row_count"], 3)
        self.assertEqual(result["duplicate_count"], 1)
        self.assertEqual(result["identity_hmacs"], [
            "digest-320000199001010011", "digest-320000199001010022",
        ])
        self.assertNotIn("320000199001010011", str(result["preview"]))

    def test_archive_workbook_keeps_sensitive_numbers_as_text(self):
        content = build_archive_workbook([{
            "physical_row": 20, "source": "测试", "name": "甲",
            "identity": "320000199001010010", "phone": "13800000000",
            "address": "测试地址", "registration": "", "result": "离苏",
            "category": "离苏",
        }], 12, __import__("datetime").datetime(2026, 8, 21, 10, 0))
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        sheet = workbook["离苏"]
        self.assertEqual(sheet.cell(2, 4).value, "320000199001010010")
        self.assertEqual(sheet.cell(2, 4).data_type, "s")
        self.assertEqual(sheet.cell(2, 5).value, "13800000000")
        self.assertEqual(sheet.cell(1, 10).value, "初步研判")
        self.assertEqual(sheet.cell(1, 22).value, "自动流转记录")
        workbook.close()

    def test_deadline_parsing_uses_business_year_and_handles_full_date(self):
        today = date(2026, 8, 21)
        self.assertEqual(_parse_deadline("08-20", today), date(2026, 8, 20))
        self.assertEqual(_parse_deadline("2026-08-22", today), date(2026, 8, 22))
        self.assertIsNone(_parse_deadline("未知", today))

    def test_preview_token_changes_with_source_revision(self):
        item = {"source_id": 1, "revision": 1, "row_hash": "a" * 64, "category": "离苏"}
        changed = dict(item, revision=2)
        self.assertNotEqual(_preview_token([item]), _preview_token([changed]))

    def test_registration_archive_uses_exact_24_hour_boundary(self):
        confirmed = datetime(2026, 8, 28, 9, 30)
        self.assertFalse(registration_archive_ready(
            confirmed, now=confirmed + timedelta(hours=24) - timedelta(seconds=1)
        ))
        self.assertTrue(registration_archive_ready(
            confirmed, now=confirmed + timedelta(hours=24)
        ))

    def test_registration_archive_evidence_columns_exist_in_bootstrap_and_runtime_schema(self):
        backend_root = Path(__file__).parents[1]
        init_sql = backend_root.joinpath("init.sql").read_text(encoding="utf-8")
        runtime_schema = backend_root.joinpath("database.py").read_text(encoding="utf-8")
        for column in (
            "registration_confirmed_at",
            "registration_status",
            "registration_identity_hmac",
            "registration_property_id",
            "registration_property_version",
            "candidate_rule_version",
        ):
            with self.subTest(column=column):
                self.assertIn(column, init_sql)
                self.assertIn(column, runtime_schema)

    def test_transfer_options_are_split_but_legacy_value_remains_readable(self):
        options = TASK_WORKFLOWS["全链条"].result_options
        self.assertIn("移交（所内）", options)
        self.assertIn("移交（所外）", options)
        self.assertIn("移交", options)

    def test_report_category_prefers_specific_transfer_value(self):
        builder = BaseReportBuilder()
        sql = builder.ledger_result_category_sql("snapshot")
        self.assertLess(sql.index("移交（所内）"), sql.index("LIKE '%%移交%%'"))

    def test_background_error_codes_never_persist_raw_exception_text(self):
        self.assertEqual(
            _safe_error_code(RuntimeError("source_row_changed"), "delete_failed"),
            "source_row_changed",
        )
        self.assertEqual(
            _safe_error_code(RuntimeError("upstream body with personal data"), "delete_failed"),
            "delete_failed",
        )

    def test_database_transaction_errors_have_distinct_safe_codes(self):
        cases = {
            1213: "archive_transaction_deadlock",
            1205: "archive_transaction_timeout",
            2006: "archive_database_unavailable",
        }
        for errno, expected in cases.items():
            with self.subTest(errno=errno):
                error = _classify_archive_error(
                    RuntimeError(errno, "sensitive database detail"),
                    stage="archive_insert",
                    fallback="archive_insert_failed",
                )
                self.assertEqual(error.code, expected)
                self.assertEqual(error.stage, "archive_insert")
                self.assertEqual(len(error.fingerprint), 64)
                self.assertNotIn("sensitive", error.fingerprint)


class _Cursor:
    def __init__(self, rows=None):
        self.rows = rows or []
        self.executions = []

    async def execute(self, sql, params=()):
        self.executions.append((sql, params))

    async def fetchall(self):
        return self.rows

    async def fetchone(self):
        return None

    async def __aenter__(self):
        return self

    async def __aexit__(self, *_args):
        return False


class _SequenceCursor(_Cursor):
    def __init__(self, result_sets):
        super().__init__()
        self.result_sets = list(result_sets)

    async def fetchall(self):
        return self.result_sets.pop(0) if self.result_sets else []


class _Connection:
    def __init__(self, cursor):
        self._cursor = cursor

    def cursor(self):
        return self._cursor


class _DeleteCursor(_Cursor):
    def __init__(self, source_row=(20, 4, "expected-hash"), *, confirm_rowcount=1):
        super().__init__()
        self.source_row = source_row
        self.rowcount = 1
        self.confirm_rowcount = confirm_rowcount

    async def execute(self, sql, params=()):
        await super().execute(sql, params)
        normalized = " ".join(str(sql).split())
        self.rowcount = self.confirm_rowcount if "external_delete_state='deleted'" in normalized else 1

    async def fetchone(self):
        return self.source_row


class _TransactionalConnection(_Connection):
    def __init__(self, cursor):
        super().__init__(cursor)
        self.events = []

    async def begin(self):
        self.events.append("begin")

    async def commit(self):
        self.events.append("commit")

    async def rollback(self):
        self.events.append("rollback")


class _ReconcileItemCursor(_Cursor):
    def __init__(self):
        super().__init__()
        self.rowcount = 1
        self.last_sql = ""

    async def execute(self, sql, params=()):
        await super().execute(sql, params)
        self.last_sql = " ".join(str(sql).split())

    async def fetchone(self):
        if "SELECT external_delete_state,status" in self.last_sql:
            return ("deleted", "conflict")
        return None


class _AcquireContext:
    def __init__(self, conn):
        self.conn = conn

    async def __aenter__(self):
        return self.conn

    async def __aexit__(self, *_args):
        return False


class _Pool:
    def __init__(self, conn):
        self.conn = conn

    def acquire(self):
        return _AcquireContext(self.conn)


class FullchainArchiveAsyncTests(unittest.IsolatedAsyncioTestCase):
    async def test_police_raw_write_endpoints_are_retired_but_explicit(self):
        for endpoint in (preview_police_raw, confirm_police_raw):
            with self.subTest(endpoint=endpoint.__name__):
                with self.assertRaises(HTTPException) as raised:
                    await endpoint(user={"id": 1})
                self.assertEqual(raised.exception.status_code, 410)
                self.assertEqual(raised.exception.detail, POLICE_RAW_RETIRED_MESSAGE)

    async def test_registered_candidate_requires_confirmed_registration_evidence(self):
        values = {
            "姓名": "测试人员",
            "身份证号": "TEST_IDENTITY_001",
            "社区": "长板社区",
            "核查结果": "已登记",
        }
        rows = await _candidate_rows(_SequenceCursor([[
            candidate_source_row(values),
        ], []]))

        self.assertEqual(len(rows), 1)
        self.assertFalse(rows[0]["eligible"])
        self.assertIn("历史已登记", rows[0]["reason"])

    async def test_confirmed_registration_waits_a_full_24_hours(self):
        values = {
            "姓名": "测试人员",
            "身份证号": "TEST_IDENTITY_001",
            "社区": "长板社区",
            "核查结果": "已登记",
        }
        confirmed_at = datetime.utcnow() - timedelta(hours=23, minutes=59)
        source_row = candidate_source_row(
            values,
            registration_status="confirmed",
            confirmed_at=confirmed_at,
            registration_source_id=7,
            registration_revision=3,
            registration_row_hash="b" * 64,
            registration_identity_hmac="c" * 64,
            task_community="长板社区",
            property_id=19,
            property_version=2,
            property_status="active",
            current_property_version=2,
        )

        rows = await _candidate_rows(_SequenceCursor([[source_row], []]))

        self.assertFalse(rows[0]["eligible"])
        self.assertIn("完整保留 24 小时", rows[0]["reason"])
        self.assertEqual(rows[0]["registration_status"], "confirmed")
        self.assertIsNotNone(rows[0]["archive_available_at"])

    async def test_confirmed_registration_is_eligible_after_24_hours(self):
        values = {
            "姓名": "测试人员",
            "身份证号": "TEST_IDENTITY_001",
            "社区": "长板社区",
            "核查结果": "已登记",
        }
        confirmed_at = datetime.utcnow() - timedelta(hours=24, seconds=1)
        source_row = candidate_source_row(
            values,
            registration_status="confirmed",
            confirmed_at=confirmed_at,
            registration_source_id=7,
            registration_revision=3,
            registration_row_hash="b" * 64,
            registration_identity_hmac="c" * 64,
            task_community="长板社区",
            property_id=19,
            property_version=2,
            property_status="active",
            current_property_version=2,
        )
        cursor = _SequenceCursor([[source_row], []])

        rows = await _candidate_rows(cursor, include_source_values=True)

        self.assertTrue(rows[0]["eligible"])
        self.assertEqual(rows[0]["candidate_rule_version"], REGISTRATION_ARCHIVE_RULE_VERSION)
        self.assertEqual(rows[0]["_registration_confirmed_at_db"], confirmed_at)
        candidate_sql = cursor.executions[0][0]
        self.assertNotIn("_fullchain_police_raw_identities", candidate_sql)
        self.assertNotIn("_fullchain_police_raw_uploads", candidate_sql)

    async def test_registration_candidate_rejects_changed_or_pending_context(self):
        values = {
            "姓名": "测试人员",
            "身份证号": "TEST_IDENTITY_001",
            "社区": "长板社区",
            "核查结果": "已登记",
        }
        base = dict(
            registration_status="confirmed",
            confirmed_at=datetime.utcnow() - timedelta(hours=25),
            registration_source_id=7,
            registration_revision=3,
            registration_row_hash="b" * 64,
            registration_identity_hmac="c" * 64,
            task_community="长板社区",
            property_id=19,
            property_version=2,
            property_status="active",
            current_property_version=2,
        )
        cases = {
            "source": ({**base, "registration_revision": 2}, "来源版本"),
            "identity": ({**base, "registration_identity_hmac": "d" * 64}, "核查对象"),
            "community": ({**base, "task_community": "龙河社区"}, "任务社区"),
            "property_status": ({**base, "property_status": "inactive"}, "房屋已停用"),
            "property_version": ({**base, "current_property_version": 3}, "房屋档案已更新"),
            "writeback": ({**base, "active_writeback": 1}, "待同步或冲突"),
            "archive": ({**base, "active_archive": 1}, "其他未终结归档批次"),
        }
        for label, (kwargs, reason) in cases.items():
            with self.subTest(label=label):
                rows = await _candidate_rows(_SequenceCursor([[
                    candidate_source_row(values, **kwargs),
                ], []]))
                self.assertFalse(rows[0]["eligible"])
                self.assertIn(reason, rows[0]["reason"])

    async def test_frozen_registration_evidence_is_rechecked_before_delete(self):
        confirmed_at = datetime(2026, 8, 27, 8, 0)
        matching = (
            "confirmed", confirmed_at, 7, 3, "b" * 64, "c" * 64,
            "长板社区", 19, 2, 3, "b" * 64, "c" * 64,
            "长板社区", 1, 0, json.dumps({"核查结果": "已登记"}, ensure_ascii=False),
            "active", 2, 1, 0,
        )
        await _validate_registration_archive_evidence(
            _DeleteCursor(source_row=matching),
            parser_type="全链条",
            row_key="a" * 32,
            source_id=7,
            expected_revision=3,
            expected_hash="b" * 64,
            registration_confirmed_at=confirmed_at,
            registration_status="confirmed",
            registration_identity_hmac="c" * 64,
            registration_property_id=19,
            registration_property_version=2,
            candidate_rule_version=REGISTRATION_ARCHIVE_RULE_VERSION,
        )

        changed = list(matching)
        changed[17] = 3
        with self.assertRaises(ArchiveStageError) as raised:
            await _validate_registration_archive_evidence(
                _DeleteCursor(source_row=tuple(changed)),
                parser_type="全链条",
                row_key="a" * 32,
                source_id=7,
                expected_revision=3,
                expected_hash="b" * 64,
                registration_confirmed_at=confirmed_at,
                registration_status="confirmed",
                registration_identity_hmac="c" * 64,
                registration_property_id=19,
                registration_property_version=2,
                candidate_rule_version=REGISTRATION_ARCHIVE_RULE_VERSION,
            )
        self.assertEqual(raised.exception.code, "registration_archive_evidence_changed")
        self.assertEqual(raised.exception.stage, "registration_evidence")

    @patch(
        "services.fullchain_archive_jobs._validate_registration_archive_evidence",
        new=AsyncMock(side_effect=ArchiveStageError(
            "registration_archive_evidence_changed", "registration_evidence"
        )),
    )
    async def test_changed_registration_evidence_never_calls_tencent_delete(self):
        conn = _TransactionalConnection(_DeleteCursor())
        client = AsyncMock()

        with self.assertRaises(ArchiveStageError):
            await _delete_source_row_once(
                conn,
                client,
                export_id=12,
                source_id=7,
                parser_type="全链条",
                spreadsheet={"file_id": "file-1"},
                sheet_id="sheet-1",
                physical_row=20,
                expected_revision=3,
                expected_hash="b" * 64,
                source_columns=["核查结果"],
                parser=object(),
                external_delete_state="pending",
                row_key="a" * 32,
                category="已登记",
                registration_confirmed_at=datetime(2026, 8, 27, 8, 0),
                registration_status="confirmed",
                registration_identity_hmac="c" * 64,
                registration_property_id=19,
                registration_property_version=2,
                candidate_rule_version=REGISTRATION_ARCHIVE_RULE_VERSION,
            )

        client.read_source_row.assert_not_awaited()
        client.batch_update.assert_not_awaited()

    async def test_archive_lock_retries_until_previous_write_releases(self):
        attempts = []

        async def acquire(_cur, spreadsheet_id, timeout):
            attempts.append((spreadsheet_id, timeout))
            return len(attempts) == 3

        sleep = AsyncMock()
        self.assertTrue(await _acquire_sheet_lock_with_retry(
            object(), 7, acquire=acquire, sleep=sleep, retry_delays=(1, 2, 3),
        ))
        self.assertEqual(len(attempts), 3)
        self.assertEqual([call.args[0] for call in sleep.await_args_list], [1, 2])
        self.assertTrue(all(timeout == 5 for _, timeout in attempts))

    async def test_archive_lock_retry_has_a_bounded_wait(self):
        acquire = AsyncMock(return_value=False)
        sleep = AsyncMock()
        self.assertFalse(await _acquire_sheet_lock_with_retry(
            object(), 7, acquire=acquire, sleep=sleep, retry_delays=(1, 2),
        ))
        self.assertEqual(acquire.await_count, 3)
        self.assertEqual([call.args[0] for call in sleep.await_args_list], [1, 2])

    async def test_duplicate_projection_is_visible_but_cannot_archive(self):
        values = {
            "姓名": "测试人员",
            "身份证号": "320000199001010011",
            "核查结果": "离苏",
        }
        cursor = _SequenceCursor([[candidate_source_row(
            values, source_count=2, conflict=1,
        )], []])

        rows = await _candidate_rows(cursor)

        self.assertEqual(len(rows), 1)
        self.assertFalse(rows[0]["eligible"])
        self.assertIn("重复或冲突来源行", rows[0]["reason"])
        self.assertNotIn("_source_values_json", rows[0])

    async def test_old_manual_archive_decision_cannot_bypass_two_stage_review(self):
        values = {"姓名": "测试人员", "核查结果": "无法核实", "截止日期": "2026-08-20"}
        source_row = candidate_source_row(values, decision="archive")
        rows = await _candidate_rows(_SequenceCursor([[source_row], []]))
        self.assertEqual(len(rows), 1)
        self.assertFalse(rows[0]["eligible"])
        self.assertIn("等待系统建立两级研判流程", rows[0]["reason"])

    async def test_only_final_unverifiable_flow_is_directly_exportable(self):
        values = {"姓名": "测试人员", "核查结果": "无法核实"}
        source_row = candidate_source_row(values)
        flow_row = (
            5, "全链条", "a" * 32, 1, 7, 3, "b" * 64,
            FINAL_UNVERIFIABLE, 4, None, "", "", 0, "", None,
            None, None, None, None,
        )
        rows = await _candidate_rows(
            _SequenceCursor([[source_row], [flow_row]]),
            include_source_values=True,
        )
        self.assertEqual(len(rows), 1)
        self.assertTrue(rows[0]["eligible"])
        self.assertEqual(rows[0]["stage"], "direct")
        self.assertEqual(rows[0]["category"], "无法核实")
        self.assertEqual(
            json.loads(rows[0]["_source_values_json"]),
            values,
        )

    async def test_final_unverifiable_export_stops_when_source_snapshot_changed(self):
        values = {"姓名": "测试人员", "核查结果": "无法核实"}
        source_row = candidate_source_row(
            values, revision=4, row_hash="c" * 64, identity_hmac="d" * 64,
        )
        stale_flow = (
            5, "全链条", "a" * 32, 1, 7, 3, "b" * 64,
            FINAL_UNVERIFIABLE, 4, None, "", "", 0, "", None,
            None, None, None, None,
        )

        rows = await _candidate_rows(_SequenceCursor([[source_row], [stale_flow]]))

        self.assertEqual(len(rows), 1)
        self.assertFalse(rows[0]["eligible"])
        self.assertEqual(rows[0]["stage"], "review")
        self.assertIn("来源已变化", rows[0]["reason"])

    @patch(
        "services.fullchain_archive_jobs.get_database_column_map",
        new=AsyncMock(return_value={"姓名": "姓名"}),
    )
    async def test_platform_archive_is_idempotent_per_export(self):
        cursor = _Cursor()
        parser = type("Parser", (), {"table_name": "t_fullchain", "COLUMNS": ["姓名"]})()

        await _stage_platform_archive(
            _Connection(cursor), parser, 123, "a" * 32, {"姓名": "测试人员"}
        )

        self.assertEqual(len(cursor.executions), 3)
        self.assertIn("DELETE FROM OnlineDataArchive.t_fullchain_archive", cursor.executions[0][0])
        self.assertEqual(cursor.executions[0][1][1], "fullchain_feedback_export:123")
        self.assertIn("_archive_reason", cursor.executions[1][0])
        self.assertIn("DELETE FROM `t_fullchain`", cursor.executions[2][0])

    @patch("services.fullchain_archive_jobs.source_row_hash", return_value="expected-hash")
    async def test_external_delete_persists_deleting_then_deleted(self, _hash):
        cursor = _DeleteCursor()
        conn = _TransactionalConnection(cursor)
        client = AsyncMock()
        client.read_source_row.return_value = {"values": {"姓名": "测试人员"}}
        client.build_delete_row_request = MagicMock(return_value={"delete": 20})
        parser = type("Parser", (), {"normalize_source_row": staticmethod(lambda values: values)})()

        await _delete_source_row_once(
            conn,
            client,
            export_id=12,
            source_id=7,
            parser_type="全链条",
            spreadsheet={"file_id": "file-1"},
            sheet_id="sheet-1",
            physical_row=20,
            expected_revision=4,
            expected_hash="expected-hash",
            source_columns=["姓名"],
            parser=parser,
            external_delete_state="pending",
        )

        self.assertEqual(conn.events, ["commit", "commit"])
        client.batch_update.assert_awaited_once()
        sql = "\n".join(item[0] for item in cursor.executions)
        self.assertIn("external_delete_state='deleting'", sql)
        self.assertIn("external_delete_state='deleted'", sql)

    async def test_confirmed_external_delete_is_never_sent_again(self):
        conn = _TransactionalConnection(_DeleteCursor())
        client = AsyncMock()

        await _delete_source_row_once(
            conn,
            client,
            export_id=12,
            source_id=7,
            parser_type="全链条",
            spreadsheet={"file_id": "file-1"},
            sheet_id="sheet-1",
            physical_row=20,
            expected_revision=4,
            expected_hash="expected-hash",
            source_columns=["姓名"],
            parser=object(),
            external_delete_state="deleted",
        )

        client.read_source_row.assert_not_awaited()
        client.batch_update.assert_not_awaited()
        self.assertEqual(conn.events, [])

    async def test_uncertain_external_delete_stops_without_retry(self):
        conn = _TransactionalConnection(_DeleteCursor())
        client = AsyncMock()

        with self.assertRaisesRegex(RuntimeError, "external_delete_outcome_unknown"):
            await _delete_source_row_once(
                conn,
                client,
                export_id=12,
                source_id=7,
                parser_type="全链条",
                spreadsheet={"file_id": "file-1"},
                sheet_id="sheet-1",
                physical_row=20,
                expected_revision=4,
                expected_hash="expected-hash",
                source_columns=["姓名"],
                parser=object(),
                external_delete_state="deleting",
            )

        client.batch_update.assert_not_awaited()
        self.assertEqual(conn.events, [])

    @patch("services.fullchain_archive_jobs.source_row_hash", return_value="expected-hash")
    async def test_external_delete_confirmation_failure_pauses_local_archive(self, _hash):
        conn = _TransactionalConnection(_DeleteCursor(confirm_rowcount=0))
        client = AsyncMock()
        client.read_source_row.return_value = {"values": {}}
        client.build_delete_row_request = MagicMock(return_value={"delete": 20})
        parser = type("Parser", (), {"normalize_source_row": staticmethod(lambda values: values)})()

        with self.assertRaisesRegex(RuntimeError, "external_delete_outcome_unknown"):
            await _delete_source_row_once(
                conn,
                client,
                export_id=12,
                source_id=7,
                parser_type="全链条",
                spreadsheet={"file_id": "file-1"},
                sheet_id="sheet-1",
                physical_row=20,
                expected_revision=4,
                expected_hash="expected-hash",
                source_columns=[],
                parser=parser,
                external_delete_state="pending",
            )

        self.assertEqual(conn.events, ["commit"])
        client.batch_update.assert_awaited_once()

    @patch("services.fullchain_archive_jobs._mark_item", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs.mark_flow_archived", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs._stage_platform_archive", new_callable=AsyncMock)
    async def test_platform_archive_and_item_success_share_transaction(
        self, stage_archive, mark_archived, mark_item,
    ):
        conn = _TransactionalConnection(_Cursor())
        await _commit_platform_archive(
            conn,
            object(),
            export_id=12,
            source_id=7,
            parser_type="全链条",
            row_key="a" * 32,
            values={"核查结果": "无法核实"},
        )
        self.assertEqual(conn.events, ["begin", "commit"])
        stage_archive.assert_awaited_once()
        mark_archived.assert_awaited_once()
        mark_item.assert_awaited_once()

    @patch("services.fullchain_archive_jobs._mark_item", new_callable=AsyncMock)
    @patch(
        "services.fullchain_archive_jobs.mark_flow_archived",
        new=AsyncMock(side_effect=RuntimeError("review_flow_state_conflict")),
    )
    @patch("services.fullchain_archive_jobs._stage_platform_archive", new_callable=AsyncMock)
    async def test_review_flow_conflict_is_not_flattened_to_platform_failure(
        self, _stage_archive, _mark_item,
    ):
        conn = _TransactionalConnection(_Cursor())
        with self.assertRaises(ArchiveStageError) as raised:
            await _commit_platform_archive(
                conn,
                object(),
                export_id=12,
                source_id=7,
                parser_type="全链条",
                row_key="a" * 32,
                values={},
            )
        self.assertEqual(raised.exception.code, "review_flow_state_conflict")
        self.assertEqual(raised.exception.stage, "review_flow_archive")
        self.assertEqual(conn.events, ["begin", "rollback"])

    @patch(
        "services.fullchain_archive_jobs.get_database_column_map",
        new=AsyncMock(return_value={"姓名": "姓名"}),
    )
    async def test_archive_insert_and_current_remove_have_distinct_stages(self):
        class StageCursor(_Cursor):
            def __init__(self, fail_on: int | None = None, remove_rowcount: int = 1):
                super().__init__()
                self.fail_on = fail_on
                self.remove_rowcount = remove_rowcount
                self.rowcount = 1

            async def execute(self, sql, params=()):
                await super().execute(sql, params)
                if self.fail_on == len(self.executions):
                    raise RuntimeError("database detail")
                self.rowcount = (
                    self.remove_rowcount if "DELETE FROM `t_fullchain`" in sql else 1
                )

        parser = type("Parser", (), {"table_name": "t_fullchain", "COLUMNS": ["姓名"]})()
        with self.assertRaises(ArchiveStageError) as insert_error:
            await _stage_platform_archive(
                _Connection(StageCursor(fail_on=2)), parser, 12, "a" * 32, {"姓名": "甲"}
            )
        self.assertEqual(insert_error.exception.code, "archive_insert_failed")
        self.assertEqual(insert_error.exception.stage, "archive_insert")

        with self.assertRaises(ArchiveStageError) as remove_error:
            await _stage_platform_archive(
                _Connection(StageCursor(remove_rowcount=0)), parser, 12, "a" * 32, {"姓名": "甲"}
            )
        self.assertEqual(remove_error.exception.code, "current_row_remove_failed")
        self.assertEqual(remove_error.exception.stage, "current_row_remove")

    @patch("services.fullchain_archive_jobs.mark_flow_archived", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs._lock_reconcile_source", new=AsyncMock(return_value=False))
    @patch("services.fullchain_archive_jobs._remove_current_platform_row", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs._insert_archive_from_snapshot", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs._stage_platform_archive", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs.get_parser")
    async def test_deleted_item_accepts_matching_sync_archive_without_redeleting(
        self, get_parser, stage_archive, insert_snapshot, remove_current, mark_archived,
    ):
        parser = type(
            "Parser",
            (),
            {
                "COLUMNS": ["姓名"],
                "normalize_source_row": staticmethod(
                    lambda values: {"姓名": str(values.get("姓名") or "")}
                ),
            },
        )()
        get_parser.return_value = parser
        conn = _TransactionalConnection(_ReconcileItemCursor())
        item = (12, 7, "全链条", "a" * 32, 4, "b" * 64, '{"姓名":"甲"}')
        with patch(
            "services.fullchain_archive_jobs._platform_rows_for_reconciliation",
            new=AsyncMock(return_value=(None, [({"姓名": "甲"}, "online_removed")])),
        ):
            result = await _reconcile_deleted_archive_item(conn, item)
        self.assertEqual(result, "reconciled_by_sync")
        self.assertEqual(conn.events, ["begin", "commit"])
        stage_archive.assert_not_awaited()
        insert_snapshot.assert_not_awaited()
        remove_current.assert_not_awaited()
        mark_archived.assert_awaited_once()
        sql = "\n".join(statement for statement, _params in conn._cursor.executions)
        self.assertIn("reconcile_state=%s", sql)
        self.assertIn("status='success'", sql)

    @patch("services.fullchain_archive_jobs.mark_flow_archived", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs._lock_reconcile_source", new=AsyncMock(return_value=False))
    @patch("services.fullchain_archive_jobs._insert_archive_from_snapshot", new_callable=AsyncMock)
    @patch("services.fullchain_archive_jobs.get_parser")
    async def test_deleted_item_rebuilds_missing_history_from_frozen_snapshot(
        self, get_parser, insert_snapshot, mark_archived,
    ):
        parser = type(
            "Parser",
            (),
            {
                "COLUMNS": ["姓名"],
                "normalize_source_row": staticmethod(
                    lambda values: {"姓名": str(values.get("姓名") or "")}
                ),
            },
        )()
        get_parser.return_value = parser
        conn = _TransactionalConnection(_ReconcileItemCursor())
        item = (12, 7, "全链条", "a" * 32, 4, "b" * 64, '{"姓名":"甲"}')
        with patch(
            "services.fullchain_archive_jobs._platform_rows_for_reconciliation",
            new=AsyncMock(return_value=(None, [])),
        ):
            result = await _reconcile_deleted_archive_item(conn, item)
        self.assertEqual(result, "reconciled_from_snapshot")
        insert_snapshot.assert_awaited_once()
        mark_archived.assert_awaited_once()

    @patch("services.fullchain_archive_jobs._lock_reconcile_source", new=AsyncMock(return_value=False))
    @patch("services.fullchain_archive_jobs.get_parser")
    async def test_deleted_item_never_overwrites_changed_archive_or_current_row(self, get_parser):
        parser = type(
            "Parser",
            (),
            {
                "COLUMNS": ["姓名"],
                "normalize_source_row": staticmethod(
                    lambda values: {"姓名": str(values.get("姓名") or "")}
                ),
            },
        )()
        get_parser.return_value = parser
        item = (12, 7, "全链条", "a" * 32, 4, "b" * 64, '{"姓名":"甲"}')
        cases = [
            ((None, [({"姓名": "乙"}, "online_removed")]), "archive_content_conflict"),
            (({"姓名": "乙"}, []), "current_row_changed_after_external_delete"),
        ]
        for platform_rows, code in cases:
            with self.subTest(code=code):
                conn = _TransactionalConnection(_ReconcileItemCursor())
                with patch(
                    "services.fullchain_archive_jobs._platform_rows_for_reconciliation",
                    new=AsyncMock(return_value=platform_rows),
                ):
                    with self.assertRaises(ArchiveStageError) as raised:
                        await _reconcile_deleted_archive_item(conn, item)
                self.assertEqual(raised.exception.code, code)
                self.assertEqual(conn.events, ["begin", "rollback"])

    async def test_reconcile_scan_only_selects_confirmed_deleted_items_and_never_uses_oauth(self):
        class ReconcileScanCursor(_Cursor):
            async def fetchall(self):
                return []

        cursor = ReconcileScanCursor()
        conn = _TransactionalConnection(cursor)
        pool = MagicMock()
        pool.acquire = AsyncMock(return_value=conn)
        with patch("services.fullchain_archive_jobs.db_manager.get_pool", return_value=pool), \
             patch("services.fullchain_archive_jobs._oauth_client", new=AsyncMock()) as oauth:
            self.assertEqual(await reconcile_deleted_archive_items(), 0)
        oauth.assert_not_awaited()
        pool.release.assert_called_once_with(conn)
        sql = "\n".join(item[0] for item in cursor.executions)
        self.assertIn("external_delete_state='deleted'", sql)
        self.assertIn("status IN ('queued','conflict','error')", sql)

    async def test_recovery_requeues_running_and_confirmed_partial_exports(self):
        class RecoveryCursor(_Cursor):
            def __init__(self):
                super().__init__()
                self.rowcount = 0
                self.queued = [(11,), (12,)]

            async def execute(self, sql, params=()):
                await super().execute(sql, params)
                normalized = " ".join(str(sql).split())
                if "WHERE status='running'" in normalized:
                    self.rowcount = 1
                elif "WHERE export_run.status='partial'" in normalized:
                    self.rowcount = 2
                else:
                    self.rowcount = 0

            async def fetchall(self):
                return self.queued

        cursor = RecoveryCursor()
        launches = []
        with patch("services.fullchain_archive_jobs.db_manager.get_pool", return_value=_Pool(_Connection(cursor))), \
             patch("services.fullchain_archive_jobs.launch_fullchain_archive_export", side_effect=launches.append):
            recovered = await recover_interrupted_fullchain_exports()
        self.assertEqual(recovered, 3)
        self.assertEqual(launches, [11, 12])
        sql = "\n".join(item[0] for item in cursor.executions)
        self.assertIn("external_delete_state='deleted'", sql)

    async def test_archive_permission_requires_allowed_position_and_all_scope(self):
        allowed = {
            "permission_scopes": {"police.dispatch.manage": "all"},
            "data_scope": "all",
            "permission_groups": [],
            "role": "user",
            "member": {"position": "所队领导"},
        }
        self.assertIs(await require_fullchain_archive(user=allowed), allowed)

        with self.assertRaises(HTTPException) as denied_scope:
            await require_fullchain_archive(user={
                **allowed,
                "permission_scopes": {"police.dispatch.manage": "own_department"},
            })
        self.assertEqual(denied_scope.exception.status_code, 403)

        with self.assertRaises(HTTPException) as denied_position:
            await require_fullchain_archive(user={
                **allowed,
                "member": {"position": "组员"},
            })
        self.assertEqual(denied_position.exception.status_code, 403)

    async def test_archive_permission_does_not_expand_dispatch_workbench_positions(self):
        user = {
            "permission_scopes": {POLICE_DISPATCH_MANAGE: "all"},
            "data_scope": "all",
            "permission_groups": [],
            "role": "user",
            "member": {"position": "所队领导"},
        }
        self.assertIs(await require_police_access(POLICE_DISPATCH_MANAGE)(user=user), user)


if __name__ == "__main__":
    unittest.main()
