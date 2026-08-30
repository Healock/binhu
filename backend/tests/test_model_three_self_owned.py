import io
import os
import zipfile
import unittest
from unittest.mock import AsyncMock, patch

from openpyxl import Workbook

os.environ.setdefault("MYSQL_PASSWORD", "test-password")
os.environ.setdefault("ENCRYPTION_KEY", "test-encryption-key")

from services.model_three_self_owned import (
    RULE_VERSION,
    ParsedSelfOwned,
    SelfOwnedImportError,
    _registry_table,
    _public_batch,
    apply_self_owned_import,
    parse_self_owned_zip,
    should_apply_self_owned_result,
)


def _xlsx_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["\ufeff社区名称", "居民证号", "住房类型", "居住处所"])
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_bytes_with_name(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "居民证号"])
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xlsx_bytes_with_person_details(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "居民证号", "个人联系电话"])
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _zip_bytes(files):
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)
    return output.getvalue()


class ModelThreeSelfOwnedTests(unittest.TestCase):
    def test_registry_tables_are_always_qualified_to_registry_database(self):
        self.assertEqual(
            _registry_table("registry_housing_people"),
            "`RegistryData`.`registry_housing_people`",
        )
        with self.assertRaisesRegex(ValueError, "unsupported registry table"):
            _registry_table("_online_source_rows")

    def test_parses_multiple_workbooks_deduplicates_and_counts_invalid_rows(self):
        content = _zip_bytes({
            "one.xlsx": _xlsx_bytes([
                ["测试社区", "11010519491231002X", "自购自住", "自购自住"],
                ["测试社区", "11010519491231002X", "自购自住", "自购自住"],
                ["测试社区", "bad", "自购自住", "自购自住"],
            ]),
            "two.xlsx": _xlsx_bytes([
                ["测试社区", "110105194912310011", "自购自住", "自购自住"],
            ]),
        })
        parsed = parse_self_owned_zip(content)
        self.assertEqual(parsed.workbook_count, 2)
        self.assertEqual(parsed.total_rows, 4)
        self.assertEqual(parsed.valid_rows, 2)
        self.assertEqual(parsed.invalid_rows, 1)
        self.assertEqual(parsed.duplicate_rows, 1)
        self.assertEqual(len(parsed.identities), 2)

    def test_requires_identity_column_and_valid_record(self):
        content = _zip_bytes({"bad.xlsx": _xlsx_bytes([["社区", "", "", ""]])})
        with self.assertRaisesRegex(SelfOwnedImportError, "没有有效身份证"):
            parse_self_owned_zip(content)

    def test_keeps_optional_names_for_personnel_archive(self):
        content = _zip_bytes({
            "named.xlsx": _xlsx_bytes_with_name([["张三", "11010519491231002X"]]),
        })
        parsed = parse_self_owned_zip(content)
        self.assertEqual(len(parsed.names), 1)
        self.assertEqual(parsed.names[0][1], "张三")

    def test_keeps_identity_and_personal_phone_for_personnel_archive(self):
        content = _zip_bytes({
            "details.xlsx": _xlsx_bytes_with_person_details([
                ["测试人员", "11010519491231002X", "138-0013-8000"],
            ]),
        })
        parsed = parse_self_owned_zip(content)
        self.assertEqual(len(parsed.people), 1)
        self.assertEqual(parsed.people[0].identity_number, "11010519491231002X")
        self.assertEqual(parsed.people[0].phone, "13800138000")
        self.assertEqual(len(parsed.people[0].phone_hmac), 64)

    def test_only_blank_or_unverifiable_result_can_be_filled(self):
        self.assertTrue(should_apply_self_owned_result("", False))
        self.assertTrue(should_apply_self_owned_result("无法核实", False))
        self.assertFalse(should_apply_self_owned_result("近期返吴", False))
        self.assertFalse(should_apply_self_owned_result("在吴", False))
        self.assertFalse(should_apply_self_owned_result("", True))
        self.assertFalse(should_apply_self_owned_result("", False, source_count=2))
        self.assertFalse(should_apply_self_owned_result("无法核实", False, conflict=True))

    def test_public_batch_does_not_expose_internal_creator_id(self):
        result = _public_batch((
            1, "名单.zip", "a" * 64, "self-owned-v1", "completed", 1,
            2, 2, 0, 0, 1, 1, 0, 99, None, None, "",
        ))
        self.assertNotIn("created_by", result)
        self.assertEqual(result["updated_tasks"], 1)


class _UpgradeCursor:
    def __init__(self):
        self.executed: list[tuple[str, object]] = []
        self.executemany_calls: list[tuple[str, object]] = []
        self.fetchone_value = None

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return False

    async def execute(self, sql, params=None):
        normalized = " ".join(str(sql).split())
        self.executed.append((normalized, params))
        if normalized.startswith("SELECT id,rule_version,status FROM _qmf_self_owned_batches"):
            self.fetchone_value = (17, "self-owned-v1", "completed")
        elif "SELECT id FROM `RegistryData`.`watch_categories`" in normalized:
            self.fetchone_value = (5,)
        else:
            self.fetchone_value = None

    async def executemany(self, sql, params):
        self.executemany_calls.append((" ".join(str(sql).split()), params))

    async def fetchone(self):
        return self.fetchone_value


class _UpgradeConnection:
    def __init__(self):
        self.cursor_instance = _UpgradeCursor()
        self.committed = False
        self.rolled_back = False

    async def begin(self):
        return None

    def cursor(self):
        return self.cursor_instance

    async def commit(self):
        self.committed = True

    async def rollback(self):
        self.rolled_back = True


class ModelThreeSelfOwnedUpgradeTests(unittest.IsolatedAsyncioTestCase):
    async def test_same_file_upgrades_v1_batch_in_place(self):
        connection = _UpgradeConnection()
        parsed = ParsedSelfOwned(
            file_sha256="a" * 64,
            total_rows=1,
            valid_rows=1,
            invalid_rows=0,
            duplicate_rows=0,
            identities=(("b" * 64, 1),),
            workbook_count=1,
        )
        people_stats = {
            "registry_people_created": 0,
            "registry_people_reused": 1,
            "tag_people_created": 0,
            "tag_people_reused": 1,
            "watch_phones_created": 0,
            "registry_phones_created": 1,
            "tag_assignments_created": 0,
        }
        match_stats = {"matched_tasks": 1, "updated_tasks": 1, "skipped_tasks": 0}
        with (
            patch(
                "services.model_three_self_owned._store_people_and_tags",
                AsyncMock(return_value=people_stats),
            ),
            patch(
                "services.model_three_self_owned.apply_self_owned_matches",
                AsyncMock(return_value=match_stats),
            ),
            patch(
                "services.model_three_self_owned.rebuild_projection",
                AsyncMock(),
            ),
        ):
            result = await apply_self_owned_import(
                connection,
                parsed=parsed,
                file_name="roster.zip",
                user_id=3,
            )

        sql_text = "\n".join(sql for sql, _params in connection.cursor_instance.executed)
        self.assertTrue(connection.committed)
        self.assertFalse(connection.rolled_back)
        self.assertTrue(result["upgraded"])
        self.assertEqual(result["batch_id"], 17)
        self.assertEqual(result["rule_version"], RULE_VERSION)
        self.assertIn("UPDATE _qmf_self_owned_batches SET file_name=%s,rule_version=%s", sql_text)
        self.assertNotIn("INSERT INTO _qmf_self_owned_batches", sql_text)


if __name__ == "__main__":
    unittest.main()
