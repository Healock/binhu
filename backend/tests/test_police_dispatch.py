from __future__ import annotations

import asyncio
import io
import os
from datetime import date, datetime
from unittest.mock import AsyncMock, MagicMock

os.environ.setdefault("MYSQL_PASSWORD", "test-password")
os.environ.setdefault("ENCRYPTION_KEY", "test-encryption-key")

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from starlette.requests import Request

from deps import require_super_admin
from services.permissions import (
    DEFAULT_PERMISSION_GROUPS,
    POLICE_ADDRESS_MANAGE,
    POLICE_DISPATCH_MANAGE,
)
from services.police_dispatch import (
    apply_clean_import_actions,
    apply_preprocessing_suggestions,
    build_feedback_workbook,
    build_publish_address,
    parse_dispatch_workbook,
    publish_business_key,
    reconcile_police_dispatch_publications,
    stable_json,
)
from routers.police_dispatch import (
    AddressCreate,
    TaskBusinessFieldsUpdate,
    TaskReview,
    TaskSearch,
    router,
    _batch_payloads,
    _clean_preview_summary,
    _clean_preview_token,
    _mark_overwrite_uncertain,
    _publish_values,
    _review_one,
    _search_tasks,
    _task_counts,
    _verify_clean_preview_token,
    delete_batch,
    delete_address,
    export_addresses,
    list_tasks,
    require_police_access,
    update_task_business_fields,
    _address_scope_community_ids,
    _assert_address_scope,
    _filter_address_rows,
    require_police_address_access,
    update_address,
)


def test_clean_preview_token_binds_file_and_metadata():
    token = _clean_preview_token("a" * 64, "clean.xlsx", "数据", 2)

    assert _verify_clean_preview_token(token, "a" * 64, "clean.xlsx", "数据", 2)
    assert not _verify_clean_preview_token(token, "b" * 64, "clean.xlsx", "数据", 2)
    assert not _verify_clean_preview_token(token, "a" * 64, "changed.xlsx", "数据", 2)


def test_clean_preview_summary_masks_sensitive_values_and_counts_actions():
    summary = _clean_preview_summary([
        {
            "source_row": 2, "person_name": "张三",
            "identity_number": "32050020000101001X", "phone": "18800000001",
            "community_name": "长板", "registration_status": "流口未登记",
            "auto_final_action": "dispatch", "auto_final_community_id": 1,
            "suggestion_reason": "可直接下发", "allocation_mode": "clean_import",
        },
        {
            "source_row": 3, "person_name": "李四",
            "identity_number": "320500200001010028", "phone": "18800000002",
            "community_name": "长板", "registration_status": "未知状态",
            "auto_final_action": "", "suggestion_reason": "需要人工确认",
            "allocation_mode": "conflict", "duplicate_group_key": "duplicate",
        },
    ])

    assert summary["counts"] == {
        "total": 2, "dispatch": 1, "no_registration": 0,
        "manual_review": 1, "invalid": 1, "duplicate": 1,
    }
    assert summary["community_distribution"] == [
        {"community_id": 1, "community_name": "长板", "count": 1},
    ]
    assert summary["rows"][0]["identity_number"] == "320500********001X"
    assert summary["rows"][0]["phone"] == "188****0001"
    assert "32050020000101001X" not in str(summary)
    assert "18800000001" not in str(summary)


def _xlsx(rows: list[list[object]], title: str = "数据") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


async def _streaming_body(response) -> bytes:
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def test_internal_business_group_gets_police_permissions():
    permissions = DEFAULT_PERMISSION_GROUPS["internal_business"]["permissions"]
    assert POLICE_DISPATCH_MANAGE in permissions
    assert POLICE_ADDRESS_MANAGE in permissions


def test_address_manager_group_is_scoped_to_member_communities():
    dependency = require_police_address_access()
    user = {
        "permissions": [POLICE_ADDRESS_MANAGE],
        "permission_scopes": {POLICE_ADDRESS_MANAGE: "own_department"},
        "data_scope": "own_department",
        "member": {"position": "组长"},
        "departments": [{"type": "community", "community_name": "冬梅"}],
    }
    assert asyncio.run(dependency(user=user)) == user


def test_address_scope_maps_formal_community_names_and_admin_sees_all():
    communities = [
        {"id": 6, "name": "冬梅", "enabled": True},
        {"id": 10, "name": "顾家荡", "enabled": True},
    ]
    member = {
        "member": {"position": "组员"},
        "permission_scopes": {POLICE_ADDRESS_MANAGE: "own_department"},
        "departments": [{"type": "community", "community_name": "冬梅社区"}],
    }
    assert _address_scope_community_ids(member, communities) == [6]
    assert _address_scope_community_ids({"member": {"position": "基础管控"}}, communities) is None
    with pytest.raises(HTTPException):
        _assert_address_scope(10, [6])


def test_address_filter_applies_keyword_and_enabled_without_broadening_scope():
    rows = [
        {"name": "甲小区", "detail_address": "一号路", "community_name": "冬梅", "aliases": ["甲"], "enabled": True},
        {"name": "乙小区", "detail_address": "二号路", "community_name": "顾家荡", "aliases": ["乙"], "enabled": False},
    ]
    assert [row["name"] for row in _filter_address_rows(rows, keyword="甲")] == ["甲小区"]
    assert [row["name"] for row in _filter_address_rows(rows, enabled=False)] == ["乙小区"]


def test_address_export_contains_expected_columns_and_scoped_rows(monkeypatch):
    rows = [{
        "id": 1,
        "name": "示例小区",
        "community_name": "冬梅",
        "address_type": "community",
        "detail_address": "示例路1号",
        "pattern": "示例模式",
        "aliases": ["示例别名一", "示例别名二"],
        "enabled": True,
    }]
    page_data = AsyncMock(return_value=(rows, [{"id": 6, "name": "冬梅"}], [6]))
    audit = AsyncMock()
    monkeypatch.setattr("routers.police_dispatch._address_page_data", page_data)
    monkeypatch.setattr("routers.police_dispatch.record_admin_audit", audit)
    conn = MagicMock()
    conn.cursor = MagicMock(return_value=_CursorContext(MagicMock()))
    request = Request({
        "type": "http", "method": "POST", "path": "/", "headers": [],
        "client": ("127.0.0.1", 1),
    })

    response = asyncio.run(export_addresses(
        data=__import__("routers.police_dispatch", fromlist=["AddressSearch"]).AddressSearch(),
        request=request,
        user={"id": 5},
        conn=conn,
    ))
    body = asyncio.run(_streaming_body(response))
    workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    sheet = workbook["小区管理"]
    assert list(next(sheet.iter_rows(values_only=True))) == [
        "名称", "正式社区", "类型", "详细地址", "模式", "别名", "状态",
    ]
    assert list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))) == [
        "示例小区", "冬梅", "居民小区", "示例路1号", "示例模式",
        "示例别名一，示例别名二", "启用",
    ]
    workbook.close()
    assert audit.await_args.kwargs["detail"] == {"row_count": 1, "file_format": "XLSX"}
    assert audit.await_args.kwargs["target_name"] == "community-scope"


def test_address_delete_removes_sources_then_entry_and_commits(monkeypatch):
    cursor = MagicMock()
    cursor.execute = AsyncMock()
    cursor.fetchone = AsyncMock(return_value=(6,))
    type(cursor).rowcount = __import__("unittest.mock", fromlist=["PropertyMock"]).PropertyMock(
        side_effect=[2, 1],
    )
    conn = MagicMock()
    conn.begin = AsyncMock()
    conn.commit = AsyncMock()
    conn.rollback = AsyncMock()
    conn.cursor = MagicMock(return_value=_CursorContext(cursor))
    audit = AsyncMock()
    monkeypatch.setattr(
        "routers.police_dispatch._communities",
        AsyncMock(return_value=[{"id": 6, "name": "冬梅", "enabled": True}]),
    )
    monkeypatch.setattr("routers.police_dispatch.record_admin_audit", audit)
    request = Request({
        "type": "http", "method": "DELETE", "path": "/", "headers": [],
        "client": ("127.0.0.1", 1),
    })
    user = {
        "id": 5,
        "member": {"position": "组长"},
        "permission_scopes": {POLICE_ADDRESS_MANAGE: "own_department"},
        "departments": [{"type": "community", "community_name": "冬梅"}],
    }

    result = asyncio.run(delete_address(1, request, user=user, conn=conn))

    assert result == {"message": "地址记录已删除"}
    statements = [call.args[0] for call in cursor.execute.await_args_list]
    assert "FOR UPDATE" in statements[0]
    assert "DELETE FROM _police_address_sources" in statements[1]
    assert "DELETE FROM _police_address_entries" in statements[2]
    conn.commit.assert_awaited_once()
    conn.rollback.assert_not_awaited()
    assert audit.await_args.kwargs["detail"] == {"source_links_removed": 2}


def test_address_update_and_delete_reject_cross_community(monkeypatch):
    communities = [
        {"id": 6, "name": "冬梅", "enabled": True},
        {"id": 10, "name": "顾家荡", "enabled": True},
    ]
    monkeypatch.setattr("routers.police_dispatch._communities", AsyncMock(return_value=communities))
    monkeypatch.setattr("routers.police_dispatch._assert_community", AsyncMock())
    monkeypatch.setattr("routers.police_dispatch.record_admin_audit", AsyncMock())
    user = {
        "id": 5,
        "member": {"position": "组员"},
        "permission_scopes": {POLICE_ADDRESS_MANAGE: "own_department"},
        "departments": [{"type": "community", "community_name": "冬梅"}],
    }
    request = Request({
        "type": "http", "method": "PUT", "path": "/", "headers": [],
        "client": ("127.0.0.1", 1),
    })
    conn, _ = _delete_batch_conn((10,))
    payload = AddressCreate(name="越权小区", community_id=10)

    with pytest.raises(HTTPException) as update_error:
        asyncio.run(update_address(9, payload, request, user=user, conn=conn))
    assert update_error.value.status_code == 403

    delete_conn, _ = _delete_batch_conn((10,))
    with pytest.raises(HTTPException) as delete_error:
        asyncio.run(delete_address(9, request, user=user, conn=delete_conn))
    assert delete_error.value.status_code == 403
    delete_conn.rollback.assert_awaited_once()


def test_police_access_requires_all_scope_and_hard_position_limit():
    dependency = require_police_access(POLICE_DISPATCH_MANAGE)
    base = {
        "permission_scopes": {POLICE_DISPATCH_MANAGE: "all"},
        "permission_groups": [{"code": "admin"}],
    }
    assert asyncio.run(dependency(user={
        **base,
        "member": {"position": "基础管控"},
    }))["member"]["position"] == "基础管控"

    with pytest.raises(HTTPException) as position_error:
        asyncio.run(dependency(user={
            **base,
            "member": {"position": "社区民警"},
        }))
    assert position_error.value.status_code == 403

    with pytest.raises(HTTPException) as scope_error:
        asyncio.run(dependency(user={
            **base,
            "permission_scopes": {POLICE_DISPATCH_MANAGE: "own_department"},
            "member": {"position": "基础管控"},
        }))
    assert scope_error.value.status_code == 403
    assert "全所数据范围" in str(scope_error.value.detail)

    system_admin = asyncio.run(dependency(user={**base, "member": None}))
    assert system_admin["member"] is None


def test_parse_dispatch_xlsx_finds_title_row_and_preserves_text_identifiers():
    content = _xlsx([
        ["全链条专项核查指令"],
        ["导出人：测试单位"],
        ["来源", "姓名", "身份证号", "手机号", "地址", "创建时间", "移交备注"],
        ["管家码", "甲", "32050020000101001X", "18800000001", "长板花园1幢", "2026-08-03", ""],
        ["来源", "姓名", "身份证号", "手机号", "地址", "创建时间", "移交备注"],
        ["平安码", "乙", "320500200001010028", "18800000002", "地址模糊", "2026-08-03", "完整移交原文"],
    ])

    sheet_name, rows = parse_dispatch_workbook(content, "系统导出.xlsx")

    assert sheet_name == "数据"
    assert len(rows) == 2
    assert rows[0].identity_number == "32050020000101001X"
    assert rows[0].phone == "18800000001"
    assert rows[1].transfer_note == "完整移交原文"


def test_clean_import_requires_clean_columns_and_maps_registration_status():
    content = _xlsx([
        ["社区", "来源", "姓名", "身份证号", "手机号", "地址", "登记情况", "创建时间", "移交备注"],
        ["长板", "平安码", "甲", "32050020000101001X", "18800000001", "长板花园", "流口未登记", "2026-08-11 09:00:00", ""],
        ["长板", "平安码", "乙", "320500200001010028", "18800000002", "长板花园", "流口已注销", "2026-08-11 09:00:01", ""],
    ])
    _, parsed = parse_dispatch_workbook(content, "clean.xlsx", require_clean_fields=True)
    rows = [{
        "source_row": item.source_row, "source_name": item.source_name,
        "community_name": item.community_name, "person_name": item.person_name,
        "identity_number": item.identity_number, "phone": item.phone,
        "original_address": item.original_address,
        "registration_status": item.registration_status,
    } for item in parsed]
    apply_clean_import_actions(rows, [{"id": 1, "name": "长板", "aliases": [], "enabled": True}])
    assert rows[0]["auto_final_action"] == "dispatch"
    assert rows[0]["auto_final_community_id"] == 1
    assert rows[1]["auto_final_action"] == "no_registration"


def test_clean_import_keeps_invalid_identity_for_manual_review():
    rows = [{
        "source_row": 2,
        "source_name": "平安码",
        "community_name": "长板",
        "person_name": "甲",
        "identity_number": "32050020000101",
        "phone": "18800000001",
        "original_address": "长板花园",
        "registration_status": "流口未登记",
    }]

    apply_clean_import_actions(
        rows,
        [{"id": 1, "name": "长板", "aliases": [], "enabled": True}],
    )

    assert rows[0]["auto_final_action"] == ""
    assert rows[0]["allocation_mode"] == "conflict"
    assert rows[0]["suggestion_reason"] == "身份证号格式异常，需要人工确认"


def test_parse_xls_uses_same_header_logic(monkeypatch):
    monkeypatch.setattr(
        "services.police_dispatch._read_xls",
        lambda _content: [("旧表", [
            ["来源", "姓名", "身份证号", "手机号", "地址"],
            ["管家码", "甲", "32050020000101001X", "18800000001", "长板花园"],
        ])],
    )

    sheet_name, rows = parse_dispatch_workbook(b"legacy", "系统导出.xls")

    assert sheet_name == "旧表"
    assert len(rows) == 1
    assert rows[0].identity_number.endswith("X")


def test_one_time_address_mapping_import_route_is_removed():
    assert not any(
        route.path == "/api/police-dispatch/addresses/import"
        for route in router.routes
    )


def test_batch_delete_route_requires_super_admin():
    route = next(
        route
        for route in router.routes
        if route.path == "/api/police-dispatch/batches/{batch_id}"
        and "DELETE" in route.methods
    )
    assert any(
        dependency.call is require_super_admin
        for dependency in route.dependant.dependencies
    )


def test_suggestions_require_review_and_balance_only_unmatched_pool():
    communities = [
        {"id": 1, "name": "长板", "aliases": ["长板社区"], "enabled": True},
        {"id": 2, "name": "龙河", "aliases": [], "enabled": True},
        {"id": 3, "name": "祥泰", "aliases": [], "enabled": True},
        {"id": 4, "name": "已停用", "aliases": [], "enabled": False},
    ]
    entries = [
        {
            "id": 1,
            "name": "芦风华庭",
            "detail_address": "芦荡路1288号",
            "aliases": [],
            "community_id": 1,
            "community_name": "长板",
            "enabled": True,
            "address_type": "community",
        },
        {
            "id": 2,
            "name": "云玺商务广场",
            "detail_address": "",
            "aliases": [],
            "community_id": 2,
            "community_name": "龙河",
            "enabled": True,
            "address_type": "apartment",
        },
        {
            "id": 3,
            "name": "停用小区",
            "detail_address": "停用路1号",
            "aliases": [],
            "community_id": 4,
            "community_name": "已停用",
            "enabled": True,
            "address_type": "community",
        },
    ]
    rows = [
        {"source_row": 4, "person_name": "甲", "identity_number": "1", "phone": "11", "original_address": "芦风华庭1幢", "transfer_note": "", "source_name": "A", "created_time": ""},
        {"source_row": 5, "person_name": "乙", "identity_number": "2", "phone": "22", "original_address": "某某酒店", "transfer_note": "", "source_name": "A", "created_time": ""},
        {"source_row": 6, "person_name": "丙", "identity_number": "3", "phone": "33", "original_address": "云玺商务广场", "transfer_note": "", "source_name": "A", "created_time": ""},
        {"source_row": 7, "person_name": "丁", "identity_number": "4", "phone": "44", "original_address": "停用小区", "transfer_note": "", "source_name": "A", "created_time": ""},
        *[
            {"source_row": 10 + i, "person_name": f"模糊{i}", "identity_number": f"9{i}", "phone": f"8{i}", "original_address": "59幢", "transfer_note": "", "source_name": "A", "created_time": ""}
            for i in range(7)
        ],
    ]

    apply_preprocessing_suggestions(rows, communities, entries)

    assert rows[0]["suggested_community_id"] == 1
    assert rows[1]["suggested_action"] == "no_registration"
    # 公寓即使模式中含“民宿较多”，地址本身不含酒店/民宿时仍只是社区映射。
    assert rows[2]["suggested_action"] == "dispatch"
    assert rows[2]["suggested_community_id"] == 2
    assert rows[3]["allocation_mode"] == "balanced"
    assert rows[3]["suggested_community_id"] in {1, 2, 3}
    balanced = [row for row in rows if row["allocation_mode"] == "balanced"]
    counts = {community_id: sum(row["suggested_community_id"] == community_id for row in balanced) for community_id in (1, 2, 3)}
    assert max(counts.values()) - min(counts.values()) <= 1
    assert all(not row.get("final_action") for row in rows)


def test_missing_phone_requires_analysis_and_is_not_assigned_to_community():
    rows = [{
        "source_row": 4,
        "person_name": "甲",
        "identity_number": "1",
        "phone": "",
        "original_address": "长板社区1号",
        "transfer_note": "",
        "source_name": "A",
        "created_time": "",
    }]

    apply_preprocessing_suggestions(
        rows,
        [{"id": 1, "name": "长板", "aliases": [], "enabled": True}],
        [],
    )

    assert rows[0]["suggested_action"] == "manual"
    assert rows[0]["suggested_community_id"] is None
    assert rows[0]["allocation_mode"] == "missing_phone"
    assert "缺少手机号" in rows[0]["suggestion_reason"]


def test_duplicate_group_marks_exact_and_conflicting_rows():
    rows = [
        {"source_row": 1, "person_name": "甲", "identity_number": "A1", "phone": "1", "original_address": "地址", "transfer_note": "", "source_name": "A", "created_time": ""},
        {"source_row": 2, "person_name": "甲", "identity_number": "A1", "phone": "1", "original_address": "地址", "transfer_note": "", "source_name": "A", "created_time": ""},
        {"source_row": 3, "person_name": "乙", "identity_number": "B1", "phone": "2", "original_address": "地址一", "transfer_note": "", "source_name": "A", "created_time": ""},
        {"source_row": 4, "person_name": "乙", "identity_number": "B1", "phone": "3", "original_address": "地址二", "transfer_note": "", "source_name": "A", "created_time": ""},
    ]

    apply_preprocessing_suggestions(rows, [{"id": 1, "name": "长板", "aliases": [], "enabled": True}], [])

    assert rows[0]["duplicate_kind"] == rows[1]["duplicate_kind"] == "exact"
    assert rows[2]["duplicate_kind"] == rows[3]["duplicate_kind"] == "conflict"


def test_publish_address_rebuilds_from_original_without_touching_analysis():
    assert build_publish_address("原地址", "完整原文") == "原地址；移交反馈：完整原文"
    assert build_publish_address("原地址；移交反馈：旧值", "") == "原地址；移交反馈：旧值"


def test_publish_values_freeze_text_dates_and_leave_analysis_blank():
    values = _publish_values({
        "source_name": "管家码",
        "person_name": "甲",
        "identity_number": "32050020000101001X",
        "phone": "18800000001",
        "original_address": "原地址",
        "transfer_note": "完整原文",
        "created_time": "2026-08-03 09:00:00",
    }, "长板", date(2026, 12, 30))

    assert values["下发日期"] == "12-30"
    assert values["截止日期"] == "01-02"
    assert values["地址"] == "原地址；移交反馈：完整原文"
    assert values["研判"] == ""
    assert values["身份证号"] == "32050020000101001X"
    assert values["电话号码"] == "18800000001"


def test_task_counts_keeps_review_and_publish_states_separate():
    counts = _task_counts([(10, 3, 7, 2, 1, 6, 4, 2, 1, 5, 1)])

    assert counts == {
        "total": 10,
        "pending_review": 3,
        "reviewed": 7,
        "no_registration": 2,
        "transfer": 1,
        "dispatch": 6,
        "balanced": 4,
        "duplicate": 2,
        "abnormal": 1,
        "pending_publish": 5,
        "published": 1,
        "retryable": 0,
        "needs_reconciliation": 0,
        "conflict": 0,
        "cache_pending": 0,
    }


def test_batch_payloads_expands_mysql_placeholders_before_execution():
    cursor = MagicMock()
    cursor.execute = AsyncMock()
    cursor.fetchall = AsyncMock(side_effect=[
        [(7, "批次.xlsx", "数据", "reviewing", 1, "{}", None, "",
          datetime(2026, 8, 5, 1, 0), datetime(2026, 8, 5, 1, 0),
          "审核员", "reviewer")],
        [(7, 1, 1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0)],
        [],
    ])

    result = asyncio.run(_batch_payloads(cursor, [7]))

    assert result[0]["id"] == 7
    first_sql = cursor.execute.await_args_list[0].args[0]
    assert "IN (%s)" in first_sql
    assert "{placeholders}" not in first_sql


def test_sensitive_task_search_uses_post_body_value_as_sql_parameter():
    cursor = MagicMock()
    cursor.execute = AsyncMock()
    cursor.fetchone = AsyncMock(return_value=(0,))
    cursor.fetchall = AsyncMock(return_value=[])

    result = asyncio.run(_search_tasks(cursor, TaskSearch(
        batch_id=7,
        keyword="32050020000101001X",
    )))

    assert result["total"] == 0
    count_params = cursor.execute.await_args_list[0].args[1]
    assert count_params == [7, "%32050020000101001X%"]


def test_get_task_list_rejects_sensitive_keyword_query_parameter():
    request = Request({
        "type": "http",
        "method": "GET",
        "path": "/api/police-dispatch/tasks",
        "query_string": b"batch_id=7&keyword=18800000001",
        "headers": [],
    })

    with pytest.raises(HTTPException) as error:
        asyncio.run(list_tasks(
            request=request,
            batch_id=7,
            status="all",
            category="all",
            page=1,
            page_size=20,
            user={},
            conn=None,
        ))

    assert error.value.status_code == 400
    assert "POST" in str(error.value.detail)


class _CursorContext:
    def __init__(self, cursor):
        self.cursor = cursor

    async def __aenter__(self):
        return self.cursor

    async def __aexit__(self, *_args):
        return False


def _delete_batch_conn(*rows: tuple):
    cursor = MagicMock()
    cursor.rowcount = 1
    cursor.execute = AsyncMock()
    cursor.fetchone = AsyncMock(side_effect=rows)
    conn = MagicMock()
    conn.begin = AsyncMock()
    conn.commit = AsyncMock()
    conn.rollback = AsyncMock()
    conn.cursor = MagicMock(return_value=_CursorContext(cursor))
    return conn, cursor


def test_super_admin_can_delete_never_published_batch(monkeypatch):
    conn, cursor = _delete_batch_conn(
        (7, None, None),
        (528, 0, 0),
    )
    audit = AsyncMock()
    monkeypatch.setattr("routers.police_dispatch.record_admin_audit", audit)
    request = Request({
        "type": "http", "method": "DELETE", "path": "/", "headers": [],
        "client": ("127.0.0.1", 1),
    })

    result = asyncio.run(delete_batch(
        batch_id=7,
        request=request,
        user={"id": 1, "username": "root", "role": "super_admin"},
        conn=conn,
    ))

    assert result == {"message": "批次已删除", "deleted_task_count": 528}
    conn.begin.assert_awaited_once()
    conn.commit.assert_awaited_once()
    conn.rollback.assert_not_awaited()
    sql = "\n".join(call.args[0] for call in cursor.execute.await_args_list)
    assert "DELETE FROM _police_dispatch_tasks" in sql
    assert "DELETE FROM _police_dispatch_batches" in sql
    assert audit.await_args.args[1] == "police_dispatch.delete"
    assert audit.await_args.kwargs["detail"] == {"row_count": 528}


def test_started_or_linked_batch_cannot_be_deleted(monkeypatch):
    monkeypatch.setattr(
        "routers.police_dispatch.record_admin_audit",
        AsyncMock(),
    )
    request = Request({
        "type": "http", "method": "DELETE", "path": "/", "headers": [],
        "client": ("127.0.0.1", 1),
    })

    published_conn, _ = _delete_batch_conn(
        (7, date(2026, 8, 6), datetime(2026, 8, 6, 1, 0)),
    )
    with pytest.raises(HTTPException) as published_error:
        asyncio.run(delete_batch(
            batch_id=7,
            request=request,
            user={"id": 1, "role": "super_admin"},
            conn=published_conn,
        ))
    assert published_error.value.status_code == 409
    published_conn.rollback.assert_awaited_once()

    linked_conn, _ = _delete_batch_conn(
        (8, None, None),
        (3, 1, 1),
    )
    with pytest.raises(HTTPException) as linked_error:
        asyncio.run(delete_batch(
            batch_id=8,
            request=request,
            user={"id": 1, "role": "super_admin"},
            conn=linked_conn,
        ))
    assert linked_error.value.status_code == 409
    linked_conn.rollback.assert_awaited_once()


def test_business_field_update_audit_contains_only_names_and_digest(monkeypatch):
    raw_values = {
        "姓名": "旧姓名",
        "身份证号": "32050020000101001X",
        "手机号": "18800000001",
        "地址": "旧地址",
    }
    cursor = MagicMock()
    cursor.rowcount = 1
    cursor.execute = AsyncMock()
    cursor.fetchone = AsyncMock(return_value=(7, 3, "not_required", stable_json(raw_values)))
    conn = MagicMock()
    conn.begin = AsyncMock()
    conn.commit = AsyncMock()
    conn.rollback = AsyncMock()
    conn.cursor = MagicMock(return_value=_CursorContext(cursor))
    audit = AsyncMock()
    monkeypatch.setattr(
        "routers.police_dispatch._recalculate_batch_tasks",
        AsyncMock(return_value={11, 12}),
    )
    monkeypatch.setattr(
        "routers.police_dispatch._refresh_batch_status",
        AsyncMock(return_value={}),
    )
    monkeypatch.setattr("routers.police_dispatch.record_admin_audit", audit)
    request = Request({
        "type": "http", "method": "PATCH", "path": "/", "headers": [],
        "client": ("127.0.0.1", 1),
    })

    result = asyncio.run(update_task_business_fields(
        task_id=11,
        data=TaskBusinessFieldsUpdate(
            expected_version=3,
            fields={"身份证号": "320500200001010028", "地址": "新地址"},
        ),
        request=request,
        user={"id": 5, "username": "reviewer"},
        conn=conn,
    ))

    assert result["affected_count"] == 2
    detail = audit.await_args.kwargs["detail"]
    assert detail["changed_fields"] == ["地址", "身份证号"]
    serialized = stable_json(detail)
    assert "320500200001010028" not in serialized
    assert "新地址" not in serialized


def test_uncertain_conflict_overwrite_is_locked_until_normal_sync(monkeypatch):
    cursor = MagicMock()
    cursor.rowcount = 1
    cursor.execute = AsyncMock()
    refresh = AsyncMock(return_value={})
    monkeypatch.setattr("routers.police_dispatch._refresh_batch_status", refresh)
    requested = {
        "下发日期": "08-05",
        "身份证号": "A1",
        "电话号码": "1",
        "社区": "长板",
    }

    marked = asyncio.run(_mark_overwrite_uncertain(
        cursor,
        task_id=11,
        batch_id=7,
        spreadsheet={"id": 5, "data_sheet_id": "sheet"},
        physical_row=20,
        requested=requested,
        error="请求结果不确定",
    ))

    assert marked is True
    sql = "\n".join(call.args[0] for call in cursor.execute.await_args_list)
    assert "publish_status='needs_reconciliation'" in sql
    assert "overwrite_uncertain" not in sql  # 错误码作为参数，避免拼接进 SQL。
    save_params = cursor.execute.await_args_list[1].args[1]
    assert "needs_reconciliation" in save_params
    assert "overwrite_uncertain" in save_params
    refresh.assert_awaited_once_with(cursor, 7)


class _ReviewCursor:
    def __init__(self, row: tuple):
        self.row = row
        self.rowcount = 0

    async def execute(self, _sql: str, _params=()):
        return None

    async def fetchone(self):
        return self.row


def test_review_returns_409_when_version_changed():
    cursor = _ReviewCursor((7, 2, "not_required", ""))

    with pytest.raises(HTTPException) as error:
        asyncio.run(_review_one(
            cursor,
            11,
            TaskReview(expected_version=1, final_action="transfer"),
            {"id": 3, "username": "reviewer"},
        ))

    assert error.value.status_code == 409


def test_unique_task_cannot_be_excluded_as_duplicate():
    cursor = _ReviewCursor((7, 1, "not_required", ""))

    with pytest.raises(HTTPException) as error:
        asyncio.run(_review_one(
            cursor,
            11,
            TaskReview(expected_version=1, final_action="duplicate_exclude"),
            {"id": 3, "username": "reviewer"},
        ))

    assert error.value.status_code == 400


def test_missing_phone_cannot_be_reviewed_for_dispatch():
    cursor = _ReviewCursor((
        7, 1, "not_required", "", "甲", "32050020000101001X", "", "长板1号",
    ))

    with pytest.raises(HTTPException) as error:
        asyncio.run(_review_one(
            cursor,
            11,
            TaskReview(
                expected_version=1,
                final_action="dispatch",
                final_community_id=1,
            ),
            {"id": 3, "username": "reviewer"},
        ))

    assert error.value.status_code == 400
    assert "手机号" in str(error.value.detail)


def test_feedback_workbook_has_three_formatted_sheets_and_text_identifiers():
    content = build_feedback_workbook(
        {"id": 7, "file_name": "原文件.xlsx", "reviewed_count": 1, "total_count": 2},
        [{
            "source_row": 4, "source_name": "管家码", "person_name": "甲",
            "identity_number": "32050020000101001X", "phone": "18800000001",
            "original_address": "原地址", "final_action": "transfer",
            "review_note": "移交", "suggestion_reason": "人工确认", "reviewer_name": "审核员",
            "reviewed_at_text": "2026-08-04 10:00:00",
        }, {
            "source_row": 5, "source_name": "平安码", "person_name": "乙",
            "identity_number": "320500200001010028", "phone": "18800000002",
            "original_address": "原地址二", "final_action": "transfer",
            "review_note": "移交", "suggestion_reason": "人工确认", "reviewer_name": "审核员",
            "reviewed_at_text": "2026-08-04 10:05:00",
        }],
        datetime(2026, 8, 4, 11, 0, 0),
    )

    workbook = load_workbook(io.BytesIO(content), data_only=False)
    try:
        assert workbook.sheetnames == ["汇总", "无需登记", "移交"]
        assert workbook["汇总"]["B6"].value == "非最终版本"
        assert workbook["移交"]["E2"].value == "32050020000101001X"
        assert workbook["移交"]["E3"].value == "320500200001010028"
        assert workbook["移交"]["E3"].data_type == "s"
        assert workbook["移交"]["E3"].number_format == "@"
        assert workbook["移交"]["E3"].quotePrefix is True
        assert workbook["移交"].auto_filter.ref == "A1:L3"
    finally:
        workbook.close()


def test_feedback_workbook_forces_formula_like_external_text_to_plain_string():
    content = build_feedback_workbook(
        {"id": 8, "file_name": "原文件.xlsx", "reviewed_count": 1, "total_count": 1},
        [{
            "source_row": 4, "source_name": "=HYPERLINK(\"bad\")",
            "person_name": "+测试", "identity_number": "-320500",
            "phone": "@18800000000", "original_address": "=1+1",
            "final_action": "transfer", "review_note": "=CMD()",
            "suggestion_reason": "人工确认", "reviewer_name": "审核员",
            "reviewed_at_text": "2026-08-05 10:00:00",
        }],
        datetime(2026, 8, 5, 11, 0, 0),
    )
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    try:
        sheet = workbook["移交"]
        for cell in sheet[2]:
            if isinstance(cell.value, str):
                assert cell.data_type == "s"
                assert cell.number_format == "@"
                assert cell.quotePrefix is True
    finally:
        workbook.close()


def test_normal_sync_reconciliation_classifies_exact_conflict_and_absent_rows():
    requested_exact = {
        "下发日期": "08-05", "身份证号": "A1", "电话号码": "1",
        "姓名": "甲", "社区": "长板",
    }
    requested_conflict = {
        "下发日期": "08-05", "身份证号": "B1", "电话号码": "2",
        "姓名": "乙", "社区": "长板",
    }
    source_exact = dict(requested_exact)
    source_conflict = {**requested_conflict, "社区": "龙河"}
    keys = [
        publish_business_key("A1", "1", "08-05"),
        publish_business_key("B1", "2", "08-05"),
        publish_business_key("C1", "3", "08-05"),
    ]
    cursor = MagicMock()
    cursor.execute = AsyncMock()
    cursor.fetchall = AsyncMock(side_effect=[
        [
            (11, 20, "a" * 64, stable_json(source_exact)),
            (12, 21, "b" * 64, stable_json(source_conflict)),
        ],
        [
            (101, keys[0], stable_json(requested_exact), 7),
            (102, keys[1], stable_json(requested_conflict), 7),
            (103, keys[2], stable_json({
                "下发日期": "08-05", "身份证号": "C1",
                "电话号码": "3", "姓名": "丙", "社区": "长板",
            }), 7),
        ],
    ])
    cursor.fetchone = AsyncMock(return_value=(3, 0, 2, 1))

    result = asyncio.run(reconcile_police_dispatch_publications(cursor, 5))

    assert result == {"success": 1, "conflict": 1, "retryable": 1}
    sql = "\n".join(call.args[0] for call in cursor.execute.await_args_list)
    assert "publish_status='success'" in sql
    assert "publish_status='conflict'" in sql
    assert "publish_status='retryable'" in sql
    assert "完整同步确认目标不存在" in sql
