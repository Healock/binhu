from datetime import date, datetime
from io import BytesIO
import json
import unittest

from openpyxl import Workbook

from deps import require_admin
from routers.grid_members import CommunityAliasesUpdate
from routers.visits import router as visits_router
from services.privacy import mask_identity_number
from services.visit_import import (
    VISIT_HEADERS,
    VisitWorkbookError,
    decide_existing_action,
    decide_member_identity,
    get_visit_coverage,
    import_parsed_workbook,
    normalize_community,
    parse_visit_workbook,
)


VALID_IDENTITY = "00000000000000000X"


def visit_row(
    *,
    community="长板社区",
    address="长板路 1 号",
    operator="张三",
    identity=VALID_IDENTITY,
    visit_at=datetime(2026, 7, 24, 10, 0),
    entry_method="扫码",
    room_count=1,
    added=0,
    changed=1,
    cancelled=0,
):
    return [
        "滨湖派出所",
        community,
        entry_method,
        address,
        operator,
        identity,
        visit_at,
        room_count,
        added,
        changed,
        cancelled,
    ]


def workbook_bytes(rows, *, headers=VISIT_HEADERS, sheet_name="走访明细"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class VisitWorkbookParsingTests(unittest.TestCase):
    def test_parses_normalizes_and_preserves_business_date(self):
        parsed = parse_visit_workbook(
            workbook_bytes([visit_row()]),
            "Asia/Shanghai",
        )

        self.assertEqual(parsed.sheet_name, "走访明细")
        self.assertEqual(parsed.total_rows, 1)
        self.assertEqual(parsed.valid_rows, 1)
        self.assertEqual(len(parsed.rows), 1)
        row = parsed.rows[0]
        self.assertEqual(row.community, "长板")
        self.assertEqual(row.visit_date, date(2026, 7, 24))
        self.assertEqual(row.visit_at, datetime(2026, 7, 24, 2, 0))
        self.assertEqual(row.changed_count, 1)
        self.assertTrue(row.operator_account_valid)

    def test_same_day_same_address_keeps_latest_time(self):
        parsed = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(visit_at=datetime(2026, 7, 24, 9, 0), changed=1),
                    visit_row(visit_at=datetime(2026, 7, 24, 11, 0), changed=3),
                ]
            ),
            "Asia/Shanghai",
        )

        self.assertEqual(len(parsed.rows), 1)
        self.assertEqual(parsed.rows[0].changed_count, 3)
        self.assertEqual(parsed.ignored_rows, 1)
        self.assertTrue(
            any(
                issue.code == "same_day_address_operator_replaced"
                for issue in parsed.issues
            )
        )

    def test_same_day_same_address_keeps_different_operators(self):
        parsed = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(operator="张三"),
                    visit_row(operator="李四"),
                ]
            ),
            "Asia/Shanghai",
        )

        self.assertEqual(len(parsed.rows), 2)
        self.assertEqual(parsed.ignored_rows, 0)
        self.assertNotEqual(parsed.rows[0].row_key, parsed.rows[1].row_key)

    def test_same_address_on_different_days_is_kept(self):
        parsed = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(visit_at=datetime(2026, 7, 24, 9, 0)),
                    visit_row(visit_at=datetime(2026, 7, 25, 9, 0)),
                ]
            ),
            "Asia/Shanghai",
        )

        self.assertEqual(len(parsed.rows), 2)
        self.assertNotEqual(parsed.rows[0].row_key, parsed.rows[1].row_key)

    def test_invalid_row_is_skipped_and_identity_is_masked(self):
        content = workbook_bytes(
            [
                visit_row(),
                visit_row(
                    address="错误地址",
                    identity="000000000000000019",
                    changed=-1,
                ),
            ]
        )
        parsed = parse_visit_workbook(content, "Asia/Shanghai")

        self.assertEqual(len(parsed.rows), 1)
        error = next(issue for issue in parsed.issues if issue.severity == "error")
        serialized = json.dumps(error.as_dict(), ensure_ascii=False)
        self.assertIn("变更必须是非负整数", serialized)
        self.assertNotIn("000000000000000019", serialized)
        self.assertIn("000000", error.row_preview["操作人账号"])

    def test_missing_headers_rejects_workbook(self):
        with self.assertRaisesRegex(VisitWorkbookError, "完整 11 列表头"):
            parse_visit_workbook(
                workbook_bytes([["值"]], headers=["地址"]),
                "Asia/Shanghai",
            )

    def test_numeric_identity_is_not_used_to_fill_member(self):
        parsed = parse_visit_workbook(
            workbook_bytes([visit_row(identity=123456789012345678)]),
            "Asia/Shanghai",
        )

        self.assertFalse(parsed.rows[0].operator_account_valid)
        self.assertTrue(
            any(issue.code == "identity_not_usable" for issue in parsed.issues)
        )

    def test_overlapping_date_ranges_are_unchanged_then_inserted(self):
        old_rows = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(
                        address=f"地址-{day}",
                        visit_at=datetime(2026, 7, day, 10, 0),
                    )
                    for day in range(24, 28)
                ]
            ),
            "Asia/Shanghai",
        ).rows
        incoming_rows = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(
                        address=f"地址-{day}",
                        visit_at=datetime(2026, 7, day, 10, 0),
                    )
                    for day in range(24, 30)
                ]
            ),
            "Asia/Shanghai",
        ).rows
        existing = {row.row_key: row.business_values() for row in old_rows}
        actions = [
            decide_existing_action(existing.get(row.row_key), row)
            for row in incoming_rows
        ]

        self.assertEqual(actions.count("unchanged"), 4)
        self.assertEqual(actions.count("insert"), 2)
        self.assertNotIn("update", actions)

    def test_existing_later_record_is_not_overwritten(self):
        incoming = parse_visit_workbook(
            workbook_bytes(
                [visit_row(visit_at=datetime(2026, 7, 24, 9, 0))]
            ),
            "Asia/Shanghai",
        ).rows[0]
        existing_values = list(incoming.business_values())
        existing_values[9] = datetime(2026, 7, 24, 3, 0)

        self.assertEqual(
            decide_existing_action(tuple(existing_values), incoming),
            "ignored",
        )

    def test_same_time_changed_content_updates_existing_record(self):
        original = parse_visit_workbook(
            workbook_bytes([visit_row(changed=1)]),
            "Asia/Shanghai",
        ).rows[0]
        corrected = parse_visit_workbook(
            workbook_bytes([visit_row(changed=2)]),
            "Asia/Shanghai",
        ).rows[0]

        self.assertEqual(
            decide_existing_action(original.business_values(), corrected),
            "update",
        )

    def test_invalid_date_and_oversized_integer_are_skipped(self):
        parsed = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(
                        address="错误日期",
                        visit_at="0001-01-01 10:00:00",
                    ),
                    visit_row(
                        address="超大数字",
                        changed=4_294_967_296,
                    ),
                ]
            ),
            "Asia/Shanghai",
        )

        self.assertEqual(parsed.valid_rows, 0)
        self.assertEqual(len(parsed.rows), 0)
        messages = " ".join(issue.message for issue in parsed.issues)
        self.assertIn("不能早于 1000-01-01", messages)
        self.assertIn("不能超过 4294967295", messages)

    def test_multiple_matching_worksheets_are_rejected(self):
        workbook = Workbook()
        first = workbook.active
        first.title = "表一"
        first.append(list(VISIT_HEADERS))
        first.append(visit_row())
        second = workbook.create_sheet("表二")
        second.append(list(VISIT_HEADERS))
        second.append(visit_row(address="另一个地址"))
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        with self.assertRaisesRegex(VisitWorkbookError, "多个走访明细工作表"):
            parse_visit_workbook(output.getvalue(), "Asia/Shanghai")


class VisitIdentityTests(unittest.TestCase):
    def test_identity_update_rules(self):
        self.assertEqual(
            decide_member_identity(
                member_name="张三",
                existing_identity="",
                incoming_identity=VALID_IDENTITY,
                id_owners={},
            ),
            "update",
        )
        self.assertEqual(
            decide_member_identity(
                member_name="张三",
                existing_identity=VALID_IDENTITY,
                incoming_identity=VALID_IDENTITY,
                id_owners={},
            ),
            "same",
        )
        self.assertEqual(
            decide_member_identity(
                member_name="张三",
                existing_identity="111111111111111111",
                incoming_identity=VALID_IDENTITY,
                id_owners={},
            ),
            "conflict",
        )
        self.assertEqual(
            decide_member_identity(
                member_name="张三",
                existing_identity="",
                incoming_identity=VALID_IDENTITY,
                id_owners={VALID_IDENTITY: "李四"},
            ),
            "used_by_other",
        )

    def test_mask_never_returns_full_identity(self):
        masked = mask_identity_number(VALID_IDENTITY)
        self.assertNotEqual(masked, VALID_IDENTITY)
        self.assertTrue(masked.startswith("000000"))
        self.assertTrue(masked.endswith("000X"))

    def test_community_suffix_is_removed_once(self):
        self.assertEqual(normalize_community(" 长板社区 "), "长板")
        self.assertEqual(normalize_community(" 南厍村 "), "南厍")
        self.assertEqual(normalize_community("社区"), "")
        self.assertEqual(normalize_community("村"), "")

    def test_community_aliases_are_normalized_and_deduplicated(self):
        payload = CommunityAliasesUpdate(
            aliases=[" 芦荡社区 ", "芦荡", "长板村"],
        )
        self.assertEqual(payload.aliases, ["芦荡", "长板"])


class CoverageCursor:
    def __init__(self):
        self.last_sql = ""

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def execute(self, sql, params=None):
        self.last_sql = " ".join(sql.split())

    async def fetchone(self):
        if self.last_sql.startswith("SELECT MIN(`业务日期`)"):
            return (date(2026, 7, 1), date(2026, 7, 4), 3, 3, 1, 2)
        return None

    async def fetchall(self):
        if self.last_sql.startswith("SELECT DISTINCT `业务日期`"):
            return [
                (date(2026, 7, 1),),
                (date(2026, 7, 2),),
                (date(2026, 7, 4),),
            ]
        if self.last_sql.startswith("SELECT import_type"):
            return [
                ("detail", datetime(2026, 7, 29, 2, 0)),
                ("rating", datetime(2026, 7, 29, 3, 0)),
            ]
        return []


class CoverageConnection:
    def __init__(self):
        self.cursor_instance = CoverageCursor()

    def cursor(self):
        return self.cursor_instance


class VisitCoverageTests(unittest.IsolatedAsyncioTestCase):
    async def test_coverage_includes_missing_dates(self):
        coverage = await get_visit_coverage(CoverageConnection())

        self.assertEqual(coverage["start_date"], "2026-07-01")
        self.assertEqual(coverage["end_date"], "2026-07-04")
        self.assertEqual(coverage["total_records"], 3)
        self.assertEqual(coverage["rated_records"], 1)
        self.assertEqual(coverage["unrated_records"], 2)
        self.assertEqual(coverage["missing_dates"], ["2026-07-03"])
        self.assertEqual(coverage["missing_date_count"], 1)
        self.assertTrue(coverage["last_import_at"].endswith("Z"))
        self.assertTrue(coverage["last_rating_import_at"].endswith("Z"))

    def test_upload_and_issue_routes_require_admin(self):
        protected_paths = {
            route.path
            for route in visits_router.routes
            if any(
                dependency.call is require_admin
                for dependency in route.dependant.dependencies
            )
        }
        self.assertIn("/api/visits/imports/detail", protected_paths)
        self.assertIn("/api/visits/imports/rating", protected_paths)
        self.assertIn("/api/visits/imports/{batch_id}/issues", protected_paths)
        self.assertNotIn("/api/visits/coverage", protected_paths)


class ImportFlowCursor:
    def __init__(self, connection):
        self.connection = connection
        self.rows = []

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def execute(self, sql, params=None):
        normalized = " ".join(sql.split())
        self.connection.executed.append((normalized, params))
        if normalized.startswith("SELECT MIN(`业务日期`), MAX(`业务日期`)"):
            self.rows = [(None, None)]
        elif normalized == "SELECT id, name FROM _communities":
            self.rows = [
                (index, name)
                for index, name in enumerate(
                    self.connection.communities,
                    start=1,
                )
            ]
        elif normalized.startswith("SELECT a.alias, c.name"):
            self.rows = list(self.connection.community_aliases)
        elif normalized.startswith(
            "SELECT id, name, community, id_card_number FROM _grid_members"
        ):
            self.rows = list(self.connection.members)
        elif (
            normalized.startswith("SELECT `_row_key`")
            and "FROM t_visit_details" in normalized
        ):
            self.rows = []
        else:
            self.rows = []

    async def executemany(self, sql, params):
        normalized = " ".join(sql.split())
        values = list(params)
        self.connection.executed_many.append((normalized, values))
        if normalized.startswith("INSERT INTO t_visit_details"):
            self.connection.visit_inserts.extend(values)
        elif normalized.startswith(
            "UPDATE _grid_members SET id_card_number"
        ):
            self.connection.member_updates.extend(values)
        elif normalized.startswith("INSERT INTO _visit_import_issues"):
            self.connection.issue_inserts.extend(values)

    async def fetchone(self):
        return self.rows[0] if self.rows else None

    async def fetchall(self):
        return list(self.rows)


class ImportFlowConnection:
    def __init__(self, *, communities=(), community_aliases=(), members=()):
        self.communities = list(communities)
        self.community_aliases = list(community_aliases)
        self.members = list(members)
        self.executed = []
        self.executed_many = []
        self.visit_inserts = []
        self.member_updates = []
        self.issue_inserts = []
        self.began = False
        self.committed = False
        self.rolled_back = False

    def cursor(self):
        return ImportFlowCursor(self)

    async def begin(self):
        self.began = True

    async def commit(self):
        self.committed = True

    async def rollback(self):
        self.rolled_back = True


class VisitImportFlowTests(unittest.IsolatedAsyncioTestCase):
    async def test_valid_rows_commit_when_another_row_is_invalid(self):
        parsed = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(),
                    visit_row(address="错误地址", changed=-1),
                ]
            ),
            "Asia/Shanghai",
        )
        connection = ImportFlowConnection(
            communities=["长板"],
            members=[(1, "张三", "长板", None)],
        )

        result = await import_parsed_workbook(
            connection,
            batch_id=7,
            parsed=parsed,
        )

        self.assertEqual(result["status"], "partial")
        self.assertEqual(result["inserted_rows"], 1)
        self.assertEqual(result["error_count"], 1)
        self.assertEqual(len(connection.visit_inserts), 1)
        self.assertEqual(len(connection.visit_inserts[0]), 19)
        self.assertEqual(connection.member_updates, [(VALID_IDENTITY, 1)])
        self.assertTrue(connection.began)
        self.assertTrue(connection.committed)
        self.assertFalse(connection.rolled_back)

    async def test_unmatched_references_warn_but_keep_visit(self):
        parsed = parse_visit_workbook(
            workbook_bytes([visit_row()]),
            "Asia/Shanghai",
        )
        connection = ImportFlowConnection()

        result = await import_parsed_workbook(
            connection,
            batch_id=8,
            parsed=parsed,
        )

        self.assertEqual(result["status"], "success")
        self.assertEqual(result["inserted_rows"], 1)
        self.assertEqual(result["warning_count"], 2)
        issue_codes = {values[2] for values in connection.issue_inserts}
        self.assertEqual(
            issue_codes,
            {"community_not_found", "member_not_found"},
        )

    async def test_alias_matches_canonical_community_without_member_mismatch(self):
        parsed = parse_visit_workbook(
            workbook_bytes(
                [
                    visit_row(
                        community="芦荡社区",
                        operator="陈亚平",
                    )
                ]
            ),
            "Asia/Shanghai",
        )
        connection = ImportFlowConnection(
            communities=["长板"],
            community_aliases=[("芦荡", "长板")],
            members=[(1, "陈亚平", "其他社区", None)],
        )

        result = await import_parsed_workbook(
            connection,
            batch_id=9,
            parsed=parsed,
        )

        self.assertEqual(result["status"], "success")
        self.assertEqual(result["warning_count"], 0)
        self.assertEqual(connection.visit_inserts[0][3], "长板")


if __name__ == "__main__":
    unittest.main()
