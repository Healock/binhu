import json
import os
import sys
from datetime import date, datetime
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch
import unittest

from openpyxl import load_workbook

os.environ.setdefault("MYSQL_PASSWORD", "test-password")
os.environ.setdefault("ENCRYPTION_KEY", "test-encryption-key")

from fastapi import HTTPException

from routers.work_logs import (
    DraftSave,
    _draft_payload,
    _draft_summary,
    _missing_items,
    _upgrade_legacy_draft,
    delete_draft,
    list_drafts,
    router,
    save_draft,
)
from services.work_log_data import (
    _community_grid_member_counts,
    _online_summary_snapshot,
    _rental_snapshot,
    _self_owned_snapshot,
    build_system_snapshot,
)
from services.work_log_pdf import _display, build_daily_pdf
from services.work_log_daily_detail import (
    build_daily_detail_data,
    build_daily_detail_workbook,
    normalize_targets,
)
from services.work_log_schema import (
    TEMPLATE_VERSION,
    default_manual_values,
    derive_values,
    effective_values,
    field_definitions,
    fill_community_grid_member_counts,
    get_schema,
    sanitize_values,
)


class CountCursor:
    def __init__(self, count=0):
        self.count = count

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def execute(self, sql, params=None):
        self.sql = sql
        self.params = params

    async def fetchone(self):
        return (self.count,)


class CountConnection:
    def __init__(self, count=0):
        self.count = count

    def cursor(self):
        return CountCursor(self.count)


class EmptyConnection:
    def cursor(self):
        return CountCursor(0)


class DraftCursor:
    def __init__(self, connection):
        self.connection = connection
        self.rowcount = 0
        self.result = None

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def execute(self, sql, params=None):
        normalized = " ".join(sql.split())
        self.rowcount = 0
        if normalized.startswith("SELECT"):
            self.result = self.connection.row
            return
        if normalized.startswith("UPDATE _work_log_drafts SET manual_values"):
            owner_id = params[4]
            version = params[5]
            if (
                self.connection.row
                and self.connection.row[3] == owner_id
                and self.connection.row[9] == version
            ):
                row = list(self.connection.row)
                row[7] = params[0]
                row[8] = params[1]
                row[9] += 1
                self.connection.row = tuple(row)
                self.rowcount = 1
            return
        if normalized.startswith("DELETE FROM _work_log_drafts"):
            if self.connection.row and self.connection.row[0] == params[0]:
                self.connection.row = None
                self.rowcount = 1

    async def fetchone(self):
        return self.result


class DraftConnection:
    def __init__(self, row):
        self.row = row

    def cursor(self):
        return DraftCursor(self)


class DraftListCursor:
    def __init__(self, connection):
        self.connection = connection
        self.result = None
        self.results = []

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def execute(self, sql, params=None):
        normalized = " ".join(sql.split())
        if normalized.startswith("SELECT COUNT(*)"):
            self.result = (len(self.connection.rows),)
            return
        if normalized.startswith("SELECT d.id"):
            page_size, offset = params[-2:]
            self.results = self.connection.rows[offset:offset + page_size]

    async def fetchone(self):
        return self.result

    async def fetchall(self):
        return self.results


class DraftListConnection:
    def __init__(self, rows):
        self.rows = rows

    def cursor(self):
        return DraftListCursor(self)


class LegacyCursor:
    def __init__(self, connection):
        self.connection = connection
        self.result = None
        self.rowcount = 0

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return None

    async def execute(self, sql, params=None):
        normalized = " ".join(sql.split())
        if normalized.startswith("UPDATE _work_log_drafts SET template_version"):
            row = list(self.connection.row)
            row[5] = params[0]
            row[6] = params[1]
            row[7] = params[2]
            row[8] = params[3]
            row[9] += 1
            self.connection.row = tuple(row)
            self.rowcount = 1
        elif normalized.startswith("SELECT"):
            self.result = self.connection.row

    async def fetchone(self):
        return self.result


class LegacyConnection:
    def __init__(self, row):
        self.row = row

    def cursor(self):
        return LegacyCursor(self)


def snapshot(values=None):
    return {
        "business_date": "2026-07-14",
        "issue_date": "2026-07-15",
        "month": 7,
        "filename_prefix": "0714",
        "communities": ["长板", "冬梅"],
        "values": {
            "meta.year": 2026,
            "meta.month": 7,
            "meta.day": 14,
            **(values or {}),
        },
        "sources": {
            "online_summary": {
                "label": "在线数据总汇总表",
                "available": False,
                "message": "该日期没有可用数据",
            },
            "rental_visit": {
                "label": "出租房走访",
                "available": False,
                "message": "该日期没有可用数据",
            },
            "self_owned_visit": {
                "label": "自购房走访",
                "available": False,
                "message": "该日期没有可用数据",
            },
        },
    }


def draft_payload(values=None):
    return {
        "id": 1,
        "report_type": "daily",
        "business_date": "2026-07-14",
        "owner": {"id": 7, "username": "admin"},
        "can_edit": True,
        "template_version": TEMPLATE_VERSION,
        "system_snapshot": snapshot(values),
        "manual_values": default_manual_values(["长板", "冬梅"]),
        "override_values": {},
        "version": 1,
        "last_export_at": None,
        "created_at": "2026-07-14T00:00:00",
        "updated_at": "2026-07-14T00:00:00",
    }


class WorkLogTests(unittest.IsolatedAsyncioTestCase):
    def sample_row(self):
        draft = draft_payload()
        return (
            4,
            "daily",
            date(2026, 7, 14),
            7,
            "admin",
            TEMPLATE_VERSION,
            json.dumps(draft["system_snapshot"], ensure_ascii=False),
            json.dumps(draft["manual_values"], ensure_ascii=False),
            "{}",
            1,
            None,
            datetime(2026, 7, 14),
            datetime(2026, 7, 14),
        )

    def test_schema_keeps_document_order_and_all_sixteen_tables(self):
        schema = get_schema()
        self.assertEqual(schema["template_version"], "daily-v2")
        self.assertEqual(
            [section["id"] for section in schema["sections"]],
            [
                "flow",
                "rental",
                "self_owned",
                "priority",
                "disputes",
                "fire",
                "security",
                "fraud",
                "notices",
                "special",
            ],
        )
        tables = [
            block["field"]["id"]
            for section in schema["sections"]
            for block in section["blocks"]
            if block["type"] == "table"
        ]
        self.assertEqual(len(tables), 16)
        self.assertIn("flow.instruction_table", tables)
        self.assertIn("rental.visit_table", tables)
        self.assertIn("special.monitor_table", tables)

    def test_requested_work_log_sentences_and_sources(self):
        schema = get_schema()
        blocks = {
            section["id"]: section["blocks"]
            for section in schema["sections"]
        }
        dingning = blocks["special"][1]
        dingning_text = "".join(
            segment
            for segment in dingning["segments"]
            if isinstance(segment, str)
        )
        self.assertIn("第三批次任务总数", dingning_text)
        self.assertIn("已核查未见面", dingning_text)
        self.assertIn("人待核查见面，已核查", dingning_text)
        self.assertIn("综合指挥室+辅警办公室宣传完成率为", dingning_text)
        rates = [
            segment
            for segment in dingning["segments"]
            if isinstance(segment, dict)
            and segment["id"].endswith("completion_rate")
        ]
        self.assertEqual([item["precision"] for item in rates], [2, 2])

        dispute_sentences = [
            block
            for block in blocks["disputes"]
            if block["type"] == "sentence"
        ]
        dispute_text = "".join(
            segment
            for block in dispute_sentences
            for segment in block["segments"]
            if isinstance(segment, str)
        )
        self.assertIn("辖区存量未决矛盾纠纷档案", dispute_text)
        self.assertIn("新下发矛盾纠纷", dispute_text)
        self.assertIn("化解矛盾纠纷", dispute_text)
        self.assertIn("未决档案化解", dispute_text)

        self_owned_table = next(
            block["field"]
            for block in blocks["self_owned"]
            if block["type"] == "table"
            and block["field"]["id"] == "self_owned.visit_table"
        )
        self.assertEqual(self_owned_table["source"], "system")
        self.assertEqual(
            self_owned_table["source_key"],
            "self_owned_visit",
        )
        self.assertEqual(
            _display(2.2, {"type": "percent", "precision": 2}),
            "2.20%",
        )

    def test_manual_tables_prefill_communities_and_fixed_categories(self):
        values = default_manual_values(
            ["冬梅", "长板"],
            {"冬梅": "张三、李四"},
            {"冬梅": 6, "长板": 4},
        )
        self.assertEqual(
            [row["responsibility_area"] for row in values["fire.table"]],
            ["冬梅", "长板"],
        )
        self.assertEqual(
            values["fire.table"][0]["community_officer"],
            "张三、李四",
        )
        self.assertEqual(
            values["fire.table"][1]["community_officer"],
            "",
        )
        registration_rows = values["flow.registration_table"]
        self.assertEqual(registration_rows[0]["grid_member_count"], 6)
        self.assertEqual(registration_rows[1]["grid_member_count"], 4)
        self.assertEqual(
            [row["venue_type"] for row in values["security.venues_table"]],
            ["足浴", "浴室", "酒吧/KTV", "宾馆", "网约房/民宿", "其他单位"],
        )

    def test_grid_member_count_refresh_only_fills_blank_cells(self):
        values = {
            "flow.registration_table": [
                {"responsibility_area": "冬梅", "grid_member_count": ""},
                {"responsibility_area": "长板", "grid_member_count": 9},
            ],
        }
        refreshed = fill_community_grid_member_counts(
            values,
            {"冬梅": 6, "长板": 4},
        )
        self.assertEqual(
            refreshed["flow.registration_table"][0]["grid_member_count"],
            6,
        )
        self.assertEqual(
            refreshed["flow.registration_table"][1]["grid_member_count"],
            9,
        )

    async def test_grid_member_counts_use_positions_attendance_and_community(self):
        context = {
            "members": {
                "甲": {"id": 1, "community": "冬梅", "position": "组员"},
                "乙": {"id": 2, "community": "冬梅", "position": "组长"},
                "丙": {"id": 3, "community": "长板", "position": "组员"},
            },
            "missing_week_starts": set(),
            "legacy_history_incomplete": False,
        }
        with (
            patch(
                "services.work_log_data.get_attendance_context",
                new=AsyncMock(return_value=context),
            ),
            patch(
                "services.work_log_data.is_member_on_duty",
                side_effect=lambda member, *_: member["id"] != 2,
            ),
        ):
            result = await _community_grid_member_counts(
                CountConnection(),
                date(2026, 7, 31),
                ["冬梅", "长板"],
            )
        self.assertTrue(result["available"])
        self.assertEqual(result["counts"], {"冬梅": 1, "长板": 1})

    async def test_grid_member_counts_stay_blank_when_weekend_duty_is_missing(self):
        with (
            patch(
                "services.work_log_data.get_attendance_context",
                new=AsyncMock(return_value={
                    "members": {},
                    "missing_week_starts": {date(2026, 7, 27)},
                    "legacy_history_incomplete": False,
                }),
            ),
        ):
            result = await _community_grid_member_counts(
                CountConnection(),
                date(2026, 8, 1),
                ["冬梅"],
            )
        self.assertFalse(result["available"])
        self.assertEqual(result["counts"], {})

    async def test_online_summary_uses_total_report_and_fixed_two_columns(self):
        report = {
            "exists": True,
            "community": {
                "columns": [
                    "社区",
                    "数据总数",
                    "未核查",
                    "已核查",
                    "已完成",
                    "核查完成率",
                    "无法见底数",
                    "核查见底率",
                    "网格员人数",
                    "当日人均核查数",
                ],
                "data": [{
                    "社区": "长板",
                    "数据总数": 10,
                    "未核查": 2,
                    "已核查": 3,
                    "已完成": 5,
                    "核查完成率": 0.5,
                    "无法见底数": 1,
                    "核查见底率": 0.8,
                    "网格员人数": 2,
                    "当日人均核查数": 2.5,
                }],
            },
        }
        with patch(
            "services.work_log_data.get_summary",
            new=AsyncMock(return_value=report),
        ):
            result = await _online_summary_snapshot(
                date(2026, 7, 14),
                {"长板": "张三、李四"},
                {"长板": 5},
            )
        row = result["values"]["flow.instruction_table"][0]
        self.assertTrue(result["available"])
        self.assertEqual(row["community_officer"], "张三、李四")
        self.assertEqual(row["grid_member_count"], 5)
        self.assertEqual(row["unchecked"], 5)
        self.assertEqual(row["checked"], 5)
        self.assertEqual(row["completion_rate"], 50.0)

    async def test_rental_snapshot_maps_community_rows(self):
        visit_result = {
            "attendance": {"complete": True},
            "community": {
                "data": [{
                    "社区": "长板",
                    "在岗人日": 2,
                    "走访户数": 12,
                    "人均日走访户数": 6,
                    "新增": 1,
                    "变更": 4,
                    "注销": 2,
                    "人均日变动数": 3.5,
                    "户均变动数": 0.6,
                    "星级评定数": 9,
                    "星级评定率": 0.75,
                }],
            },
        }
        with patch(
            "services.work_log_data.get_visit_summary",
            new=AsyncMock(return_value=visit_result),
        ):
            result = await _rental_snapshot(
                CountConnection(count=1),
                date(2026, 7, 14),
                {"长板": "张三、李四"},
                {"长板": 5},
            )
        row = result["values"]["rental.visit_table"][0]
        self.assertEqual(row["community_officer"], "张三、李四")
        self.assertEqual(row["grid_member_count"], 5)
        self.assertEqual(row["total_changes"], 7)
        self.assertEqual(row["rating_rate"], 75.0)

    async def test_self_owned_snapshot_maps_inspector_rows(self):
        visit_result = {
            "inspector": {
                "data": [{
                    "社区": "社区甲",
                    "姓名": "网格员甲",
                    "走访户数": 11,
                    "新增": 2,
                    "变更": 5,
                    "注销": 1,
                }],
            },
        }
        with patch(
            "services.work_log_data.get_visit_summary",
            new=AsyncMock(return_value=visit_result),
        ) as summary:
            result = await _self_owned_snapshot(
                CountConnection(),
                date(2026, 7, 14),
            )
        summary.assert_awaited_once()
        row = result["values"]["self_owned.visit_table"][0]
        self.assertEqual(
            row,
            {
                "grid_member": "网格员甲",
                "visits": 11,
                "changed": 5,
                "cancelled": 1,
            },
        )

    async def test_snapshot_dates_and_missing_sources_are_explicit(self):
        with (
            patch(
                "services.work_log_data._community_names",
                new=AsyncMock(return_value=["长板"]),
            ),
            patch(
                "services.work_log_data._community_officers",
                new=AsyncMock(return_value={"长板": "张三、李四"}),
            ),
            patch(
                "services.work_log_data._community_grid_member_counts",
                new=AsyncMock(return_value={
                    "available": True,
                    "message": "",
                    "counts": {"长板": 2},
                }),
            ),
            patch(
                "services.work_log_data._online_summary_snapshot",
                new=AsyncMock(return_value={
                    "available": False,
                    "message": "无日报",
                    "values": {},
                }),
            ),
            patch(
                "services.work_log_data._rental_snapshot",
                new=AsyncMock(return_value={
                    "available": False,
                    "message": "无走访",
                    "values": {},
                }),
            ),
            patch(
                "services.work_log_data._self_owned_snapshot",
                new=AsyncMock(return_value={
                    "available": False,
                    "message": "无自购房走访",
                    "values": {},
                }),
            ),
        ):
            result = await build_system_snapshot(
                CountConnection(),
                date(2026, 7, 14),
            )
        self.assertEqual(result["business_date"], "2026-07-14")
        self.assertEqual(result["issue_date"], "2026-07-15")
        self.assertEqual(result["communities"], ["长板"])
        self.assertEqual(
            result["community_officers"],
            {"长板": "张三、李四"},
        )
        self.assertEqual(
            result["community_grid_member_counts"],
            {"长板": 2},
        )
        self.assertEqual(result["values"]["meta.year"], 2026)
        self.assertFalse(result["sources"]["online_summary"]["available"])

    def test_edited_tables_recalculate_overview_until_overridden(self):
        draft = draft_payload({
            "flow.instruction_table": [{
                "grid_member_count": 2,
                "total": 10,
                "unchecked": 4,
                "checked": 6,
                "unable": 1,
            }],
            "rental.visit_table": [{
                "grid_member_count": 2,
                "visits": 8,
                "added": 1,
                "changed": 3,
                "cancelled": 0,
                "rated": 4,
            }],
        })
        values = effective_values(draft)
        self.assertEqual(values["flow.instruction.completion_rate"], 60.0)
        self.assertEqual(values["rental.visit.average_visits"], 4.0)
        draft["override_values"]["rental.visit.average_visits"] = 9
        self.assertEqual(
            effective_values(draft)["rental.visit.average_visits"],
            9,
        )

    def test_schema_sanitizes_tables_and_wrong_sources(self):
        manual = sanitize_values(
            {
                "flow.population.total": 100,
                "flow.instruction_table": [{"total": 9}],
                "unknown": "secret",
            },
            source="manual",
        )
        overrides = sanitize_values(
            {
                "flow.instruction_table": [{
                    "responsibility_area": "长板",
                    "total": "9",
                    "unknown": "secret",
                }],
                "flow.population.total": 100,
            },
            source="system",
        )
        self.assertEqual(manual, {"flow.population.total": 100.0})
        self.assertEqual(
            overrides["flow.instruction_table"][0]["total"],
            9.0,
        )
        self.assertNotIn(
            "unknown",
            overrides["flow.instruction_table"][0],
        )

    def test_draft_row_reports_creator_editing_rights(self):
        row = list(self.sample_row())
        row[9] = 3
        own = _draft_payload(tuple(row), {"id": 7})
        other = _draft_payload(tuple(row), {"id": 8})
        self.assertTrue(own["can_edit"])
        self.assertFalse(other["can_edit"])

    def test_draft_summary_includes_creator_and_current_editor(self):
        row = (
            4,
            "daily",
            date(2026, 7, 14),
            7,
            "current-editor",
            TEMPLATE_VERSION,
            3,
            None,
            datetime(2026, 7, 14),
            datetime(2026, 7, 15),
            8,
            "creator",
        )
        summary = _draft_summary(row)
        self.assertEqual(summary["business_date"], "2026-07-14")
        self.assertEqual(summary["owner"]["username"], "current-editor")
        self.assertEqual(summary["creator"]["username"], "creator")
        self.assertEqual(summary["version"], 3)

    async def test_admin_can_delete_draft_without_touching_source_data(self):
        connection = DraftConnection(self.sample_row())
        request = SimpleNamespace(headers={}, client=None)
        audit = AsyncMock()
        with patch("routers.work_logs.record_admin_audit", new=audit):
            result = await delete_draft(
                4,
                request,
                user={"id": 8, "username": "another-admin"},
                conn=connection,
            )
        self.assertEqual(result, {"message": "草稿已删除", "id": 4})
        self.assertIsNone(connection.row)
        audit.assert_awaited_once()
        self.assertEqual(audit.await_args.args[1], "work_log.delete")
        self.assertEqual(
            audit.await_args.kwargs["detail"]["business_date"],
            "2026-07-14",
        )

    async def test_draft_list_returns_paginated_summaries(self):
        rows = [(
            4,
            "daily",
            date(2026, 7, 14),
            7,
            "current-editor",
            TEMPLATE_VERSION,
            3,
            None,
            datetime(2026, 7, 14),
            datetime(2026, 7, 15),
            8,
            "creator",
        )]
        result = await list_drafts(
            start_date=None,
            end_date=None,
            keyword=None,
            page=1,
            page_size=20,
            user={"id": 7},
            conn=DraftListConnection(rows),
        )
        self.assertEqual(result["total"], 1)
        self.assertEqual(result["data"][0]["business_date"], "2026-07-14")
        self.assertEqual(result["data"][0]["creator"]["username"], "creator")

    async def test_legacy_draft_is_backed_up_before_v2_mapping(self):
        row = list(self.sample_row())
        row[5] = "daily-v1"
        row[6] = json.dumps({"values": {"old.system": 1}}, ensure_ascii=False)
        row[7] = json.dumps(
            {"basic.total_population": 321},
            ensure_ascii=False,
        )
        connection = LegacyConnection(tuple(row))
        current_snapshot = snapshot()
        with patch(
            "routers.work_logs.build_system_snapshot",
            new=AsyncMock(return_value=current_snapshot),
        ):
            upgraded = await _upgrade_legacy_draft(
                connection,
                connection.row,
                {"id": 7},
            )
        self.assertEqual(upgraded[5], TEMPLATE_VERSION)
        new_snapshot = json.loads(upgraded[6])
        self.assertEqual(
            new_snapshot["legacy_v1"]["system_snapshot"]["values"]["old.system"],
            1,
        )
        self.assertEqual(
            json.loads(upgraded[7])["flow.population.total"],
            321.0,
        )

    async def test_autosave_increments_version_and_rejects_stale_version(self):
        connection = DraftConnection(self.sample_row())
        saved = await save_draft(
            4,
            DraftSave(
                version=1,
                manual_values={"flow.population.total": 120},
                override_values={},
            ),
            user={"id": 7},
            conn=connection,
        )
        self.assertEqual(saved["version"], 2)
        self.assertEqual(
            saved["manual_values"]["flow.population.total"],
            120.0,
        )
        with self.assertRaises(HTTPException) as raised:
            await save_draft(
                4,
                DraftSave(
                    version=1,
                    manual_values={"flow.population.total": 121},
                    override_values={},
                ),
                user={"id": 7},
                conn=connection,
            )
        self.assertEqual(raised.exception.status_code, 409)

    def test_missing_items_explain_unavailable_system_data(self):
        missing = _missing_items(draft_payload())
        instruction = next(
            item for item in missing
            if item["field_id"] == "flow.instruction_table"
        )
        self.assertEqual(instruction["reason"], "该日期没有可用数据")

    def test_every_work_log_route_requires_work_log_permission(self):
        for route in router.routes:
            dependency_names = {
                dependency.call.__name__
                for dependency in route.dependant.dependencies
                if dependency.call
            }
            self.assertIn("require_worklog_manage", dependency_names, route.path)

    def test_pdf_export_uses_pdf_filename_and_clean_html(self):
        captured = {}

        class FakeHTML:
            def __init__(self, *, string):
                captured["html"] = string

            def write_pdf(self):
                return b"%PDF-1.7 fake"

        fake_module = SimpleNamespace(HTML=FakeHTML)
        draft = draft_payload()
        draft["manual_values"]["flow.population.total"] = 123
        draft["system_snapshot"]["legacy_v1"] = {
            "manual_values": {"old.private": "PRIVATE_MARKER_SHOULD_NOT_RENDER"}
        }
        with patch.dict(sys.modules, {"weasyprint": fake_module}):
            content, filename = build_daily_pdf(
                draft,
                get_schema(),
                effective_values(draft),
            )
        self.assertTrue(content.startswith(b"%PDF"))
        self.assertEqual(
            filename,
            "0714日报滨湖新城派出所社区警务工作日志.pdf",
        )
        self.assertIn("流动人口采集", captured["html"])
        self.assertIn("7", captured["html"])
        self.assertIn("14", captured["html"])
        self.assertIn("123", captured["html"])
        self.assertIn('class="document-title"', captured["html"])
        self.assertIn('<div class="document-number">一</div>', captured["html"])
        self.assertIn("color: #ff0000", captured["html"])
        self.assertIn("margin: 37mm 27mm 35mm", captured["html"])
        self.assertIn("<strong>问题分析：</strong>", captured["html"])
        self.assertIn('class="blank blank-filled"', captured["html"])
        self.assertLess(
            captured["html"].index("基础数据"),
            captured["html"].index("一、流动人口采集"),
        )
        self.assertNotIn("background: #eef2f6", captured["html"])
        self.assertNotIn("{{", captured["html"])
        self.assertNotIn("PRIVATE_MARKER_SHOULD_NOT_RENDER", captured["html"])

    def test_daily_detail_workbook_keeps_layout_and_clears_other_sections(self):
        rows = [
            {
                "group_key": "area:东片",
                "group_label": "东片\n片长甲",
                "community": "长板",
                "name": "网格员甲",
                "attendance": "是",
                "visit_target": 10,
                "visits": 8,
                "ratings": 6,
                "checked_instructions": 5,
                "completed_instructions": 3,
            },
            {
                "group_key": "area:东片",
                "group_label": "东片\n片长甲",
                "community": "长板",
                "name": "网格员乙",
                "attendance": "否",
                "visit_target": 10,
                "visits": 0,
                "ratings": 0,
                "checked_instructions": 2,
                "completed_instructions": 1,
            },
        ]
        content, filename = build_daily_detail_workbook(
            date(2026, 8, 9),
            rows,
        )
        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Sheet1"])
        sheet = workbook["Sheet1"]
        self.assertEqual(filename, "0809滨湖网格工作每日明细.xlsx")
        self.assertEqual(sheet["B2"].value, "8月9号")
        self.assertEqual(sheet["A4"].value, 1)
        self.assertEqual(sheet["J4"].value, 8)
        self.assertEqual(sheet["L4"].value, 5)
        self.assertEqual(sheet["M4"].value, 3)
        self.assertIsNone(sheet["F4"].value)
        self.assertIsNone(sheet["N4"].value)
        self.assertIsNone(sheet["V4"].value)
        self.assertIn("B4:B5", {str(item) for item in sheet.merged_cells.ranges})
        self.assertIn("C4:C5", {str(item) for item in sheet.merged_cells.ranges})
        self.assertEqual(sheet.max_row, 67)
        for row in sheet.iter_rows(min_row=1, max_row=67, min_col=1, max_col=22):
            for cell in row:
                self.assertIsNone(cell.fill.fill_type)

    async def test_daily_detail_uses_role_specific_visits_and_snapshot_tasks(self):
        roster = [
            {
                "name": "网格员甲",
                "position": "组员",
                "group_key": "area:东片",
                "group_label": "东片",
                "community": "长板",
                "attendance": "是",
            },
            {
                "name": "自购房乙",
                "position": "自购房",
                "group_key": "internal:自购房",
                "group_label": "自购房",
                "community": "",
                "attendance": "是",
            },
        ]
        rental = {
            "inspector": {"data": [{
                "姓名": "网格员甲",
                "走访户数": 4,
                "星级评定数": 3,
            }]},
        }
        self_owned = {
            "inspector": {"data": [{
                "姓名": "自购房乙",
                "走访户数": 6,
                "星级评定数": 5,
            }]},
        }
        online = {
            "exists": True,
            "inspector": {"data": [{
                "姓名": "网格员甲",
                "已核查": 2,
                "已完成": 3,
            }]},
        }
        with (
            patch(
                "services.work_log_daily_detail._load_roster",
                new=AsyncMock(return_value=roster),
            ),
            patch(
                "services.work_log_daily_detail.get_known_personnel_positions",
                new=AsyncMock(return_value={}),
            ),
            patch(
                "services.work_log_daily_detail.get_configured_positions",
                new=AsyncMock(return_value=["组员"]),
            ),
            patch(
                "services.work_log_daily_detail.get_visit_summary",
                new=AsyncMock(side_effect=[rental, self_owned]),
            ),
            patch(
                "services.work_log_daily_detail.get_summary",
                new=AsyncMock(return_value=online),
            ),
        ):
            result = await build_daily_detail_data(
                EmptyConnection(),
                date(2026, 8, 9),
                rental_target=10,
                self_owned_target=15,
            )
        self.assertEqual(result[0]["visit_target"], 10)
        self.assertEqual(result[0]["visits"], 4)
        self.assertEqual(result[0]["ratings"], 3)
        self.assertEqual(result[0]["checked_instructions"], 5)
        self.assertEqual(result[0]["completed_instructions"], 3)
        self.assertEqual(result[1]["visit_target"], 15)
        self.assertEqual(result[1]["visits"], 6)
        self.assertIsNone(result[1]["checked_instructions"])

    def test_daily_detail_targets_are_bounded(self):
        self.assertEqual(
            normalize_targets({
                "rental_target": -1,
                "self_owned_target": 1200,
            }),
            {"rental_target": 0, "self_owned_target": 999},
        )


if __name__ == "__main__":
    unittest.main()
