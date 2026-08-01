"""从私密 XLSX 名单预览或创建账号。

默认只预览。正式执行必须额外提供 ``--apply``，新账号密码只从指定
环境变量读取，脚本不会输出密码，也不会把名单文件复制进仓库。
"""

from __future__ import annotations

import argparse
import asyncio
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_HEADERS = {"用户名", "姓名", "职位"}

UNLINKED_POSITION_GROUPS = {
    "社区民警": "admin",
    "所（队）领导": "admin",
    "所队领导": "admin",
    "内勤岗": "internal_business",
}
PLACEHOLDER_POSITIONS = {
    "流口岗": "组员",
}
DEFAULT_INITIAL_PASSWORD_LENGTH = 8
CONFIRMED_SHORT_PASSWORD_LENGTH = 5


def _legacy_role(group_code: str) -> str:
    if group_code == "super_admin":
        return "super_admin"
    if group_code in {"admin", "internal_business"}:
        return "admin"
    if group_code == "global_viewer":
        return "leader"
    return "member"


def validate_initial_password(
    password: str,
    *,
    allow_short_password: bool = False,
) -> None:
    minimum = (
        CONFIRMED_SHORT_PASSWORD_LENGTH
        if allow_short_password
        else DEFAULT_INITIAL_PASSWORD_LENGTH
    )
    if len(password) < minimum:
        qualifier = "（已启用短密码确认开关）" if allow_short_password else ""
        raise ValueError(f"统一初始密码至少需要 {minimum} 个字符{qualifier}")


def read_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(values)
        except StopIteration as exc:
            raise ValueError("名单为空") from exc
        headers = [str(value or "").strip() for value in raw_headers]
        missing = REQUIRED_HEADERS - set(headers)
        if missing:
            raise ValueError(f"名单缺少列：{'、'.join(sorted(missing))}")
        username_index = headers.index("用户名")
        name_index = headers.index("姓名")
        position_index = headers.index("职位")
        rows: list[dict[str, str]] = []
        for source_row, values_row in enumerate(values, start=2):
            username = str(values_row[username_index] or "").strip()
            name = str(values_row[name_index] or "").strip()
            position = str(values_row[position_index] or "").strip()
            if not username and not name and not position:
                continue
            if not username or not name:
                raise ValueError(f"第 {source_row} 行用户名或姓名为空")
            rows.append({
                "username": username,
                "name": name,
                "position": position,
            })
        if not rows:
            raise ValueError("名单没有有效数据")
        return rows
    finally:
        workbook.close()


def reject_duplicates(rows: list[dict[str, str]]) -> None:
    for key, label in (("username", "用户名"), ("name", "姓名")):
        seen: set[str] = set()
        duplicates: set[str] = set()
        for row in rows:
            value = row[key]
            if value in seen:
                duplicates.add(value)
            seen.add(value)
        if duplicates:
            raise ValueError(f"{label}重复：{'、'.join(sorted(duplicates))}")


async def build_preview(cur, rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    await cur.execute(
        """
        SELECT member.id, member.name, member.position,
               department.name, mapping.permission_group_id,
               permission_group.name, department.department_type,
               permission_group.code
        FROM _grid_members AS member
        LEFT JOIN _departments AS department
          ON department.id=member.department_id
        LEFT JOIN _position_permission_groups AS mapping
          ON mapping.position=member.position
        LEFT JOIN _permission_groups AS permission_group
          ON permission_group.id=mapping.permission_group_id
        """
    )
    members_by_name: dict[str, list[tuple[Any, ...]]] = {}
    for member in await cur.fetchall():
        members_by_name.setdefault(str(member[1]), []).append(member)

    await cur.execute(
        """
        SELECT user.id, user.username, user.member_id, user.role,
               user.permission_group_id, user.group_assignment_mode,
               permission_group.name
        FROM _users AS user
        LEFT JOIN _permission_groups AS permission_group
          ON permission_group.id=user.permission_group_id
        """
    )
    users = {str(row[1]): row for row in await cur.fetchall()}
    linked_members = {
        int(row[2]): str(row[1])
        for row in users.values()
        if row[2] is not None
    }

    await cur.execute(
        "SELECT id, code, name FROM _permission_groups"
    )
    permission_groups = {
        str(row[1]): {
            "id": int(row[0]),
            "code": str(row[1]),
            "name": str(row[2]),
        }
        for row in await cur.fetchall()
    }
    permission_groups_by_id = {
        group["id"]: group for group in permission_groups.values()
    }
    await cur.execute(
        "SELECT position, permission_group_id "
        "FROM _position_permission_groups"
    )
    position_groups = {
        str(row[0]): int(row[1]) for row in await cur.fetchall()
    }

    preview: list[dict[str, Any]] = []
    errors: list[str] = []
    for item in rows:
        existing = users.get(item["username"])
        if existing:
            preview.append({
                "action": "skip_existing",
                "username": item["username"],
                "member_id": existing[2],
                "name": item["name"],
                "position": item["position"],
                "department": "保持原账号不变",
                "permission_group_id": existing[4],
                "permission_group_code": None,
                "permission_group": str(existing[6] or "沿用原权限"),
                "assignment_mode": str(existing[5] or "inherited"),
                "legacy_role": str(existing[3]),
                "create_member": False,
            })
            continue

        matches = members_by_name.get(item["name"], [])
        if len(matches) > 1:
            errors.append(
                f"{item['name']}：人员姓名不唯一"
            )
            continue
        if matches:
            member = matches[0]
            if member[4] is None or member[7] is None:
                errors.append(
                    f"{item['name']}：岗位“{member[2]}”没有默认权限组"
                )
                continue
            if member[2] in {"组长", "组员"} and member[6] != "community":
                errors.append(f"{item['name']}：组长或组员尚未选择社区部门")
                continue
            if (
                member[2] in {"片长", "中队长", "基础管控"}
                and member[6] != "internal"
            ):
                errors.append(f"{item['name']}：内勤岗位的所属部门不正确")
                continue
            owner = linked_members.get(int(member[0]))
            if owner:
                errors.append(f"{item['name']}：已经关联账号 {owner}")
                continue
            group_code = str(member[7])
            preview.append({
                "action": "create",
                "username": item["username"],
                "member_id": int(member[0]),
                "name": item["name"],
                "position": str(member[2]),
                "department": str(member[3] or "未分配部门"),
                "permission_group_id": int(member[4]),
                "permission_group_code": group_code,
                "permission_group": str(member[5]),
                "assignment_mode": "inherited",
                "legacy_role": _legacy_role(group_code),
                "create_member": False,
            })
            continue

        special_group_code = UNLINKED_POSITION_GROUPS.get(item["position"])
        placeholder_position = PLACEHOLDER_POSITIONS.get(item["position"])
        if special_group_code:
            group = permission_groups.get(special_group_code)
            if not group:
                errors.append(
                    f"{item['name']}：权限组“{special_group_code}”不存在"
                )
                continue
            preview.append({
                "action": "create_unlinked",
                "username": item["username"],
                "member_id": None,
                "name": item["name"],
                "position": item["position"],
                "department": "不关联人员资料",
                "permission_group_id": group["id"],
                "permission_group_code": group["code"],
                "permission_group": group["name"],
                "assignment_mode": "custom",
                "legacy_role": _legacy_role(group["code"]),
                "create_member": False,
            })
            continue
        if placeholder_position:
            group_id = position_groups.get(placeholder_position)
            group = permission_groups_by_id.get(group_id) if group_id else None
            if not group:
                errors.append(
                    f"{item['name']}：岗位“{placeholder_position}”没有默认权限组"
                )
                continue
            preview.append({
                "action": "create_placeholder",
                "username": item["username"],
                "member_id": None,
                "name": item["name"],
                "position": placeholder_position,
                "department": "待分配部门",
                "permission_group_id": group["id"],
                "permission_group_code": group["code"],
                "permission_group": group["name"],
                "assignment_mode": "inherited",
                "legacy_role": _legacy_role(group["code"]),
                "create_member": True,
            })
            continue
        errors.append(
            f"{item['name']}：找不到人员，且职位“{item['position']}”没有导入规则"
        )
    if errors:
        raise ValueError("名单校验失败：\n- " + "\n- ".join(errors))
    return preview


async def apply_preview(
    cur,
    preview: list[dict[str, Any]],
    password_hash: str,
) -> None:
    for item in preview:
        if item["action"] == "skip_existing":
            continue
        member_id = item["member_id"]
        if item["create_member"]:
            await cur.execute(
                "INSERT INTO _grid_members "
                "(name, community, department_id, position) "
                "VALUES (%s, '', NULL, %s)",
                (item["name"], item["position"]),
            )
            member_id = int(cur.lastrowid)
        await cur.execute(
            """
            INSERT INTO _users (
                username, password_hash, role, member_id,
                permission_group_id, group_assignment_mode,
                password_is_temporary
            ) VALUES (%s, %s, %s, %s, %s, %s, 1)
            """,
            (
                item["username"], password_hash, item["legacy_role"],
                member_id, item["permission_group_id"],
                item["assignment_mode"],
            ),
        )


async def run(args) -> None:
    from database import close_db, db_manager, init_db

    rows = read_rows(Path(args.file).resolve())
    reject_duplicates(rows)
    await init_db()
    pool = db_manager.get_pool("online_data")
    conn = await pool.acquire()
    try:
        async with conn.cursor() as cur:
            preview = await build_preview(cur, rows)
            print("动作\t用户名\t姓名\t所属部门\t岗位\t权限组")
            for item in preview:
                print(
                    f"{item['action']}\t{item['username']}\t{item['name']}\t"
                    f"{item['department']}\t{item['position']}\t"
                    f"{item['permission_group']}"
                )
            print(
                "预览完成：新增 "
                f"{sum(item['action'].startswith('create') for item in preview)}，"
                "其中待分配部门 "
                f"{sum(item['action'] == 'create_placeholder' for item in preview)}，"
                "跳过已有账号 "
                f"{sum(item['action'] == 'skip_existing' for item in preview)}"
            )
            if not args.apply:
                print("当前为预览模式，数据库未修改。")
                return

            import bcrypt

            password = os.environ.get(args.password_env, "")
            if not password:
                raise ValueError(
                    f"正式执行前请通过环境变量 {args.password_env} "
                    "提供统一初始密码"
                )
            validate_initial_password(
                password,
                allow_short_password=args.allow_short_password,
            )
            password_hash = bcrypt.hashpw(
                password.encode(), bcrypt.gensalt()
            ).decode()
            await conn.begin()
            try:
                await apply_preview(cur, preview, password_hash)
                if args.publish_announcement:
                    await cur.execute(
                        """
                        INSERT INTO _announcements (
                            severity, title, content, published_at
                        ) VALUES ('warning', %s, %s, UTC_TIMESTAMP())
                        """,
                        (
                            "新账号初始密码通知",
                            "新账号已创建，统一初始密码为："
                            + password
                            + "。请登录后尽快在账号菜单中修改密码。",
                        ),
                    )
                if args.enable_permissions:
                    await cur.execute(
                        """
                        INSERT INTO _system_config (config_key, config_value)
                        VALUES ('permission_enforcement_enabled', '1')
                        ON DUPLICATE KEY UPDATE config_value='1'
                        """
                    )
                await conn.commit()
            except Exception:
                await conn.rollback()
                raise
            print("正式执行完成。初始密码未输出到终端或日志。")
    finally:
        pool.release(conn)
        await close_db()


def main() -> None:
    parser = argparse.ArgumentParser(description="预览或导入用户名单")
    parser.add_argument("--file", required=True, help="私密 XLSX 名单路径")
    parser.add_argument("--apply", action="store_true", help="正式写入数据库")
    parser.add_argument(
        "--password-env",
        default="BINHU_INITIAL_PASSWORD",
        help="保存统一初始密码的环境变量名",
    )
    parser.add_argument("--publish-announcement", action="store_true")
    parser.add_argument("--enable-permissions", action="store_true")
    parser.add_argument(
        "--allow-short-password",
        action="store_true",
        help="仅在项目管理人明确确认后，允许 5 至 7 位统一初始密码",
    )
    args = parser.parse_args()
    if (args.publish_announcement or args.enable_permissions) and not args.apply:
        parser.error("发布公告或启用权限前必须同时提供 --apply")
    asyncio.run(run(args))


if __name__ == "__main__":
    main()
